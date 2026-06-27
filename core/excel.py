# -*- coding: utf-8 -*-
"""Parsers de archivos Excel — boletines, listados, registros."""

import sqlite3
import logging
import re

from .constants import DATABASE
from .helpers import limpiar_v

logger = logging.getLogger("axula")

__all__ = [
    "_buscar_o_crear_estudiante",
    "_detectar_mencion_listado",
    "_limpiar_nota",
    "_parsear_boletin_bj",
]


def _detectar_mencion_listado(ws):
    """
    FIX 1 — Lee el encabezado de sección del LISTADO buscando la fila que
    contiene 'DATOS DEL ALUMNO'. La mención está en la columna 8 de esa fila:
    MULTIMEDIA, MÚSICA, TEATRO, ARTES VISUALES, etc.
    Se usa al inicio de cada hoja para determinar la mención antes de
    iterar estudiantes — más robusto que leerla inline fila a fila.
    Retorna la mención en MAYÚSCULAS o '' si no se encuentra.
    """
    for r in range(1, 12):
        v1 = str(ws.cell(r, 1).value or '').strip().upper()
        if 'DATOS DEL ALUMNO' in v1:
            mencion_raw = str(ws.cell(r, 8).value or '').strip().upper()
            return mencion_raw
    # Fallback: escanear columna 8 en las primeras 15 filas
    PALABRAS_MENCION = ['MULTIMEDIA', 'MUSICA', 'MÚSICA', 'TEATRO',
                        'VISUAL', 'DANZA', 'ARTES']
    for r in range(1, 15):
        v8 = str(ws.cell(r, 8).value or '').strip().upper()
        if any(p in v8 for p in PALABRAS_MENCION):
            return v8
    return ''


def _limpiar_nota(v):
    """
    Convierte celda de Excel a float entre 1-100, o 0 si no es válida.
    Acepta: int, float, strings '75', '73.5', '85,0' (coma decimal).
    Ignora: None, booleans, textos, errores Excel.
    """
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            return round(f, 2) if 1 <= f <= 100 else 0.0
        except Exception:
            return 0.0
    # String: limpiar espacios y convertir coma decimal a punto
    s = str(v).strip().replace(',', '.')
    if not s or s in ('#N/A', '#REF!', '#VALUE!', '#DIV/0!', '-', 'N/A', 'None', 'False', 'True'):
        return 0.0
    if s.isalpha():   # texto puro como 'P1', 'PC'
        return 0.0
    try:
        f = float(s)
        return round(f, 2) if 1 <= f <= 100 else 0.0
    except (ValueError, TypeError):
        return 0.0


def _parsear_boletin_bj(file_bytes):
    """
    Parser del Boletín Oficial del C.E. Benito Juárez.
    Estructura verificada con archivos reales:
      - Bloque de 43 filas por estudiante (4to-6to) / 41 filas (1ro-3ro)
      - col[2]=='ALUMNO/A:' → nombre en col[6]
      - col[2]=='MAESTRO/A:' → maestro en col[6]
      - Académicas: col[2]=materia, P1=col[3], P2=col[5], P3=col[7], P4=col[9]
      - Técnicas (2do ciclo): col[1]='COMPONENTE TÉCNICO' marca inicio,
        luego col[1]=materia, P1=col[3], P2=col[4]
    """
    from openpyxl import load_workbook
    import io as _io, re as _re

    SKIP_MATERIAS = {
        'promedio de grupos', 'calificación final', 'calificacion final',
        'condición', 'condicion', 'el estudiante', 'felicidades',
        'competencias', 'períodos', 'periodos', 'pc1', 'pc2', 'pc3', 'pc4',
        'componete académico', 'componete academico',
        'áreas curriculares', 'areas curriculares',
        'componente técnico', 'componete tecnico',
        'ninguno', 'yordania', 'maestro/a encargado',
        '50%', 'c.e.c', 'c.c.f', 'promedio',
    }

    wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    estudiantes = []

    for sheet_name in wb.sheetnames:
        sn = sheet_name.upper()

        # Detectar grado desde nombre de hoja
        grado = ''
        for g in ['6TO', '5TO', '4TO', '3RO', '2DO', '1RO']:
            if g in sn:
                grado = g.lower()
                break
        if not grado:
            continue

        ciclo = 'primer_ciclo' if grado in ['1ro', '2do', '3ro'] else 'segundo_ciclo'

        # Mención (solo 2do ciclo)
        mencion = ''
        if ciclo == 'segundo_ciclo':
            for m in ['MULTIMEDIA', 'TEATRO', 'VISUALES', 'VISUAL', 'MÚSICA', 'MUSICA']:
                if m in sn:
                    mencion = m
                    break
            if 'VISUAL' in mencion:
                mencion = 'ARTES VISUALES'
            elif 'MUSICA' in mencion or 'MÚSICA' in mencion:
                mencion = 'MÚSICA'

        # Sección (letra A-E del nombre de hoja)
        seccion = 'A'
        ms = _re.search(r'\b([A-E])\b', sheet_name)
        if ms:
            seccion = ms.group(1)

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        i = 0

        while i < len(rows):
            r = rows[i]
            # Inicio de bloque de estudiante
            if len(r) > 6 and str(r[2] or '').strip().upper() in ('ALUMNO/A:', 'ALUMNO/A', 'ALUMNA/O:') and r[6]:
                nombre_completo = ' '.join(str(r[6]).strip().split())  # FIX 2: elimina dobles espacios

                # Maestro en la fila siguiente
                maestro = ''
                if i + 1 < len(rows) and str(rows[i + 1][2] or '').strip().upper() in ('MAESTRO/A:', 'MAESTRA/O:', 'MAESTRO/A'):
                    maestro = str(rows[i + 1][6] or '').strip()

                # Separar nombre y apellido (nombres hispanos dominicanos)
                # Patrón: [Nombre1] [Nombre2?] [Apellido1] [Apellido2?]
                partes = nombre_completo.split()
                if len(partes) >= 4:
                    # Ej: "Miguel Angel Martinez Medina" → nombre="Miguel Angel" apellido="Martinez Medina"
                    nombre   = ' '.join(partes[:2])
                    apellido = ' '.join(partes[2:])
                elif len(partes) == 3:
                    # Ambiguo: puede ser "Jose Junior Perez" (nombre compuesto + 1 apellido)
                    # o "Maria Garcia Lopez" (1 nombre + 2 apellidos)
                    # Guardamos nombre_completo para fuzzy matching flexible
                    nombre   = ' '.join(partes[:2])   # "Jose Junior"
                    apellido = partes[2]              # "Perez"
                elif len(partes) == 2:
                    nombre, apellido = partes[0], partes[1]
                else:
                    nombre, apellido = nombre_completo, ''

                materias = []
                en_tecnico = False
                j = i + 2

                while j < min(i + 52, len(rows)):
                    rj = rows[j]
                    if not rj or len(rj) < 3:
                        j += 1
                        continue

                    c0 = str(rj[0] or '').strip().upper()
                    c1 = str(rj[1] or '').strip()
                    c2 = str(rj[2] or '').strip()

                    # Fin de bloque por CONDICIÓN FINAL
                    if c0 and ('CONDICIÓN' in c0 or 'CONDICION' in c0):
                        break

                    # Inicio de siguiente estudiante
                    if str(c2).strip().upper() in ('ALUMNO/A:', 'ALUMNO/A', 'ALUMNA/O:'):
                        break

                    # Detectar sección técnica
                    if 'COMPONENTE TÉCNICO' in c1.upper() or 'COMPONETE TÉCNICO' in c1.upper():
                        en_tecnico = True
                        j += 1
                        continue

                    if not en_tecnico:
                        # — Materias académicas —
                        # Nombre en col[2], notas en cols 3,5,7,9
                        if c2 and c2 != ' ':
                            c2_low = c2.lower()
                            skip = any(s in c2_low for s in SKIP_MATERIAS)
                            if not skip and len(c2) > 1:
                                p1 = _limpiar_nota(rj[3] if len(rj) > 3 else None)
                                p2 = _limpiar_nota(rj[5] if len(rj) > 5 else None)
                                p3 = _limpiar_nota(rj[7] if len(rj) > 7 else None)
                                p4 = _limpiar_nota(rj[9] if len(rj) > 9 else None)
                                notas = [n for n in [p1, p2, p3, p4] if n > 0]
                                if notas:
                                    # Deduplicar: si ya existe la materia con notas, no sobreescribir
                                    nom_key = c2.lower().strip()
                                    ya_existe = any(m['nombre'].lower().strip() == nom_key for m in materias)
                                    if not ya_existe:
                                        materias.append({
                                            'nombre':   c2,
                                            'tipo':     'académico',
                                            'p1': p1, 'p2': p2, 'p3': p3, 'p4': p4,
                                            'promedio': round(sum(notas) / len(notas), 2)
                                        })
                    else:
                        # — Materias técnicas —
                        # Nombre en col[1], notas en cols 3 y 4
                        if c1 and c1 != ' ':
                            c1_low = c1.lower()
                            skip = any(s in c1_low for s in SKIP_MATERIAS)
                            if not skip and len(c1) > 3:
                                p1 = _limpiar_nota(rj[3] if len(rj) > 3 else None)
                                p2 = _limpiar_nota(rj[4] if len(rj) > 4 else None)
                                notas = [n for n in [p1, p2] if n > 0]
                                if notas:
                                    nom_key = c1.lower().strip()
                                    ya_existe = any(m['nombre'].lower().strip() == nom_key for m in materias)
                                    if not ya_existe:
                                        materias.append({
                                            'nombre':   c1,
                                            'tipo':     'técnico',
                                            'p1': p1, 'p2': p2, 'p3': 0.0, 'p4': 0.0,
                                            'promedio': round(sum(notas) / len(notas), 2)
                                        })
                    j += 1

                if nombre and materias:
                    estudiantes.append({
                        'nombre':   nombre.strip(),
                        'apellido': apellido.strip(),
                        'grado':    grado,
                        'seccion':  seccion,
                        'mencion':  mencion,
                        'ciclo':    ciclo,
                        'maestro':  maestro,
                        'materias': materias
                    })
                i = j
            else:
                i += 1

    wb.close()
    return estudiantes


def _buscar_o_crear_estudiante(conn, nombre, apellido, grado, ciclo, seccion, mencion):
    """
    Busca un estudiante usando _buscar_estudiante_bd.
    Intenta múltiples combinaciones de nombre/apellido para cubrir casos como
    'Jose Junior Perez' que puede estar guardado como 'Jose Junior / Perez'
    o 'Jose / Junior Perez'.
    Si no existe, lo crea. Retorna el id del estudiante.
    """
    nombre_raw   = nombre.strip()
    apellido_raw = apellido.strip()
    nombre_completo = f"{nombre_raw} {apellido_raw}".strip()

    # Generar todas las combinaciones posibles de split nombre/apellido
    partes = nombre_completo.split()
    candidatos = []
    if len(partes) >= 2:
        for split_at in range(1, len(partes)):
            n = ' '.join(partes[:split_at])
            a = ' '.join(partes[split_at:])
            candidatos.append((n, a))
    else:
        candidatos = [(nombre_raw, apellido_raw)]

    # Probar cada combinación
    for n, a in candidatos:
        est = _buscar_estudiante_bd(conn, n, a, filtro_grado=grado)
        if est:
            # Actualizar seccion, ciclo y curso (mención) si cambiaron
            curso_nuevo = f"{grado} {mencion}".strip() if mencion else grado
            conn.execute(
                "UPDATE estudiantes SET seccion=?, ciclo=?, curso=? WHERE id=?",
                (seccion, ciclo, curso_nuevo, est['id'])
            )
            return est['id']

    # Último intento: buscar por nombre completo sin split (fuzzy global)
    est = _buscar_estudiante_bd(conn, nombre_completo, '', filtro_grado=grado)
    if est:
        curso_nuevo = f"{grado} {mencion}".strip() if mencion else grado
        conn.execute(
            "UPDATE estudiantes SET seccion=?, ciclo=?, curso=? WHERE id=?",
            (seccion, ciclo, curso_nuevo, est['id'])
        )
        return est['id']

    # No encontrado — crear perfil nuevo con la separación original del parser
    curso = f"{grado} {mencion}".strip() if mencion else grado
    conn.execute(
        """INSERT INTO estudiantes (nombre, apellido, grado, curso, ciclo, seccion, condicion)
           VALUES (?, ?, ?, ?, ?, ?, 'ACTIVO')""",
        (nombre_raw, apellido_raw, grado, curso, ciclo, seccion)
    )
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return new_id


