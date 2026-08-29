# -*- coding: utf-8 -*-
"""
core/importar_listado.py
=========================
Lógica compartida para leer y aplicar CUALQUIER "Listado de Estudiantes"
oficial contra la tabla `estudiantes`. Soporta dos formas del mismo tipo
de archivo, detectadas automáticamente:

  1. Un solo bloque por archivo (una hoja, un grado/mención) — el formato
     "Listado_Estudiantes_4TO_A_Multimedia.xlsx" que entrega un profesor.
  2. Varios bloques por hoja (el LISTADO institucional completo del año,
     "LISTADO AÑO 2026-2027.xlsx") — una hoja POR GRADO, y dentro de cada
     hoja un bloque "DATOS DEL ALUMNO..." por cada mención (2do ciclo) o
     por cada sección (1er ciclo), uno debajo del otro. Verificado contra
     el archivo real: el grado y la mención/sección no tienen posición fija
     ni formato consistente entre bloques —
       '4TO A' + 'MUSICA'      (sección pegada al grado)
       '5TO'   + 'A Musica'    (sección pegada a la mención, al revés)
       '6TO'   + 'B Teatro'
       '3ERO'  + 'SECCION B'   (1er ciclo: no hay mención, solo sección)
     — así que se busca la celda que contiene "GRADO" en cada fila
     encabezado y se leen las siguientes celdas no vacías en orden, sin
     asumir columna fija ni si la sección viene con el grado o la mención.

Ambos casos se resuelven con el MISMO algoritmo: se buscan todas las filas
"DATOS DEL ALUMNO" en TODAS las hojas del archivo — cada una es el inicio
de un bloque — y cada bloque se procesa de forma independiente. Un archivo
de un solo bloque simplemente da como resultado una lista de longitud 1.

Usada por:
  - scripts/cargar_listado_estudiantes.py (CLI, para Render Shell)
  - routes/profesor.py (subida desde el navegador en /profesor → Cargar Excel)

Una sola implementación — evita el patrón "una copia evoluciona, la otra
fosiliza" que ya causó bugs reales en este proyecto (ver CLAUDE.md sesión 8).
"""
import unicodedata
from difflib import SequenceMatcher

import openpyxl

GRADOS_VALIDOS = ["1ro", "2do", "3ro", "4to", "5to", "6to"]
_GRADO_ALIAS = {"1ero": "1ro", "1er": "1ro", "2do": "2do", "2er": "2do", "3ero": "3ro", "3er": "3ro"}
# Forma en que estudiantes.grado guarda cada grado en producción — verificado
# contra la BD real (diag_roster_profesor.py): '1ERO'/'3ERO' con 3 letras de
# sufijo pero '2DO'/'4TO'/'5TO'/'6TO' con 2, no un patrón uniforme.
_GRADO_SALIDA = {"1ro": "1ERO", "2do": "2DO", "3ro": "3ERO", "4to": "4TO", "5to": "5TO", "6to": "6TO"}

_MENCIONES_CANONICAS = {
    "multimedia": "MULTIMEDIA",
    "musica": "MÚSICA",
    "teatro": "TEATRO",
    "danza": "DANZA",
    "visuales": "ARTES VISUALES",
    "artes visuales": "ARTES VISUALES",
    "artes": "ARTES VISUALES",
}

# nombre normalizado de encabezado → clave interna. Cubre las variantes más
# comunes; agregar aquí si aparece un listado con un nombre de columna nuevo.
HEADER_MAP = {
    "NO": "no", "NO.": "no", "NUM": "no", "NUMERO": "no", "N": "no", "#": "no",
    "NOMBRE": "nombre", "NOMBRES": "nombre",
    "APELLIDO": "apellido", "APELLIDOS": "apellido",
    "EDAD": "edad",
    "ID": "cedula", "CEDULA": "cedula", "IDENTIFICACION": "cedula", "NO IDENTIFICACION": "cedula",
    "SEXO": "sexo", "GENERO": "sexo",
    "FECHA DE NACIMIENTO": "nacimiento", "FECHA NACIMIENTO": "nacimiento", "NACIMIENTO": "nacimiento",
    "TELEFONO": "telefono", "TELEFONO CONTACTO": "telefono", "CONTACTO": "telefono", "TEL": "telefono",
}
COLUMNAS_REQUERIDAS = {"no", "nombre", "apellido"}


def norm(s):
    s = (s or "").strip().lower()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _norm_header(s):
    s = str(s or "").strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.rstrip(".").strip()


def _parsear_grado_seccion(valor):
    """
    '4TO' → ('4TO', None) · '4TO A' → ('4TO', 'A') · '1er.' → ('1ERO', None)
    · '3ERO' → ('3ERO', None)

    Devuelve el grado en el formato exacto que ya usan los ~800 estudiantes
    existentes en `estudiantes.grado` (verificado contra la BD real de
    producción con diag_roster_profesor.py): '4TO'/'5TO'/'6TO'/'2DO' con
    sufijo de 2 letras, pero '1ERO'/'3ERO' con 3 — no un patrón uniforme,
    así que se mapea explícito por _GRADO_SALIDA en vez de solo poner en
    mayúscula lo que vino del archivo.
    """
    grado, seccion = None, None
    resto = []
    for tok in str(valor or "").split():
        t = tok.strip().lower().rstrip(".")
        t = _GRADO_ALIAS.get(t, t)
        if t in GRADOS_VALIDOS:
            grado = _GRADO_SALIDA[t]
        else:
            resto.append(tok.strip())
    if resto:
        cand = resto[0].strip().upper().rstrip(".")
        if cand and len(cand) <= 2 and cand.isalpha():
            seccion = cand
    return grado, seccion


def _normalizar_mencion(texto):
    """'Musica'/'MUSICA'/'A Musica' (ya sin la 'A')/'Visuales ' → forma canónica."""
    t = norm(texto)
    if not t:
        return None
    if t in _MENCIONES_CANONICAS:
        return _MENCIONES_CANONICAS[t]
    for clave, canon in _MENCIONES_CANONICAS.items():
        if clave in t or t in clave:
            return canon
    return texto.strip().upper() or None


def _parsear_bloque_encabezado(grado_raw, resto_raw):
    """
    Combina el valor de grado y el siguiente valor no vacío de la misma
    fila (que puede ser una mención, una sección, o ambas mezcladas) en
    (grado, seccion, mencion). Ver casos reales soportados en el docstring
    del módulo.
    """
    grado, seccion = _parsear_grado_seccion(grado_raw)
    mencion = None

    resto = (resto_raw or "").strip()
    resto_up = _norm_header(resto)

    if not resto:
        return grado, seccion, None

    if resto_up.startswith("SECCION"):
        partes = resto.split()
        if len(partes) >= 2:
            seccion = partes[-1].strip().upper().rstrip(".")
        return grado, seccion, None

    partes = resto.split()
    if len(partes) >= 2 and len(partes[0].rstrip(".")) <= 2 and partes[0].rstrip(".").isalpha():
        # sección pegada delante de la mención: 'A Musica', 'B Teatro'
        seccion = partes[0].strip().upper().rstrip(".")
        mencion = _normalizar_mencion(" ".join(partes[1:]))
    else:
        mencion = _normalizar_mencion(resto)

    return grado, seccion, mencion


def _localizar_bloques(wb):
    """
    Devuelve [(worksheet, fila_inicio, fila_fin_exclusiva), ...] — cada
    tupla es un bloque "DATOS DEL ALUMNO..." independiente, en TODAS las
    hojas del archivo, en el orden en que aparecen.
    """
    bloques = []
    for ws in wb.worksheets:
        marcadores = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            v0 = str(row[0].value or "").strip().upper()
            if "DATOS DEL ALUMNO" in v0:
                marcadores.append(row[0].row)
        for i, inicio in enumerate(marcadores):
            fin = marcadores[i + 1] if i + 1 < len(marcadores) else ws.max_row + 1
            bloques.append((ws, inicio, fin))
    return bloques


_ETIQUETAS_A_SALTAR = {"GRADO", "AREA", "MENCION"}


def _extraer_grado_mencion_fila(ws, fila_marker):
    """
    Busca la celda que contiene 'GRADO' en la fila del marcador y toma las
    siguientes DOS celdas con VALOR (no otra etiqueta) de esa misma fila,
    en orden — sin asumir en qué columna exacta caen. Dos formatos reales
    conviven en distintos archivos:
      'GRADO:' '4TO' 'ÁREA:' 'MULTIMEDIA'   → 2 etiquetas + 2 valores
      'GRADO'  '4TO A' 'MUSICA'             → 1 etiqueta + 2 valores
    por eso se saltan las celdas que son ellas mismas una etiqueta conocida
    (con o sin ':') en vez de asumir que las siguientes 2 celdas no vacías
    son siempre valores.
    """
    celdas = list(ws[fila_marker])
    idx_label = None
    for i, cell in enumerate(celdas):
        if "GRADO" in _norm_header(cell.value):
            idx_label = i
            break
    if idx_label is None:
        return "", ""

    valores = []
    for cell in celdas[idx_label + 1:]:
        v = str(cell.value or "").strip()
        if not v:
            continue
        if _norm_header(v).rstrip(":") in _ETIQUETAS_A_SALTAR:
            continue
        valores.append(v)
        if len(valores) == 2:
            break
    grado_raw  = valores[0] if len(valores) > 0 else ""
    resto_raw  = valores[1] if len(valores) > 1 else ""
    return grado_raw, resto_raw


def _detectar_fila_encabezado(ws, fila_desde, fila_hasta):
    """
    Busca, dentro de [fila_desde, fila_hasta), la fila con los encabezados
    de la tabla de alumnos (No./Nombre/Apellido obligatorios) y arma
    {clave_interna: numero_de_columna}. No asume columnas fijas.

    Fallback: en el archivo institucional real, algunas hojas traen la
    columna NOMBRE con el título corrompido (ej. "Columna1" — nombre por
    defecto de Excel cuando el encabezado original se borró por error).
    Si falta 'nombre' pero 'no' y 'apellido' sí se reconocieron, y la
    columna justo entre medio tiene CUALQUIER texto, se asume que es
    'nombre' — es la única columna que puede ir ahí en este formato.
    """
    mejor_fila, mejor_cols = None, {}
    for row in ws.iter_rows(min_row=fila_desde, max_row=min(fila_hasta - 1, ws.max_row)):
        cols = {}
        cols_con_texto = set()
        for cell in row:
            texto = str(cell.value or "").strip()
            if texto:
                cols_con_texto.add(cell.column)
            clave = HEADER_MAP.get(_norm_header(cell.value))
            if clave and clave not in cols:
                cols[clave] = cell.column

        if "no" in cols and "apellido" in cols and "nombre" not in cols:
            col_candidata = cols["no"] + 1
            if cols["apellido"] - cols["no"] == 2 and col_candidata in cols_con_texto:
                cols["nombre"] = col_candidata

        if COLUMNAS_REQUERIDAS.issubset(cols.keys()) and len(cols) > len(mejor_cols):
            mejor_fila, mejor_cols = row[0].row, cols
    return mejor_fila, mejor_cols


def _leer_alumnos_bloque(ws, fila_header, fila_fin_exclusiva, cols):
    def _val(row, clave):
        col = cols.get(clave)
        return row[col - 1].value if col else None

    alumnos = []
    for row in ws.iter_rows(min_row=fila_header + 1, max_row=fila_fin_exclusiva - 1):
        no = _val(row, "no")
        if not isinstance(no, (int, float)) or not (1 <= no <= 999):
            continue
        nombre   = str(_val(row, "nombre") or "").strip()
        apellido = str(_val(row, "apellido") or "").strip()
        if not nombre or not apellido:
            continue

        edad_raw = _val(row, "edad")
        try:
            edad = int(edad_raw) if edad_raw not in (None, "") else None
        except (TypeError, ValueError):
            edad = None

        cedula_raw = str(_val(row, "cedula") or "").strip()
        cedula_raw = cedula_raw.replace(".0", "").replace(",", "").replace("-", "").strip()
        cedula = cedula_raw if cedula_raw.isdigit() else None

        alumnos.append({
            "no": int(no), "nombre": nombre, "apellido": apellido,
            "edad": edad, "cedula": cedula,
        })
    return alumnos


def leer_listado_workbook(wb):
    """
    Recibe un Workbook de openpyxl ya cargado (desde archivo o BytesIO).
    Devuelve una lista de bloques: [{"grado","seccion","mencion","alumnos"}].
    Un archivo de un solo grado/mención da una lista de longitud 1.
    """
    bloques_crudos = _localizar_bloques(wb)
    if not bloques_crudos:
        raise ValueError(
            "No se encontró ningún encabezado 'DATOS DEL ALUMNO' en el archivo. "
            "Verifica que sea el listado oficial."
        )

    resultado = []
    errores_grado = []
    for ws, inicio, fin in bloques_crudos:
        grado_raw, resto_raw = _extraer_grado_mencion_fila(ws, inicio)
        grado, seccion, mencion = _parsear_bloque_encabezado(grado_raw, resto_raw)
        if not grado:
            errores_grado.append(f"{ws.title!r} fila {inicio} (GRADO={grado_raw!r})")
            continue

        fila_header, cols = _detectar_fila_encabezado(ws, inicio, fin)
        if not fila_header:
            continue  # bloque sin tabla de alumnos debajo (encabezado suelto) — se ignora

        alumnos = _leer_alumnos_bloque(ws, fila_header, fin, cols)
        if not alumnos:
            continue
        alumnos = resolver_duplicados_intra_bloque(alumnos)

        resultado.append({
            "grado": grado, "seccion": seccion, "mencion": mencion, "alumnos": alumnos,
        })

    if not resultado:
        detalle = f" ({'; '.join(errores_grado)})" if errores_grado else ""
        raise ValueError(
            "No se pudo determinar el grado de ningún bloque del archivo"
            + detalle
            + ". Verifica que cada encabezado tenga 'GRADO' seguido de "
              "1ro/2do/3ro/4to/5to/6to."
        )
    return resultado


def leer_listado(path):
    """Variante para CLI: recibe una ruta de archivo."""
    wb = openpyxl.load_workbook(path, data_only=True)
    return leer_listado_workbook(wb)


def _similitud_nombre(alumno, row):
    return SequenceMatcher(
        None,
        norm(alumno["nombre"] + " " + alumno["apellido"]),
        norm(row["nombre"] + " " + row["apellido"]),
    ).ratio()


def resolver_duplicados_intra_bloque(alumnos):
    """
    Si dentro del MISMO bloque (misma hoja/sección del archivo) una cédula
    aparece más de una vez: si es la misma persona (nombre parecido) se
    conserva solo la primera aparición — el resto se descarta en silencio,
    no se duplica. Si son dos personas DISTINTAS con la misma cédula (error
    de tecleo en el archivo — pasó en el LISTADO institucional real: dos
    pares así dentro de la misma hoja), a AMBAS se les quita la cédula del
    archivo para no fusionarlas por error; quedan con ID provisional.

    Necesario porque cedula_en_conflicto()/buscar_existente() comparan
    contra lo que YA está en la BD — dos ocurrencias dentro del mismo
    archivo, antes de escribir nada, no se ven entre sí sin este paso.
    """
    por_cedula = {}
    for a in alumnos:
        if a["cedula"]:
            por_cedula.setdefault(a["cedula"], []).append(a)

    cedulas_conflictivas = set()
    for cedula, ocurrencias in por_cedula.items():
        if len(ocurrencias) < 2:
            continue
        base = ocurrencias[0]
        for otro in ocurrencias[1:]:
            if _similitud_nombre(base, otro) < 0.5:
                cedulas_conflictivas.add(cedula)
                break

    resultado = []
    cedulas_vistas = set()
    for a in alumnos:
        cedula = a["cedula"]
        if cedula and cedula in cedulas_conflictivas:
            a = dict(a)
            a["cedula"] = None
            a["_cedula_conflicto_archivo"] = True
            resultado.append(a)
            continue
        if cedula:
            if cedula in cedulas_vistas:
                continue  # misma persona repetida en el archivo — se omite
            cedulas_vistas.add(cedula)
        resultado.append(a)
    return resultado


def cedula_en_conflicto(conn, alumno):
    """
    Si la cédula del archivo ya existe en la BD pero con un nombre muy
    distinto, devuelve ese registro — probable error de tecleo en el
    archivo (un dígito de la cédula mal escrito), NO la misma persona.
    Encontrado en el LISTADO institucional real: 5 pares de estudiantes
    con nombres completamente distintos comparten cédula. Sin este chequeo,
    aplicar_carga() fusionaría a esos 5 pares en un solo registro por
    error. Retorna None si no hay conflicto (cédula libre, o coincide con
    un nombre razonablemente parecido — la misma persona).
    """
    if not alumno["cedula"]:
        return None
    row = conn.execute(
        "SELECT id, nombre, apellido FROM estudiantes WHERE cedula=?",
        (alumno["cedula"],)
    ).fetchone()
    if not row:
        return None
    return row if _similitud_nombre(alumno, row) < 0.5 else None


def buscar_existente(conn, alumno, grado, mencion):
    """
    cédula exacta (global, solo si el nombre coincide razonablemente —
    ver cedula_en_conflicto) → nombre+apellido exacto (mismo grado+mención)
    → fuzzy (mismo grado+mención).

    El nombre+apellido exacto SIEMPRE debe ir acotado al grado/mención
    objetivo — sin eso, un alumno sin cédula en el archivo puede coincidir
    por nombre con un estudiante de OTRO grado que ya está en la BD (pasó
    en producción: 27 alumnos de "4to Multimedia" con nombre exacto igual
    a estudiantes ya cargados en '3ERO', que habrían quedado reasignados
    por error). La comparación de grado es case-insensitive porque los
    datos reales usan MAYÚSCULA ('4TO') pero eso no debe ser un requisito
    silencioso — mejor comparar sin depender de la case.
    """
    if alumno["cedula"]:
        row = conn.execute(
            "SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes WHERE cedula=?",
            (alumno["cedula"],)
        ).fetchone()
        if row and _similitud_nombre(alumno, row) >= 0.5:
            return row

    if mencion:
        scope_sql = "UPPER(grado)=UPPER(?) AND (mencion=? OR mencion IS NULL OR mencion='')"
        scope_params = (grado, mencion)
    else:
        scope_sql = "UPPER(grado)=UPPER(?)"
        scope_params = (grado,)

    row = conn.execute(
        f"""SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes
           WHERE {scope_sql} AND lower(nombre)=lower(?) AND lower(apellido)=lower(?)""",
        scope_params + (alumno["nombre"], alumno["apellido"])
    ).fetchone()
    if row:
        return row

    clave = norm(alumno["nombre"] + " " + alumno["apellido"])
    candidatos = conn.execute(
        f"SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes WHERE {scope_sql}",
        scope_params
    ).fetchall()
    mejor, mejor_score = None, 0.0
    for c in candidatos:
        score = SequenceMatcher(None, clave, norm(c["nombre"] + " " + c["apellido"])).ratio()
        if score > mejor_score:
            mejor, mejor_score = c, score
    if mejor and mejor_score >= 0.82:
        return mejor
    return None


def construir_plan(conn, grado, seccion, mencion, alumnos):
    """
    Calcula, SIN escribir en BD, qué haría la carga de UN bloque (un
    grado/sección/mención). Devuelve (curso, ciclo, plan) donde plan es
    una lista de dicts:
      {no, nombre, apellido, cedula, accion: 'nuevo'|'actualiza',
       cambios: [...], advertencia: str|None, estudiante_id: int|None}
    """
    curso = f"{grado} {mencion}".strip() if mencion else grado
    ciclo = "primer_ciclo" if grado.upper() in ("1ERO", "2DO", "3ERO") else "segundo_ciclo"

    plan = []
    for alumno in alumnos:
        existente = buscar_existente(conn, alumno, grado, mencion)
        item = {
            "no": alumno["no"], "nombre": alumno["nombre"], "apellido": alumno["apellido"],
            "cedula": alumno["cedula"], "edad": alumno["edad"],
            "advertencia": None, "estudiante_id": None,
        }
        if existente:
            cambios = []
            if existente["grado"] != grado: cambios.append(f"grado {existente['grado']!r}→{grado!r}")
            if existente["curso"] != curso: cambios.append(f"curso {existente['curso']!r}→{curso!r}")
            if not existente["cedula"] and alumno["cedula"]:
                cambios.append(f"cédula → {alumno['cedula']}")
            item.update(accion="actualiza", cambios=cambios, estudiante_id=existente["id"])
            if existente["condicion"] in ("RETIRADO", "TRANSFERIDO"):
                item["advertencia"] = (
                    f"Ya existe con condición {existente['condicion']} — "
                    "no se reactiva automáticamente, revisar a mano"
                )
        else:
            item.update(accion="nuevo", cambios=[])
            if alumno.get("_cedula_conflicto_archivo"):
                item["advertencia"] = (
                    "Esta cédula aparece más de una vez en el archivo con nombres distintos "
                    "— posible error de tecleo. Se carga sin cédula (ID provisional); revisar a mano."
                )
            elif not alumno["cedula"]:
                item["advertencia"] = "Sin cédula/ID en el archivo — queda con ID provisional"
            else:
                conflicto = cedula_en_conflicto(conn, alumno)
                if conflicto:
                    item["advertencia"] = (
                        f"La cédula {alumno['cedula']} ya está registrada a nombre de "
                        f"{conflicto['nombre']} {conflicto['apellido']} — posible error de "
                        "tecleo en el archivo. Se crea como alumno nuevo sin tocar ese registro; revisar a mano."
                    )
        plan.append(item)
    return curso, ciclo, plan


def aplicar_carga(conn, grado, seccion, mencion, alumnos):
    """Escribe en BD UN bloque. Devuelve (nuevos, actualizados)."""
    curso, ciclo, plan = construir_plan(conn, grado, seccion, mencion, alumnos)
    nuevos = actualizados = 0

    for alumno, item in zip(alumnos, plan):
        if item["accion"] == "actualiza":
            actualizados += 1
            # No se toca 'condicion' en un UPDATE — si el alumno ya existía
            # como RETIRADO/TRANSFERIDO, reactivarlo es una decisión humana.
            conn.execute(
                """UPDATE estudiantes
                      SET nombre=?, apellido=?, grado=?, curso=?, mencion=?, ciclo=?,
                          seccion=COALESCE(?, seccion),
                          cedula=COALESCE(NULLIF(cedula,''), ?),
                          edad=COALESCE(?, edad)
                    WHERE id=?""",
                (alumno["nombre"], alumno["apellido"], grado, curso, mencion, ciclo,
                 seccion, alumno["cedula"] or "", alumno["edad"], item["estudiante_id"])
            )
        else:
            nuevos += 1
            # Si la cédula ya la usa otro estudiante con nombre muy distinto
            # (error de tecleo en el archivo — ver cedula_en_conflicto), no
            # se graba esa cédula en el registro nuevo: dejaría dos
            # estudiantes distintos compartiendo cédula, ambiguo para
            # cualquier búsqueda futura por cédula. Queda con ID provisional
            # (ya se avisó en el plan) hasta que se corrija a mano.
            cedula_usar = alumno["cedula"]
            if cedula_usar and cedula_en_conflicto(conn, alumno):
                cedula_usar = None
            conn.execute(
                """INSERT INTO estudiantes
                       (id_evaluacion, cedula, nombre, apellido, curso, grado,
                        mencion, ciclo, seccion, condicion, edad)
                   VALUES (?,?,?,?,?,?,?,?,COALESCE(?,'A'),'ACTIVO',?)""",
                (cedula_usar or f"PROV_{alumno['nombre'].split()[0]}_{alumno['apellido'].split()[0]}",
                 cedula_usar or "", alumno["nombre"], alumno["apellido"],
                 curso, grado, mencion, ciclo, seccion, alumno["edad"])
            )
    conn.commit()
    return nuevos, actualizados


def construir_plan_multi(conn, bloques):
    """
    Igual que construir_plan pero para una LISTA de bloques (un archivo
    institucional con varios grados/menciones). Devuelve una lista de
    dicts por bloque: {grado, seccion, mencion, curso, nuevos, actualizados,
    plan}.
    """
    resultado = []
    for b in bloques:
        curso, ciclo, plan = construir_plan(conn, b["grado"], b["seccion"], b["mencion"], b["alumnos"])
        resultado.append({
            "grado": b["grado"], "seccion": b["seccion"], "mencion": b["mencion"], "curso": curso,
            "nuevos": sum(1 for p in plan if p["accion"] == "nuevo"),
            "actualizados": sum(1 for p in plan if p["accion"] == "actualiza"),
            "plan": plan,
        })
    return resultado


def aplicar_carga_multi(conn, bloques):
    """Igual que aplicar_carga pero para una lista de bloques. Devuelve (nuevos, actualizados) totales."""
    nuevos_total = actualizados_total = 0
    for b in bloques:
        n, a = aplicar_carga(conn, b["grado"], b["seccion"], b["mencion"], b["alumnos"])
        nuevos_total += n
        actualizados_total += a
    return nuevos_total, actualizados_total
