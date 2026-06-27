# -*- coding: utf-8 -*-
"""
Parser de Registros de Calificaciones del Coordinador (formato .xlsm BJ).

Detecta automáticamente:
  - Grado, sección y mención desde el nombre de la hoja  ("CF 4TO A MÚSICA")
  - Materia desde las filas de encabezado              ("ÁREAS: IDIOMA INGLÉS")
  - Columnas de períodos desde el panel derecho de resumen
  - Calificación final (CF o Promedio CF)

Compatible con ambos formatos detectados:
  • Con columnas RP intercaladas (INGLÉS)     → panel derecho: PC1-PC4 + CF
  • Sin columnas RP (DIBUJO y similares)      → panel derecho: Periodo I-IV + Promedio CF
"""

import re
import unicodedata

__all__ = ["parsear_registro_coordinador"]

# ── Normalización de texto ─────────────────────────────────────────────────────
def _norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


MATS_ACADEMICAS_NORM = {
    _norm(m) for m in [
        "Lengua Española", "Idioma Inglés", "Idioma Frances",
        "Matemática", "Matematica", "Ciencias Sociales",
        "Ciencias Naturales", "Educación Física", "Educacion Fisica", "FIHR",
    ]
}

def _es_academica(nombre_materia: str) -> bool:
    """Devuelve True si la materia pertenece al componente académico."""
    n = _norm(nombre_materia)
    return any(n.startswith(m) or m.startswith(n) for m in MATS_ACADEMICAS_NORM)


# ── Parser principal ──────────────────────────────────────────────────────────
def parsear_registro_coordinador(filepath: str) -> dict:
    """
    Lee un archivo .xlsm/.xlsx de registro del coordinador.

    Retorna:
    {
        "materia":   str,
        "hojas": [
            {
                "hoja":    str,
                "grado":   str,
                "seccion": str,
                "mencion": str,
                "tipo":    "académico" | "técnico",
                "alumnos": [
                    {
                        "nombre":   str,   # Nombres + Apellidos
                        "p1": float, "p2": float, "p3": float, "p4": float,
                        "promedio": float,
                        "estado":   str,   # ACTIVO / RETIRADO
                    },
                    ...
                ]
            },
            ...
        ],
        "advertencias": [str]
    }
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no instalado — ejecuta: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    resultado = {"materia": "", "hojas": [], "advertencias": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ── 1. Grado / Sección / Mención desde nombre de hoja ─────────────────
        # Patrón 2do ciclo: "CF 4TO A MÚSICA" | "CF 5TO B TEATRO"
        # Patrón 1er ciclo: "CF 1ERO A" | "CF 2DO B" (sin mención)
        m = re.match(r"CF\s+(\w+)\s+([A-Ea-e])(?:\s+(.+))?", sheet_name.strip(), re.IGNORECASE)
        if not m:
            resultado["advertencias"].append(
                f"Hoja '{sheet_name}' ignorada — nombre no reconocido (esperado: CF GRADO SECC [MENCIÓN])"
            )
            continue

        grado   = m.group(1).upper()
        seccion = m.group(2).upper()
        mencion = (m.group(3) or "").strip().upper()  # vacío en primer ciclo

        # ── 2. Nombre de la materia desde filas de encabezado ─────────────────
        materia_hoja = ""
        for row_idx in range(1, 12):
            for col_idx in range(1, 10):
                val = ws.cell(row_idx, col_idx).value
                if not val:
                    continue
                s = str(val).strip()
                # Buscar celdas con "ÁREA" o "MATERIA" o "ASIGNATURA"
                if re.search(r"ÁREA|AREA|ASIGNATURA|MATERIA", s, re.IGNORECASE):
                    partes = re.split(r"[:：]", s, 1)
                    if len(partes) == 2:
                        materia_hoja = partes[1].strip()
                    elif not re.search(r"ÁREA|AREA|ASIGNATURA|MATERIA", partes[0], re.IGNORECASE):
                        materia_hoja = s
                    if materia_hoja:
                        break
            if materia_hoja:
                break

        if not resultado["materia"] and materia_hoja:
            resultado["materia"] = materia_hoja

        # ── 3. Fila de encabezado: buscar "Nombres" en col 3 ─────────────────
        header_row = None
        for row_idx in range(1, 16):
            val = ws.cell(row_idx, 3).value
            if val and str(val).strip().lower() in ("nombres", "nombre"):
                header_row = row_idx
                break

        if header_row is None:
            resultado["advertencias"].append(
                f"Hoja '{sheet_name}': no se encontró la fila de encabezados (buscando 'Nombres' en col C)"
            )
            continue

        # ── 4. Panel derecho: segunda aparición de "Nombres" en col 3 ────────
        right_nombres_col = None
        found_first = False
        for col_idx in range(2, (ws.max_column or 110) + 1):
            val = ws.cell(header_row, col_idx).value
            if val and str(val).strip().lower() in ("nombres", "nombre"):
                if found_first:
                    right_nombres_col = col_idx
                    break
                found_first = True

        if right_nombres_col is None:
            resultado["advertencias"].append(
                f"Hoja '{sheet_name}': no se encontró el panel de resumen (segunda columna 'Nombres')"
            )
            continue

        # ── 5. Buscar columnas de períodos y CF en el panel derecho ──────────
        # Los labels pueden estar en el header_row (ej: PC1-PC4 en INGLÉS) o
        # en filas superiores con celdas combinadas (ej: CF en fila 8 para INGLÉS,
        # Periodo I-IV y Promedio CF en fila 8 para DIBUJO).
        # Solución: escanear TODAS las filas desde 1 hasta header_row (inclusive)
        # en el rango de columnas del panel derecho.
        p1_col = p2_col = p3_col = p4_col = cf_col = None
        max_col = (ws.max_column or 110) + 1
        # El panel derecho empieza en right_nombres_col; aceptar hasta 5 cols antes
        # por si el label de la col de datos está ligeramente a la izquierda.
        scan_start_col = max(1, right_nombres_col - 5)

        for scan_row in range(1, header_row + 1):
            for col_idx in range(scan_start_col, max_col):
                val = ws.cell(scan_row, col_idx).value
                if not val:
                    continue
                label = re.sub(r"\s+", " ", str(val)).strip().upper()

                if label in ("PC1", "PERIODO I", "P I", "PI") and p1_col is None:
                    p1_col = col_idx
                elif label in ("PC2", "PERIODO II", "P II", "PII") and p2_col is None:
                    p2_col = col_idx
                elif label in ("PC3", "PERIODO III", "P III", "PIII") and p3_col is None:
                    p3_col = col_idx
                elif label in ("PC4", "PERIODO IV", "P IV", "PIV") and p4_col is None:
                    p4_col = col_idx
                elif label in ("CF", "PROMEDIO CF", "CALIFICACION FINAL", "CALIFICACIÓN FINAL") and cf_col is None:
                    cf_col = col_idx

        if cf_col is None:
            resultado["advertencias"].append(
                f"Hoja '{sheet_name}': no se encontró columna CF/Promedio CF en el panel derecho"
            )
            continue

        # ── 6. Leer filas de alumnos ──────────────────────────────────────────
        alumnos = []
        for row_idx in range(header_row + 1, header_row + 60):
            num = ws.cell(row_idx, 2).value
            # Fin de datos: celda vacía o cero
            if num is None or num == 0:
                break

            nombres   = str(ws.cell(row_idx, 3).value or "").strip()
            apellidos = str(ws.cell(row_idx, 4).value or "").strip()
            estado    = str(ws.cell(row_idx, 5).value or "ACTIVO").strip().upper()

            nombre_completo = f"{nombres} {apellidos}".strip()
            if not nombre_completo:
                continue

            def _num(col):
                if col is None:
                    return 0.0
                v = ws.cell(row_idx, col).value
                try:
                    return round(float(v), 2) if v is not None else 0.0
                except (TypeError, ValueError):
                    return 0.0

            alumnos.append({
                "nombre":   nombre_completo,
                "p1":       _num(p1_col),
                "p2":       _num(p2_col),
                "p3":       _num(p3_col),
                "p4":       _num(p4_col),
                "promedio": _num(cf_col),
                "estado":   estado,
            })

        tipo = "académico" if _es_academica(materia_hoja) else "técnico"

        resultado["hojas"].append({
            "hoja":    sheet_name,
            "grado":   grado,
            "seccion": seccion,
            "mencion": mencion,
            "tipo":    tipo,
            "alumnos": alumnos,
        })

    wb.close()
    return resultado
