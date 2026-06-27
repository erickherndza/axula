# -*- coding: utf-8 -*-
"""Blueprint: reportes"""

import sqlite3
import logging
import json as _json
import os
import re
import time as _time
from datetime import datetime, date, timedelta
from io import BytesIO
from flask import (
    Blueprint, render_template, request, jsonify, session,
    redirect, url_for, g, send_from_directory, Response, send_file,
)

from core.constants import *
from core.database import get_db, cache_get, cache_set, cache_bust, _CACHE
from core.auth import (
    _hash, _check_password, _normalizar_rol, _ciclo_del_rol,
    login_required, coord_required, admin_required, directora_required,
    _csrf_token, _csrf_check, csrf_protected, rate_limited,
    get_usuario,
)
from core.helpers import *
from core.helpers import _notificar_reporte_nuevo
from core.ia import _get_groq_client, groq_client, construir_prompt, construir_prompt_planificacion, construir_prompt_rubrica, construir_prompt_estrategia
from core.excel import _parsear_boletin_bj, _buscar_o_crear_estudiante, _detectar_mencion_listado, _limpiar_nota
from core.pdf import _generar_pdf_acuerdo

logger = logging.getLogger("axula")

reportes_bp = Blueprint("reportes_bp", __name__)

@reportes_bp.route("/reportes")
@login_required
def vista_reportes():
    return render_template("reportes.html", current_user=get_usuario())


@reportes_bp.route("/api/reportes", methods=["GET"])
@login_required
def listar_reportes():
    tipo      = request.args.get("tipo", "")
    estado    = request.args.get("estado", "")
    est_id    = request.args.get("estudiante_id", "")
    severidad = request.args.get("severidad", "")

    q = """
        SELECT r.*, e.nombre, e.apellido, e.grado, e.curso
        FROM reportes r
        JOIN estudiantes e ON e.id = r.estudiante_id
        WHERE 1=1
    """
    # caso_id ya es columna de reportes (migración)
    params = []
    if tipo:      q += " AND r.tipo=?";             params.append(tipo)
    if estado:    q += " AND r.estado=?";           params.append(estado)
    if est_id:    q += " AND r.estudiante_id=?";    params.append(int(est_id))
    if severidad: q += " AND r.severidad=?";        params.append(severidad)
    limit = request.args.get("limit", "")
    q += " ORDER BY r.fecha DESC, r.id DESC"
    if limit.isdigit():
        q += f" LIMIT {int(limit)}"

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])


@reportes_bp.route("/api/reportes", methods=["POST"])
@login_required
def crear_reporte():
    d = request.get_json(silent=True) or {}
    est_id      = d.get("estudiante_id")
    tipo        = d.get("tipo", "").strip()
    descripcion = d.get("descripcion", "").strip()
    if not est_id or not tipo or not descripcion:
        return jsonify({"error": "estudiante_id, tipo y descripcion son requeridos"}), 400

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            INSERT INTO reportes
                (estudiante_id, tipo, subtipo, titulo, descripcion,
                 severidad, reportado_por, estado)
            VALUES (?,?,?,?,?,?,?,?)
        """, (est_id, tipo,
              d.get("subtipo",""),
              d.get("titulo",""),
              descripcion,
              d.get("severidad","Media"),
              d.get("reportado_por",""),
              "Abierto"))
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # ── Notificar automáticamente a psicóloga ─────────────────────────────
    with sqlite3.connect(DATABASE, timeout=10) as conn2:
        conn2.row_factory = sqlite3.Row
        _notificar_reporte_nuevo(
            conn2, est_id, rid,
            tipo,
            d.get("severidad", "Media"),
            d.get("titulo", ""),
            d.get("reportado_por", "")
        )
        conn2.commit()

    return jsonify({"ok": True, "id": rid})


@reportes_bp.route("/api/reportes/<int:rid>", methods=["PATCH"])
@login_required
def actualizar_reporte(rid):
    d = request.get_json(silent=True) or {}
    campos = {}
    for k in ["estado","seguimiento","fecha_cierre","severidad","titulo","descripcion"]:
        if k in d: campos[k] = d[k]
    if not campos:
        return jsonify({"error": "Sin campos"}), 400
    sets   = ", ".join(f"{k}=?" for k in campos)
    vals   = list(campos.values()) + [rid]
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute(f"UPDATE reportes SET {sets} WHERE id=?", vals)
        conn.commit()
    return jsonify({"ok": True})


@reportes_bp.route("/api/reportes/<int:rid>/escalar-caso", methods=["POST"])
@login_required
def escalar_reporte_a_caso(rid):
    """
    Convierte un reporte en un Caso formal.
    - Crea un caso prerellenado con los datos del reporte.
    - Marca el reporte con el caso_id creado.
    - Solo coordinadores, psicólogas y directora.
    """
    u = get_usuario()
    rol = _normalizar_rol(u.get("rol", ""))
    if rol not in ROLES_PSICOLOGA and rol not in ROLES_COORD and not u.get("es_directora"):
        return jsonify({"error": "Sin permisos para escalar casos."}), 403

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rep = conn.execute(
            "SELECT * FROM reportes WHERE id=?", (rid,)
        ).fetchone()
        if not rep:
            return jsonify({"error": "Reporte no encontrado"}), 404
        rep = dict(rep)

        if rep.get("caso_id"):
            return jsonify({"error": "Este reporte ya fue escalado", "caso_id": rep["caso_id"]}), 400

        # Mapeo tipo reporte → tipo caso
        TIPO_MAP = {
            "conducta":        "conducta",
            "psicologico":     "familiar",
            "academico":       "academico",
            "incidente_grave": "conducta",
        }
        tipo_caso = TIPO_MAP.get(rep["tipo"], "conducta")
        titulo_caso = rep.get("titulo") or f"Caso originado en reporte {rep['tipo']}"
        desc_caso = (
            f"Escalado desde Reporte #{rid} ({rep['tipo']} — {rep.get('severidad','')}).\n\n"
            f"{rep.get('descripcion','')}"
        )

        conn.execute("""
            INSERT INTO casos
                (estudiante_id, abierto_por, tipo, titulo, descripcion,
                 estado, nivel_escala, origen_tipo, origen_id)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            rep["estudiante_id"], u["id"], tipo_caso,
            titulo_caso, desc_caso,
            "Abierto", 1, "reporte", rid
        ))
        conn.commit()
        caso_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Vincular reporte al caso
        conn.execute(
            "UPDATE reportes SET caso_id=?, estado='En seguimiento' WHERE id=?",
            (caso_id, rid)
        )
        conn.commit()

    logger.info(f"[CASOS] Reporte #{rid} escalado a Caso #{caso_id} por {u['username']}")
    return jsonify({"ok": True, "caso_id": caso_id})


@reportes_bp.route("/api/reportes/<int:rid>", methods=["DELETE"])
@login_required
def eliminar_reporte(rid):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("DELETE FROM reportes WHERE id=?", (rid,))
        conn.commit()
    return jsonify({"ok": True})


@reportes_bp.route("/api/reportes/resumen")
@login_required
def resumen_reportes():
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        total  = conn.execute("SELECT COUNT(*) FROM reportes").fetchone()[0]
        abiertos= conn.execute("SELECT COUNT(*) FROM reportes WHERE estado='Abierto'").fetchone()[0]
        graves  = conn.execute("SELECT COUNT(*) FROM reportes WHERE tipo='incidente_grave'").fetchone()[0]
        por_tipo= conn.execute("""
            SELECT tipo, COUNT(*) n FROM reportes GROUP BY tipo ORDER BY n DESC
        """).fetchall()
        recientes = conn.execute("""
            SELECT r.id, r.tipo, r.titulo, r.severidad, r.fecha, r.estado,
                   e.nombre, e.apellido
            FROM reportes r JOIN estudiantes e ON e.id=r.estudiante_id
            ORDER BY r.fecha DESC, r.id DESC LIMIT 10
        """).fetchall()
    return jsonify({
        "total": total, "abiertos": abiertos, "graves": graves,
        "por_tipo": [{"tipo":r[0],"n":r[1]} for r in por_tipo],
        "recientes": [dict(zip(
            ["id","tipo","titulo","severidad","fecha","estado","nombre","apellido"], r
        )) for r in recientes]
    })


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO ML — Clustering K-Means + Patrones de comportamiento
# ══════════════════════════════════════════════════════════════════════════════

# Etiquetas y colores para cada cluster (se asignan dinámicamente)


@reportes_bp.route("/api/reportes/exportar-xlsx")
@login_required
def exportar_reportes_xlsx():
    """Descarga los reportes filtrados como archivo Excel."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    tipo      = request.args.get("tipo","")
    estado    = request.args.get("estado","")
    severidad = request.args.get("severidad","")

    q = """
        SELECT r.fecha, e.nombre||' '||e.apellido AS estudiante,
               e.grado, r.tipo, r.subtipo, r.titulo, r.descripcion,
               r.severidad, r.estado, r.reportado_por,
               r.seguimiento, r.fecha_cierre
        FROM reportes r JOIN estudiantes e ON e.id=r.estudiante_id
        WHERE 1=1
    """
    params = []
    if tipo:      q += " AND r.tipo=?";      params.append(tipo)
    if estado:    q += " AND r.estado=?";    params.append(estado)
    if severidad: q += " AND r.severidad=?"; params.append(severidad)
    q += " ORDER BY r.fecha DESC"

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        rows = conn.execute(q, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ["Fecha","Estudiante","Grado","Tipo","Subtipo","Título",
               "Descripción","Severidad","Estado","Reportado por","Seguimiento","Fecha cierre"]

    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="C8F060", size=10)

    sev_colors = {"Alta":"FF6B6B","Media":"FFC44D","Baja":"4DFFB4"}

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val or "")
        sev = row[7] or ""
        if sev in sev_colors:
            fill = PatternFill("solid", fgColor=sev_colors[sev])
            ws.cell(row=ri, column=8).fill = fill

    col_widths = [12,25,12,14,16,28,40,10,14,20,30,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import Response
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reportes_multimediatrack.xlsx"}
    )


# ── GESTIÓN DE BASE DE DATOS (panel admin) ───────────────────────────────────


@reportes_bp.route("/api/reportes/exportar-pdf")
@login_required
def exportar_reportes_pdf():
    """Genera un PDF printable de los reportes filtrados usando HTML → PDF."""
    from flask import make_response
    import datetime

    tipo      = request.args.get("tipo", "")
    estado    = request.args.get("estado", "")
    severidad = request.args.get("severidad", "")

    q = """
        SELECT r.fecha, e.nombre||' '||e.apellido AS estudiante,
               e.grado, r.tipo, r.titulo, r.descripcion,
               r.severidad, r.estado, r.reportado_por, r.seguimiento
        FROM reportes r JOIN estudiantes e ON e.id=r.estudiante_id
        WHERE 1=1
    """
    params = []
    if tipo:      q += " AND r.tipo=?";      params.append(tipo)
    if estado:    q += " AND r.estado=?";    params.append(estado)
    if severidad: q += " AND r.severidad=?"; params.append(severidad)
    q += " ORDER BY r.fecha DESC"

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]

    tipo_label = {
        "conducta": "Conducta", "psicologico": "Psicológico",
        "academico": "Académico", "incidente_grave": "Incidente Grave"
    }
    sev_color = {"Alta": "#ff6b6b", "Media": "#ffc44d", "Baja": "#4dffb4"}

    filas_html = ""
    for r in rows:
        sev_c = sev_color.get(r.get("severidad",""), "#888")
        filas_html += f"""
        <tr>
          <td>{r.get('fecha','')}</td>
          <td><b>{r.get('estudiante','')}</b><br><small>{r.get('grado','')}</small></td>
          <td>{tipo_label.get(r.get('tipo',''), r.get('tipo',''))}</td>
          <td>{r.get('titulo','')}</td>
          <td style="color:{sev_c};font-weight:700;">{r.get('severidad','')}</td>
          <td>{r.get('estado','')}</td>
          <td style="font-size:10px;">{(r.get('descripcion') or '')[:120]}</td>
          <td style="font-size:10px;">{r.get('reportado_por','')}</td>
        </tr>"""

    now = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11pt; color: #111; margin: 20px; }}
  h1   {{ font-size: 16pt; margin-bottom: 2px; }}
  .sub {{ font-size: 10pt; color: #666; margin-bottom: 16px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  thead th {{ background: #1a1a1a; color: #c8f060; padding: 6px 8px;
              font-size: 9pt; text-align: left; }}
  tbody tr:nth-child(even) {{ background: #f9f9f9; }}
  td {{ padding: 5px 8px; vertical-align: top; border-bottom: 1px solid #eee; font-size: 9pt; }}
  @page {{ margin: 15mm; size: A4 landscape; }}
  @media print {{ button {{ display:none; }} }}
</style>
</head><body>
<h1>📋 Reportes — Axula</h1>
<div class="sub">Generado: {now} · {len(rows)} registros{' · Tipo: '+tipo_label.get(tipo,tipo) if tipo else ''}{' · Estado: '+estado if estado else ''}{' · Severidad: '+severidad if severidad else ''}</div>
<button onclick="window.print()" style="margin-bottom:14px;padding:7px 18px;background:#1a1a1a;
  color:#c8f060;border:1px solid #333;border-radius:6px;cursor:pointer;font-size:12px;">
  🖨 Imprimir / Guardar PDF
</button>
<table>
  <thead>
    <tr>
      <th>Fecha</th><th>Estudiante</th><th>Tipo</th><th>Título</th>
      <th>Severidad</th><th>Estado</th><th>Descripción</th><th>Reportado por</th>
    </tr>
  </thead>
  <tbody>{filas_html}</tbody>
</table>
</body></html>"""

    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp



# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO PROFESORES — Asignación, Validación y Asistencia
# Normativa: Ordenanza 1'96 mod. 1'98 + 04-2023 MINERD
# • Escala 0-100, mínimo aprobatorio 70 puntos
# • Art. 51 mod. 1'98: reprueba con >20% inasistencias injustificadas
# • Asistencia: (horas_presentes / horas_totales) × 100
# ══════════════════════════════════════════════════════════════════════════════

# Plan de estudio oficial MINERD — referencia al dict ya definido arriba en PLAN_ARTES
# (No redefinir aquí — usar PLAN_ARTES["MULTIMEDIA"] que ya incluye todos los grados)


