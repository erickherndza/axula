# -*- coding: utf-8 -*-
"""
Blueprint: expediente
Digitalización y consulta de expedientes históricos (25+ años).
Acceso: secretaria, secretaria_docente, digitador, coordinadores, directora.
"""

import sqlite3
import json as _json
import logging
import re
from datetime import date
from io import BytesIO

from flask import Blueprint, request, jsonify, session, send_file

from core.constants import DATABASE
from core.auth import login_required, get_usuario, _normalizar_rol, rate_limited
from core import rls as _rls
from core.helpers import _validar_magic_excel, _anio_escolar_actual, _audit

logger = logging.getLogger("axula")
expediente_bp = Blueprint("expediente_bp", __name__)

# ── Roles con acceso ──────────────────────────────────────────────────────────
_ROLES_EXPEDIENTE = {
    "secretaria", "secretaria_docente", "digitador",
    "coordinador_general", "coordinador_primer_ciclo",
    "coordinador_segundo_ciclo", "directora",
}

def _exp_required(f):
    import functools
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return jsonify({"error": "No autenticado"}), 401
        rol = _normalizar_rol(session.get("rol", ""))
        if rol not in _ROLES_EXPEDIENTE:
            return jsonify({"error": "Sin permisos"}), 403
        return f(*args, **kwargs)
    return decorated


# ── Helpers ───────────────────────────────────────────────────────────────────

def _detectar_sistema(anio_escolar: str) -> str:
    """Pre-2016-2017 = bachillerato, desde 2016-2017 = secundaria."""
    try:
        anio_inicio = int(str(anio_escolar).split("-")[0].strip())
        return "bachillerato" if anio_inicio < 2016 else "secundaria"
    except Exception:
        return "secundaria"


GRADOS_BACHILLERATO = ["1ro Bachiller", "2do Bachiller", "3ro Bachiller", "4to Bachiller"]
GRADOS_SECUNDARIA   = ["1ro", "2do", "3ro", "4to", "5to", "6to"]


def _normalizar_grado(grado: str, sistema: str) -> str:
    """Normaliza texto libre a grado estándar según sistema educativo."""
    g = str(grado).strip().lower()
    if sistema == "bachillerato":
        mapa = {"1": "1ro Bachiller", "2": "2do Bachiller",
                "3": "3ro Bachiller", "4": "4to Bachiller",
                "primero": "1ro Bachiller", "segundo": "2do Bachiller",
                "tercero": "3ro Bachiller", "cuarto": "4to Bachiller"}
        for k, v in mapa.items():
            if k in g:
                return v
        return grado  # devolver tal cual si no reconoce
    else:
        mapa = {"1": "1ro", "2": "2do", "3": "3ro",
                "4": "4to", "5": "5to", "6": "6to",
                "primero": "1ro", "segundo": "2do", "tercero": "3ro",
                "cuarto": "4to", "quinto": "5to", "sexto": "6to"}
        for k, v in mapa.items():
            if k in g:
                return v
        return grado


def _buscar_estudiante_vinculable(conn, cedula, nombre, apellido):
    """Intenta encontrar el id del estudiante en la tabla estudiantes."""
    if cedula:
        row = conn.execute(
            "SELECT id FROM estudiantes WHERE cedula=? LIMIT 1", (cedula,)
        ).fetchone()
        if row:
            return row["id"]
    # Búsqueda por nombre+apellido exacto
    row = conn.execute(
        "SELECT id FROM estudiantes WHERE LOWER(TRIM(nombre))=? AND LOWER(TRIM(apellido))=? LIMIT 1",
        (nombre.strip().lower(), apellido.strip().lower())
    ).fetchone()
    return row["id"] if row else None


# ── ENDPOINTS ─────────────────────────────────────────────────────────────────

@expediente_bp.route("/api/expediente/grados")
@login_required
@_exp_required
def listar_grados():
    """Devuelve los grados válidos según el año escolar."""
    anio = request.args.get("anio", _anio_escolar_actual())
    sistema = _detectar_sistema(anio)
    grados = GRADOS_BACHILLERATO if sistema == "bachillerato" else GRADOS_SECUNDARIA
    return jsonify({"sistema": sistema, "grados": grados})


@expediente_bp.route("/api/expediente/buscar")
@login_required
@_exp_required
@rate_limited(max_calls=40, window=60)
def buscar_expediente():
    """Búsqueda por cédula o nombre+apellido. Devuelve todos los registros del estudiante."""
    cedula  = request.args.get("cedula", "").strip()
    nombre  = request.args.get("nombre", "").strip()
    apellido= request.args.get("apellido", "").strip()

    if not cedula and not (nombre or apellido):
        return jsonify({"error": "Proporciona cédula o nombre/apellido"}), 400

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        if cedula:
            rows = conn.execute(
                "SELECT * FROM expedientes_historicos WHERE cedula=? ORDER BY anio_escolar ASC",
                (cedula,)
            ).fetchall()
        else:
            q = f"%{nombre.lower()}%"
            qa = f"%{apellido.lower()}%"
            sql = """SELECT * FROM expedientes_historicos
                     WHERE (? = '' OR LOWER(nombre) LIKE ?)
                       AND (? = '' OR LOWER(apellido) LIKE ?)
                     ORDER BY apellido, nombre, anio_escolar ASC"""
            rows = conn.execute(sql, (nombre, q, apellido, qa)).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        try:
            d["materias"] = _json.loads(d.get("materias_json") or "[]")
        except Exception:
            d["materias"] = []
        result.append(d)

    return jsonify(result)


@expediente_bp.route("/api/expediente/estudiante/<int:est_id>")
@login_required
@_exp_required
def historial_por_estudiante(est_id):
    """Devuelve el historial completo de un estudiante vinculado + notas actuales.
    RLS: coordinadores de ciclo solo ven estudiantes de su ciclo.
    """
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        _rls.verificar_acceso_estudiante(conn, est_id)  # RLS — 403 si fuera de ciclo
        est = conn.execute("SELECT * FROM estudiantes WHERE id=?", (est_id,)).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        historico = conn.execute(
            "SELECT * FROM expedientes_historicos WHERE estudiante_id=? ORDER BY anio_escolar ASC",
            (est_id,)
        ).fetchall()

        actual = conn.execute(
            """SELECT materia, p1, p2, p3, p4, promedio, tipo
               FROM materias_calificaciones WHERE estudiante_id=? ORDER BY materia""",
            (est_id,)
        ).fetchall()

    hist_list = []
    for r in historico:
        d = dict(r)
        try:
            d["materias"] = _json.loads(d.get("materias_json") or "[]")
        except Exception:
            d["materias"] = []
        hist_list.append(d)

    return jsonify({
        "estudiante": dict(est),
        "historico": hist_list,
        "actual": [dict(r) for r in actual],
        "anio_actual": _anio_escolar_actual(),
    })


@expediente_bp.route("/api/expediente", methods=["POST"])
@login_required
@_exp_required
@rate_limited(max_calls=30, window=60)
def crear_expediente():
    """Crea un registro histórico manualmente."""
    u = get_usuario()
    d = request.get_json(silent=True) or {}

    nombre   = d.get("nombre", "").strip()
    apellido = d.get("apellido", "").strip()
    anio     = d.get("anio_escolar", "").strip()
    grado    = d.get("grado", "").strip()

    if not nombre or not apellido or not anio or not grado:
        return jsonify({"error": "nombre, apellido, anio_escolar y grado son requeridos"}), 400

    sistema = _detectar_sistema(anio)
    grado   = _normalizar_grado(grado, sistema)
    materias = d.get("materias", [])
    if not isinstance(materias, list):
        materias = []

    prom_vals = [m.get("promedio", 0) for m in materias if m.get("promedio")]
    prom_general = round(sum(prom_vals) / len(prom_vals), 2) if prom_vals else None

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        cedula = d.get("cedula", "").strip() or None
        est_id = _buscar_estudiante_vinculable(conn, cedula, nombre, apellido)

        conn.execute("""
            INSERT INTO expedientes_historicos
            (cedula, nombre, apellido, fecha_nacimiento, estudiante_id,
             anio_escolar, grado, sistema_educativo, seccion, mencion,
             centro_educativo, es_externo, materias_json, promedio_general,
             condicion, fuente, digitado_por, observaciones)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            cedula, nombre, apellido,
            d.get("fecha_nacimiento", "").strip() or None,
            est_id, anio, grado, sistema,
            d.get("seccion", "").strip() or None,
            d.get("mencion", "").strip() or None,
            d.get("centro_educativo", "Centro Educativo en Artes Benito Juárez"),
            1 if d.get("es_externo") else 0,
            _json.dumps(materias, ensure_ascii=False),
            prom_general,
            d.get("condicion", "").strip() or None,
            "manual", u["id"],
            d.get("observaciones", "").strip() or None,
        ))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()

    _audit("expediente_creado", f"{nombre} {apellido} — {anio} — {grado}", "expedientes_historicos", new_id)
    return jsonify({"ok": True, "id": new_id, "vinculado": est_id is not None})


@expediente_bp.route("/api/expediente/<int:exp_id>", methods=["PATCH"])
@login_required
@_exp_required
def editar_expediente(exp_id):
    """Edita un registro histórico existente."""
    u = get_usuario()
    d = request.get_json(silent=True) or {}
    permitidos = ["nombre", "apellido", "cedula", "fecha_nacimiento", "anio_escolar",
                  "grado", "seccion", "mencion", "centro_educativo", "es_externo",
                  "materias_json", "promedio_general", "condicion", "observaciones"]
    campos, vals = [], []

    for k in permitidos:
        if k in d:
            if k == "materias_json" and isinstance(d[k], list):
                campos.append(f"{k}=?")
                vals.append(_json.dumps(d[k], ensure_ascii=False))
            else:
                campos.append(f"{k}=?")
                vals.append(d[k])

    if not campos:
        return jsonify({"error": "Nada que actualizar"}), 400

    vals.append(exp_id)
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute(f"UPDATE expedientes_historicos SET {','.join(campos)} WHERE id=?", vals)
        conn.commit()

    return jsonify({"ok": True})


@expediente_bp.route("/api/expediente/<int:exp_id>/vincular", methods=["POST"])
@login_required
@_exp_required
def vincular_expediente(exp_id):
    """Vincula un expediente histórico a un estudiante activo."""
    d = request.get_json(silent=True) or {}
    est_id = d.get("estudiante_id")
    if not est_id:
        return jsonify({"error": "estudiante_id requerido"}), 400

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("UPDATE expedientes_historicos SET estudiante_id=? WHERE id=?", (est_id, exp_id))
        conn.commit()

    return jsonify({"ok": True})


@expediente_bp.route("/api/expediente/<int:exp_id>", methods=["DELETE"])
@login_required
def eliminar_expediente(exp_id):
    """Elimina un registro — solo directora y coordinador_general."""
    rol = _normalizar_rol(session.get("rol", ""))
    if rol not in {"directora", "coordinador_general"}:
        return jsonify({"error": "Sin permisos"}), 403

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("DELETE FROM expedientes_historicos WHERE id=?", (exp_id,))
        conn.commit()

    return jsonify({"ok": True})


# ── IMPORTACIÓN MASIVA DESDE EXCEL ────────────────────────────────────────────

@expediente_bp.route("/api/expediente/importar-excel", methods=["POST"])
@login_required
@_exp_required
@rate_limited(max_calls=5, window=60)
def importar_expediente_excel():
    """
    Importa registros históricos desde Excel.
    Columnas: cedula(opt), nombre*, apellido*, fecha_nacimiento(opt),
              anio_escolar*, grado*, seccion(opt), mencion(opt),
              materia*, p1, p2, p3, p4, promedio(opt), tipo(opt),
              condicion(opt), centro_educativo(opt), es_externo(opt)
    Una fila por materia. El sistema agrupa por (cedula|nombre+apellido + anio_escolar).
    """
    u = get_usuario()
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"error": "No se recibió archivo"}), 400

    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls", "csv"}:
        return jsonify({"error": "Solo .xlsx, .xls o .csv"}), 400

    try:
        data = archivo.read()
        if not _validar_magic_excel(data, ext):
            return jsonify({"error": "El archivo no es un Excel válido"}), 400

        from io import BytesIO as _BIO
        import openpyxl

        if ext == "csv":
            import csv, io
            texto = data.decode("utf-8-sig", errors="ignore")
            filas = list(csv.DictReader(io.StringIO(texto)))
        else:
            wb = openpyxl.load_workbook(_BIO(data), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            filas = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)
                     if any(v is not None for v in row)]
    except Exception as e:
        return jsonify({"error": f"Error leyendo archivo: {e}"}), 400

    def _g(row, *keys):
        for k in keys:
            for rk in row:
                if rk and k in str(rk).lower():
                    v = row[rk]
                    return str(v).strip() if v not in (None, "") else ""
        return ""

    def _f(v):
        try:
            return round(float(str(v).replace(",", "").strip()), 2)
        except Exception:
            return None

    # Agrupar filas por (cedula|nombre+apellido, anio_escolar)
    grupos = {}
    errores = []
    for i, row in enumerate(filas, 2):
        cedula   = _g(row, "cedula", "ced")
        nombre   = _g(row, "nombre", "name")
        apellido = _g(row, "apellido", "lastname", "surname")
        anio     = _g(row, "anio_escolar", "año", "year", "periodo")
        grado    = _g(row, "grado", "grade", "nivel")
        materia  = _g(row, "materia", "asignatura", "subject")

        if not nombre or not apellido or not anio or not grado:
            errores.append(f"Fila {i}: faltan nombre, apellido, año o grado")
            continue
        if not materia:
            errores.append(f"Fila {i}: falta materia")
            continue

        clave = (cedula or f"{nombre.lower()}_{apellido.lower()}", anio)
        if clave not in grupos:
            grupos[clave] = {
                "cedula": cedula or None,
                "nombre": nombre, "apellido": apellido,
                "fecha_nacimiento": _g(row, "nacimiento", "birth"),
                "anio_escolar": anio,
                "grado": grado,
                "seccion": _g(row, "seccion", "section"),
                "mencion": _g(row, "mencion"),
                "centro_educativo": _g(row, "centro", "school") or "Centro Educativo en Artes Benito Juárez",
                "es_externo": 1 if _g(row, "externo", "external").lower() in ("1","si","sí","yes","true") else 0,
                "condicion": _g(row, "condicion", "condition"),
                "materias": [],
            }

        p1 = _f(_g(row, "p1", "periodo1", "primer"))
        p2 = _f(_g(row, "p2", "periodo2", "segundo"))
        p3 = _f(_g(row, "p3", "periodo3", "tercer"))
        p4 = _f(_g(row, "p4", "periodo4", "cuarto"))
        prom_raw = _g(row, "promedio", "average", "final")
        prom = _f(prom_raw) if prom_raw else None
        if prom is None:
            nums = [x for x in [p1, p2, p3, p4] if x is not None]
            prom = round(sum(nums) / len(nums), 2) if nums else None

        grupos[clave]["materias"].append({
            "materia": materia,
            "p1": p1, "p2": p2, "p3": p3, "p4": p4,
            "promedio": prom,
            "tipo": _g(row, "tipo", "type") or "académico",
        })

    # Insertar en BD
    importados = 0
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        for (_, anio), datos in grupos.items():
            try:
                sistema = _detectar_sistema(datos["anio_escolar"])
                grado_n = _normalizar_grado(datos["grado"], sistema)
                prom_vals = [m["promedio"] for m in datos["materias"] if m.get("promedio")]
                prom_gen  = round(sum(prom_vals) / len(prom_vals), 2) if prom_vals else None
                est_id    = _buscar_estudiante_vinculable(
                    conn, datos["cedula"], datos["nombre"], datos["apellido"]
                )

                # Verificar si ya existe para no duplicar
                existe = conn.execute("""
                    SELECT id FROM expedientes_historicos
                    WHERE LOWER(TRIM(nombre))=? AND LOWER(TRIM(apellido))=? AND anio_escolar=?
                """, (datos["nombre"].lower(), datos["apellido"].lower(), datos["anio_escolar"])
                ).fetchone()

                if existe:
                    # Actualizar materias si ya existe
                    conn.execute("""
                        UPDATE expedientes_historicos
                        SET materias_json=?, promedio_general=?
                        WHERE id=?
                    """, (_json.dumps(datos["materias"], ensure_ascii=False), prom_gen, existe["id"]))
                else:
                    conn.execute("""
                        INSERT INTO expedientes_historicos
                        (cedula, nombre, apellido, fecha_nacimiento, estudiante_id,
                         anio_escolar, grado, sistema_educativo, seccion, mencion,
                         centro_educativo, es_externo, materias_json, promedio_general,
                         condicion, fuente, digitado_por)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        datos["cedula"], datos["nombre"], datos["apellido"],
                        datos["fecha_nacimiento"] or None, est_id,
                        datos["anio_escolar"], grado_n, sistema,
                        datos["seccion"] or None, datos["mencion"] or None,
                        datos["centro_educativo"], datos["es_externo"],
                        _json.dumps(datos["materias"], ensure_ascii=False),
                        prom_gen, datos["condicion"] or None, "excel", u["id"],
                    ))
                importados += 1
            except Exception as ex:
                errores.append(f"{datos.get('nombre','')} {datos.get('apellido','')} {anio}: {ex}")

        conn.commit()

    return jsonify({
        "ok": True,
        "importados": importados,
        "errores": errores[:15],
        "total_filas": len(filas),
    })


# ── PLANTILLA EXCEL PARA DESCARGAR ───────────────────────────────────────────

@expediente_bp.route("/api/expediente/plantilla-excel")
@login_required
@_exp_required
def descargar_plantilla():
    """Genera y descarga la plantilla Excel para importar expedientes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl no instalado"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    encabezados = [
        "cedula", "nombre", "apellido", "fecha_nacimiento",
        "anio_escolar", "grado", "seccion", "mencion",
        "centro_educativo", "es_externo",
        "materia", "p1", "p2", "p3", "p4", "promedio", "tipo",
        "condicion", "observaciones",
    ]
    descripciones = [
        "Opcional (si existe)", "Requerido", "Requerido", "Opcional (YYYY-MM-DD)",
        "Requerido (ej: 2015-2016)", "Requerido (ej: 3ro Bachiller)", "Opcional", "Opcional",
        "Dejar en blanco si es este centro", "0=este centro, 1=externo",
        "Requerido", "Nota P1 (0-100)", "Nota P2", "Nota P3", "Nota P4",
        "Se calcula si está vacío", "académico o técnico",
        "Aprobado / Reprobado", "Opcional",
    ]

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    desc_fill = PatternFill("solid", fgColor="E8F0FE")
    desc_font = Font(color="444444", italic=True, size=9)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, (h, desc) in enumerate(zip(encabezados, descripciones), 1):
        c1 = ws.cell(row=1, column=col, value=h.upper())
        c1.font = hdr_font; c1.fill = hdr_fill
        c1.alignment = Alignment(horizontal="center"); c1.border = thin
        c2 = ws.cell(row=2, column=col, value=desc)
        c2.font = desc_font; c2.fill = desc_fill
        c2.alignment = Alignment(wrap_text=True); c2.border = thin

    # Ejemplos — bachillerato y secundaria
    ejemplos = [
        ["00112345678", "Juan", "Pérez", "1998-03-15", "2000-2001",
         "1ro Bachiller", "A", "", "", "0",
         "Lengua Española", 85, 78, 90, 82, "", "académico", "Aprobado", ""],
        ["00112345678", "Juan", "Pérez", "1998-03-15", "2000-2001",
         "1ro Bachiller", "A", "", "", "0",
         "Matemáticas", 72, 68, 75, 80, "", "académico", "Aprobado", ""],
        ["", "María", "García", "2005-07-20", "2019-2020",
         "4to", "B", "Multimedia", "", "0",
         "Fotografía", 90, 88, 92, 95, "", "técnico", "Aprobado", ""],
    ]
    data_font = Font(size=10)
    for r, row in enumerate(ejemplos, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font; cell.border = thin

    ws.freeze_panes = "A3"
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[col[0].column_letter].width = 18

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR EXPEDIENTES HISTÓRICOS"],
        [""],
        ["1. Una fila por MATERIA por estudiante por año escolar"],
        ["2. Si un estudiante tiene 8 materias en un año → 8 filas con los mismos datos personales"],
        ["3. El sistema agrupa automáticamente por (cedula o nombre+apellido) + anio_escolar"],
        ["4. Si el estudiante ya existe en el sistema (tiene cédula), se vincula automáticamente"],
        [""],
        ["GRADOS VÁLIDOS PRE-2016-2017 (Bachillerato):"],
        ["  1ro Bachiller, 2do Bachiller, 3ro Bachiller, 4to Bachiller"],
        [""],
        ["GRADOS VÁLIDOS DESDE 2016-2017 (Secundaria):"],
        ["  1ro, 2do, 3ro, 4to, 5to, 6to"],
        [""],
        ["FORMATO AÑO ESCOLAR: AAAA-AAAA  (ej: 2015-2016, 2016-2017, 2023-2024)"],
        [""],
        ["CONDICIÓN: Aprobado | Reprobado | Promovido | En proceso"],
        ["TIPO MATERIA: académico | técnico"],
        ["ES_EXTERNO: 0 = este centro | 1 = viene de otro centro"],
    ]
    for r, row in enumerate(instrucciones, 1):
        ws2.cell(row=r, column=1, value=row[0])
    ws2.column_dimensions["A"].width = 70

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name="plantilla_expedientes_historicos.xlsx",
        as_attachment=True,
    )
