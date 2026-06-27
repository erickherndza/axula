# -*- coding: utf-8 -*-
"""Blueprint: ocr — Escáner y digitalización de documentos."""

import base64
import json
import logging
import os
import re
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint, jsonify, redirect, render_template,
    request, send_file, session,
)

from core.auth import get_usuario, login_required, _normalizar_rol, rate_limited
from core.database import get_db, cache_bust
from core.ocr import escanear_documento, TIPOS_DOCUMENTO
from core.helpers import _validar_magic_imagen

logger = logging.getLogger("axula")

ocr_bp = Blueprint("ocr_bp", __name__)


# Extensiones permitidas para subir
_MIME_MAP = {
    "jpg":  "image/jpeg",
    "jpeg": "image/jpeg",
    "png":  "image/png",
    "webp": "image/webp",
    "gif":  "image/gif",
    "bmp":  "image/bmp",
}
MAX_SIZE_MB = 10


# ── PÁGINA PRINCIPAL ─────────────────────────────────────────────────────────

@ocr_bp.route("/escaner")
@login_required
def pagina_escaner():
    u = get_usuario()
    return render_template("escaner.html", usuario=u, current_user=u,
                           tipos_documento=TIPOS_DOCUMENTO)


# ── API: ESCANEAR IMAGEN ─────────────────────────────────────────────────────

@ocr_bp.route("/api/ocr/escanear", methods=["POST"])
@login_required
@rate_limited(max_calls=10, window=60)   # máx 10 escaneos/minuto por IP
def api_escanear():
    """
    Recibe imagen (multipart/form-data o JSON base64), la procesa con
    Groq Vision y devuelve texto + datos estructurados.
    Guarda el escaneo en historial.
    """
    u = get_usuario()
    usuario_id = session.get("user_id")

    # ── Obtener imagen ────────────────────────────────────────────────────────
    imagen_b64 = None
    mime_type = "image/jpeg"
    nombre_archivo = "documento"

    if request.content_type and "multipart" in request.content_type:
        archivo = request.files.get("imagen")
        if not archivo or archivo.filename == "":
            return jsonify({"error": "No se recibió ninguna imagen"}), 400

        ext = archivo.filename.rsplit(".", 1)[-1].lower()
        if ext not in _MIME_MAP:
            return jsonify({"error": f"Formato no soportado: .{ext}. Usa JPG, PNG o WEBP"}), 400

        datos = archivo.read()
        if len(datos) > MAX_SIZE_MB * 1024 * 1024:
            return jsonify({"error": f"La imagen supera los {MAX_SIZE_MB} MB permitidos"}), 413
        if not _validar_magic_imagen(datos, ext):
            return jsonify({"error": "El archivo no es una imagen válida"}), 400

        imagen_b64 = base64.b64encode(datos).decode("utf-8")
        mime_type = _MIME_MAP[ext]
        nombre_archivo = archivo.filename

    elif request.is_json:
        body = request.get_json(force=True)
        imagen_b64 = body.get("imagen_base64", "").strip()
        mime_type = body.get("mime_type", "image/jpeg")
        nombre_archivo = body.get("nombre_archivo", "documento")
        if not imagen_b64:
            return jsonify({"error": "Campo 'imagen_base64' requerido"}), 400
        # Limpiar prefijo data: si viene incluido
        if imagen_b64.startswith("data:"):
            partes = imagen_b64.split(",", 1)
            if len(partes) == 2:
                meta = partes[0]          # "data:image/jpeg;base64"
                imagen_b64 = partes[1]
                if "image/" in meta:
                    mime_type = meta.split(";")[0].replace("data:", "")
    else:
        return jsonify({"error": "Envía la imagen como multipart/form-data o JSON con imagen_base64"}), 400

    # ── Llamar al OCR ─────────────────────────────────────────────────────────
    try:
        resultado = escanear_documento(imagen_b64, mime_type)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        logger.error(f"[OCR] Error inesperado: {e}")
        return jsonify({"error": "Error al procesar el documento. Intenta con otra imagen."}), 500

    # ── Guardar en historial ──────────────────────────────────────────────────
    try:
        with get_db() as db:
            db.execute(
                """INSERT INTO escaneos_documentos
                   (usuario_id, tipo_documento, confianza, nombre_archivo,
                    texto_extraido, datos_json, creado_en)
                   VALUES (?, ?, ?, ?, ?, ?, datetime('now'))""",
                (
                    usuario_id,
                    resultado.get("tipo", "otro"),
                    resultado.get("confianza", "media"),
                    nombre_archivo,
                    resultado.get("texto_completo", ""),
                    json.dumps(resultado.get("datos", {}), ensure_ascii=False),
                ),
            )
            db.commit()
    except Exception as e:
        logger.warning(f"[OCR] No se pudo guardar historial: {e}")

    return jsonify({
        "ok": True,
        "tipo": resultado.get("tipo", "otro"),
        "tipo_label": TIPOS_DOCUMENTO.get(resultado.get("tipo", "otro"), "Documento"),
        "confianza": resultado.get("confianza", "media"),
        "texto_completo": resultado.get("texto_completo", ""),
        "datos": resultado.get("datos", {}),
    })


# ── API: EXPORTAR FACTURA A EXCEL ────────────────────────────────────────────

@ocr_bp.route("/api/ocr/exportar-excel", methods=["POST"])
@login_required
def api_exportar_excel():
    """Recibe datos de cualquier documento escaneado y devuelve un archivo Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl no está instalado. Ejecuta: pip install openpyxl"}), 500

    body = request.get_json(force=True) or {}
    datos = body.get("datos", {})
    texto_completo = body.get("texto_completo", "")
    # Pasar tipo como campo interno para enrutamiento
    datos["_tipo"] = body.get("tipo", "")

    try:
        return _generar_excel(datos, texto_completo)
    except Exception as exc:
        logger.error(f"[OCR Excel] Error: {exc}", exc_info=True)
        return jsonify({"error": "Error al generar el Excel. Intenta de nuevo."}), 500


def _generar_excel(datos, texto_completo):
    """Genera Excel para cualquier tipo de documento escaneado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Detectar tipo para enrutar a generador específico
    tipo = datos.get("_tipo", "")
    movimientos = datos.get("movimientos", [])
    if movimientos or datos.get("banco") or datos.get("numero_cuenta") or datos.get("saldo_inicial"):
        return _excel_estado_cuenta(datos, texto_completo)

    wb = Workbook()
    ws = wb.active

    fill_header = PatternFill("solid", fgColor="1E3A5F")
    fill_sub    = PatternFill("solid", fgColor="2E75B6")
    fill_alt    = PatternFill("solid", fgColor="EBF3FB")
    fill_total  = PatternFill("solid", fgColor="D6E4F0")
    borde = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    st_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    st_sub    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    st_label  = Font(name="Calibri", size=10, bold=True)
    st_valor  = Font(name="Calibri", size=10)
    st_total  = Font(name="Calibri", size=11, bold=True, color="1E3A5F")

    def celda(fila, col, valor, fuente=None, relleno=None, alin=None, fmt=None):
        c = ws.cell(row=fila, column=col, value=valor)
        if fuente:  c.font = fuente
        if relleno: c.fill = relleno
        if alin:    c.alignment = alin
        if fmt:     c.number_format = fmt
        c.border = borde
        return c

    def seccion(fila, titulo):
        ws.merge_cells(f"A{fila}:G{fila}")
        c = ws.cell(row=fila, column=1, value=titulo)
        c.font = st_sub; c.fill = fill_sub
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[fila].height = 22
        return fila + 1

    # Detectar si es factura para título específico
    es_factura = bool(datos.get("numero_factura") or datos.get("total") or datos.get("items"))
    titulo_doc = "COMPROBANTE DE FACTURA" if es_factura else "DOCUMENTO DIGITALIZADO"
    ws.title = "Factura" if es_factura else "Documento"

    # ── Título principal ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = titulo_doc
    c.font = st_titulo; c.fill = fill_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    fila = 2

    # ── Campos generales (todo lo que no sea items/lista) ─────────────────────
    SKIP = {"items", "condiciones_clave", "partes"}
    campos = [(k, v) for k, v in datos.items()
              if k not in SKIP and v and not isinstance(v, list)]

    if campos:
        fila = seccion(fila, "DATOS DEL DOCUMENTO")
        for k, v in campos:
            etiqueta = k.replace("_", " ").title()
            ws.merge_cells(f"A{fila}:B{fila}")
            celda(fila, 1, etiqueta, fuente=st_label, relleno=fill_alt,
                  alin=Alignment(horizontal="left", vertical="center"))
            ws.merge_cells(f"C{fila}:G{fila}")
            celda(fila, 3, str(v), fuente=st_valor,
                  alin=Alignment(horizontal="left", vertical="center"))
            fila += 1

    # ── Tabla de ítems (solo si hay) ──────────────────────────────────────────
    items = datos.get("items", [])
    if items:
        fila += 1
        fila = seccion(fila, "DETALLE DE PRODUCTOS / SERVICIOS")
        # Encabezados
        hdrs = ["#", "Descripción", "Cantidad", "Precio Unit.", "Subtotal", "", ""]
        anchos = [5, 35, 12, 18, 18, 5, 5]
        for i, (h, w) in enumerate(zip(hdrs, anchos), 1):
            c = ws.cell(row=fila, column=i, value=h)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center")
            c.border = borde
            ws.column_dimensions[get_column_letter(i)].width = w
        fila += 1
        for idx, item in enumerate(items, 1):
            rel = fill_alt if idx % 2 == 0 else None
            celda(fila, 1, idx, fuente=st_valor, relleno=rel,
                  alin=Alignment(horizontal="center"))
            celda(fila, 2, item.get("descripcion", ""), fuente=st_valor, relleno=rel)
            celda(fila, 3, _to_num(item.get("cantidad")), fuente=st_valor, relleno=rel,
                  alin=Alignment(horizontal="center"))
            celda(fila, 4, _to_num(item.get("precio_unitario")), fuente=st_valor,
                  relleno=rel, alin=Alignment(horizontal="right"), fmt='#,##0.00')
            celda(fila, 5, _to_num(item.get("subtotal")), fuente=st_valor,
                  relleno=rel, alin=Alignment(horizontal="right"), fmt='#,##0.00')
            fila += 1

        # Totales
        fila += 1
        for lbl, key in [("Subtotal","subtotal"),("ITBIS (18%)","itbis"),
                          ("Descuento","descuento"),("TOTAL","total")]:
            val = datos.get(key, "")
            if not val:
                continue
            es_tot = lbl == "TOTAL"
            ws.merge_cells(f"C{fila}:D{fila}")
            c = ws.cell(row=fila, column=3, value=lbl)
            c.font = st_total if es_tot else st_label
            if es_tot: c.fill = fill_total
            c.alignment = Alignment(horizontal="right")
            c.border = borde
            c2 = ws.cell(row=fila, column=5, value=_to_num(val))
            c2.font = st_total if es_tot else st_valor
            if es_tot: c2.fill = fill_total
            c2.alignment = Alignment(horizontal="right")
            c2.number_format = '#,##0.00'
            c2.border = borde
            if es_tot: ws.row_dimensions[fila].height = 20
            fila += 1

    # ── Texto completo ────────────────────────────────────────────────────────
    if texto_completo:
        fila += 2
        fila = seccion(fila, "TEXTO COMPLETO EXTRAÍDO")
        ws.merge_cells(f"A{fila}:G{fila + 4}")
        c = ws.cell(row=fila, column=1, value=texto_completo)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(name="Calibri", size=9)
        ws.row_dimensions[fila].height = 90

    # ── Anchos por defecto si no hubo items ──────────────────────────────────
    if not items:
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        for col in ["C","D","E","F","G"]:
            ws.column_dimensions[col].width = 15

    # ── Guardar y devolver ────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_base = (datos.get("numero_factura") or datos.get("nombre_beneficiario")
                   or datos.get("nombre") or "documento") or "documento"
    nombre_base = re.sub(r"[^\w\-]", "_", str(nombre_base))[:40]
    nombre_archivo = f"documento_{nombre_base}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo,
    )


# ── API: MARCAR ESCANEO COMO "EN PERFIL" ────────────────────────────────────

@ocr_bp.route("/api/ocr/guardar-perfil", methods=["POST"])
@login_required
def api_guardar_perfil():
    """Marca el escaneo más reciente del usuario como guardado en perfil."""
    usuario_id = session.get("user_id")

    with get_db() as db:
        row = db.execute(
            "SELECT id FROM escaneos_documentos WHERE usuario_id=? ORDER BY creado_en DESC LIMIT 1",
            (usuario_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "No se encontró el escaneo"}), 404
        db.execute(
            "UPDATE escaneos_documentos SET en_perfil=1, metodo_uso='perfil' WHERE id=?",
            (row["id"],)
        )
        db.commit()
    return jsonify({"ok": True})


# ── API: CATEGORIZAR Y ARCHIVAR (secretaria / coordinador / directora) ────────

CATEGORIAS_ARCHIVO = {
    "acta_nacimiento":      "Acta de Nacimiento",
    "licencia":             "Licencia de Personal",
    "permiso":              "Permiso Institucional",
    "certificado":          "Certificado",
    "documento_institucional": "Documento Institucional",
    "factura":              "Factura / Recibo",
    "contrato":             "Contrato / Acuerdo",
    "expediente_personal":  "Expediente Personal",
    "general":              "General",
}

@ocr_bp.route("/api/ocr/categorizar", methods=["POST"])
@login_required
def api_categorizar():
    """
    Categoriza y archiva el escaneo más reciente.
    Opcionalmente asocia a un miembro del personal.
    """
    usuario_id = session.get("user_id")
    body = request.get_json(silent=True) or {}
    categoria   = body.get("categoria", "general")
    personal_id = body.get("personal_id")  # opcional
    etiqueta    = body.get("etiqueta", "")  # nombre libre adicional

    if categoria not in CATEGORIAS_ARCHIVO:
        categoria = "general"

    with get_db() as db:
        row = db.execute(
            "SELECT id FROM escaneos_documentos WHERE usuario_id=? ORDER BY creado_en DESC LIMIT 1",
            (usuario_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "No se encontró el escaneo"}), 404

        db.execute(
            """UPDATE escaneos_documentos
               SET categoria=?, personal_id=?, metodo_uso='archivo',
                   nombre_archivo=COALESCE(NULLIF(nombre_archivo,''), ?)
               WHERE id=?""",
            (categoria, personal_id, etiqueta or categoria, row["id"])
        )
        db.commit()
    return jsonify({"ok": True, "categoria": CATEGORIAS_ARCHIVO[categoria]})


@ocr_bp.route("/api/ocr/personal-lista")
@login_required
def api_personal_lista():
    """Lista de personal activo para asociar documentos."""
    with get_db() as db:
        rows = db.execute(
            "SELECT id, nombre, rol FROM usuarios WHERE activo=1 ORDER BY nombre"
        ).fetchall()
    return jsonify([{"id": r["id"], "nombre": r["nombre"], "rol": r["rol"]} for r in rows])


# ── API: PASE DE LISTA DESDE ESCANEO (profesor) ───────────────────────────────

_MAX_USOS_SEMANA = 3   # máximo de pases de lista escaneados por semana

@ocr_bp.route("/api/ocr/pase-lista-preview", methods=["POST"])
@login_required
def api_pase_lista_preview():
    """
    Extrae nombres de un texto escaneado y los cruza con los estudiantes
    asignados al profesor. Devuelve: coincidencias, sin_match, usos_semana.
    """
    u          = get_usuario()
    prof_id    = u.get("id")
    body       = request.get_json(silent=True) or {}
    texto      = body.get("texto_completo", "")

    with get_db() as db:
        # ── Límite semanal ────────────────────────────────────────────────────
        usos = db.execute(
            """SELECT COUNT(*) as n FROM asistencia
               WHERE profesor_id=? AND metodo='escaner'
                 AND fecha >= date('now','-7 days')""",
            (prof_id,)
        ).fetchone()
        usos_semana = usos["n"] if usos else 0

        if usos_semana >= _MAX_USOS_SEMANA:
            return jsonify({
                "ok": False,
                "limite": True,
                "usos_semana": usos_semana,
                "max": _MAX_USOS_SEMANA,
                "mensaje": f"Alcanzaste el límite de {_MAX_USOS_SEMANA} pases de lista escaneados por semana. "
                           "Usa el pase de lista del sistema para continuar.",
            })

        # ── Estudiantes del profesor ──────────────────────────────────────────
        ids_notas = [r[0] for r in db.execute(
            "SELECT DISTINCT estudiante_id FROM calificaciones_periodo WHERE profesor_id=?",
            (prof_id,)
        ).fetchall()]
        ids_asist = [r[0] for r in db.execute(
            "SELECT DISTINCT estudiante_id FROM asistencia WHERE profesor_id=?",
            (prof_id,)
        ).fetchall()]
        all_ids = list(set(ids_notas + ids_asist))

        if not all_ids:
            return jsonify({"ok": False, "mensaje": "No tienes estudiantes asignados en el sistema."})

        ph = ",".join("?" * len(all_ids))
        estudiantes = db.execute(
            f"SELECT id, nombre, apellido, grado, curso FROM estudiantes WHERE id IN ({ph})",
            all_ids
        ).fetchall()

    # ── Extraer nombres del texto escaneado ───────────────────────────────────
    # Buscar líneas que parezcan nombres (2+ palabras, sin números dominantes)
    import re as _re
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    candidatos = []
    for linea in lineas:
        # Filtrar líneas que parezcan encabezados, fechas o números
        if _re.search(r'\d{2,}', linea):     continue  # tiene números largos
        if len(linea) < 4 or len(linea) > 60: continue
        palabras = linea.split()
        if len(palabras) < 2:                 continue
        # Capitalizar para normalizar
        candidatos.append(" ".join(w.capitalize() for w in palabras))

    # ── Cruzar con estudiantes ────────────────────────────────────────────────
    def _similitud(a, b):
        """Similitud simple: cuántas palabras de 'a' aparecen en 'b'."""
        wa = set(a.lower().split())
        wb = set(b.lower().split())
        comunes = wa & wb
        return len(comunes) / max(len(wa), 1)

    coincidencias = []
    sin_match     = []
    usados        = set()

    for cand in candidatos:
        mejor = None
        mejor_score = 0
        for est in estudiantes:
            nombre_completo = f"{est['nombre']} {est['apellido']}"
            score = _similitud(cand, nombre_completo)
            if score > mejor_score:
                mejor_score = score
                mejor = est
        if mejor and mejor_score >= 0.5 and mejor["id"] not in usados:
            usados.add(mejor["id"])
            coincidencias.append({
                "id":       mejor["id"],
                "nombre":   mejor["nombre"],
                "apellido": mejor["apellido"],
                "grado":    mejor["grado"],
                "curso":    mejor["curso"],
                "score":    round(mejor_score, 2),
                "texto_original": cand,
            })
        else:
            sin_match.append(cand)

    return jsonify({
        "ok": True,
        "limite": False,
        "usos_semana": usos_semana,
        "max": _MAX_USOS_SEMANA,
        "coincidencias": coincidencias,
        "sin_match": sin_match[:20],
    })


@ocr_bp.route("/api/ocr/registrar-asistencia-escaner", methods=["POST"])
@login_required
def api_registrar_asistencia_escaner():
    """
    Registra asistencia basada en escaneo. Marca metodo='escaner'.
    Solo profesores. Límite: _MAX_USOS_SEMANA por semana.
    """
    u       = get_usuario()
    prof_id = u.get("id")
    rol     = _normalizar_rol(u.get("rol", ""))

    if rol != "profesor":
        return jsonify({"error": "Solo los profesores pueden usar esta función"}), 403

    body       = request.get_json(silent=True) or {}
    presentes  = body.get("presentes", [])    # lista de estudiante_id
    ausentes   = body.get("ausentes", [])
    materia    = body.get("materia", "").strip()
    fecha      = body.get("fecha", datetime.now().strftime("%Y-%m-%d"))
    periodo    = body.get("periodo", 1)

    if not materia:
        return jsonify({"error": "Debes indicar la materia"}), 400
    if not presentes and not ausentes:
        return jsonify({"error": "No hay estudiantes para registrar"}), 400

    with get_db() as db:
        # Verificar límite semanal
        usos = db.execute(
            """SELECT COUNT(*) as n FROM asistencia
               WHERE profesor_id=? AND metodo='escaner'
                 AND fecha >= date('now','-7 days')""",
            (prof_id,)
        ).fetchone()
        if (usos["n"] if usos else 0) >= _MAX_USOS_SEMANA:
            return jsonify({"error": f"Límite de {_MAX_USOS_SEMANA} pases escaneados por semana alcanzado."}), 429

        # Insertar registros
        total = 0
        for est_id in presentes:
            db.execute(
                """INSERT OR REPLACE INTO asistencia
                   (estudiante_id, profesor_id, materia, fecha, periodo, estado, metodo)
                   VALUES (?,?,?,?,?,'P','escaner')""",
                (est_id, prof_id, materia, fecha, periodo)
            )
            total += 1
        for est_id in ausentes:
            db.execute(
                """INSERT OR REPLACE INTO asistencia
                   (estudiante_id, profesor_id, materia, fecha, periodo, estado, metodo)
                   VALUES (?,?,?,?,?,'A','escaner')""",
                (est_id, prof_id, materia, fecha, periodo)
            )
            total += 1
        db.commit()

    return jsonify({"ok": True, "registrados": total})


def _excel_estado_cuenta(datos, texto_completo):
    """Genera Excel específico para estados de cuenta / movimientos bancarios."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # ── Estilos ───────────────────────────────────────────────────────────────
    fill_header  = PatternFill("solid", fgColor="1E3A5F")
    fill_sub     = PatternFill("solid", fgColor="2E75B6")
    fill_alt     = PatternFill("solid", fgColor="EBF3FB")
    fill_debito  = PatternFill("solid", fgColor="FDE7E9")   # rojo claro
    fill_credito = PatternFill("solid", fgColor="DDF3D9")   # verde claro
    fill_saldo   = PatternFill("solid", fgColor="D6E4F0")
    borde = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    st_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    st_sub    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    st_label  = Font(name="Calibri", size=10, bold=True)
    st_valor  = Font(name="Calibri", size=10)
    st_num    = Font(name="Calibri", size=10)
    st_total  = Font(name="Calibri", size=10, bold=True, color="1E3A5F")
    CENTRO    = Alignment(horizontal="center", vertical="center")
    IZQ       = Alignment(horizontal="left",   vertical="center")
    DER       = Alignment(horizontal="right",  vertical="center")
    NUM_FMT   = '#,##0.00'

    def c(fila, col, valor, fuente=None, relleno=None, alin=None, fmt=None):
        cell = ws.cell(row=fila, column=col, value=valor)
        if fuente:  cell.font = fuente
        if relleno: cell.fill = relleno
        if alin:    cell.alignment = alin
        if fmt:     cell.number_format = fmt
        cell.border = borde
        return cell

    # ── Anchos de columnas (7 columnas) ──────────────────────────────────────
    anchos = {"A":13, "B":14, "C":40, "D":16, "E":16, "F":16, "G":16}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "ESTADO DE CUENTA / MOVIMIENTOS"
    cell.font = st_titulo; cell.fill = fill_header
    cell.alignment = CENTRO
    ws.row_dimensions[1].height = 28

    # ── Info del banco y cuenta ───────────────────────────────────────────────
    fila = 2
    info = [
        ("Banco",          datos.get("banco", "")),
        ("Titular",        datos.get("titular", "")),
        ("N° de Cuenta",   datos.get("numero_cuenta", "")),
        ("Tipo de Cuenta", datos.get("tipo_cuenta", "")),
        ("Período",        f"{datos.get('periodo_inicio','')} — {datos.get('periodo_fin','')}".strip(" —")),
        ("Saldo Inicial",  datos.get("saldo_inicial", "")),
        ("Saldo Final",    datos.get("saldo_final", "")),
    ]
    for etiqueta, valor in info:
        if not valor:
            continue
        ws.merge_cells(f"A{fila}:B{fila}")
        c(fila, 1, etiqueta, fuente=st_label, relleno=fill_alt, alin=IZQ)
        ws.merge_cells(f"C{fila}:F{fila}")
        c(fila, 3, str(valor), fuente=st_valor, alin=IZQ)
        fila += 1

    # ── Encabezado de tabla de movimientos ───────────────────────────────────
    fila += 1
    ws.merge_cells(f"A{fila}:F{fila}")
    cell = ws.cell(row=fila, column=1, value="DETALLE DE MOVIMIENTOS")
    cell.font = st_sub; cell.fill = fill_sub
    cell.alignment = CENTRO
    ws.row_dimensions[fila].height = 20
    fila += 1

    hdrs = ["Fecha", "Referencia", "Descripción / Concepto", "Tipo", "Débito", "Crédito", "Saldo"]
    for i, h in enumerate(hdrs, 1):
        cell = ws.cell(row=fila, column=i, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = CENTRO
        cell.border = borde
    fila += 1

    # ── Filas de movimientos ──────────────────────────────────────────────────
    movimientos = datos.get("movimientos", [])
    if not movimientos:
        ws.merge_cells(f"A{fila}:G{fila}")
        cell = ws.cell(row=fila, column=1,
                       value="No se detectaron movimientos individuales. Ver texto completo al final.")
        cell.font = Font(name="Calibri", size=9, italic=True, color="605E5C")
        cell.alignment = CENTRO
        fila += 1
    else:
        for idx, mov in enumerate(movimientos):
            debito  = _to_num(mov.get("debito"))
            credito = _to_num(mov.get("credito"))
            saldo   = _to_num(mov.get("saldo"))

            if isinstance(debito, float) and debito and not credito:
                fill_fila = fill_debito
            elif isinstance(credito, float) and credito and not debito:
                fill_fila = fill_credito
            else:
                fill_fila = fill_alt if idx % 2 == 0 else None

            c(fila, 1, mov.get("fecha", ""),        fuente=st_num,   relleno=fill_fila, alin=CENTRO)
            c(fila, 2, mov.get("referencia", ""),   fuente=st_num,   relleno=fill_fila, alin=CENTRO)
            c(fila, 3, mov.get("descripcion", ""),  fuente=st_valor, relleno=fill_fila, alin=IZQ)
            c(fila, 4, mov.get("tipo", ""),         fuente=st_num,   relleno=fill_fila, alin=CENTRO)
            c(fila, 5, debito,  fuente=st_num, relleno=fill_fila, alin=DER, fmt=NUM_FMT)
            c(fila, 6, credito, fuente=st_num, relleno=fill_fila, alin=DER, fmt=NUM_FMT)
            c(fila, 7, saldo,   fuente=st_num, relleno=fill_fila, alin=DER, fmt=NUM_FMT)
            fila += 1

    # ── Totales ───────────────────────────────────────────────────────────────
    if movimientos:
        fila += 1
        ws.merge_cells(f"A{fila}:D{fila}")
        cell = ws.cell(row=fila, column=1, value="TOTALES")
        cell.font = st_total; cell.fill = fill_saldo
        cell.alignment = DER; cell.border = borde

        total_deb  = sum((_to_num(m.get("debito"))  or 0) for m in movimientos if isinstance(_to_num(m.get("debito")), float))
        total_cred = sum((_to_num(m.get("credito")) or 0) for m in movimientos if isinstance(_to_num(m.get("credito")), float))
        c(fila, 5, total_deb  or None, fuente=st_total, relleno=fill_saldo, alin=DER, fmt=NUM_FMT)
        c(fila, 6, total_cred or None, fuente=st_total, relleno=fill_saldo, alin=DER, fmt=NUM_FMT)
        c(fila, 7, None, relleno=fill_saldo)
        fila += 1

    # ── Texto completo ────────────────────────────────────────────────────────
    if texto_completo:
        fila += 2
        ws.merge_cells(f"A{fila}:G{fila}")
        cell = ws.cell(row=fila, column=1, value="TEXTO COMPLETO EXTRAÍDO DEL DOCUMENTO")
        cell.font = st_sub; cell.fill = fill_sub
        cell.alignment = CENTRO
        fila += 1
        ws.merge_cells(f"A{fila}:G{fila + 10}")
        cell = ws.cell(row=fila, column=1, value=texto_completo)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Calibri", size=9)
        ws.row_dimensions[fila].height = 120

    # ── Guardar ───────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = datos.get("numero_cuenta") or datos.get("banco") or "estado_cuenta"
    nombre = re.sub(r"[^\w\-]", "_", str(nombre))[:40]
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"movimientos_{nombre}.xlsx",
    )


def _to_num(valor):
    """Convierte string numérico a float, o devuelve el valor si ya es número."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return valor
    try:
        return float(str(valor).replace(",", "").replace("$", "").replace("RD$", "").strip())
    except (ValueError, TypeError):
        return str(valor)


# ── API: HISTORIAL ───────────────────────────────────────────────────────────

@ocr_bp.route("/api/ocr/historial")
@login_required
def api_historial():
    usuario_id = session.get("user_id")
    u = get_usuario()
    rol = _normalizar_rol(u.get("rol", ""))

    # Admins y coordinadores ven todo; los demás, solo lo suyo
    roles_admin = {"directora", "coordinador_general", "coordinador_primer_ciclo",
                   "coordinador_segundo_ciclo", "secretaria", "secretaria_docente"}

    with get_db() as db:
        if rol in roles_admin:
            rows = db.execute(
                """SELECT e.id, e.usuario_id, u.nombre as usuario_nombre,
                          e.tipo_documento, e.confianza, e.nombre_archivo,
                          e.texto_extraido, e.datos_json, e.creado_en
                   FROM escaneos_documentos e
                   LEFT JOIN usuarios u ON u.id = e.usuario_id
                   ORDER BY e.creado_en DESC LIMIT 200"""
            ).fetchall()
        else:
            rows = db.execute(
                """SELECT e.id, e.usuario_id, u.nombre as usuario_nombre,
                          e.tipo_documento, e.confianza, e.nombre_archivo,
                          e.texto_extraido, e.datos_json, e.creado_en
                   FROM escaneos_documentos e
                   LEFT JOIN usuarios u ON u.id = e.usuario_id
                   WHERE e.usuario_id = ?
                   ORDER BY e.creado_en DESC LIMIT 100""",
                (usuario_id,),
            ).fetchall()

    result = []
    for r in rows:
        try:
            datos = json.loads(r["datos_json"] or "{}")
        except Exception:
            datos = {}
        result.append({
            "id":              r["id"],
            "usuario_nombre":  r["usuario_nombre"] or "—",
            "tipo_documento":  r["tipo_documento"],
            "tipo_label":      TIPOS_DOCUMENTO.get(r["tipo_documento"], "Documento"),
            "confianza":       r["confianza"],
            "nombre_archivo":  r["nombre_archivo"],
            "texto_extraido":  (r["texto_extraido"] or "")[:300],
            "datos":           datos,
            "creado_en":       r["creado_en"],
        })

    return jsonify(result)


@ocr_bp.route("/api/ocr/historial/<int:escaneo_id>")
@login_required
def api_historial_detalle(escaneo_id):
    usuario_id = session.get("user_id")
    with get_db() as db:
        row = db.execute(
            """SELECT e.*, u.nombre as usuario_nombre
               FROM escaneos_documentos e
               LEFT JOIN usuarios u ON u.id = e.usuario_id
               WHERE e.id = ?""",
            (escaneo_id,),
        ).fetchone()
    if not row:
        return jsonify({"error": "Escaneo no encontrado"}), 404
    if row["usuario_id"] != usuario_id:
        u = get_usuario()
        rol = _normalizar_rol(u.get("rol", ""))
        if rol not in {"directora", "coordinador_general", "secretaria", "secretaria_docente"}:
            return jsonify({"error": "Sin permisos"}), 403
    try:
        datos = json.loads(row["datos_json"] or "{}")
    except Exception:
        datos = {}
    return jsonify({
        "id":             row["id"],
        "usuario_nombre": row["usuario_nombre"],
        "tipo_documento": row["tipo_documento"],
        "tipo_label":     TIPOS_DOCUMENTO.get(row["tipo_documento"], "Documento"),
        "confianza":      row["confianza"],
        "nombre_archivo": row["nombre_archivo"],
        "texto_completo": row["texto_extraido"],
        "datos":          datos,
        "creado_en":      row["creado_en"],
    })


@ocr_bp.route("/api/ocr/historial/<int:escaneo_id>", methods=["DELETE"])
@login_required
def api_eliminar_escaneo(escaneo_id):
    usuario_id = session.get("user_id")
    with get_db() as db:
        row = db.execute(
            "SELECT usuario_id FROM escaneos_documentos WHERE id = ?", (escaneo_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "No encontrado"}), 404
        if row["usuario_id"] != usuario_id:
            u = get_usuario()
            if _normalizar_rol(u.get("rol", "")) not in {"directora", "coordinador_general"}:
                return jsonify({"error": "Sin permisos"}), 403
        db.execute("DELETE FROM escaneos_documentos WHERE id = ?", (escaneo_id,))
        db.commit()
    cache_bust()
    return jsonify({"ok": True})
