# -*- coding: utf-8 -*-
"""Blueprint: digitador — Portal de Digitación de datos académicos"""

import sqlite3
import logging
import json as _json
import os
from datetime import datetime, date, timedelta
from flask import (
    Blueprint, render_template, request, jsonify, session,
    redirect, url_for, Response,
)

from core.constants import *
from core.database import get_db, cache_get, cache_set, cache_bust
from core.auth import (
    _normalizar_rol, login_required, get_usuario,
    _csrf_token, _csrf_check,
)
from core.helpers import *
from core.tasks import encolar_tarea, obtener_tarea, actualizar_progreso
from core.helpers import _anio_escolar_actual, _periodo_actual, _audit, _nota_estado, _color_nota

logger = logging.getLogger("axula")

digitador_bp = Blueprint("digitador_bp", __name__)

def _digitador_required(f):
    """Solo digitador, secretaria_docente, coordinadores y directora."""
    import functools
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect("/login")
        rol = _normalizar_rol(session.get("rol", ""))
        permitidos = {"digitador", "secretaria_docente", "directora",
                      "coordinador_general", "coordinador_primer_ciclo",
                      "coordinador_segundo_ciclo", "asistente_directora"}
        if rol not in permitidos:
            return jsonify({"error": "Sin permisos para esta sección"}), 403
        return f(*args, **kwargs)
    return decorated


@digitador_bp.route("/digitador")
@login_required
def portal_digitador():
    u = get_usuario()
    rol = _normalizar_rol(u.get("rol", ""))
    if rol not in {"digitador", "secretaria_docente", "directora",
                   "coordinador_general", "coordinador_primer_ciclo",
                   "coordinador_segundo_ciclo", "asistente_directora"}:
        return redirect("/")
    return render_template("digitador.html", usuario=u, current_user=u)


# ── LISTAR ESTUDIANTES PARA DIGITACIÓN ───────────────────────────────────────

@digitador_bp.route("/api/digitador/estudiantes")
@login_required
@_digitador_required
def listar_estudiantes_digitador():
    grado = request.args.get("grado", "")
    mencion = request.args.get("mencion", "")
    seccion = request.args.get("seccion", "")
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = """
            SELECT id, nombre, apellido, grado, curso, condicion, cedula,
                   ciclo, seccion, tiene_notas
            FROM estudiantes WHERE condicion='ACTIVO'
        """
        params = []
        if grado:
            sql += " AND grado=?"
            params.append(grado)
        if mencion:
            sql += " AND curso=?"
            params.append(mencion)
        if seccion:
            sql += " AND seccion=?"
            params.append(seccion)
        sql += " ORDER BY apellido, nombre"
        rows = conn.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


# ── DIGITAR CALIFICACIÓN ─────────────────────────────────────────────────────

@digitador_bp.route("/api/digitador/calificacion", methods=["POST"])
@login_required
@_digitador_required
def digitar_calificacion():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    est_id = d.get("estudiante_id")
    materia = d.get("materia", "").strip()
    periodo = d.get("periodo", "").strip()
    nota = d.get("nota")

    if not all([est_id, materia, periodo, nota is not None]):
        return jsonify({"error": "Faltan campos obligatorios"}), 400

    try:
        nota = float(nota)
    except (ValueError, TypeError):
        return jsonify({"error": "Nota debe ser numérica"}), 400

    anio = _anio_escolar_actual()

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        # Verificar si ya existe
        existe = conn.execute("""
            SELECT id FROM calificaciones_periodo
            WHERE estudiante_id=? AND materia=? AND periodo=? AND anio_escolar=?
        """, (est_id, materia, periodo, anio)).fetchone()

        if existe:
            conn.execute("""
                UPDATE calificaciones_periodo SET nota=?, registrado_por=?,
                       fecha_registro=datetime('now')
                WHERE id=?
            """, (nota, u["id"], existe[0]))
        else:
            conn.execute("""
                INSERT INTO calificaciones_periodo
                (estudiante_id, materia, periodo, nota, anio_escolar, registrado_por)
                VALUES (?,?,?,?,?,?)
            """, (est_id, materia, periodo, nota, anio, u["id"]))

        conn.commit()
        cache_bust()

    _audit("calificacion_digitada",
           f"Nota {nota} en {materia} P{periodo} para est#{est_id}",
           "calificaciones_periodo", est_id)
    return jsonify({"ok": True, "estado": _nota_estado(nota)})


# ── DIGITAR LOTE (múltiples notas de una vez) ────────────────────────────────

@digitador_bp.route("/api/digitador/calificacion-lote", methods=["POST"])
@login_required
@_digitador_required
def digitar_calificacion_lote():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    notas = d.get("notas", [])
    if not notas:
        return jsonify({"error": "No hay notas para guardar"}), 400

    anio = _anio_escolar_actual()
    guardadas = 0
    errores = []

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        for item in notas:
            try:
                est_id = item["estudiante_id"]
                materia = item["materia"]
                periodo = item["periodo"]
                nota = float(item["nota"])

                existe = conn.execute("""
                    SELECT id FROM calificaciones_periodo
                    WHERE estudiante_id=? AND materia=? AND periodo=? AND anio_escolar=?
                """, (est_id, materia, periodo, anio)).fetchone()

                if existe:
                    conn.execute(
                        "UPDATE calificaciones_periodo SET nota=?, registrado_por=? WHERE id=?",
                        (nota, u["id"], existe[0]))
                else:
                    conn.execute("""
                        INSERT INTO calificaciones_periodo
                        (estudiante_id, materia, periodo, nota, anio_escolar, registrado_por)
                        VALUES (?,?,?,?,?,?)
                    """, (est_id, materia, periodo, nota, anio, u["id"]))
                guardadas += 1
            except Exception as ex:
                errores.append(f"Error en est#{item.get('estudiante_id','?')}: {ex}")

        conn.commit()
        cache_bust()

    return jsonify({"ok": True, "guardadas": guardadas, "errores": errores})


# ── VER NOTAS DE UN ESTUDIANTE ───────────────────────────────────────────────

@digitador_bp.route("/api/digitador/notas/<int:est_id>")
@login_required
@_digitador_required
def ver_notas_estudiante(est_id):
    anio = request.args.get("anio", _anio_escolar_actual())
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        est = conn.execute(
            "SELECT id, nombre, apellido, grado, curso FROM estudiantes WHERE id=?",
            (est_id,)
        ).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        notas = conn.execute("""
            SELECT materia, periodo, nota, fecha_registro
            FROM calificaciones_periodo
            WHERE estudiante_id=? AND anio_escolar=?
            ORDER BY materia, periodo
        """, (est_id, anio)).fetchall()

    return jsonify({
        "estudiante": dict(est),
        "notas": [dict(n) for n in notas],
        "anio": anio,
    })


# ── VERIFICAR DATOS DE ESTUDIANTES ──────────────────────────────────────────

@digitador_bp.route("/api/digitador/verificar-datos")
@login_required
@_digitador_required
def verificar_datos():
    """Detecta estudiantes con datos incompletos o inconsistentes."""
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row

        # Sin cédula
        sin_cedula = conn.execute("""
            SELECT id, nombre, apellido, grado, curso
            FROM estudiantes WHERE condicion='ACTIVO'
            AND (cedula IS NULL OR cedula = '')
        """).fetchall()

        # Sin grado
        sin_grado = conn.execute("""
            SELECT id, nombre, apellido, grado, curso
            FROM estudiantes WHERE condicion='ACTIVO'
            AND (grado IS NULL OR grado = '')
        """).fetchall()

        # Duplicados potenciales (mismo nombre+apellido)
        duplicados = conn.execute("""
            SELECT nombre, apellido, COUNT(*) as total
            FROM estudiantes WHERE condicion='ACTIVO'
            GROUP BY LOWER(nombre), LOWER(apellido)
            HAVING total > 1
        """).fetchall()

        # Sin notas registradas
        sin_notas = conn.execute("""
            SELECT e.id, e.nombre, e.apellido, e.grado, e.curso
            FROM estudiantes e
            WHERE e.condicion='ACTIVO' AND e.tiene_notas = 0
            ORDER BY e.grado, e.apellido
        """).fetchall()

    return jsonify({
        "sin_cedula": [dict(r) for r in sin_cedula],
        "sin_grado": [dict(r) for r in sin_grado],
        "duplicados": [dict(r) for r in duplicados],
        "sin_notas": [dict(r) for r in sin_notas],
    })


# ── RESUMEN DIGITADOR ───────────────────────────────────────────────────────

@digitador_bp.route("/api/digitador/resumen")
@login_required
@_digitador_required
def resumen_digitador():
    anio = _anio_escolar_actual()
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        total_est = conn.execute(
            "SELECT COUNT(*) FROM estudiantes WHERE condicion='ACTIVO'"
        ).fetchone()[0]

        con_notas = conn.execute("""
            SELECT COUNT(DISTINCT estudiante_id) FROM calificaciones_periodo
            WHERE anio_escolar=?
        """, (anio,)).fetchone()[0]

        total_calif = conn.execute(
            "SELECT COUNT(*) FROM calificaciones_periodo WHERE anio_escolar=?",
            (anio,)
        ).fetchone()[0]

        por_periodo = conn.execute("""
            SELECT periodo, COUNT(*) as total
            FROM calificaciones_periodo WHERE anio_escolar=?
            GROUP BY periodo ORDER BY periodo
        """, (anio,)).fetchall()

    return jsonify({
        "total_estudiantes": total_est,
        "con_notas": con_notas,
        "sin_notas": total_est - con_notas,
        "total_calificaciones": total_calif,
        "por_periodo": [dict(r) for r in por_periodo],
    })


# ═══════════════════════════════════════════════════════════════════════════
#  MVP: INSCRIPCIÓN Y REGISTRO DE ESTUDIANTES
# ═══════════════════════════════════════════════════════════════════════════

@digitador_bp.route("/api/digitador/registrar-estudiante", methods=["POST"])
@login_required
@_digitador_required
def registrar_estudiante():
    """Registra un nuevo estudiante en el sistema."""
    d = request.get_json(silent=True) or {}
    u = get_usuario()

    nombre = d.get("nombre", "").strip()
    apellido = d.get("apellido", "").strip()
    if not nombre or not apellido:
        return jsonify({"error": "Nombre y apellido son requeridos"}), 400

    cedula = d.get("cedula", "").strip()
    grado = d.get("grado", "4to")
    mencion = d.get("mencion", "").strip()
    seccion = d.get("seccion", "A").strip()
    edad = d.get("edad", 0)
    sexo = d.get("sexo", "")
    telefono = d.get("telefono", "")
    direccion = d.get("direccion", "")
    nombre_tutor = d.get("nombre_tutor", "")
    telefono_tutor = d.get("telefono_tutor", "")

    # Curso = "4TO MULTIMEDIA" format
    curso = f"{grado.upper()} {mencion.upper()}" if mencion else grado.upper()

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        # Check duplicate cedula
        if cedula:
            dup = conn.execute(
                "SELECT id FROM estudiantes WHERE cedula=?", (cedula,)
            ).fetchone()
            if dup:
                return jsonify({"error": f"Ya existe un estudiante con cédula {cedula}"}), 400

        conn.execute("""
            INSERT INTO estudiantes
            (nombre, apellido, cedula, grado, curso, seccion, edad, condicion, ciclo)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (nombre, apellido, cedula, grado, curso, seccion,
              edad, "ACTIVO", "segundo_ciclo"))
        est_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create inscription record
        anio = _anio_escolar_actual()
        conn.execute("""
            INSERT INTO inscripciones
            (estudiante_id, anio_escolar, grado, mencion, seccion, tipo, registrado_por)
            VALUES (?,?,?,?,?,?,?)
        """, (est_id, anio, grado, mencion, seccion, "nueva", u["id"]))

        conn.commit()
        cache_bust()

    _audit("estudiante_registrado", f"Nuevo: {nombre} {apellido} ({grado} {mencion})",
           "estudiantes", est_id)
    return jsonify({"ok": True, "id": est_id, "mensaje": f"Estudiante {nombre} {apellido} registrado"})


@digitador_bp.route("/api/digitador/editar-estudiante/<int:est_id>", methods=["PATCH"])
@login_required
@_digitador_required
def editar_estudiante_digitador(est_id):
    """Edita datos de un estudiante existente."""
    d = request.get_json(silent=True) or {}
    campos = []
    params = []
    permitidos = ["nombre", "apellido", "cedula", "grado", "curso", "seccion",
                  "edad", "condicion", "ciclo"]
    for campo in permitidos:
        if campo in d:
            campos.append(f"{campo}=?")
            params.append(d[campo])

    # Si cambian grado+mencion, recalcular curso
    if "mencion" in d and "grado" in d:
        curso = f"{d['grado'].upper()} {d['mencion'].upper()}" if d["mencion"] else d["grado"].upper()
        campos.append("curso=?")
        params.append(curso)

    if not campos:
        return jsonify({"error": "Nada que actualizar"}), 400

    params.append(est_id)
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute(f"UPDATE estudiantes SET {','.join(campos)} WHERE id=?", params)
        conn.commit()
        cache_bust()

    return jsonify({"ok": True})


@digitador_bp.route("/api/digitador/asignar-mencion-bulk", methods=["POST"])
@login_required
@_digitador_required
def asignar_mencion_bulk():
    """Asigna mención en masa a estudiantes de un grado/sección sin mención asignada."""
    d = request.get_json(silent=True) or {}
    grado   = d.get("grado", "").strip().upper()
    seccion = d.get("seccion", "").strip().upper()
    mencion = d.get("mencion", "").strip()
    solo_vacias = d.get("solo_vacias", True)   # por defecto solo actualiza los que no tienen

    MENCIONES_VALIDAS = ["Multimedia", "Teatro", "Música", "Artes Visuales", "Danza"]
    GRADOS_ARTES = {"4TO", "5TO", "6TO"}

    if not grado:
        return jsonify({"error": "Grado requerido"}), 400
    if mencion not in MENCIONES_VALIDAS:
        return jsonify({"error": f"Mención inválida. Válidas: {', '.join(MENCIONES_VALIDAS)}"}), 400
    if grado not in GRADOS_ARTES:
        return jsonify({"error": "Mención solo aplica a 4TO, 5TO y 6TO (Bachillerato en Artes)"}), 400

    curso = f"{grado} {mencion.upper()}"
    params = [mencion, curso, grado]
    cond = "WHERE grado=?"
    if solo_vacias:
        cond += " AND (mencion IS NULL OR mencion='')"
    if seccion:
        cond += " AND seccion=?"
        params.append(seccion)

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        cur = conn.execute(f"UPDATE estudiantes SET mencion=?, curso=? {cond}", params)
        conn.commit()
        n = cur.rowcount
        cache_bust()

    logger.info(f"asignar_mencion_bulk: {n} estudiantes → {grado}/{seccion or 'todas'} = {mencion}")
    return jsonify({"ok": True, "actualizados": n, "grado": grado, "seccion": seccion, "mencion": mencion})


@digitador_bp.route("/api/digitador/cargar-notas-coordinador", methods=["POST"])
@login_required
@_digitador_required
def cargar_notas_coordinador():
    """
    Carga un archivo .xlsm/.xlsx de Registro del Coordinador (formato BJ).
    Detecta automáticamente materia, grado, sección, mención y períodos.
    Inserta/actualiza en materias_calificaciones.
    """
    import tempfile, unicodedata
    from difflib import SequenceMatcher
    from core.excel_notas import parsear_registro_coordinador

    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm"):
        return jsonify({"ok": False, "error": "Formato no válido. Usa .xlsx o .xlsm"}), 400

    # Guardar temporalmente
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        datos = parsear_registro_coordinador(tmp_path)
    except Exception as e:
        os.unlink(tmp_path)
        logger.error(f"[EXCEL] Error parseando {f.filename}: {e}")
        return jsonify({"ok": False, "error": f"Error al leer el archivo: {e}"}), 400
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if not datos["hojas"]:
        return jsonify({
            "ok": False,
            "error": "No se encontraron hojas con formato reconocido. Se esperan hojas nombradas como 'CF 1ERO A' (primer ciclo) o 'CF 4TO A MÚSICA' (segundo ciclo).",
            "advertencias": datos.get("advertencias", [])
        }), 400

    materia = datos["materia"]
    anio    = _anio_escolar_actual()

    # ── Helper de normalización para matching ──────────────────────────────────
    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or ""))
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        import re
        return re.sub(r"\s+", " ", s).strip().lower()

    def _similar(a, b):
        return SequenceMatcher(None, _norm(a), _norm(b)).ratio()

    resumen = {
        "materia":     materia,
        "hojas":       len(datos["hojas"]),
        "cargados":    0,
        "actualizados": 0,
        "sin_match":   [],
        "advertencias": datos.get("advertencias", []),
    }

    with sqlite3.connect(DATABASE, timeout=15) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")

        for hoja in datos["hojas"]:
            grado   = hoja["grado"]
            seccion = hoja["seccion"]
            mencion = hoja["mencion"]
            tipo    = hoja["tipo"]
            ciclo   = "primer_ciclo" if grado in ("1ERO","2DO","3ERO") else "segundo_ciclo"

            # Índice de estudiantes del grado en BD
            rows_bd = conn.execute(
                "SELECT id, nombre, apellido FROM estudiantes WHERE grado=? AND condicion='ACTIVO'",
                (grado,)
            ).fetchall()
            idx_bd = {_norm(f"{r['nombre']} {r['apellido']}"): r["id"] for r in rows_bd}

            for alumno in hoja["alumnos"]:
                if alumno["estado"] == "RETIRADO":
                    continue

                nombre_pdf = alumno["nombre"]
                clave      = _norm(nombre_pdf)

                # Búsqueda exacta primero
                est_id = idx_bd.get(clave)

                # Fuzzy si no encontró exacto
                if est_id is None:
                    mejor = max(idx_bd.items(), key=lambda kv: _similar(kv[0], clave), default=(None, 0))
                    if mejor[0] and _similar(mejor[0], clave) >= 0.82:
                        est_id = mejor[1]

                if est_id is None:
                    resumen["sin_match"].append(f"[{grado}] {nombre_pdf}")
                    continue

                # INSERT OR REPLACE en materias_calificaciones
                existente = conn.execute(
                    "SELECT id FROM materias_calificaciones WHERE estudiante_id=? AND materia=? AND grado=? AND anio_escolar=?",
                    (est_id, materia, grado, anio)
                ).fetchone()

                if existente:
                    conn.execute("""
                        UPDATE materias_calificaciones
                        SET p1=?, p2=?, p3=?, p4=?, promedio=?, tipo=?, ciclo=?
                        WHERE id=?
                    """, (alumno["p1"], alumno["p2"], alumno["p3"], alumno["p4"],
                          alumno["promedio"], tipo, ciclo, existente["id"]))
                    resumen["actualizados"] += 1
                else:
                    conn.execute("""
                        INSERT INTO materias_calificaciones
                            (estudiante_id, materia, grado, p1, p2, p3, p4, promedio, tipo, ciclo, anio_escolar)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """, (est_id, materia, grado,
                          alumno["p1"], alumno["p2"], alumno["p3"], alumno["p4"],
                          alumno["promedio"], tipo, ciclo, anio))
                    resumen["cargados"] += 1

        conn.commit()

    total = resumen["cargados"] + resumen["actualizados"]
    return jsonify({
        "ok":      True,
        "mensaje": (
            f"✓ {materia} cargada — "
            f"{resumen['cargados']} nuevos, {resumen['actualizados']} actualizados "
            f"en {resumen['hojas']} hoja(s)."
        ),
        **resumen,
    })


@digitador_bp.route("/api/digitador/preview-notas-coordinador", methods=["POST"])
@login_required
@_digitador_required
def preview_notas_coordinador():
    """
    Parsea el archivo del coordinador y devuelve un preview con:
    - matched: lista de {est_id, nombre, grado, p1_actual..p4_actual, p1_nuevo..p4_nuevo}
    - unmatched: nombres sin match en BD
    No escribe nada en la BD.
    """
    import tempfile, unicodedata
    from difflib import SequenceMatcher
    from core.excel_notas import parsear_registro_coordinador

    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm"):
        return jsonify({"ok": False, "error": "Formato no válido. Usa .xlsx o .xlsm"}), 400

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        datos = parsear_registro_coordinador(tmp_path)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al leer el archivo: {e}"}), 400
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if not datos["hojas"]:
        return jsonify({
            "ok": False,
            "error": "No se encontraron hojas con formato reconocido.",
            "advertencias": datos.get("advertencias", [])
        }), 400

    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or ""))
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        import re
        return re.sub(r"\s+", " ", s).strip().lower()

    def _similar(a, b):
        return SequenceMatcher(None, _norm(a), _norm(b)).ratio()

    materia = datos["materia"]
    anio    = _anio_escolar_actual()
    matched   = []
    unmatched = []

    with sqlite3.connect(DATABASE, timeout=15) as conn:
        conn.row_factory = sqlite3.Row

        for hoja in datos["hojas"]:
            grado = hoja["grado"]

            rows_bd = conn.execute(
                "SELECT id, nombre, apellido FROM estudiantes WHERE grado=? AND condicion='ACTIVO'",
                (grado,)
            ).fetchall()
            idx_bd = {_norm(f"{r['nombre']} {r['apellido']}"): r for r in rows_bd}

            for alumno in hoja["alumnos"]:
                if alumno["estado"] == "RETIRADO":
                    continue

                clave = _norm(alumno["nombre"])
                row_bd = idx_bd.get(clave)

                if row_bd is None:
                    mejor_score = 0
                    mejor_row   = None
                    for k, v in idx_bd.items():
                        s = _similar(k, clave)
                        if s > mejor_score:
                            mejor_score = s
                            mejor_row   = v
                    if mejor_row and mejor_score >= 0.72:
                        row_bd = mejor_row

                if row_bd is None:
                    unmatched.append({"nombre": alumno["nombre"], "grado": grado})
                    continue

                est_id = row_bd["id"]
                nombre_bd = f"{row_bd['nombre']} {row_bd['apellido']}"

                # Notas actuales en calificaciones_periodo
                actuales = {}
                for p_str in ("P1", "P2", "P3", "P4"):
                    row_p = conn.execute(
                        "SELECT calificacion FROM calificaciones_periodo "
                        "WHERE estudiante_id=? AND materia=? AND periodo=? AND anio_escolar=?",
                        (est_id, materia, p_str, anio)
                    ).fetchone()
                    actuales[p_str.lower()] = float(row_p["calificacion"]) if row_p else None

                matched.append({
                    "est_id":    est_id,
                    "nombre_bd": nombre_bd,
                    "nombre_excel": alumno["nombre"],
                    "grado":     grado,
                    "hoja":      hoja["hoja"],
                    "p1_actual": actuales["p1"],
                    "p2_actual": actuales["p2"],
                    "p3_actual": actuales["p3"],
                    "p4_actual": actuales["p4"],
                    "p1_nuevo":  alumno["p1"],
                    "p2_nuevo":  alumno["p2"],
                    "p3_nuevo":  alumno["p3"],
                    "p4_nuevo":  alumno["p4"],
                    "cf_nuevo":  alumno["promedio"],
                })

    return jsonify({
        "ok":          True,
        "materia":     materia,
        "anio":        anio,
        "matched":     matched,
        "unmatched":   unmatched,
        "advertencias": datos.get("advertencias", []),
    })


@digitador_bp.route("/api/digitador/confirmar-notas-coordinador", methods=["POST"])
@login_required
@_digitador_required
def confirmar_notas_coordinador():
    """
    Guarda los registros preview confirmados en calificaciones_periodo (origen='importacion').
    Solo sobreescribe si no hay nota manual preexistente.
    Body JSON: {materia, anio, records: [{est_id, p1_nuevo, p2_nuevo, p3_nuevo, p4_nuevo}]}
    """
    from core.helpers import recalcular_kpis_estudiante

    d = request.get_json(silent=True) or {}
    materia = (d.get("materia") or "").strip()
    anio    = (d.get("anio") or _anio_escolar_actual()).strip()
    records = d.get("records", [])

    if not materia or not records:
        return jsonify({"ok": False, "error": "Faltan datos (materia o records)"}), 400

    prof_id = get_usuario().get("id")
    guardados   = 0
    omitidos    = 0
    actualizados = 0

    with sqlite3.connect(DATABASE, timeout=15) as conn:
        conn.row_factory = sqlite3.Row

        for rec in records:
            est_id = rec.get("est_id")
            if not est_id:
                continue

            for p_num, campo in enumerate(["p1_nuevo", "p2_nuevo", "p3_nuevo", "p4_nuevo"], start=1):
                nota = rec.get(campo)
                if nota is None:
                    continue
                periodo_str = f"P{p_num}"

                existing = conn.execute(
                    "SELECT id, origen FROM calificaciones_periodo "
                    "WHERE estudiante_id=? AND materia=? AND periodo=? AND anio_escolar=?",
                    (est_id, materia, periodo_str, anio)
                ).fetchone()

                if existing is None:
                    conn.execute(
                        "INSERT INTO calificaciones_periodo "
                        "(estudiante_id, profesor_id, materia, periodo, calificacion, "
                        " anio_escolar, observacion, origen) "
                        "VALUES (?,?,?,?,?,?,'Cargado desde Excel del coordinador','importacion')",
                        (est_id, prof_id, materia, periodo_str, nota, anio)
                    )
                    guardados += 1
                elif existing["origen"] == "manual":
                    omitidos += 1  # Nota manual tiene precedencia — no tocar
                else:
                    conn.execute(
                        "UPDATE calificaciones_periodo "
                        "SET calificacion=?, profesor_id=?, origen='importacion', "
                        "    actualizado=datetime('now') "
                        "WHERE id=?",
                        (nota, prof_id, existing["id"])
                    )
                    actualizados += 1

        conn.commit()

        # Recalcular KPIs de todos los estudiantes afectados
        est_ids = list({rec["est_id"] for rec in records if rec.get("est_id")})
        for eid in est_ids:
            try:
                recalcular_kpis_estudiante(conn, eid, anio)
            except Exception:
                pass
        conn.commit()

    return jsonify({
        "ok":         True,
        "mensaje":    f"✓ {materia} — {guardados} nuevos, {actualizados} actualizados, {omitidos} omitidos (nota manual).",
        "guardados":  guardados,
        "actualizados": actualizados,
        "omitidos":   omitidos,
    })


@digitador_bp.route("/api/digitador/buscar-estudiante")
@login_required
@_digitador_required
def buscar_estudiante_digitador():
    """Búsqueda rápida de estudiantes."""
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, nombre, apellido, grado, curso, cedula, condicion, seccion
            FROM estudiantes
            WHERE (nombre LIKE ? OR apellido LIKE ? OR cedula LIKE ?)
            ORDER BY apellido, nombre LIMIT 30
        """, (f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
    return jsonify([dict(r) for r in rows])


# ── CARGA DE NOTAS POR LISTADO SIMPLE ────────────────────────────────────────

@digitador_bp.route("/api/digitador/plantilla-notas")
@login_required
@_digitador_required
def descargar_plantilla_notas():
    """
    Genera un Excel descargable con los estudiantes que NO tienen notas,
    listo para que el coordinador lo rellene y lo devuelva.

    Parámetros opcionales:
      ?grado=3ERO   → filtra por grado
      ?seccion=A    → filtra por sección
      ?todos=1      → incluye todos los estudiantes (con y sin notas)
    """
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl no disponible"}), 500

    grado   = request.args.get("grado", "").strip().upper()
    seccion = request.args.get("seccion", "").strip().upper()
    todos   = request.args.get("todos", "0") == "1"

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        q = "SELECT id, nombre, apellido, grado, seccion, mencion FROM estudiantes WHERE 1=1"
        p = []
        if grado:
            q += " AND upper(grado) LIKE upper(?)"
            p.append(f"%{grado}%")
        if seccion:
            q += " AND upper(seccion)=upper(?)"
            p.append(seccion)
        if not todos:
            q += " AND id NOT IN (SELECT DISTINCT estudiante_id FROM materias_calificaciones)"
        q += " ORDER BY grado, apellido, nombre"
        estudiantes = conn.execute(q, p).fetchall()

    if not estudiantes:
        return jsonify({"error": "No hay estudiantes que coincidan con el filtro"}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Notas"

    # ── Estilos ───────────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="024959")
    sub_fill  = PatternFill("solid", fgColor="038C8C")
    sub_font  = Font(bold=True, color="FFFFFF", size=10)
    inst_fill = PatternFill("solid", fgColor="FFF3CD")
    inst_font = Font(bold=False, color="856404", size=9, italic=True)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Fila de instrucciones ─────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    ws["A1"] = (
        "INSTRUCCIONES: Completa las columnas P1-P4 para cada materia. "
        "Una fila por alumno/materia. NO cambies las columnas ID, Nombre ni Grado. "
        "Guarda como .xlsx y sube al sistema."
    )
    ws["A1"].font      = inst_font
    ws["A1"].fill      = inst_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # ── Encabezados ───────────────────────────────────────────────────────────
    headers = [
        ("ID",       8),  ("Nombre Completo", 28), ("Grado",   8),
        ("Sección",  8),  ("Mención",         12), ("Materia", 28),
        ("P1",       7),  ("P2",               7), ("P3",      7),
        ("P4",       7),  ("Promedio CF",      11),
    ]
    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 22

    # ── Materias por defecto según ciclo ─────────────────────────────────────
    MATS_1ER_CICLO = [
        "Lengua Española", "Matemática", "Ciencias Naturales",
        "Ciencias Sociales", "Idioma Inglés", "Educación Física",
        "FIHR", "Educación Artística",
    ]
    MATS_2DO_CICLO_BASE = [
        "Lengua Española", "Matemática", "Ciencias Naturales",
        "Ciencias Sociales", "Idioma Inglés", "Idioma Francés",
        "Educación Física", "FIHR", "Identidad, Cultura y Emprendimiento",
    ]

    def _materias_para(grado_est, mencion_est):
        g = str(grado_est or "").upper()
        if any(x in g for x in ("1ER", "2DO", "3ER")):
            return MATS_1ER_CICLO
        m = str(mencion_est or "").upper()
        base = list(MATS_2DO_CICLO_BASE)
        if "USIC" in m or "ÚSIC" in m:
            base += ["Instrumento I", "Canto Coral I", "Lenguaje Musical, Teoría y Entrenamiento"]
        elif "ATRO" in m:
            base += ["Entrenamiento Rítmico, Corporal y Vocal I", "Expresión Corporal", "Historia del Teatro"]
        elif "ISUAL" in m:
            base += ["Dibujo", "Pintura y Técnicas Mixtas", "Historia del Arte Universal"]
        else:
            base += ["Fotografía", "Lenguaje Visual", "Diseño Básico"]
        return base

    # ── Filas de datos ────────────────────────────────────────────────────────
    row_num   = 3
    fill_alt  = PatternFill("solid", fgColor="F0FAFA")
    fill_norm = PatternFill("solid", fgColor="FFFFFF")

    for i, est in enumerate(estudiantes):
        nombre_completo = f"{est['nombre']} {est['apellido']}"
        materias = _materias_para(est['grado'], est['mencion'])
        alt = (i % 2 == 0)

        for mat in materias:
            fondo = fill_alt if alt else fill_norm
            vals = [
                est['id'], nombre_completo, est['grado'],
                est['seccion'] or "", est['mencion'] or "", mat,
                "", "", "", "", "",   # P1-P4, Promedio (vacíos)
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill      = fondo
                cell.border    = border
                cell.alignment = center if col in (1, 3, 4, 7, 8, 9, 10, 11) else left
                cell.font      = Font(size=9)
            row_num += 1

    # Fijar encabezados al hacer scroll
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = f"_{grado}" if grado else "_todos"
    fname  = f"plantilla_notas{suffix}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _procesar_excel_notas(raw: bytes, anio_escolar: str, task_id: str) -> dict:
    """
    Función pura que corre en background thread.
    NO usa Flask g ni get_db() — crea su propia conexión.
    Retorna el dict de resultado que queda en la tarea.
    """
    import io, unicodedata, re
    from difflib import SequenceMatcher
    from openpyxl import load_workbook

    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or ""))
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return re.sub(r"\s+", " ", s).strip().lower()

    def _fl(v):
        try:
            x = float(v)
            return round(x, 1) if 0 < x <= 100 else None
        except Exception:
            return None

    # ── 1. Parsear Excel ─────────────────────────────────────────────────────
    actualizar_progreso(task_id, 5)
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    header_row = None
    col_id = col_nombre = col_materia = col_p1 = col_p2 = col_p3 = col_p4 = col_prom = None

    for ri in range(1, min(6, ws.max_row + 1)):
        for ci in range(1, ws.max_column + 1):
            v = _norm(ws.cell(ri, ci).value)
            if v in ("id", "codigo"):                               col_id      = ci
            elif "nombre" in v and "completo" in v:                 col_nombre  = ci
            elif v in ("nombre completo", "alumno", "estudiante"):  col_nombre  = ci
            elif v in ("materia", "asignatura"):                    col_materia = ci
            elif v in ("p1", "periodo 1", "pc1"):                   col_p1      = ci
            elif v in ("p2", "periodo 2", "pc2"):                   col_p2      = ci
            elif v in ("p3", "periodo 3", "pc3"):                   col_p3      = ci
            elif v in ("p4", "periodo 4", "pc4"):                   col_p4      = ci
            elif "prom" in v or "cf" in v:                          col_prom    = ci
        if col_nombre and col_materia:
            header_row = ri
            break

    if not header_row or not col_nombre or not col_materia:
        raise ValueError(
            "No se encontraron columnas 'Nombre Completo' y 'Materia'. "
            "Verifica que el archivo tenga el formato correcto."
        )

    filas = []
    for ri in range(header_row + 1, ws.max_row + 1):
        nombre  = str(ws.cell(ri, col_nombre).value  or "").strip()
        materia = str(ws.cell(ri, col_materia).value or "").strip()
        if not nombre or not materia:
            continue
        est_id_hint = None
        if col_id:
            try:
                est_id_hint = int(ws.cell(ri, col_id).value)
            except Exception:
                pass
        filas.append({
            "nombre":      nombre,
            "materia":     materia,
            "p1":          _fl(ws.cell(ri, col_p1).value)   if col_p1   else None,
            "p2":          _fl(ws.cell(ri, col_p2).value)   if col_p2   else None,
            "p3":          _fl(ws.cell(ri, col_p3).value)   if col_p3   else None,
            "p4":          _fl(ws.cell(ri, col_p4).value)   if col_p4   else None,
            "promedio":    _fl(ws.cell(ri, col_prom).value) if col_prom else None,
            "est_id_hint": est_id_hint,
        })

    if not filas:
        raise ValueError("El archivo no tiene filas de datos")

    actualizar_progreso(task_id, 15)

    # ── 2. Cargar índice de estudiantes ──────────────────────────────────────
    with sqlite3.connect(DATABASE, timeout=15) as conn:
        conn.row_factory = sqlite3.Row
        todos_est = conn.execute(
            "SELECT id, nombre, apellido FROM estudiantes"
        ).fetchall()

    indice = {r["id"]: f"{r['nombre']} {r['apellido']}" for r in todos_est}

    def _similar(a, b):
        return SequenceMatcher(None, _norm(a), _norm(b)).ratio()

    def _buscar_id(nombre_xlsx, id_hint=None):
        if id_hint and id_hint in indice:
            return id_hint
        mejor_id, mejor_score = None, 0.0
        for eid, nombre_bd in indice.items():
            s = _similar(nombre_xlsx, nombre_bd)
            if s > mejor_score:
                mejor_score, mejor_id = s, eid
        return mejor_id if mejor_score >= 0.82 else None

    # ── 3. Insertar / actualizar (con progreso) ──────────────────────────────
    insertados   = 0
    actualizados = 0
    sin_match    = []
    sin_nota     = 0
    ids_afectados: set[int] = set()
    total = len(filas)

    with sqlite3.connect(DATABASE, timeout=30) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")

        for i, fila in enumerate(filas):
            # Progreso: de 20% a 85%
            if i % 50 == 0:
                pct = 20 + int((i / total) * 65)
                actualizar_progreso(task_id, pct)

            vals = [v for v in [fila["p1"], fila["p2"], fila["p3"], fila["p4"]] if v]
            if not vals and not fila["promedio"]:
                sin_nota += 1
                continue

            est_id = _buscar_id(fila["nombre"], fila["est_id_hint"])
            if not est_id:
                sin_match.append(fila["nombre"])
                continue

            ids_afectados.add(est_id)
            prom = fila["promedio"] or (round(sum(vals) / len(vals), 1) if vals else None)

            existente = conn.execute(
                "SELECT rowid FROM materias_calificaciones "
                "WHERE estudiante_id=? AND LOWER(materia)=LOWER(?) AND anio_escolar=?",
                (est_id, fila["materia"], anio_escolar)
            ).fetchone()

            if existente:
                conn.execute("""
                    UPDATE materias_calificaciones SET
                        p1      = CASE WHEN (p1 IS NULL OR p1=0) AND ? IS NOT NULL THEN ? ELSE p1 END,
                        p2      = CASE WHEN (p2 IS NULL OR p2=0) AND ? IS NOT NULL THEN ? ELSE p2 END,
                        p3      = CASE WHEN (p3 IS NULL OR p3=0) AND ? IS NOT NULL THEN ? ELSE p3 END,
                        p4      = CASE WHEN (p4 IS NULL OR p4=0) AND ? IS NOT NULL THEN ? ELSE p4 END,
                        promedio= CASE WHEN (promedio IS NULL OR promedio=0) AND ? IS NOT NULL THEN ? ELSE promedio END
                    WHERE estudiante_id=? AND LOWER(materia)=LOWER(?) AND anio_escolar=?
                """, (
                    fila["p1"], fila["p1"], fila["p2"], fila["p2"],
                    fila["p3"], fila["p3"], fila["p4"], fila["p4"],
                    prom, prom, est_id, fila["materia"], anio_escolar,
                ))
                actualizados += 1
            else:
                conn.execute("""
                    INSERT INTO materias_calificaciones
                        (estudiante_id, materia, tipo, p1, p2, p3, p4, promedio,
                         anio_escolar, fecha_carga)
                    VALUES (?, ?, 'académico', ?, ?, ?, ?, ?, ?, date('now'))
                """, (est_id, fila["materia"],
                      fila["p1"], fila["p2"], fila["p3"], fila["p4"],
                      prom, anio_escolar))
                insertados += 1

        # ── 4. Recalcular p_acad ─────────────────────────────────────────────
        actualizar_progreso(task_id, 88)
        for eid in ids_afectados:
            proms = [
                r["promedio"] for r in conn.execute(
                    "SELECT promedio FROM materias_calificaciones "
                    "WHERE estudiante_id=? AND anio_escolar=? AND promedio>0",
                    (eid, anio_escolar)
                ).fetchall()
            ]
            if proms:
                p_acad_nuevo = round(sum(proms) / len(proms), 1)
                conn.execute(
                    "UPDATE estudiantes SET p_acad=?, tiene_notas=1 WHERE id=?",
                    (p_acad_nuevo, eid)
                )

        conn.commit()

    cache_bust()
    actualizar_progreso(task_id, 99)

    return {
        "ok":              True,
        "anio_escolar":    anio_escolar,
        "insertados":      insertados,
        "actualizados":    actualizados,
        "sin_match":       sin_match[:50],
        "sin_match_total": len(sin_match),
        "sin_nota_vacios": sin_nota,
        "filas_totales":   total,
        "mensaje": (
            f"{insertados} materias nuevas · {actualizados} actualizadas"
            + (f" · {len(sin_match)} sin match" if sin_match else "")
        ),
    }


@digitador_bp.route("/api/digitador/cargar-notas-listado", methods=["POST"])
@login_required
@_digitador_required
def cargar_notas_listado():
    """
    Acepta el Excel con columnas: Nombre Completo | Materia | P1 | P2 | P3 | P4
    Parámetro opcional: anio_escolar (default = año escolar actual)

    El procesamiento corre en background — devuelve task_id para polling.
    Sondear: GET /api/digitador/tarea/<task_id>
    """
    try:
        from openpyxl import load_workbook as _lw  # noqa: F401 — verificar disponibilidad
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl no disponible"}), 500

    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
        return jsonify({"ok": False, "error": "Formato no válido. Usa .xlsx"}), 400

    # Leer bytes en el contexto del request (antes de pasar al thread)
    raw = f.read()
    if len(raw) > 20 * 1024 * 1024:
        return jsonify({"ok": False, "error": "Archivo demasiado grande (máx 20 MB)"}), 413

    anio_escolar = (
        request.form.get("anio_escolar")
        or request.args.get("anio_escolar")
        or _anio_escolar_actual()
    )

    task_id = encolar_tarea(_procesar_excel_notas, raw, anio_escolar)
    return jsonify({
        "ok":          True,
        "task_id":     task_id,
        "anio_escolar": anio_escolar,
        "mensaje":     "Procesamiento iniciado. Sondea /api/digitador/tarea/<task_id> para el resultado.",
    })


@digitador_bp.route("/api/digitador/tarea/<task_id>")
@login_required
@_digitador_required
def estado_tarea(task_id: str):
    """
    GET /api/digitador/tarea/<task_id>
    Polling endpoint para tareas de carga de Excel.
    Retorna: {status, progreso, resultado, error}
    """
    tarea = obtener_tarea(task_id)
    if not tarea:
        return jsonify({"ok": False, "error": "Tarea no encontrada o expirada"}), 404
    return jsonify({"ok": True, **tarea})
