# -*- coding: utf-8 -*-
"""Blueprint: estudiantes"""

import sqlite3
import logging
import json as _json
import os
import re
import time as _time
from datetime import datetime, date, timedelta
from io import BytesIO
from flask import (
    Blueprint, render_template, request, jsonify, session,
    redirect, url_for, g, send_from_directory, Response, send_file,
)

from core.constants import *
from core.database import get_db, cache_get, cache_set, cache_bust, _CACHE
from core.auth import (
    _hash, _check_password, _normalizar_rol, _ciclo_del_rol,
    login_required, coord_required, admin_required, directora_required,
    _csrf_token, _csrf_check, csrf_protected, rate_limited,
    get_usuario,
)
from core.helpers import *
from core.helpers import _anonimizar_estudiante, _buscar_estudiante_bd, _calcular_bienestar_emocional, _calcular_indice_conductual, _features_para_clustering, _get_profesor, _recalcular_indicadores, _validar_materia_profesor, _validar_magic_imagen, _normalizar_clave_materia, _dedup_materias
from core import rls as _rls
from core.ia import _get_groq_client, groq_client, construir_prompt, construir_prompt_planificacion, construir_prompt_rubrica, construir_prompt_estrategia
from core.excel import _parsear_boletin_bj, _buscar_o_crear_estudiante, _detectar_mencion_listado, _limpiar_nota
from core.pdf import _generar_pdf_acuerdo

logger = logging.getLogger("axula")

estudiantes_bp = Blueprint("estudiantes_bp", __name__)

@estudiantes_bp.route("/cargar", methods=["POST"])
@login_required
def cargar():
    """
    Carga la Plantilla Axula (cualquier version).
    Detecta header buscando la celda 'Nombre' en cols 2-6, filas 1-8.
    UPSERT: preserva todos los estudiantes del LISTADO.
    """
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "No se recibio ningún archivo"}), 400

    try:
        from openpyxl import load_workbook
        import io as _io, unicodedata

        raw = file.read()
        wb  = load_workbook(_io.BytesIO(raw), data_only=True)
        hoja = next((s for s in wb.sheetnames
                     if 'SEGUIMIENTO' in s.upper() or 'ESTUDIANTE' in s.upper()),
                    wb.sheetnames[0])
        ws = wb[hoja]

        # ── Detectar fila de encabezado ───────────────────────────────────
        # Buscamos la celda exacta que diga 'Nombre' en cols 2-6
        header_row = None
        nombre_col = None
        for row in range(1, 9):
            for col in range(2, 7):
                v = str(ws.cell(row=row, column=col).value or '').strip()
                if v == 'Nombre':
                    header_row = row
                    nombre_col = col
                    break
            if header_row:
                break

        if not header_row:
            return jsonify({
                "error": "No se encontro la columna 'Nombre' en las primeras 8 filas. "
                         "Verifica que estas subiendo la Plantilla Axula."
            }), 400

        # Columnas basicas relativas al nombre_col
        col_nombre   = nombre_col        # 'Nombre'
        col_apellido = nombre_col + 1    # 'Apellido'
        col_curso    = nombre_col + 2    # 'Curso'
        col_grado    = nombre_col + 3    # 'Grado'
        col_condicion= nombre_col + 4    # 'Condición'
        col_edad     = nombre_col + 5    # 'Edad'
        col_id       = nombre_col - 2    # 'ID' (2 antes de Nombre)

        data_start = header_row + 1

        # ── Detectar grupos de columnas buscando en fila anterior ────────
        grupo_row = header_row - 1
        col_foto = col_lv = col_diseno = col_acad_prom = None
        col_conductual = col_emocional = col_analisis = None

        for col in range(1, ws.max_column + 1):
            gv = str(ws.cell(row=grupo_row, column=col).value or '').upper().strip()
            if not gv:
                continue
            if 'FOTOGRAF' in gv:                    col_foto       = col
            elif 'LENGUAJE' in gv:                  col_lv         = col
            elif 'DISE' in gv:                      col_diseno     = col
            elif 'PROM' in gv and 'ACAD' in gv:    col_acad_prom  = col
            elif 'CONDUCTUAL' in gv:                col_conductual = col
            elif 'EMOCIONAL' in gv:                 col_emocional  = col
            elif 'AN' in gv and ('LISIS' in gv or 'ÁLISIS' in gv): col_analisis = col

        # Fallback por version si no se detectaron grupos
        if col_foto is None:
            if header_row == 3:   # Axula
                col_foto=29; col_lv=33; col_diseno=37; col_acad_prom=41
                col_conductual=18; col_emocional=24; col_analisis=48
            else:                  # Modelo v3 (header_row == 4)
                col_foto=14; col_lv=18; col_diseno=22; col_acad_prom=30
                col_conductual=35; col_emocional=42; col_analisis=48

        def fl(v):
            try:
                f = float(v)
                return round(f, 2) if f == f else None
            except Exception:
                return None

        def norm(s):
            return ''.join(c for c in unicodedata.normalize('NFD', str(s).lower())
                           if unicodedata.category(c) != 'Mn')

        def gc(base, offset):
            if base is None: return None
            return fl(ws.cell(row=row, column=base + offset).value)

        # ── Leer estudiantes ─────────────────────────────────────────────
        students = []
        for row in range(data_start, ws.max_row + 1):
            nombre = ws.cell(row=row, column=col_nombre).value
            if not nombre or str(nombre).strip() in ('', 'nan', 'None', 'Nombre'):
                continue

            id_bj = str(ws.cell(row=row, column=col_id).value or '').strip()

            # Conductual base (Puntualidad-Rendimiento, cols 9-13 en ambas versiones)
            punt = fl(ws.cell(row=row, column=9).value)
            tar  = fl(ws.cell(row=row, column=10).value)
            part = fl(ws.cell(row=row, column=11).value)
            comp = fl(ws.cell(row=row, column=12).value)
            rend = fl(ws.cell(row=row, column=13).value)

            foto_p1=gc(col_foto,0); foto_p2=gc(col_foto,1)
            foto_p3=gc(col_foto,2); foto_p4=gc(col_foto,3)
            lv_p1=gc(col_lv,0);    lv_p2=gc(col_lv,1)
            lv_p3=gc(col_lv,2);    lv_p4=gc(col_lv,3)
            dis_p1=gc(col_diseno,0); dis_p2=gc(col_diseno,1)
            dis_p3=gc(col_diseno,2); dis_p4=gc(col_diseno,3)

            acad_p1=gc(col_acad_prom,0); acad_p2=gc(col_acad_prom,1)
            acad_p3=gc(col_acad_prom,2); acad_p4=gc(col_acad_prom,3)
            p_acad = gc(col_acad_prom,4)  # Final

            intr   = gc(col_conductual,0)
            uso_cel= str(ws.cell(row=row, column=col_conductual+1).value or 'No') if col_conductual else 'No'
            conf   = gc(col_conductual,2)
            desaf  = gc(col_conductual,3)
            distr  = gc(col_conductual,4)
            falt   = gc(col_conductual,5)
            p_cond = gc(col_conductual,6)

            motiv  = gc(col_emocional,0)
            p_auto = gc(col_emocional,1)
            est_em = gc(col_emocional,2)
            int_fut= gc(col_emocional,3)
            apoy   = gc(col_emocional,4)
            p_emoc = gc(col_emocional,5)

            idx_r  = gc(col_analisis,0)
            niv_r  = str(ws.cell(row=row, column=col_analisis+1).value or 'N/D') if col_analisis else 'N/D'
            tend   = str(ws.cell(row=row, column=col_analisis+2).value or 'igual') if col_analisis else 'igual'
            proy   = gc(col_analisis,3)

            fotos=[x for x in [foto_p1,foto_p2,foto_p3,foto_p4] if x]
            lvs  =[x for x in [lv_p1,lv_p2,lv_p3,lv_p4] if x]
            diss =[x for x in [dis_p1,dis_p2,dis_p3,dis_p4] if x]
            p_foto   = round(sum(fotos)/len(fotos),2) if fotos else None
            p_lv     = round(sum(lvs)/len(lvs),2)     if lvs   else None
            p_diseno = round(sum(diss)/len(diss),2)   if diss  else None
            mods=[x for x in [p_foto,p_lv,p_diseno] if x]
            prom_mod = round(sum(mods)/len(mods),2) if mods else None
            if not p_acad and prom_mod: p_acad = prom_mod

            students.append({
                'id_bj':id_bj,'nombre':str(nombre).strip(),
                'apellido':str(ws.cell(row=row,column=col_apellido).value or '').strip(),
                'curso':str(ws.cell(row=row,column=col_curso).value or '').strip(),
                'grado':str(ws.cell(row=row,column=col_grado).value or '4to').strip(),
                'condicion':str(ws.cell(row=row,column=col_condicion).value or 'ACTIVO'),
                'edad':fl(ws.cell(row=row,column=col_edad).value),
                'puntualidad':punt,'tareas':tar,'participacion':part,
                'comprension':comp,'rendimiento':rend,
                'interrupciones':intr,'uso_celular':uso_cel,'conflictos':conf,
                'desafia_autoridad':desaf,'distraccion':distr,'falta_respeto':falt,
                'motivacion':motiv,'p_auto':p_auto,'estado_emocional':est_em,
                'interes_futuro':int_fut,'apoyo_familiar':apoy,
                'fotografia_p1':foto_p1,'fotografia_p2':foto_p2,
                'fotografia_p3':foto_p3,'fotografia_p4':foto_p4,
                'lv_p1':lv_p1,'lv_p2':lv_p2,'lv_p3':lv_p3,'lv_p4':lv_p4,
                'diseno_p1':dis_p1,'diseno_p2':dis_p2,
                'diseno_p3':dis_p3,'diseno_p4':dis_p4,
                'acad_p1':acad_p1,'acad_p2':acad_p2,'acad_p3':acad_p3,'acad_p4':acad_p4,
                'p_acad':p_acad,'p_cond':p_cond,'p_emocional':p_emoc,
                'p_foto':p_foto,'p_lv':p_lv,'p_diseno':p_diseno,'prom_modulos':prom_mod,
                'indice_riesgo':idx_r,'nivel_riesgo':niv_r,'tendencia':tend,'proyeccion':proy,
            })

        # ── UPSERT ───────────────────────────────────────────────────────
        actualizados=0; nuevos=0

        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row
            conn.create_function("norm", 1, norm)

            for s in students:
                est = None
                # FIX 4: Intento 1 — por ID/cédula (BJ-... o numérico)
                if s['id_bj']:
                    est = conn.execute(
                        "SELECT id FROM estudiantes WHERE cedula=? OR id_evaluacion=? LIMIT 1",
                        (s['id_bj'], s['id_bj'])
                    ).fetchone()
                # FIX 4: Intento 2 — por cédula numérica si viene en campo separado
                if not est and s.get('cedula') and s.get('cedula') != s.get('id_bj'):
                    est = conn.execute(
                        "SELECT id FROM estudiantes WHERE cedula=? LIMIT 1",
                        (str(s['cedula']),)
                    ).fetchone()
                # Intento 3 — por nombre normalizado (cubre casos sin ID)
                if not est:
                    nom1 = norm(s['nombre']).split()[0] if s['nombre'] else ''
                    ape1 = norm(s['apellido']).split()[0] if s['apellido'] else ''
                    if nom1 and ape1:
                        est = conn.execute(
                            "SELECT id FROM estudiantes "
                            "WHERE norm(nombre) LIKE ? AND norm(apellido) LIKE ? LIMIT 1",
                            (f"%{nom1}%", f"%{ape1}%")
                        ).fetchone()
                if not est:
                    conn.execute(
                        "INSERT INTO estudiantes "
                        "(id_evaluacion,cedula,nombre,apellido,curso,grado,condicion) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (s['id_bj'],s['id_bj'],s['nombre'],s['apellido'],
                         s['curso'],s['grado'],s['condicion'])
                    )
                    est_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    nuevos += 1
                else:
                    est_id = est['id']

                # ── Leer valores existentes del estudiante para no pisar notas ya cargadas
                ex = conn.execute("SELECT * FROM estudiantes WHERE id=?", (est_id,)).fetchone()
                ex = dict(ex) if ex else {}

                def _keep(new_val, old_key):
                    """Usa new_val si tiene dato real; si no, conserva el existente."""
                    if new_val is not None and new_val != 0:
                        return new_val
                    return ex.get(old_key)

                conn.execute("""
                    UPDATE estudiantes SET
                        id_evaluacion=?,cedula=?,nombre=?,apellido=?,
                        curso=?,grado=?,condicion=?,edad=?,
                        puntualidad=?,tareas=?,participacion=?,comprension=?,
                        rendimiento=?,interrupciones=?,uso_celular=?,
                        conflictos=?,desafia_autoridad=?,distraccion=?,falta_respeto=?,
                        motivacion=?,p_auto=?,estado_emocional=?,
                        interes_futuro=?,apoyo_familiar=?,
                        fotografia_p1=?,fotografia_p2=?,fotografia_p3=?,fotografia_p4=?,
                        lv_p1=?,lv_p2=?,lv_p3=?,lv_p4=?,
                        diseno_p1=?,diseno_p2=?,diseno_p3=?,diseno_p4=?,
                        acad_p1=?,acad_p2=?,acad_p3=?,acad_p4=?,
                        p_acad=?,p_cond=?,p_emocional=?,
                        p_foto=?,p_lv=?,p_diseno=?,prom_modulos=?,
                        indice_riesgo=?,nivel_riesgo=?,tendencia=?,proyeccion=?
                    WHERE id=?
                """, (
                    s['id_bj'],s['id_bj'],s['nombre'],s['apellido'],
                    s['curso'],s['grado'],s['condicion'],s['edad'],
                    s['puntualidad'],s['tareas'],s['participacion'],s['comprension'],
                    s['rendimiento'],s['interrupciones'],s['uso_celular'],
                    s['conflictos'],s['desafia_autoridad'],s['distraccion'],s['falta_respeto'],
                    s['motivacion'],s['p_auto'],s['estado_emocional'],
                    s['interes_futuro'],s['apoyo_familiar'],
                    _keep(s['fotografia_p1'],'fotografia_p1'),
                    _keep(s['fotografia_p2'],'fotografia_p2'),
                    _keep(s['fotografia_p3'],'fotografia_p3'),
                    _keep(s['fotografia_p4'],'fotografia_p4'),
                    _keep(s['lv_p1'],'lv_p1'), _keep(s['lv_p2'],'lv_p2'),
                    _keep(s['lv_p3'],'lv_p3'), _keep(s['lv_p4'],'lv_p4'),
                    _keep(s['diseno_p1'],'diseno_p1'), _keep(s['diseno_p2'],'diseno_p2'),
                    _keep(s['diseno_p3'],'diseno_p3'), _keep(s['diseno_p4'],'diseno_p4'),
                    _keep(s['acad_p1'],'acad_p1'), _keep(s['acad_p2'],'acad_p2'),
                    _keep(s['acad_p3'],'acad_p3'), _keep(s['acad_p4'],'acad_p4'),
                    s['p_acad'],s['p_cond'],s['p_emocional'],
                    s['p_foto'],s['p_lv'],s['p_diseno'],s['prom_modulos'],
                    s['indice_riesgo'],s['nivel_riesgo'],s['tendencia'],s['proyeccion'],
                    est_id
                ))
                actualizados += 1

            conn.commit()

        with sqlite3.connect(DATABASE, timeout=10) as c2:
            total     = c2.execute("SELECT COUNT(*) FROM estudiantes").fetchone()[0]
            con_notas = c2.execute(
                "SELECT COUNT(*) FROM estudiantes WHERE p_acad>0 OR acad_p1>0"
            ).fetchone()[0]

        cache_bust()
        return jsonify({
            "status":"success",
            "version_detectada": f"header fila {header_row}, Nombre en C{nombre_col}",
            "actualizados":actualizados,"nuevos":nuevos,
            "total_sistema":total,"con_notas":con_notas,
            "mensaje":(f"{actualizados} estudiantes actualizados. "
                       f"Total en sistema: {total} ({con_notas} con notas).")
        })

    except Exception as ex:
        logger.error(f"[ML] Error clustering: {ex}", exc_info=True)
        return jsonify({"error": "Error al calcular métricas. Intenta de nuevo."}), 500


@estudiantes_bp.route("/api/datos")
@login_required
def api_datos():
    """
    Devuelve todos los estudiantes con tiene_notas=True si tienen
    calificaciones en la Plantilla O en materias_calificaciones dinamicas.
    Soporta filtros: grado, mencion, ciclo (primer_ciclo/segundo_ciclo).
    """
    grado      = request.args.get("grado", "").strip()
    mencion    = request.args.get("mencion", "").strip()
    ciclo      = request.args.get("ciclo", "").strip()
    solo_notas = request.args.get("solo_con_notas", "0") == "1"

    # ── Aplicar restricciones de acceso por rol ──────────────────────────
    _usr = _get_profesor()
    es_profesor = False
    cache_key = None
    _prof_grados = []
    _prof_menciones = []
    _prof_filtro_men = False

    if _usr:
        rol_n = _normalizar_rol(_usr.get("rol",""))
        ciclo_usr = _ciclo_del_rol(rol_n)

        if rol_n == "profesor":
            # Profesor: VER SOLO sus estudiantes — filtro forzado, sin caché
            # Usa _resolver_alcance_profesor para manejar multi-grado y mención con acento
            es_profesor = True
            _alcance_prof = _resolver_alcance_profesor(_usr)
            _prof_grados   = _alcance_prof["grados"]    # lista de grados canónicos
            _prof_menciones = _alcance_prof["menciones"] # lista de menciones canónicas
            _prof_filtro_men = _alcance_prof["filtro_mencion"]
        elif ciclo_usr:
            # Coord/psicóloga de ciclo: restringir a su ciclo
            if not ciclo:
                ciclo = ciclo_usr
            # Caché por ciclo
            if not grado and not mencion and not solo_notas:
                cache_key = f"api_datos_ciclo_{ciclo}"
        else:
            # Directora / coordinador_general / admin: ven TODO
            if not grado and not mencion and not solo_notas and not ciclo:
                cache_key = "api_datos_all"

    # Intentar servir desde caché (solo para roles sin filtro forzado)
    if cache_key and not es_profesor:
        cached = cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)

    seccion = request.args.get("seccion", "").strip()

    query  = "SELECT * FROM estudiantes WHERE 1=1"
    params = []

    # Profesores: usar alcance resuelto (multi-grado + mención canónica con acento)
    if es_profesor:
        if _prof_grados:
            query += " AND (" + " OR ".join(["upper(grado) LIKE upper(?)" for _ in _prof_grados]) + ")"
            params.extend([f"%{g}%" for g in _prof_grados])
        if _prof_filtro_men and _prof_menciones:
            query += " AND (" + " OR ".join(["upper(curso) LIKE upper(?)" for _ in _prof_menciones]) + ")"
            params.extend([f"%{m}%" for m in _prof_menciones])
    else:
        if grado:
            query  += " AND upper(grado) LIKE upper(?)"
            params.append(f"%{grado}%")
        if mencion:
            query  += " AND upper(curso) LIKE upper(?)"
            params.append(f"%{mencion}%")
    if ciclo:
        query  += " AND ciclo=?"
        params.append(ciclo)
    if seccion:
        query  += " AND upper(seccion)=upper(?)"
        params.append(seccion)

    query += " ORDER BY apellido, nombre"

    import time as _t
    _t0 = _t.time()
    logger.info(f"[api_datos] inicio — query: {query!r} params: {params}")
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(query, params).fetchall()
            logger.info(f"[api_datos] estudiantes: {len(rows)} ({_t.time()-_t0:.2f}s)")

            # IDs con materias dinamicas cargadas
            ids_con_materias = set(
                r[0] for r in conn.execute(
                    "SELECT DISTINCT estudiante_id FROM materias_calificaciones"
                ).fetchall()
            )
            logger.info(f"[api_datos] ids_con_materias: {len(ids_con_materias)} ({_t.time()-_t0:.2f}s)")

            resultado = []
            for r in rows:
                d = dict(r)
                tiene_plantilla = bool(
                    (d.get('p_acad') or 0) > 0 or
                    (d.get('acad_p1') or 0) > 0 or
                    (d.get('fotografia_p1') or 0) > 0
                )
                tiene_materias = d.get('id') in ids_con_materias
                d['tiene_notas']    = tiene_plantilla or tiene_materias
                d['tiene_plantilla'] = tiene_plantilla
                d['tiene_materias']  = tiene_materias

                if solo_notas and not d['tiene_notas']:
                    continue
                resultado.append(d)

            # ── Calcular categoría de alerta dinámica ─────────────
            for d in resultado:
                p = float(d.get('p_acad') or 0)
                if not d.get('tiene_notas') or p == 0:
                    d['categoria']    = 'SIN DATOS'
                    d['alerta_nivel'] = 0
                    d['alerta_color'] = ''
                elif p < 70:
                    d['categoria']    = 'ALERTA CRÍTICA'
                    d['alerta_nivel'] = 2
                    d['alerta_color'] = 'rojo'
                elif p < 80:
                    d['categoria']    = 'ESTUDIANTE EN OBSERVACIÓN'
                    d['alerta_nivel'] = 1
                    d['alerta_color'] = 'naranja'
                elif p >= 85:
                    d['categoria']    = 'EXCELENTE'
                    d['alerta_nivel'] = 0
                    d['alerta_color'] = 'verde'
                else:
                    d['categoria']    = 'REGULAR'
                    d['alerta_nivel'] = 0
                    d['alerta_color'] = ''

            # Ordenar: alertas críticas primero, luego observación, luego el resto
            resultado.sort(key=lambda x: (
                0 if x['tiene_notas'] else 1,
                -(x.get('alerta_nivel') or 0),
                -(x.get('p_acad') or 0),
                x.get('apellido','')
            ))

            # Guardar en caché SOLO si es resultado completo (no filtrado por profesor)
            if cache_key and not es_profesor:
                cache_set(cache_key, resultado)

            return jsonify(resultado)
        except Exception as ex:
            logger.error(f"api_datos error: {ex}")
            return jsonify([])


@estudiantes_bp.route("/perfil/<int:id>")
@login_required
def perfil_estudiante(id):
    # Si es padre, verificar que tiene vínculo con este estudiante
    u = get_usuario()
    if u.get("rol") == "padre":
        with sqlite3.connect(DATABASE, timeout=10) as _vc:
            vinculo = _vc.execute(
                "SELECT id FROM vinculos_padre_estudiante WHERE padre_id=? AND estudiante_id=?",
                (u["id"], id)
            ).fetchone()
        if not vinculo:
            return redirect("/portal-padres")

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        estudiante = conn.execute(
            "SELECT * FROM estudiantes WHERE id = ?", (id,)
        ).fetchone()
    if estudiante:
        e = dict(estudiante)

        # ── Completar notas desde materias_calificaciones si campos directos están vacíos ──
        with sqlite3.connect(DATABASE, timeout=10) as conn2:
            conn2.row_factory = sqlite3.Row
            mats = conn2.execute("""
                SELECT materia, p1, p2, p3, p4, promedio, fecha_carga, profesor
                FROM materias_calificaciones
                WHERE estudiante_id = ?
            """, (id,)).fetchall()

        if mats:
            # ══ MAPA UNIVERSAL DE MATERIAS ══════════════════════════════════
            # Clave: nombre en minúsculas como puede venir en la BD
            # Valor: prefijo del campo directo en la tabla estudiantes
            # Cubre todas las menciones, variantes con/sin tilde, typos comunes
            MODULO_MAP_PERFIL = {
                # ── Fotografía (Multimedia + Artes Visuales 5to) ───────────
                'fotografía':                       'fotografia',
                'fotografia':                       'fotografia',
                'foto':                             'fotografia',
                'fotografía digital':               'fotografia',
                'fotografia digital':               'fotografia',
                'introducción fotografía':          'fotografia',
                'introduccion fotografia':          'fotografia',
                'introducción fotografía digital':  'fotografia',
                'introduccion fotografia digital':  'fotografia',
                'introducción a la fotografía':     'fotografia',
                'introduccion a la fotografia':     'fotografia',
                'introducción a la fotografía digital': 'fotografia',
                'introduccion a la fotografia digital': 'fotografia',
                'fotografía artística':             'fotografia',
                'fotografia artistica':             'fotografia',
                # ── Lenguaje Visual / Lenguaje Plástico ────────────────────
                'lenguaje visual':                                          'lv',
                'lenguaje visual, dibujo y creación de personajes':         'lv',
                'lenguaje visual dibujo y creacion de personajes':          'lv',
                'lenguaje visual, dibujo y creacion de personajes':         'lv',
                'lenguaje visual y principios de diseño':                   'lv',
                'lenguaje visual y principios de diseno':                   'lv',
                'lenguaje visual y principios del diseño':                  'lv',
                'lenguaje visual y principios del diseno':                  'lv',
                # Variantes reales vistas en BD ↓
                'lenguaje visual y principios del diseño artesanal':        'lv',
                'lenguaje visual y principios del diseno artesanal':        'lv',
                'lenguaje visual artesanal':                                'lv',
                'lenguaje plástico y visual':                               'lv',
                'lenguaje plastico y visual':                               'lv',
                'lv':                                                       'lv',
                # ── Diseño Básico (Multimedia) ──────────────────────────────
                'diseño básico y expresión visual':                  'diseno',
                'diseño basico y expresion visual':                  'diseno',
                'diseño básico y expresion visual':                  'diseno',
                'diseno basico y expresion visual':                  'diseno',
                'diseño básico':                                     'diseno',
                'diseño basico':                                     'diseno',
                'diseño':                                            'diseno',
                'diseno':                                            'diseno',
                # Variantes reales vistas en BD ↓
                'dibujo técnico':                                    'diseno',
                'dibujo tecnico':                                    'diseno',
                'dibujo técnico y artístico':                        'diseno',
                'dibujo tecnico y artistico':                        'diseno',
                'principios de dibujo, pintura y creatividad':       'diseno',
                'principios de dibujo pintura y creatividad':        'diseno',
                'pintura y técnicas mixtas':                         'diseno',
                'pintura y tecnicas mixtas':                         'diseno',
                'pintura':                                           'diseno',
                'lenguaje plástico':                                 'diseno',
                'lenguaje plastico':                                 'diseno',
            }

            # ══ MATERIAS ACADÉMICAS ESTÁNDAR (7 áreas para todas menciones) ═
            # Cualquier variante de nombre que represente estas 7 materias
            MATS_ACAD_STD = {
                # Lengua Española
                'lengua española', 'lengua espanola', 'español', 'espanol',
                'lengua', 'lengua y literatura', 'castellano',
                # Inglés — incluyendo "Idioma Inglés" que usan algunos maestros
                'inglés', 'ingles', 'english', 'idioma extranjero',
                'idioma inglés', 'idioma ingles', 'idioma inglés',
                # Matemática
                'matemática', 'matematica', 'matemáticas', 'matematicas', 'math',
                'matemática general', 'matematica general',
                # Ciencias Sociales — SOLO sociales, no confundir con naturales
                'ciencias sociales', 'sociales', 'cs. sociales', 'cs sociales',
                'historia y geografía', 'historia y geografia',
                # Ciencias de la Naturaleza — variantes comunes
                'ciencias de la naturaleza', 'ciencias naturales', 'naturales',
                'cs. naturales', 'cs naturales', 'ciencias nat.',
                'ciencias nat', 'biología', 'biologia',
                # Formación Integral Humana y Religiosa — TODAS las variantes reales
                'formación integral humana y religiosa',
                'formacion integral humana y religiosa',
                'formación humana y religiosa',
                'formacion humana y religiosa',
                'formación religiosa', 'formacion religiosa',
                'religión', 'religion',
                'f.i.h.r.', 'fihr',   # ← abreviatura real usada en la BD
                'formacion integral', 'formación integral',
                'ed. religiosa', 'educacion religiosa', 'educación religiosa',
                'formación integral humana',
                # Educación Física — variantes comunes
                'educación física', 'educacion fisica',
                'ed. física', 'ed fisica', 'ed. fis.',
                'educacion fiscia', 'educacion fis',
                'fisica', 'física',
            }

            materias_extras = {}
            promedios_acad  = []
            ASISTENCIA_KEYS = {'asistencia', 'asistencias', 'asist.', 'attendence'}

            for mat in mats:
                mn = mat['materia'].lower().strip()

                # 1. ¿Es asistencia?
                if any(a in mn for a in ASISTENCIA_KEYS):
                    for pi in [1,2,3,4]:
                        key = f'asistencia_p{pi}'
                        if not e.get(key):
                            e[key] = mat[f'p{pi}'] or 0
                    continue

                # 2. ¿Es una materia técnica de Multimedia con campo directo?
                col = MODULO_MAP_PERFIL.get(mn)
                if not col:
                    for k, v in MODULO_MAP_PERFIL.items():
                        if k in mn or mn in k:
                            col = v
                            break

                if col:
                    p_col = 'p_foto' if col == 'fotografia' else f'p_{col}'
                    for pi in [1,2,3,4]:
                        key = f'{col}_p{pi}'
                        if not e.get(key):
                            e[key] = mat[f'p{pi}'] or 0
                    if not e.get(p_col) and mat['promedio']:
                        e[p_col] = mat['promedio']
                    continue

                # 3. ¿Es materia académica estándar?
                es_acad = any(a in mn or mn in a for a in MATS_ACAD_STD)
                if es_acad:
                    if mat['promedio'] and mat['promedio'] > 0:
                        promedios_acad.append(mat['promedio'])
                    for pi in [1,2,3,4]:
                        key = f'acad_p{pi}'
                        if not e.get(key) and mat[f'p{pi}']:
                            e[key] = mat[f'p{pi}']
                    continue

                # 4. Materia técnica de otra mención (Artes Visuales, Teatro, Música, Danza)
                # Guardar individualmente con el nombre real para el boletín
                nombre_mat = mat['materia'].strip()
                materias_extras[nombre_mat] = {
                    'p1': mat['p1'], 'p2': mat['p2'],
                    'p3': mat['p3'], 'p4': mat['p4'],
                    'promedio': mat['promedio'],
                }
                if mat['promedio'] and mat['promedio'] > 0:
                    promedios_acad.append(mat['promedio'])

            # Exponer materias_extras en el dict del estudiante
            if materias_extras:
                e['materias_extras'] = materias_extras

            # Recalcular prom_modulos si fue completado desde Excel
            mods = [e.get(k) or 0 for k in ['p_foto','p_lv','p_diseno'] if (e.get(k) or 0) > 0]
            if not mods and e.get('materias_extras'):
                mods = [v.get('promedio') or 0 for v in e['materias_extras'].values() if (v.get('promedio') or 0) > 0]
            if mods and not e.get('prom_modulos'):
                e['prom_modulos'] = round(sum(mods)/len(mods), 1)

            # Calcular p_acad desde materias regulares si estaba en 0
            if not e.get('p_acad') and promedios_acad:
                e['p_acad'] = round(sum(promedios_acad)/len(promedios_acad), 1)

        # Campos de texto con fallback
        texto_defaults = {
            'uso_celular': 'No', 'nivel_riesgo': 'N/D',
            'grado': '4to', 'condicion': 'ACTIVO',
            'tendencia': 'igual', 'categoria': '',
            'reporte': '', 'color': '', 'ia_analisis': None,
            'nombre': '', 'apellido': '', 'curso': '',
        }
        for k, v in texto_defaults.items():
            if e.get(k) is None:
                e[k] = v

        # Todos los campos numéricos: None → 0.0
        # Esto evita el error ">= no compatible con NoneType"
        campos_numericos = [
            'p_acad','p_foto','p_lv','p_diseno','p_cond','p_auto',
            'p_emocional','prom_modulos','asistencia','proyeccion',
            'puntualidad','tareas','participacion','comprension','rendimiento',
            'interrupciones','conflictos','desafia_autoridad','distraccion',
            'falta_respeto','motivacion','estado_emocional','interes_futuro',
            'apoyo_familiar','indice_riesgo','edad','silencioso',
            'fotografia_p1','fotografia_p2','fotografia_p3','fotografia_p4',
            'lv_p1','lv_p2','lv_p3','lv_p4',
            'diseno_p1','diseno_p2','diseno_p3','diseno_p4',
            'asistencia_p1','asistencia_p2','asistencia_p3','asistencia_p4',
            'acad_p1','acad_p2','acad_p3','acad_p4',
        ]
        for k in campos_numericos:
            if e.get(k) is None:
                e[k] = 0.0

        # ── Calcular proyección dinamicamente ─────────────────────────────
        # No depende de la columna de la Plantilla (que el maestro no llena)
        # Usa los periodos reales para proyectar la tendencia al cierre del año
        acad_p1 = e.get('acad_p1') or 0
        acad_p2 = e.get('acad_p2') or 0
        acad_p3 = e.get('acad_p3') or 0
        acad_p4 = e.get('acad_p4') or 0
        p_acad  = e.get('p_acad')  or 0

        periodos_con_nota = [p for p in [acad_p1,acad_p2,acad_p3,acad_p4] if p > 0]

        if len(periodos_con_nota) >= 2:
            # Tendencia entre los dos últimos períodos disponibles
            ultimo    = periodos_con_nota[-1]
            penultimo = periodos_con_nota[-2]
            delta     = ultimo - penultimo
            # Proyectar con amortiguación: la tendencia se modera con más períodos
            amort     = 0.6 if len(periodos_con_nota) >= 3 else 0.8
            proyeccion_calc = round(min(max(ultimo + delta * amort, 0), 100), 1)
        elif len(periodos_con_nota) == 1:
            # Solo un período: proyectar igual + ajuste emocional/conductual
            base = periodos_con_nota[0]
            p_cond  = e.get('p_cond') or 0
            p_auto  = e.get('p_auto') or 0
            # Si conducta y autoestima son buenas, proyectar ligeramente al alza
            boost = 0
            if p_cond  >= 75: boost += 1.5
            if p_auto  >= 70: boost += 1.0
            if p_cond  <  60: boost -= 2.0
            proyeccion_calc = round(min(max(base + boost, 0), 100), 1)
        elif p_acad > 0:
            # Tiene promedio final pero sin desglose por período
            proyeccion_calc = p_acad
        else:
            # Sin datos académicos — proyectar con módulos técnicos si existen
            mods = [e.get(m) or 0 for m in ['p_foto','p_lv','p_diseno'] if (e.get(m) or 0) > 0]
            proyeccion_calc = round(sum(mods)/len(mods), 1) if mods else 0.0

        # Solo sobreescribir si la DB tiene 0 o None
        if not e.get('proyeccion') or e.get('proyeccion', 0) == 0:
            e['proyeccion'] = proyeccion_calc

        # Recalcular tendencia
        if len(periodos_con_nota) >= 2:
            if periodos_con_nota[-1] > periodos_con_nota[-2]:
                e['tendencia'] = 'subiendo'
            elif periodos_con_nota[-1] < periodos_con_nota[-2]:
                e['tendencia'] = 'bajando'
            else:
                e['tendencia'] = 'igual'

        # Marcar si tiene notas para que el template pueda mostrar aviso
        e['tiene_notas'] = bool(e.get('p_acad', 0) > 0 or e.get('acad_p1', 0) > 0)

        # Calcular índices dinámicos desde reportes y cuaderno
        with sqlite3.connect(DATABASE, timeout=10) as _conn:
            _conn.row_factory = sqlite3.Row
            ind_cond  = _calcular_indice_conductual(_conn, id)
            ind_psico = _calcular_bienestar_emocional(_conn, id)
        if not e.get('p_cond') or e['p_cond'] == 0:
            e['p_cond'] = ind_cond
        else:
            e['p_cond'] = round(min(e['p_cond'], ind_cond), 1)
        if not e.get('p_auto') or e['p_auto'] == 0:
            e['p_auto'] = ind_psico

        # ── Restringir campos sensibles según rol del visualizador ──────────
        viewer_rol = _normalizar_rol(u.get("rol", ""))
        _ROLES_COORD = {"coordinador_general", "coordinador_primer_ciclo", "coordinador_segundo_ciclo"}
        _ROLES_PSICO_DIR = {"psicologa_primer_ciclo", "psicologa_segundo_ciclo", "directora"}
        if viewer_rol in _ROLES_COORD:
            # Coordinadores solo ven información académica, no psicológica/conductual sensible
            _CAMPOS_SENSIBLES = [
                "p_emocional", "p_auto", "motivacion", "estado_emocional",
                "silencioso", "indice_riesgo", "nivel_riesgo",
                "apoyo_familiar", "ind_conducta", "ind_psico",
                "interrupciones", "conflictos", "desafia_autoridad",
                "distraccion", "falta_respeto", "ia_analisis",
            ]
            for campo in _CAMPOS_SENSIBLES:
                e[campo] = None

        # Detectar si el estudiante fue promovido (notas del año pertenecen a otro grado)
        grado_est = (e.get("grado") or "").upper()
        anio_act  = _anio_escolar_actual()
        with sqlite3.connect(DATABASE, timeout=10) as _dc:
            _dc.row_factory = sqlite3.Row
            _fue_promovido = _dc.execute("""
                SELECT 1 FROM materias_calificaciones
                WHERE estudiante_id=? AND anio_escolar=?
                  AND grado IS NOT NULL AND UPPER(grado) != ?
                LIMIT 1
            """, (id, anio_act, grado_est)).fetchone()
            if not _fue_promovido:
                _fue_promovido = _dc.execute("""
                    SELECT 1 FROM promociones
                    WHERE estudiante_id=? AND anio_escolar=? AND estado='PROMOVIDO'
                    LIMIT 1
                """, (id, anio_act)).fetchone()

        return render_template("perfil.html", e=e, current_user=get_usuario(),
                               fue_promovido=bool(_fue_promovido),
                               anio_escolar=anio_act)
    return "Estudiante no encontrado", 404


# ── GROQ: Generar análisis pedagógico IA ────────────────────────────────────


@estudiantes_bp.route("/api/analisis-ia/<int:id>", methods=["POST"])
@login_required
@rate_limited(max_calls=20, window=60)
def generar_analisis_ia(id):
    _viewer = get_usuario()
    _vrol = _normalizar_rol(_viewer.get("rol", ""))
    if _vrol in {"coordinador_general", "coordinador_primer_ciclo", "coordinador_segundo_ciclo"}:
        return jsonify({"error": "Sin permisos para ver análisis psicopedagógico"}), 403
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        e = conn.execute(
            "SELECT * FROM estudiantes WHERE id = ?", (id,)
        ).fetchone()

    if not e:
        return jsonify({"error": "Estudiante no encontrado"}), 404

    e = dict(e)

    if e.get("ia_analisis"):
        return jsonify({"analisis": e["ia_analisis"], "cached": True})

    try:
        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Eres un orientador pedagógico experto en la Modalidad de Artes "
                        "del bachillerato dominicano. Respondes siempre en español, "
                        "de forma estructurada y profesional."
                    )
                },
                {
                    "role": "user",
                    "content": construir_prompt(e)
                }
            ],
            temperature=0.6,
            max_tokens=600,
            top_p=0.9
        )

        analisis = completion.choices[0].message.content.strip()

        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.execute(
                "UPDATE estudiantes SET ia_analisis = ? WHERE id = ?",
                (analisis, id)
            )
            conn.commit()

        return jsonify({"analisis": analisis, "cached": False})

    except Exception as ex:
        logger.error(f"[IA] Error Groq estudiantes: {ex}")
        return jsonify({"error": "Error al generar el análisis. Intenta de nuevo."}), 500


@estudiantes_bp.route("/api/analisis-ia/<int:id>", methods=["DELETE"])
@login_required
def limpiar_analisis_ia(id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute(
            "UPDATE estudiantes SET ia_analisis = NULL WHERE id = ?", (id,)
        )
        conn.commit()
    return jsonify({"status": "cleared"})


# ── SERVICIO DE PLANIFICACIÓN DOCENTE ───────────────────────────────────────

# Base curricular oficial MINERD extraída del documento
# ══════════════════════════════════════════════════════════════════════════════
#  CURRÍCULO OFICIAL MINERD — MODALIDAD ARTES — C.E. Benito Juárez
#  Menciones: Multimedia, Artes Visuales, Música, Teatro, Danza
#  Segundo Ciclo: 4TO, 5TO, 6TO
#  Fuente: DEMA - Dirección de Educación Modalidad en Artes
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/competencias-ia/<int:id>", methods=["POST"])
@login_required
def evaluar_competencias_ia(id):
    """Genera evaluación por competencias usando datos reales de indicadores."""
    try:
        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row
            e = conn.execute("SELECT * FROM estudiantes WHERE id=?", (id,)).fetchone()

        if not e:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        e = dict(e)

        # Construir resumen de notas por período
        def pv(campo):
            v = e.get(campo)
            return round(float(v), 1) if v and float(v) > 0 else None

        bloques = []

        # Fotografía
        f_vals = [(f"P{i+1}", pv(f"fotografia_p{i+1}")) for i in range(4)]
        f_con  = [(l,v) for l,v in f_vals if v]
        if f_con:
            bloques.append("FOTOGRAFÍA: " + " | ".join(f"{l}={v}" for l,v in f_con))

        # Lenguaje Visual
        lv_vals = [(f"P{i+1}", pv(f"lv_p{i+1}")) for i in range(4)]
        lv_con  = [(l,v) for l,v in lv_vals if v]
        if lv_con:
            bloques.append("LENGUAJE VISUAL: " + " | ".join(f"{l}={v}" for l,v in lv_con))

        # Diseño
        d_vals = [(f"P{i+1}", pv(f"diseno_p{i+1}")) for i in range(4)]
        d_con  = [(l,v) for l,v in d_vals if v]
        if d_con:
            bloques.append("DISEÑO: " + " | ".join(f"{l}={v}" for l,v in d_con))

        # Asistencia
        a_vals = [(f"P{i+1}", pv(f"asistencia_p{i+1}")) for i in range(4)]
        a_con  = [(l,v) for l,v in a_vals if v]
        if a_con:
            bloques.append("ASISTENCIA: " + " | ".join(f"{l}={v}%" for l,v in a_con))

        # Promedio Académico por período
        ac_vals = [(f"P{i+1}", pv(f"acad_p{i+1}")) for i in range(4)]
        ac_con  = [(l,v) for l,v in ac_vals if v]
        if ac_con:
            bloques.append("PROM. ACADÉMICO: " + " | ".join(f"{l}={v}" for l,v in ac_con))

        # Comportamiento
        comp_items = [
            ("Puntualidad", pv("puntualidad")), ("Tareas", pv("tareas")),
            ("Participación", pv("participacion")), ("Comprensión", pv("comprension")),
            ("Rendimiento", pv("rendimiento")),
        ]
        comp_con = [(l,v) for l,v in comp_items if v]
        if comp_con:
            bloques.append("COMPORTAMIENTO ACADÉMICO: " + " | ".join(f"{l}={v}" for l,v in comp_con))

        # Emocional
        emoc_items = [
            ("Motivación", pv("motivacion")), ("Autoestima", pv("p_auto")),
            ("Estado Emocional", pv("estado_emocional")), ("Interés Futuro", pv("interes_futuro")),
            ("Apoyo Familiar", pv("apoyo_familiar")),
        ]
        emoc_con = [(l,v) for l,v in emoc_items if v]
        if emoc_con:
            bloques.append("INDICADORES EMOCIONALES: " + " | ".join(f"{l}={v}" for l,v in emoc_con))

        if not bloques:
            return jsonify({"error": "Este estudiante no tiene datos de calificaciones cargados aún."}), 400

        datos_txt = "\n".join(bloques)
        grado = e.get("grado", "4to")
        nombre_completo = _anonimizar_estudiante(e.get('id', 0))  # ANON: nunca nombre real a IA

        prompt = f"""Eres un especialista en evaluación por competencias del bachillerato dominicano,
Modalidad de Artes, mención Multimedia. Genera un informe de competencias basado en datos reales.

ESTUDIANTE: {nombre_completo}
GRADO: {grado} Multimedia
CATEGORÍA: {e.get('categoria','Estable')} | Riesgo: {e.get('nivel_riesgo','Bajo')}
PROMEDIO FINAL: {round(float(e.get('p_acad') or 0), 1)} | Conductual: {round(float(e.get('p_cond') or 0), 1)} | Emocional: {round(float(e.get('p_emocional') or 0), 1)}

CALIFICACIONES:
{datos_txt}

Genera un informe con EXACTAMENTE estas 4 secciones:

1. NIVEL DE COMPETENCIA ALCANZADO
Por cada materia presente, indica el nivel real: Inicial / En desarrollo / Competente / Destacado.
Basa el nivel en los promedios reales. Menciona los datos específicos.

2. PROGRESIÓN ENTRE PERÍODOS
Analiza la evolución P1→P2 (y P3/P4 si existen). ¿Qué mejoró? ¿Qué bajó?
¿La tendencia es positiva o negativa? Sé específico con los números.

3. INDICADORES CRÍTICOS
Lista los 2-3 aspectos con notas más bajas o mayor preocupación.
Para cada uno sugiere una acción concreta y práctica para el docente.

4. PROYECCIÓN P3 Y P4
Con base en la tendencia actual, ¿qué puede lograr al final del año?
¿Qué debe priorizar el docente en los próximos períodos?

Usa lenguaje técnico-pedagógico. Sé preciso con los datos. Máximo 380 palabras."""

        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "Eres un evaluador pedagógico experto en la Modalidad de Artes del bachillerato dominicano. Respondes en español con análisis precisos basados en datos reales."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5,
            max_tokens=700
        )
        resultado = completion.choices[0].message.content.strip()
        return jsonify({"resultado": resultado})

    except Exception as ex:
        logger.error(f"[IA] Error generando análisis: {ex}")
        return jsonify({"error": "Error al generar el análisis. Intenta de nuevo."}), 500


@estudiantes_bp.route("/api/promover/<int:id>", methods=["POST"])
@login_required
def promover_estudiante(id):
    """Cambia el grado del estudiante (4to ↔ 5to)."""
    try:
        data = request.json or {}
        nuevo_grado = data.get("grado", "5to")
        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.execute("UPDATE estudiantes SET grado=? WHERE id=?", (nuevo_grado, id))
            conn.commit()
        return jsonify({"ok": True, "grado": nuevo_grado})
    except Exception as ex:
        logger.error(f"[ESTUDIANTES] Error actualizando grado: {ex}")
        return jsonify({"error": "Error al actualizar el grado. Intenta de nuevo."}), 500



# ── CARGA DE LISTADO MAESTRO DEL LICEO ───────────────────────────────────────


@estudiantes_bp.route("/api/cargar-listado", methods=["POST"])
@login_required
@rate_limited(max_calls=10, window=3600)
def cargar_listado():
    """
    Carga el Listado oficial del liceo (LISTADO-AÑO-XXXX.xlsx).
    - Lee hojas 4TO, 5TO, 6TO con todas sus menciones
    - Guarda identidad en registro_liceo (cédula, nombre, apellido, grado, mención)
    - Crea perfil vacío en 'estudiantes' para TODOS los alumnos
      (aparecen en dashboard aunque no tengan notas todavía)
    - Sin cédula → ID provisional, sistema no se detiene
    - Carga segura: re-ejecutable sin duplicar ni borrar datos
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({"error": "Formato no válido. Sube el archivo LISTADO Excel del liceo"}), 400

    # ── Filtro de perfil de profesor ─────────────────────────────────────────
    # Si el usuario es profesor, el sistema solo cargará hojas/alumnos
    # correspondientes al grado y mención asignados a su perfil.
    _prof_listado = _get_profesor()
    _prof_grado   = ""
    _prof_mencion = ""
    if _prof_listado and _prof_listado.get("rol") == "profesor":
        _prof_grado   = (_prof_listado.get("grado") or "").upper()
        _prof_mencion = (_prof_listado.get("mencion") or "").upper()

    try:
        from openpyxl import load_workbook
        from datetime import datetime as dt
        import io

        wb = load_workbook(io.BytesIO(file.read()), data_only=True)

        # Detectar hojas del listado — segundo ciclo (4to-6to) y primer ciclo (1ro-3ro)
        grados_objetivo = [s for s in wb.sheetnames
                           if any(g in s.upper() for g in
                                  ['4TO', '5TO', '6TO', '1RO', '2DO', '3RO', '1ER', '2DO', '3ER'])]

        if not grados_objetivo:
            return jsonify({"error": "No se encontraron hojas de grados reconocidos. "
                                     "Verifica que es el Listado correcto del liceo."}), 400

        # Si es profesor, filtrar solo las hojas de su grado y mención
        if _prof_grado:
            grados_filtrados = [s for s in grados_objetivo if _prof_grado in s.upper()]
            if not grados_filtrados:
                return jsonify({
                    "error": f"Esta lista no contiene hojas del grado asignado a tu perfil ({_prof_grado}). "
                             f"Solo puedes cargar listas de tu grado. Verifica el archivo."
                }), 403
            grados_objetivo = grados_filtrados
        if _prof_mencion:
            grados_filtrados = [s for s in grados_objetivo if _prof_mencion in s.upper()]
            if grados_filtrados:  # Only restrict if matching sheets exist
                grados_objetivo = grados_filtrados

        resumen = {}
        sin_cedula_lista = []

        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row

            for sheet_name in grados_objetivo:
                ws    = wb[sheet_name]
                grado = sheet_name.replace('(2)', '').replace('(3)', '').strip()
                # FIX 1: leer mención desde encabezado antes de iterar filas
                mencion_header = _detectar_mencion_listado(ws)
                mencion_actual = mencion_header if mencion_header else 'GENERAL'
                conteo = 0

                for row in range(1, ws.max_row + 1):
                    col1 = ws.cell(row=row, column=1).value
                    col8 = ws.cell(row=row, column=8).value

                    # ── Detectar cambio de mención ────────────────────────
                    if isinstance(col8, str) and len(col8) > 2:
                        palabras_clave = ['MUSICA', 'MÚSICA', 'MULTIMEDIA',
                                          'TEATRO', 'VISUAL', 'DANZA',
                                          'ARTES', 'BACHILLER']
                        if any(p in col8.upper() for p in palabras_clave):
                            mencion_actual = col8.strip()
                            continue

                    # ── Fila de estudiante ────────────────────────────────
                    if not (isinstance(col1, (int, float)) and 1 <= col1 <= 60):
                        continue

                    nombre   = str(ws.cell(row=row, column=2).value or '').strip()
                    apellido = str(ws.cell(row=row, column=3).value or '').strip()
                    if not nombre or nombre.startswith('='):
                        continue

                    sexo     = str(ws.cell(row=row, column=4).value or '').strip()
                    nac_raw  = ws.cell(row=row, column=5).value
                    ced_raw  = ws.cell(row=row, column=7).value
                    telefono = str(ws.cell(row=row, column=8).value or '').strip()

                    # Fecha y edad
                    nac_str   = ''
                    edad_calc = None
                    if isinstance(nac_raw, dt):
                        nac_str   = nac_raw.strftime('%d/%m/%Y')
                        edad_calc = (dt.now() - nac_raw).days // 365
                    elif nac_raw:
                        nac_str = str(nac_raw)

                    # Cédula real o provisional
                    es_provisional = 0
                    ced_str = str(ced_raw or '').replace('.0','').replace(',','').strip()
                    if ced_str.isdigit():
                        cedula = ced_str
                    else:
                        nom1 = nombre.split()[0]  if nombre  else 'X'
                        ape1 = apellido.split()[0] if apellido else 'X'
                        cedula = f"PROV_{nom1}_{ape1}"
                        es_provisional = 1
                        sin_cedula_lista.append(f"{nombre} {apellido} ({grado} {mencion_actual})")

                    # ── UPSERT en registro_liceo ──────────────────────────
                    existing = conn.execute(
                        "SELECT cedula FROM registro_liceo WHERE cedula=?", (cedula,)
                    ).fetchone()

                    if existing:
                        conn.execute("""
                            UPDATE registro_liceo
                               SET nombre=?, apellido=?, sexo=?, nacimiento=?,
                                   edad=?, telefono=?, grado=?, mencion=?, es_provisional=?
                             WHERE cedula=?
                        """, (nombre, apellido, sexo, nac_str,
                              edad_calc, telefono, grado, mencion_actual,
                              es_provisional, cedula))
                    else:
                        conn.execute("""
                            INSERT INTO registro_liceo
                                (cedula, nombre, apellido, sexo, nacimiento,
                                 edad, telefono, grado, mencion, es_provisional)
                            VALUES (?,?,?,?,?,?,?,?,?,?)
                        """, (cedula, nombre, apellido, sexo, nac_str,
                              edad_calc, telefono, grado, mencion_actual,
                              es_provisional))

                    # ── Crear perfil vacío en 'estudiantes' si no existe ──
                    # Así aparecen en el dashboard aunque no tengan notas
                    ya_existe = conn.execute(
                        """SELECT id FROM estudiantes
                           WHERE cedula=?
                              OR (lower(nombre)=lower(?) AND lower(apellido)=lower(?))
                           LIMIT 1""",
                        (cedula, nombre, apellido)
                    ).fetchone()

                    # Detectar ciclo según grado
                    _ciclo_listado = 'segundo_ciclo'
                    for _g1c in ['1RO', '2DO', '3RO', '1ER', '2ER', '3ER']:
                        if _g1c in grado.upper():
                            _ciclo_listado = 'primer_ciclo'
                            break

                    if not ya_existe:
                        conn.execute("""
                            INSERT INTO estudiantes
                                (id_evaluacion, cedula, nombre, apellido,
                                 curso, grado, condicion, edad, ciclo)
                            VALUES (?, ?, ?, ?, ?, ?, 'ACTIVO', ?, ?)
                        """, (
                            cedula,          # id_evaluacion = cédula hasta tener BJ
                            cedula,
                            nombre,
                            apellido,
                            f"{grado} {mencion_actual}",
                            grado,
                            edad_calc or 0,
                            _ciclo_listado
                        ))
                    else:
                        # Actualizar cédula si el perfil existía con BJ* o vacío
                        conn.execute("""
                            UPDATE estudiantes
                               SET cedula=?, grado=?, curso=?, edad=?
                             WHERE (cedula IS NULL OR cedula='' OR cedula LIKE 'BJ%')
                               AND lower(nombre)=lower(?)
                               AND lower(apellido)=lower(?)
                        """, (cedula, grado, f"{grado} {mencion_actual}",
                              edad_calc or 0, nombre, apellido))

                    conteo += 1

                resumen[f"{grado} {mencion_actual}"] = conteo

            conn.commit()

        # Totales
        with sqlite3.connect(DATABASE, timeout=10) as conn2:
            total_liceo  = conn2.execute("SELECT COUNT(*) FROM registro_liceo").fetchone()[0]
            total_perfiles = conn2.execute("SELECT COUNT(*) FROM estudiantes").fetchone()[0]
            con_notas    = conn2.execute(
                "SELECT COUNT(*) FROM estudiantes WHERE p_acad > 0 OR acad_p1 > 0"
            ).fetchone()[0]

        return jsonify({
            "ok": True,
            "resumen_por_seccion": resumen,
            "totales": {
                "en_registro_liceo":    total_liceo,
                "perfiles_en_sistema":  total_perfiles,
                "con_notas_cargadas":   con_notas,
                "solo_identidad":       total_perfiles - con_notas
            },
            "sin_cedula": sin_cedula_lista,
            "mensaje": (
                f"{total_liceo} estudiantes en el registro del liceo. "
                f"{total_perfiles} perfiles en el sistema "
                f"({con_notas} con notas, {total_perfiles - con_notas} solo identidad)."
                + (f" {len(sin_cedula_lista)} sin cédula (ID provisional)."
                   if sin_cedula_lista else "")
            )
        })

    except Exception as ex:
        logger.error(f"[ESTUDIANTES] Error procesando listado: {ex}", exc_info=True)
        return jsonify({"error": "Error al procesar el listado. Intenta de nuevo."}), 500


@estudiantes_bp.route("/api/registro-liceo", methods=["GET"])
@login_required
def get_registro_liceo():
    """Devuelve el registro de todos los estudiantes del liceo para consulta."""
    grado  = request.args.get("grado", "")
    mencion = request.args.get("mencion", "")

    query = "SELECT * FROM registro_liceo WHERE 1=1"
    params = []
    if grado:
        query += " AND upper(grado) LIKE upper(?)"
        params.append(f"%{grado}%")
    if mencion:
        query += " AND upper(mencion) LIKE upper(?)"
        params.append(f"%{mencion}%")
    query += " ORDER BY grado, mencion, apellido"

    try:
        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(query, params).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as ex:
        logger.error(f"[ESTUDIANTES] Error en búsqueda: {ex}")
        return jsonify({"error": "Error al buscar estudiantes. Intenta de nuevo."}), 500



# ══════════════════════════════════════════════════════════════════════════════
# CARGA DINÁMICA DE MATERIAS — LLaMA interpreta cualquier Excel de maestro
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/interpretar-excel", methods=["POST"])
@login_required
def interpretar_excel():
    """
    Analiza el Excel del maestro con lógica inteligente en 3 pasos:
    1. Busca zona de resumen (Periodo I/II/III/IV) → lectura directa sin IA
    2. Si no encuentra, usa LLaMA para interpretar columnas
    3. Devuelve mapeo para confirmación del usuario
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file     = request.files["file"]
    materia  = request.form.get("materia", "").strip()
    if not materia:
        return jsonify({"error": "Debes indicar el nombre de la materia"}), 400

    nombre_archivo = file.filename
    raw_bytes = file.read()

    # ── Verificar mapeo guardado ──────────────────────────────────
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        mapeo_guardado = conn.execute(
            "SELECT * FROM mapeos_excel WHERE nombre_archivo=?", (nombre_archivo,)
        ).fetchone()
    if mapeo_guardado:
        m = dict(mapeo_guardado)
        return jsonify({
            "mapeo": {
                "col_nombre": m["col_nombre"], "col_apellido": m["col_apellido"],
                "col_nombre_completo": m["col_nombre_completo"],
                "col_p1": m["col_p1"], "col_p2": m["col_p2"],
                "col_p3": m["col_p3"], "col_p4": m["col_p4"],
                "modo": m.get("col_nombre_completo") or "openpyxl",
            },
            "materia": materia, "archivo": nombre_archivo,
            "confianza": "guardado",
            "mensaje": f"Mapeo recordado de carga anterior.",
            "columnas_disponibles": [], "muestra": [],
            "modo_lectura": "guardado"
        })

    # ── MODO 1: Lector inteligente openpyxl ──────────────────────
    # Detecta zona de resumen (Periodo I/II/III/IV) automáticamente
    try:
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(raw_bytes), data_only=True)

        hojas_validas = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            col_p1=col_p2=col_p3=col_p4=col_nom=col_ape=None
            header_row=None

            # Buscar fila con Periodo I, II, III, IV
            for row in range(1, min(20, ws.max_row+1)):
                for col in range(1, ws.max_column+1):
                    v = str(ws.cell(row=row, column=col).value or '').strip().upper()
                    v = v.replace('\xa0',' ').strip()
                    if v in ('PERIODO I','PERIODO 1') or v.endswith('PERIODO I'):
                        col_p1 = col; header_row = row
                    elif v in ('PERIODO II','PERIODO 2'):
                        col_p2 = col
                    elif v in ('PERIODO III','PERIODO 3'):
                        col_p3 = col
                    elif v in ('PERIODO IV','PERIODO 4'):
                        col_p4 = col

            if not header_row:
                continue

            # Buscar Nombres/Apellidos en zona del resumen
            for row in range(max(1,header_row-2), header_row+4):
                for col in range(max(1,(col_p1 or 1)-15), ws.max_column+1):
                    v = str(ws.cell(row=row, column=col).value or '').strip().upper()
                    if 'NOMBRE' in v and col_nom is None: col_nom = col
                    if 'APELLIDO' in v and col_ape is None: col_ape = col

            # Encontrar fila de inicio de datos
            data_start = None
            for row in range(header_row+1, ws.max_row+1):
                v = ws.cell(row=row, column=col_nom or 3).value
                if v and str(v).strip() not in ('','None','nan','Nombres','NOMBRES'):
                    # Verificar que no es otra fila de encabezado
                    if not any(kw in str(v).upper() for kw in ['NOMBRE','ALUMNO','#']):
                        data_start = row
                        break

            if col_nom and data_start:
                hojas_validas.append({
                    "hoja": sname, "header_row": header_row,
                    "data_start": data_start,
                    "col_nom": col_nom, "col_ape": col_ape,
                    "col_p1": col_p1, "col_p2": col_p2,
                    "col_p3": col_p3, "col_p4": col_p4,
                })

        if hojas_validas:
            h = hojas_validas[0]
            muestra = []
            ws0 = wb[h["hoja"]]
            for row in range(h["data_start"], min(h["data_start"]+4, ws0.max_row+1)):
                nom = ws0.cell(row=row, column=h["col_nom"]).value
                ape = ws0.cell(row=row, column=h["col_ape"]).value if h["col_ape"] else ''
                p1  = ws0.cell(row=row, column=h["col_p1"]).value if h["col_p1"] else None
                if nom:
                    muestra.append([str(nom), str(ape or ''), str(p1 or '-')])

            n_hojas = len(hojas_validas)
            return jsonify({
                "mapeo": {
                    "col_nombre":          h["col_nom"],
                    "col_apellido":        h["col_ape"],
                    "col_nombre_completo": None,
                    "col_p1": h["col_p1"], "col_p2": h["col_p2"],
                    "col_p3": h["col_p3"], "col_p4": h["col_p4"],
                    "header_row":   h["header_row"],
                    "data_start":   h["data_start"],
                    "fila_inicio_datos": h["data_start"],
                    "modo":         "openpyxl_resumen",
                    "hojas":        [x["hoja"] for x in hojas_validas],
                    "notas": f"Zona de resumen detectada automaticamente. "
                             f"{n_hojas} hoja(s): {', '.join(x['hoja'] for x in hojas_validas)}. "
                             f"Periodo I→IV mapeados directamente."
                },
                "materia": materia, "archivo": nombre_archivo,
                "confianza": "alta",
                "mensaje": f"Estructura de RAEs detectada. {n_hojas} seccion(es) encontrada(s).",
                "columnas_disponibles": ["(auto - zona resumen)"],
                "muestra": muestra,
                "modo_lectura": "openpyxl_resumen",
                "hojas_info": hojas_validas
            })

    except Exception as ex_openpyxl:
        logger.info(f"openpyxl fallback: {ex_openpyxl}")

    # ── MODO 2: LLaMA interpreta columnas simples ─────────────────
    try:
        import pandas as pd, io as _io, json as _json
        df = pd.read_excel(_io.BytesIO(raw_bytes), header=None, nrows=12)
        mejor_fila = max(range(min(8,len(df))),
                        key=lambda i: sum(1 for v in df.iloc[i]
                                          if isinstance(v,str) and len(str(v).strip())>0))
        df2 = pd.read_excel(_io.BytesIO(raw_bytes), header=mejor_fila)
        columnas = [str(c) for c in df2.columns if str(c).strip() and str(c)!='nan']
        muestra  = df2.head(4).fillna('').astype(str).values.tolist()

        prompt = f"""Eres un asistente que analiza archivos Excel de maestros dominicanos.
Materia: {materia} | Archivo: {nombre_archivo}
Columnas: {columnas}
Primeras filas: {muestra[:3]}

Responde SOLO con JSON:
{{"col_nombre":null,"col_apellido":null,"col_nombre_completo":null,
"col_p1":null,"col_p2":null,"col_p3":null,"col_p4":null,
"fila_inicio_datos":1,"confianza":"media","notas":""}}

Reglas: P1/P2 pueden llamarse Periodo 1, Trimestre 1, Parcial 1, PROM P1, etc.
Si nombre+apellido van juntos usa col_nombre_completo."""

        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role":"system","content":"Responde SOLO con JSON válido."},
                {"role":"user","content":prompt}
            ],
            temperature=0.1, max_tokens=400,
        )
        resp = completion.choices[0].message.content.strip()
        resp = resp.replace("```json","").replace("```","").strip()
        mapeo = _json.loads(resp)
        mapeo["modo"] = "llama"

        return jsonify({
            "mapeo": mapeo, "materia": materia, "archivo": nombre_archivo,
            "confianza": mapeo.get("confianza","media"),
            "mensaje": mapeo.get("notas","LLaMA interpreto el archivo."),
            "columnas_disponibles": columnas, "muestra": muestra[:3],
            "modo_lectura": "llama"
        })

    except Exception as ex_llama:
        return jsonify({
            "mapeo": None, "materia": materia, "archivo": nombre_archivo,
            "confianza": "manual",
            "mensaje": "No se pudo interpretar automaticamente. Selecciona las columnas manualmente.",
            "columnas_disponibles": [], "muestra": [], "modo_lectura": "manual",
            "error_ia": str(ex_llama)
        })


@estudiantes_bp.route("/api/cargar-materia", methods=["POST"])
@login_required
@rate_limited(max_calls=10, window=3600)
def cargar_materia():
    """
    Carga notas de una materia usando el mapeo confirmado.
    Solo profesores, coordinadores, digitadores y directora.
    Modo openpyxl_resumen: lee zona Periodo I/II/III/IV en todas las hojas.
    Modo pandas: lee con nombres de columnas (LLaMA o manual).
    """
    u = get_usuario()
    rol = _normalizar_rol(u.get("rol", ""))
    _ROLES_PERMITIDOS = {"profesor", "coordinador_general", "coordinador_primer_ciclo",
                         "coordinador_segundo_ciclo", "digitador", "secretaria_docente",
                         "directora"}
    if rol not in _ROLES_PERMITIDOS:
        return jsonify({"error": "Sin permisos para cargar calificaciones."}), 403
    if "file" not in request.files:
        return jsonify({"error": "No se recibio archivo"}), 400

    import io as _io, json as _json

    file      = request.files["file"]
    # Normalizar casing para evitar duplicados (ej: "lenguaje visual" vs "Lenguaje Visual")
    _mat_raw = request.form.get("materia","").strip()
    materia  = " ".join(w.capitalize() for w in _mat_raw.lower().split()) if _mat_raw else ""
    mapeo_str = request.form.get("mapeo","{}")
    raw_bytes = file.read()

    try:
        mapeo = _json.loads(mapeo_str)
    except Exception:
        return jsonify({"error": "Mapeo JSON invalido"}), 400

    if not materia:
        return jsonify({"error": "Materia requerida"}), 400

    # ── VALIDACIÓN DE PERFIL DE PROFESOR ────────────────────────────────────
    # Si el usuario autenticado es un profesor (no coordinador),
    # verificar que la materia y el curso correspondan a su perfil asignado.
    _prof_actual = _get_profesor()
    if _prof_actual and _prof_actual.get("rol") == "profesor":
        _curso_archivo = request.form.get("grado", "") + " " + request.form.get("mencion", "")
        _ok_val, _msg_val = _validar_materia_profesor(materia, _curso_archivo, _prof_actual)
        if not _ok_val:
            return jsonify({"error": _msg_val, "tipo": "validacion_perfil"}), 403

    # Nombre del profesor — desde form o extraído del nombre del archivo
    profesor_nombre = request.form.get("profesor", "").strip()
    if not profesor_nombre and file:
        import re as _re2
        fn = file.filename or ""
        fn_clean = _re2.sub(r"\.(xlsx?|xlsm)$", "", fn, flags=_re2.I)
        fn_clean = _re2.sub(r"(6to|5to|4to|[AB]_|grado|teatro|musica|multimedia|artes|diseno|diseño)",
                            " ", fn_clean, flags=_re2.I)
        fn_clean = fn_clean.replace("_", " ").replace("-", " ").strip()
        if fn_clean:
            profesor_nombre = fn_clean[:50]

    modo = mapeo.get("modo","llama")

    def limpiar_num(v):
        try:
            if v is None or str(v).strip() in ('','nan','None','-'):
                return None
            return round(float(v), 1)
        except Exception:
            return None

    registros = []

    # ── MODO openpyxl_resumen ─────────────────────────────────────
    if modo == "openpyxl_resumen":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(raw_bytes), data_only=True)
            hojas_raw  = mapeo.get("hojas", wb.sheetnames)
            hojas_usar = hojas_raw if isinstance(hojas_raw, list) else wb.sheetnames

            for sname in hojas_usar:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                col_p1=col_p2=col_p3=col_p4=col_nom=col_ape=None
                header_row=None

                for row in range(1, min(20, ws.max_row+1)):
                    for col in range(1, ws.max_column+1):
                        v = str(ws.cell(row=row,column=col).value or '').strip().upper()
                        if v in ('PERIODO I','PERIODO 1') or v.endswith('PERIODO I'):
                            col_p1=col; header_row=row
                        elif v in ('PERIODO II','PERIODO 2'): col_p2=col
                        elif v in ('PERIODO III','PERIODO 3'): col_p3=col
                        elif v in ('PERIODO IV','PERIODO 4'): col_p4=col

                if not header_row:
                    continue

                for row in range(max(1,header_row-2), header_row+4):
                    for col in range(max(1,(col_p1 or 1)-15), ws.max_column+1):
                        v = str(ws.cell(row=row,column=col).value or '').strip().upper()
                        if 'NOMBRE' in v and col_nom is None: col_nom=col
                        if 'APELLIDO' in v and col_ape is None: col_ape=col

                if not col_nom:
                    continue

                data_start = None
                for row in range(header_row+1, ws.max_row+1):
                    v = ws.cell(row=row, column=col_nom).value
                    vs = str(v or '').strip()
                    if vs and vs not in ('None','nan','Nombres','NOMBRES'):
                        if not any(kw in vs.upper() for kw in ['NOMBRE','ALUMNO','#']):
                            data_start=row; break

                if not data_start:
                    continue

                for row in range(data_start, ws.max_row+1):
                    nom = ws.cell(row=row, column=col_nom).value
                    if not nom or str(nom).strip() in ('','None','nan'):
                        continue
                    ape = ws.cell(row=row, column=col_ape).value if col_ape else ''
                    p1  = limpiar_num(ws.cell(row=row,column=col_p1).value) if col_p1 else None
                    p2  = limpiar_num(ws.cell(row=row,column=col_p2).value) if col_p2 else None
                    p3  = limpiar_num(ws.cell(row=row,column=col_p3).value) if col_p3 else None
                    p4  = limpiar_num(ws.cell(row=row,column=col_p4).value) if col_p4 else None
                    registros.append({
                        "nombre":   str(nom).strip(),
                        "apellido": str(ape or '').strip(),
                        "p1":p1,"p2":p2,"p3":p3,"p4":p4
                    })

        except Exception as ex:
            logger.warning(f"[ESTUDIANTES] Error leyendo Excel: {ex}")
        return jsonify({"error": "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)."}), 400

    # ── MODO pandas ───────────────────────────────────────────────
    else:
        import pandas as pd
        col_nombre          = mapeo.get("col_nombre")
        col_apellido        = mapeo.get("col_apellido")
        col_nombre_completo = mapeo.get("col_nombre_completo")
        col_p1 = mapeo.get("col_p1")
        col_p2 = mapeo.get("col_p2")
        col_p3 = mapeo.get("col_p3")
        col_p4 = mapeo.get("col_p4")
        fila_h = max(0, int(mapeo.get("fila_inicio_datos",1))-1)

        try:
            df = pd.read_excel(_io.BytesIO(raw_bytes), header=fila_h)
            df.columns = [str(c).strip() for c in df.columns]
        except Exception as ex:
            logger.warning(f"[ESTUDIANTES] Error leyendo Excel: {ex}")
        return jsonify({"error": "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)."}), 400

        for _, fila in df.iterrows():
            d = fila.to_dict()
            if col_nombre_completo and col_nombre_completo in d:
                nc = str(d[col_nombre_completo] or '').strip()
                partes = nc.split()
                nom = ' '.join(partes[:max(1,len(partes)//2)])
                ape = ' '.join(partes[max(1,len(partes)//2):])
            else:
                nom = str(d.get(col_nombre,'') or '').strip()
                ape = str(d.get(col_apellido,'') or '').strip()
            if not nom or nom.lower() in ('nan','none',''):
                continue
            registros.append({
                "nombre":nom,"apellido":ape,
                "p1":limpiar_num(d.get(col_p1)) if col_p1 else None,
                "p2":limpiar_num(d.get(col_p2)) if col_p2 else None,
                "p3":limpiar_num(d.get(col_p3)) if col_p3 else None,
                "p4":limpiar_num(d.get(col_p4)) if col_p4 else None,
            })

    # ── Vincular a estudiantes y guardar ─────────────────────────
    cargados = 0
    no_encontrados = []

    # Filtro de grado y mención: evita vincular "María García 4to" con "María García 5to"
    filtro_grado   = request.form.get("grado",   "").strip()
    filtro_mencion = request.form.get("mencion", "").strip()

    # Búsqueda delegada al helper de módulo _buscar_estudiante_bd
    def buscar_estudiante(conn, nom1, ape1):
        return _buscar_estudiante_bd(conn, nom1, ape1,
                                     filtro_grado=filtro_grado,
                                     filtro_mencion=filtro_mencion)
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row

        for reg in registros:
            nom1 = reg["nombre"].lower().split()[0]  if reg["nombre"]  else ''
            ape1 = reg["apellido"].lower().split()[0] if reg["apellido"] else ''
            if not nom1:
                continue

            est = _buscar_estudiante_bd(conn, reg["nombre"], reg["apellido"])

            if not est:
                no_encontrados.append(f"{reg['nombre']} {reg['apellido']}")
                continue

            notas = [x for x in [reg["p1"],reg["p2"],reg["p3"],reg["p4"]] if x is not None]
            prom  = round(sum(notas)/len(notas),1) if notas else None

            conn.execute("""
                INSERT INTO materias_calificaciones
                    (estudiante_id,cedula,materia,p1,p2,p3,p4,promedio,fuente,profesor)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(estudiante_id,materia) DO UPDATE SET
                    p1      = CASE WHEN excluded.p1 IS NOT NULL AND excluded.p1 > 0
                                   THEN excluded.p1 ELSE materias_calificaciones.p1 END,
                    p2      = CASE WHEN excluded.p2 IS NOT NULL AND excluded.p2 > 0
                                   THEN excluded.p2 ELSE materias_calificaciones.p2 END,
                    p3      = CASE WHEN excluded.p3 IS NOT NULL AND excluded.p3 > 0
                                   THEN excluded.p3 ELSE materias_calificaciones.p3 END,
                    p4      = CASE WHEN excluded.p4 IS NOT NULL AND excluded.p4 > 0
                                   THEN excluded.p4 ELSE materias_calificaciones.p4 END,
                    promedio= excluded.promedio,
                    fuente  = excluded.fuente,
                    profesor= CASE WHEN excluded.profesor IS NOT NULL AND excluded.profesor != ''
                                   THEN excluded.profesor ELSE materias_calificaciones.profesor END,
                    fecha_carga = date('now')
            """, (est["id"],est["cedula"],_normalizar_materia(materia),
                  reg["p1"],reg["p2"],reg["p3"],reg["p4"],prom,
                  file.filename, profesor_nombre))
            cargados += 1

            # ── Sincronizar módulos técnicos Multimedia y asistencia ──────────
            # Si la materia coincide con un módulo de Multimedia, actualizar
            # también los campos directos en la tabla estudiantes
            mat_norm = materia.lower().strip()

            # Mapeo de nombre de materia → columna base en estudiantes
            MODULO_MAP = {
                'fotografía': 'fotografia', 'fotografia': 'fotografia',
                'foto':       'fotografia',
                'lenguaje visual': 'lv', 'lenguaje_visual': 'lv', 'lv': 'lv',
                'diseño': 'diseno', 'diseno': 'diseno', 'diseño básico': 'diseno',
                'diseño basico': 'diseno',
            }
            ASISTENCIA_KEYS = {'asistencia', 'asistencias', 'attend', 'attendance'}

            col_base = MODULO_MAP.get(mat_norm)
            es_asistencia = mat_norm in ASISTENCIA_KEYS

            if col_base or es_asistencia:
                def _v(val):
                    return val if (val is not None and val != 0) else None

                if col_base:
                    # Módulo técnico → actualizar p1-p4 y promedio del módulo
                    p_col = f'p_{col_base}' if col_base != 'fotografia' else 'p_foto'
                    notas_mod = [x for x in [reg['p1'],reg['p2'],reg['p3'],reg['p4']] if x]
                    prom_mod  = round(sum(notas_mod)/len(notas_mod),1) if notas_mod else None

                    # Build SET clause only for non-null values
                    sets, vals = [], []
                    for pi, pv in enumerate([reg['p1'],reg['p2'],reg['p3'],reg['p4']], 1):
                        if _v(pv) is not None:
                            sets.append(f"{col_base}_p{pi} = CASE WHEN {col_base}_p{pi} IS NULL OR {col_base}_p{pi}=0 THEN ? ELSE {col_base}_p{pi} END")
                            vals.append(pv)
                    if prom_mod and sets:
                        sets.append(f"{p_col} = ?")
                        vals.append(prom_mod)
                    if sets:
                        vals.append(est['id'])
                        conn.execute(
                            f"UPDATE estudiantes SET {', '.join(sets)} WHERE id=?", vals
                        )
                        # Recalculate prom_modulos
                        e2 = conn.execute(
                            "SELECT p_foto,p_lv,p_diseno FROM estudiantes WHERE id=?",
                            (est['id'],)
                        ).fetchone()
                        if e2:
                            mods = [x for x in e2 if x and x > 0]
                            pm2  = round(sum(mods)/len(mods),1) if mods else None
                            conn.execute(
                                "UPDATE estudiantes SET prom_modulos=? WHERE id=?",
                                (pm2, est['id'])
                            )

                elif es_asistencia:
                    # Asistencia → asistencia_p1..p4 y asistencia general
                    sets, vals = [], []
                    for pi, pv in enumerate([reg['p1'],reg['p2'],reg['p3'],reg['p4']], 1):
                        if _v(pv) is not None:
                            sets.append(f"asistencia_p{pi} = CASE WHEN asistencia_p{pi} IS NULL OR asistencia_p{pi}=0 THEN ? ELSE asistencia_p{pi} END")
                            vals.append(pv)
                    if sets:
                        vals.append(est['id'])
                        conn.execute(
                            f"UPDATE estudiantes SET {', '.join(sets)} WHERE id=?", vals
                        )

        conn.commit()

        # Guardar mapeo
        try:
            conn.execute("""
                INSERT INTO mapeos_excel
                    (nombre_archivo,materia,col_nombre,col_apellido,
                     col_nombre_completo,col_p1,col_p2,col_p3,col_p4)
                VALUES (?,?,?,?,?,?,?,?,?)
                ON CONFLICT(nombre_archivo) DO UPDATE SET
                    materia=excluded.materia,
                    col_nombre_completo=excluded.col_nombre_completo,
                    fecha_uso=date('now')
            """, (file.filename, materia,
                  str(mapeo.get("col_nombre","") or ""),
                  str(mapeo.get("col_apellido","") or ""),
                  str(mapeo.get("modo","llama")),
                  str(mapeo.get("col_p1","") or ""),
                  str(mapeo.get("col_p2","") or ""),
                  str(mapeo.get("col_p3","") or ""),
                  str(mapeo.get("col_p4","") or "")))
            conn.commit()
        except Exception as _e:
            logger.warning(f"[estudiantes] Excepción silenciada")

    return jsonify({
        "ok":             True,
        "materia":        materia,
        "cargados":       cargados,
        "total_excel":    len(registros),
        "no_encontrados": no_encontrados,
        "mensaje": (f"{cargados} de {len(registros)} estudiantes actualizados en '{materia}'." +
                    (f" {len(no_encontrados)} no encontrados." if no_encontrados else ""))
    })



@estudiantes_bp.route("/api/materias/<int:estudiante_id>")
@login_required
def get_materias_estudiante(estudiante_id):
    """Devuelve materias cargadas para un estudiante. Deduplica y filtra por asignatura del prof."""
    prof = _get_profesor()
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        anio = _anio_escolar_actual()

        # Obtener grado actual del estudiante
        est_row = conn.execute(
            "SELECT grado FROM estudiantes WHERE id=?", (estudiante_id,)
        ).fetchone()
        grado_actual = (est_row["grado"] if est_row else "") or ""

        rows = conn.execute("""
            SELECT materia, p1, p2, p3, p4, promedio, fecha_carga, fuente,
                   COALESCE(tipo, 'académico') as tipo,
                   COALESCE(profesor, '') as profesor
            FROM materias_calificaciones
            WHERE estudiante_id = ?
              AND (anio_escolar = ? OR anio_escolar IS NULL)
              AND (grado IS NULL OR UPPER(grado) = UPPER(?))
            ORDER BY CASE COALESCE(tipo,'académico')
                          WHEN 'académico' THEN 0 ELSE 1 END,
                     materia
        """, (estudiante_id, anio, grado_actual)).fetchall()
    resultado = _dedup_materias([dict(r) for r in rows])
    resultado.sort(key=lambda r: (0 if r["tipo"] == "académico" else 1, r["materia"].lower()))
    resultado = _rls.filtrar_materias_profesor(resultado)
    return jsonify(resultado)


@estudiantes_bp.route("/api/materias-disponibles")
@login_required
def get_materias_disponibles():
    """Lista todas las materias que han sido cargadas al sistema."""
    grado = request.args.get("grado", "")
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT DISTINCT mc.materia, COUNT(DISTINCT mc.estudiante_id) as total_estudiantes,
                   mc.fecha_carga
            FROM materias_calificaciones mc
            JOIN estudiantes e ON e.id = mc.estudiante_id
            WHERE (? = '' OR e.grado LIKE ?)
            GROUP BY mc.materia
            ORDER BY mc.materia
        """, (grado, f"%{grado}%")).fetchall()
    return jsonify([dict(r) for r in rows])

# ── Indicadores por período (para gráficas del perfil) ───────────────────────


@estudiantes_bp.route("/api/indicadores/materias/<int:estudiante_id>")
@login_required
def get_indicadores_materias(estudiante_id):
    """Lista de materias disponibles para el estudiante con promedio."""
    from core.helpers import _get_profesor
    from core.auth import _normalizar_rol
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT materia,
                   COALESCE(p1,0) p1, COALESCE(p2,0) p2,
                   COALESCE(p3,0) p3, COALESCE(p4,0) p4,
                   COALESCE(promedio,0) promedio, fecha_carga
            FROM materias_calificaciones
            WHERE estudiante_id = ?
            ORDER BY materia
        """, (estudiante_id,)).fetchall()

    result = _rls.filtrar_materias_profesor([dict(r) for r in rows])
    return jsonify(result)


@estudiantes_bp.route("/api/indicadores/<int:estudiante_id>")
@login_required
def get_indicadores(estudiante_id):
    """Indicadores de una materia específica por período."""
    materia = request.args.get("materia", "")
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("""
            SELECT materia, p1, p2, p3, p4, promedio, fecha_carga
            FROM materias_calificaciones
            WHERE estudiante_id=? AND materia=?
        """, (estudiante_id, materia)).fetchone()
    if not row:
        return jsonify({"indicadores": [], "materia": materia, "promedio": 0})
    r = dict(row)
    # Build period rows for chart
    periodos = []
    for i, (p, label) in enumerate([("p1","P1"),("p2","P2"),("p3","P3"),("p4","P4")], 1):
        val = r.get(p)
        if val and val > 0:
            periodos.append({
                "periodo": label,
                "indicador_texto": f"Período {i}",
                "p1": val if i==1 else None,
                "p2": val if i==2 else None,
                "p3": val if i==3 else None,
                "p4": val if i==4 else None,
            })
    return jsonify({
        "materia":    r["materia"],
        "promedio":   r["promedio"] or 0,
        "indicadores": periodos,
        "p1": r["p1"], "p2": r["p2"], "p3": r["p3"], "p4": r["p4"],
    })


# ══════════════════════════════════════════════════════════════════════════════
# CRUD DE PERFILES — Eliminar, Retirar, Agregar, Editar campos manuales
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/estudiante/<int:id>", methods=["DELETE"])
@login_required
def eliminar_estudiante(id):
    """Elimina un perfil completo y sus materias dinámicas. Solo directora/admin."""
    u = get_usuario()
    rol = _normalizar_rol(u.get("rol", ""))
    from core.constants import ROLES_DIRECTORA
    if rol not in ROLES_DIRECTORA:
        return jsonify({"error": "Sin permisos. Solo la directora puede eliminar estudiantes."}), 403
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("DELETE FROM materias_calificaciones WHERE estudiante_id=?", (id,))
        conn.execute("DELETE FROM estudiantes WHERE id=?", (id,))
        conn.commit()
    return jsonify({"ok": True, "mensaje": "Perfil eliminado."})


@estudiantes_bp.route("/api/estudiante/<int:id>/condicion", methods=["POST"])
@login_required
def cambiar_condicion(id):
    """Cambia la condición del estudiante: ACTIVO / RETIRADO / GRADUADO."""
    data      = request.get_json(silent=True) or {}
    condicion = str(data.get("condicion", "ACTIVO")).upper().strip()
    if condicion not in ("ACTIVO", "RETIRADO", "GRADUADO"):
        return jsonify({"error": "Condición inválida. Usa ACTIVO, RETIRADO o GRADUADO"}), 400
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("UPDATE estudiantes SET condicion=? WHERE id=?", (condicion, id))
        conn.commit()
    return jsonify({"ok": True, "condicion": condicion})


@estudiantes_bp.route("/api/estudiante", methods=["POST"])
@login_required
def agregar_estudiante():
    """Crea un perfil nuevo manualmente."""
    data = request.get_json(silent=True) or {}
    nombre   = str(data.get("nombre",   "")).strip()
    apellido = str(data.get("apellido", "")).strip()
    if not nombre or not apellido:
        return jsonify({"error": "Nombre y apellido son requeridos"}), 400

    grado   = str(data.get("grado",   "4to")).strip()
    curso   = str(data.get("curso",   "")).strip()
    cedula  = str(data.get("cedula",  "")).strip()
    edad    = data.get("edad")

    if not cedula:
        cedula = f"MANUAL_{nombre[:4].upper()}_{apellido[:4].upper()}"

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        existing = conn.execute(
            "SELECT id FROM estudiantes WHERE cedula=? OR (lower(nombre)=lower(?) AND lower(apellido)=lower(?))",
            (cedula, nombre, apellido)
        ).fetchone()
        if existing:
            return jsonify({"error": "Ya existe un estudiante con ese nombre o cédula"}), 409

        conn.execute("""
            INSERT INTO estudiantes
                (id_evaluacion, cedula, nombre, apellido, curso, grado, condicion, edad)
            VALUES (?,?,?,?,?,?,?,?)
        """, (cedula, cedula, nombre, apellido, curso, grado, "ACTIVO",
              float(edad) if edad else None))
        conn.commit()
        nuevo_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    return jsonify({"ok": True, "id": nuevo_id,
                    "mensaje": f"Perfil de {nombre} {apellido} creado."})


@estudiantes_bp.route("/api/estudiante/<int:id>", methods=["PATCH"])
@login_required
def editar_estudiante(id):
    """
    Actualiza campos específicos de un estudiante.
    Acepta cualquier campo numérico o de texto de la tabla estudiantes.
    Usado para ingresar datos manuales (conductual, emocional, módulos).
    """
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"error": "Sin datos"}), 400

    # Campos permitidos (seguridad: no se puede cambiar id ni cedula via PATCH)
    CAMPOS_TEXTO = {
        'nombre','apellido','curso','grado','condicion','uso_celular',
        'nivel_riesgo','tendencia','reporte','color','ia_analisis'
    }
    CAMPOS_NUM = {
        'edad','p_acad','p_cond','p_auto','p_emocional','prom_modulos',
        'puntualidad','tareas','participacion','comprension','rendimiento',
        'interrupciones','conflictos','desafia_autoridad','distraccion',
        'falta_respeto','motivacion','estado_emocional','interes_futuro',
        'apoyo_familiar','indice_riesgo','proyeccion','asistencia',
        'fotografia_p1','fotografia_p2','fotografia_p3','fotografia_p4',
        'lv_p1','lv_p2','lv_p3','lv_p4',
        'diseno_p1','diseno_p2','diseno_p3','diseno_p4',
        'acad_p1','acad_p2','acad_p3','acad_p4',
        'asistencia_p1','asistencia_p2','asistencia_p3','asistencia_p4',
        'silencioso',
    }

    set_parts  = []
    set_values = []

    for k, v in data.items():
        if k in CAMPOS_NUM:
            try:
                set_parts.append(f"{k}=?")
                set_values.append(round(float(v), 2) if v not in (None, '') else None)
            except (ValueError, TypeError):
                pass
        elif k in CAMPOS_TEXTO:
            set_parts.append(f"{k}=?")
            set_values.append(str(v).strip() if v is not None else None)

    if not set_parts:
        return jsonify({"error": "Ningún campo válido para actualizar"}), 400

    # Recalcular promedios derivados si se actualizaron módulos
    campos_recibidos = set(data.keys())
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute(
            f"UPDATE estudiantes SET {', '.join(set_parts)} WHERE id=?",
            set_values + [id]
        )

        # ── Releer estudiante actualizado para recalcular promedios ──────────────
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM estudiantes WHERE id=?", (id,)).fetchone()
        e   = dict(row) if row else {}

        def _pm(vals):
            v = [x for x in vals if x and float(x) > 0]
            return round(sum(float(x) for x in v) / len(v), 2) if v else None

        updates = {}

        # Recalcular p_foto, p_lv, p_diseno, prom_modulos
        if campos_recibidos & {'fotografia_p1','fotografia_p2','fotografia_p3','fotografia_p4',
                                'lv_p1','lv_p2','lv_p3','lv_p4',
                                'diseno_p1','diseno_p2','diseno_p3','diseno_p4'}:
            pf   = _pm([e.get('fotografia_p1'),e.get('fotografia_p2'),e.get('fotografia_p3'),e.get('fotografia_p4')])
            pl   = _pm([e.get('lv_p1'),e.get('lv_p2'),e.get('lv_p3'),e.get('lv_p4')])
            pd   = _pm([e.get('diseno_p1'),e.get('diseno_p2'),e.get('diseno_p3'),e.get('diseno_p4')])
            mods = [x for x in [pf, pl, pd] if x]
            pm2  = round(sum(mods)/len(mods), 2) if mods else None
            updates.update({'p_foto': pf, 'p_lv': pl, 'p_diseno': pd, 'prom_modulos': pm2})

        # Recalcular p_cond desde campos conductuales individuales
        CAMPOS_COND = ['puntualidad','tareas','participacion','comprension','rendimiento',
                       'interrupciones','conflictos','desafia_autoridad','distraccion','falta_respeto']
        if campos_recibidos & set(CAMPOS_COND):
            vals_cond = [e.get(c) for c in CAMPOS_COND]
            updates['p_cond'] = _pm(vals_cond)

        # Recalcular p_auto / p_emocional desde campos emocionales
        CAMPOS_EMOC = ['motivacion','p_auto','estado_emocional','interes_futuro','apoyo_familiar']
        if campos_recibidos & set(CAMPOS_EMOC):
            vals_emoc = [e.get(c) for c in CAMPOS_EMOC]
            updates['p_emocional'] = _pm(vals_emoc)
            # p_auto viene directo del campo, no recalculado

        # También recalcular indice_riesgo básico si hay datos suficientes
        if updates.get('p_cond') is not None or updates.get('p_emocional') is not None:
            p_cond_v = updates.get('p_cond') or e.get('p_cond') or 0
            p_emoc_v = updates.get('p_emocional') or e.get('p_emocional') or 0
            p_acad_v = e.get('p_acad') or 0
            if p_cond_v > 0 or p_emoc_v > 0:
                # Riesgo = inverso ponderado: mala conducta y bajo desempeño emocional suben el riesgo
                riesgo = round(
                    max(0, min(100,
                        (100 - p_cond_v) * 0.5 +
                        (100 - p_emoc_v) * 0.3 +
                        (100 - p_acad_v) * 0.2
                    )), 1)
                updates['indice_riesgo'] = riesgo
                updates['nivel_riesgo']  = (
                    'Alto'  if riesgo >= 50 else
                    'Medio' if riesgo >= 30 else 'Bajo'
                )

        if updates:
            sets = ', '.join(f"{k}=?" for k in updates)
            conn.execute(f"UPDATE estudiantes SET {sets} WHERE id=?",
                         list(updates.values()) + [id])

        conn.commit()

    cache_bust()
    return jsonify({"ok": True, "mensaje": "Perfil actualizado."})




# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO REPORTES — Conducta · Psicológico · Académico · Incidente
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/ml/calcular", methods=["POST"])
@login_required
def calcular_clusters():
    """K-Means puro con numpy — compatible con cualquier Python."""
    try:
        import numpy as np

        def kmeans_numpy(X, k, max_iter=300, n_init=10, seed=42):
            """K-Means implementado con numpy puro."""
            rng = np.random.RandomState(seed)
            best_inertia = float('inf')
            best_labels  = None
            best_centers = None

            for _ in range(n_init):
                # Inicialización K-Means++
                idx = [rng.randint(0, len(X))]
                for _ in range(k - 1):
                    dists = np.array([min(np.sum((x - X[i])**2) for i in idx) for x in X])
                    probs = dists / dists.sum()
                    idx.append(rng.choice(len(X), p=probs))
                centers = X[idx].copy()

                labels = np.zeros(len(X), dtype=int)
                for iteration in range(max_iter):
                    # Asignar clusters
                    dists  = np.array([[np.sum((x - c)**2) for c in centers] for x in X])
                    new_labels = np.argmin(dists, axis=1)
                    if np.all(new_labels == labels):
                        break
                    labels = new_labels
                    # Actualizar centros
                    for ci in range(k):
                        mask = labels == ci
                        if mask.sum() > 0:
                            centers[ci] = X[mask].mean(axis=0)

                inertia = sum(np.sum((X[labels == ci] - centers[ci])**2)
                              for ci in range(k) if (labels == ci).sum() > 0)
                if inertia < best_inertia:
                    best_inertia  = inertia
                    best_labels   = labels.copy()
                    best_centers  = centers.copy()

            return best_labels, best_centers, best_inertia

        def silhouette(X, labels):
            """Silhouette score simplificado."""
            n = len(X)
            if n < 4:
                return 0.0
            scores = []
            unique = list(set(labels))
            if len(unique) < 2:
                return 0.0
            for i in range(n):
                same  = X[labels == labels[i]]
                a     = np.mean([np.sqrt(np.sum((X[i]-x)**2)) for x in same if not np.all(x==X[i])]) if len(same) > 1 else 0
                other_means = []
                for c in unique:
                    if c != labels[i]:
                        grp = X[labels == c]
                        if len(grp) > 0:
                            other_means.append(np.mean([np.sqrt(np.sum((X[i]-x)**2)) for x in grp]))
                b = min(other_means) if other_means else 0
                m = max(a, b)
                scores.append((b - a) / m if m > 0 else 0)
            return float(np.mean(scores))

        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row
            rows = _features_para_clustering(conn)

        if len(rows) < 10:
            return jsonify({"error": f"Solo {len(rows)} estudiantes con datos. "
                                      "Se necesitan al menos 10 para clustering."}), 400

        ids   = [r[0] for r in rows]
        X_raw = np.array([[r[1],r[2],r[3],r[4],r[5],r[6],r[7],r[8],r[9],r[10],r[11],r[12]]
                           for r in rows], dtype=float)

        # Estandarizar (Z-score manual)
        mean = X_raw.mean(axis=0)
        std  = X_raw.std(axis=0)
        std[std == 0] = 1
        Xs = (X_raw - mean) / std

        # Buscar k óptimo (2-5)
        max_k  = min(5, len(rows) // 5)
        max_k  = max(2, max_k)
        best_k = 2
        best_s = -1
        best_labels = None
        best_centers = None

        for k in range(2, max_k + 1):
            lbs, ctrs, _ = kmeans_numpy(Xs, k)
            if len(set(lbs.tolist())) > 1:
                s = silhouette(Xs, lbs)
                if s > best_s:
                    best_s = s; best_k = k
                    best_labels  = lbs
                    best_centers = ctrs

        if best_labels is None:
            best_labels, best_centers, _ = kmeans_numpy(Xs, 2)
            best_k = 2

        # Ordenar clusters por riesgo compuesto
        cluster_risk = {}
        for ci in range(best_k):
            mask = best_labels == ci
            if mask.sum() == 0:
                cluster_risk[ci] = 0
                continue
            avg_riesgo = float(X_raw[mask, 6].mean())
            avg_acad   = float(X_raw[mask, 0].mean())
            avg_auto   = float(X_raw[mask, 2].mean())
            cluster_risk[ci] = avg_riesgo - avg_acad * 0.3 - avg_auto * 0.2

        sorted_clusters = sorted(cluster_risk.keys(), key=lambda c: cluster_risk[c])
        cluster_map     = {orig: pos for pos, orig in enumerate(sorted_clusters)}

        # Calcular distancia al centroide para score
        distances = np.array([[np.sqrt(np.sum((Xs[i] - best_centers[ci])**2))
                                for ci in range(best_k)]
                               for i in range(len(Xs))])

        meta_map = {}
        for orig_ci in range(best_k):
            mask = best_labels == orig_ci
            if mask.sum() == 0:
                continue
            avg_acad   = float(X_raw[mask, 0].mean())
            avg_cond   = float(X_raw[mask, 1].mean())
            avg_auto   = float(X_raw[mask, 2].mean())
            avg_riesgo = float(X_raw[mask, 6].mean())
            avg_conflic= float(X_raw[mask, 8].mean())

            if avg_acad >= 75 and avg_auto >= 70 and avg_riesgo < 25:
                meta_idx = 0
            elif avg_acad >= 75 and avg_auto < 60:
                meta_idx = 3
            elif avg_riesgo >= 45 or avg_conflic >= 50:
                meta_idx = 4
            elif avg_conflic >= 35 or avg_cond < 55:
                meta_idx = 2
            else:
                meta_idx = 1

            meta = CLUSTER_META[min(meta_idx, len(CLUSTER_META) - 1)]
            meta_map[orig_ci] = {
                "label":       meta["label"],
                "color":       meta["color"],
                "icon":        meta["icon"],
                "accion":      meta["accion"],
                "desc":        meta["desc"],
                "n":           int(mask.sum()),
                "avg_acad":    round(avg_acad, 1),
                "avg_cond":    round(avg_cond, 1),
                "avg_auto":    round(avg_auto, 1),
                "avg_riesgo":  round(avg_riesgo, 1),
            }

        # Guardar en DB
        with sqlite3.connect(DATABASE, timeout=10) as conn:
            for i, (est_id, orig_ci) in enumerate(zip(ids, best_labels.tolist())):
                meta      = meta_map.get(int(orig_ci), {})
                dist_min  = float(distances[i][orig_ci])
                score     = round(max(0, 100 - dist_min * 10), 1)
                conn.execute("""
                    UPDATE estudiantes
                    SET cluster_id=?, cluster_label=?, cluster_color=?, cluster_score=?
                    WHERE id=?
                """, (int(cluster_map[int(orig_ci)]), meta.get("label",""),
                      meta.get("color","#888"), score, est_id))
            conn.execute("""
                INSERT INTO ml_clusters (n_clusters, features_usadas, resumen)
                VALUES (?,?,?)
            """, (best_k,
                  "p_acad,p_cond,p_auto,p_emocional,motivacion,apoyo_familiar,"
                  "indice_riesgo,interrupciones,conflictos,falta_respeto,distraccion,prom_modulos",
                  str({v["label"]: v["n"] for v in meta_map.values()})))
            conn.commit()

        return jsonify({
            "ok":    True,
            "k":     best_k,
            "silhouette": round(best_s, 3),
            "estudiantes_analizados": len(ids),
            "clusters": [meta_map[ci] for ci in sorted(meta_map.keys(),
                          key=lambda c: cluster_risk[c])]
        })

    except Exception as ex:
        logger.error(f"[ML] Error clustering: {ex}", exc_info=True)
        return jsonify({"error": "Error al calcular métricas. Intenta de nuevo."}), 500


@estudiantes_bp.route("/api/ml/patrones")
@login_required
def get_patrones():
    """Devuelve resumen de clusters + estudiantes por cluster."""
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        clusters = conn.execute("""
            SELECT cluster_id, cluster_label, cluster_color,
                   COUNT(*) n,
                   ROUND(AVG(p_acad),1)       avg_acad,
                   ROUND(AVG(p_cond),1)       avg_cond,
                   ROUND(AVG(p_auto),1)       avg_auto,
                   ROUND(AVG(indice_riesgo),1) avg_riesgo,
                   ROUND(AVG(cluster_score),1) avg_score
            FROM estudiantes
            WHERE cluster_id IS NOT NULL
            GROUP BY cluster_id
            ORDER BY cluster_id
        """).fetchall()

        estudiantes = conn.execute("""
            SELECT id, nombre, apellido, grado, curso,
                   cluster_id, cluster_label, cluster_color, cluster_score,
                   p_acad, p_cond, indice_riesgo
            FROM estudiantes
            WHERE cluster_id IS NOT NULL
            ORDER BY cluster_id, indice_riesgo DESC
        """).fetchall()

        ultimo = conn.execute("""
            SELECT fecha_calculo, n_clusters FROM ml_clusters
            ORDER BY id DESC LIMIT 1
        """).fetchone()

    # Añadir accion/desc/icon de CLUSTER_META por label
    label_to_meta = {m["label"]: m for m in CLUSTER_META}

    clusters_out = []
    for c in clusters:
        d = dict(c)
        meta = label_to_meta.get(d["cluster_label"], {})
        d["icon"]   = meta.get("icon",  "📌")
        d["accion"] = meta.get("accion","")
        d["desc"]   = meta.get("desc",  "")
        clusters_out.append(d)

    return jsonify({
        "clusters":     clusters_out,
        "estudiantes":  [dict(e) for e in estudiantes],
        "ultimo_calculo": dict(ultimo) if ultimo else None
    })


@estudiantes_bp.route("/api/ml/similar/<int:est_id>")
@login_required
def estudiantes_similares(est_id):
    """Devuelve los 5 estudiantes más similares al perfil dado."""
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        target = conn.execute(
            "SELECT * FROM estudiantes WHERE id=?", (est_id,)
        ).fetchone()
        if not target:
            return jsonify([])
        t = dict(target)

        same_cluster = conn.execute("""
            SELECT id, nombre, apellido, grado, curso,
                   p_acad, p_cond, indice_riesgo, cluster_label, cluster_score
            FROM estudiantes
            WHERE cluster_id=? AND id!=?
            ORDER BY ABS(COALESCE(p_acad,0) - ?)
                   + ABS(COALESCE(indice_riesgo,0) - ?)
            LIMIT 5
        """, (t.get("cluster_id"), est_id,
              t.get("p_acad") or 0, t.get("indice_riesgo") or 0)).fetchall()

    return jsonify([dict(e) for e in same_cluster])


@estudiantes_bp.route("/patrones")
@login_required
def vista_patrones():
    return render_template("patrones.html", current_user=get_usuario())


# ── Búsqueda rápida de estudiantes (autocomplete) ────────────────────────────


@estudiantes_bp.route("/api/cargar-registro", methods=["POST"])
@login_required
def cargar_registro():
    """
    Carga el 'Registro de Calificaciones' oficial (formato MINERD).
    Detecta automáticamente:
      - Nombre del profesor (fila 4)
      - Materia / área (fila 3)
      - Columnas de calificaciones por período (sección resumen derecha)
      - Columnas de porcentaje de asistencia por período
    Funciona con múltiples hojas (ej: 6TO A, 6TO B).
    No sobreescribe notas que ya existan.
    """
    import openpyxl, re as _re

    if "file" not in request.files:
        return jsonify({"error": "Sin archivo"}), 400

    file   = request.files["file"]
    raw    = file.read()
    materia_override = request.form.get("materia", "").strip()

    try:
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, keep_vba=False)
    except Exception as ex:
        return jsonify({"error": f"No se pudo leer el archivo: {ex}"}), 400

    def limpiar_num(v):
        try:
            if v is None or str(v).strip() in ('', 'nan', 'None', '-', '0'):
                return None
            f = float(v)
            return round(f, 1) if f > 0 else None
        except Exception:
            return None

    def limpiar_str(v):
        return str(v).strip() if v else ''

    resultados = []   # por hoja
    total_ok = 0
    total_no = []

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # ── 1. Detectar profesor (fila 4, col 2)
            cell_prof = limpiar_str(ws.cell(4, 2).value)
            profesor  = _re.sub(r'(?i)maestro[/a]*\s*:\s*', '', cell_prof).strip()
            if not profesor:
                profesor = file.filename

            # ── 2. Detectar materia (fila 3, col 2)
            cell_mat = limpiar_str(ws.cell(3, 2).value)
            materia  = materia_override or _re.sub(r'(?i)área[/técnica]*\s*:\s*', '', cell_mat).strip()
            if not materia:
                materia = sheet_name

            # ── 3. Localizar columna "PORCENTAJE DE ASISTENCIA" (fila 6)
            # y columnas del resumen de períodos (fila 8: "Periodo I/II/III/IV")
            col_asist_p1 = None
            col_resumen_p1 = None

            for c in range(1, ws.max_column + 1):
                v6 = limpiar_str(ws.cell(6, c).value).lower()
                if 'porcentaje' in v6 and 'asistencia' in v6:
                    col_asist_p1 = c   # P1 asistencia, P2=c+1, P3=c+2, P4=c+3

                v8 = limpiar_str(ws.cell(8, c).value).lower().strip()
                # Exacto: debe ser "periodo i" (no "periodo ii", "periodo iv"...)
                if v8 in ('periodo i', 'período i', 'periodo  i'):
                    col_resumen_p1 = c  # P1, P2=c+1, P3=c+2, P4=c+3

            sheet_ok  = 0
            sheet_no  = []

            # ── 4. Recorrer estudiantes (fila 10 en adelante, col 3=nombre, 4=apellido)
            for row in range(10, ws.max_row + 1):
                nombre_raw   = ws.cell(row, 3).value
                apellido_raw = ws.cell(row, 4).value

                if not nombre_raw or str(nombre_raw).strip() in ('', '0', 'Nombres'):
                    continue

                nombre_full   = limpiar_str(nombre_raw)
                apellido_full = limpiar_str(apellido_raw) if apellido_raw else ''
                nombre   = nombre_full   # pass full name for better fuzzy match
                apellido = apellido_full

                est = _buscar_estudiante_bd(conn, nombre, apellido)
                if not est:
                    sheet_no.append(f"{limpiar_str(nombre_raw)} {limpiar_str(apellido_raw)}")
                    continue

                est_id  = est['id']
                cedula  = est['cedula']

                # ── Leer calificaciones de la sección resumen
                p1n = p2n = p3n = p4n = None
                if col_resumen_p1:
                    p1n = limpiar_num(ws.cell(row, col_resumen_p1).value)
                    p2n = limpiar_num(ws.cell(row, col_resumen_p1 + 1).value)
                    p3n = limpiar_num(ws.cell(row, col_resumen_p1 + 2).value)
                    p4n = limpiar_num(ws.cell(row, col_resumen_p1 + 3).value)

                notas = [x for x in [p1n, p2n, p3n, p4n] if x]
                if not notas:
                    sheet_no.append(f"{limpiar_str(nombre_raw)} (sin notas)")
                    continue

                prom = round(sum(notas) / len(notas), 1)

                # ── Guardar en materias_calificaciones (sin sobreescribir)
                conn.execute("""
                    INSERT INTO materias_calificaciones
                        (estudiante_id, cedula, materia, p1, p2, p3, p4, promedio, fuente, profesor)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(estudiante_id, materia) DO UPDATE SET
                        p1       = CASE WHEN excluded.p1 IS NOT NULL THEN excluded.p1
                                        ELSE materias_calificaciones.p1 END,
                        p2       = CASE WHEN excluded.p2 IS NOT NULL THEN excluded.p2
                                        ELSE materias_calificaciones.p2 END,
                        p3       = CASE WHEN excluded.p3 IS NOT NULL THEN excluded.p3
                                        ELSE materias_calificaciones.p3 END,
                        p4       = CASE WHEN excluded.p4 IS NOT NULL THEN excluded.p4
                                        ELSE materias_calificaciones.p4 END,
                        promedio = excluded.promedio,
                        fuente   = excluded.fuente,
                        profesor = CASE WHEN excluded.profesor IS NOT NULL AND excluded.profesor != ''
                                        THEN excluded.profesor
                                        ELSE materias_calificaciones.profesor END,
                        fecha_carga = date('now')
                """, (est_id, cedula, _normalizar_materia(materia), p1n, p2n, p3n, p4n, prom, file.filename, profesor))

                # ── Sincronizar módulos Multimedia si aplica
                mat_norm = materia.lower().strip()
                MODULO_MAP = {
                    'fotografía':'fotografia','fotografia':'fotografia','foto':'fotografia',
                    'lenguaje visual':'lv','lv':'lv',
                    'diseño':'diseno','diseno':'diseno','diseño básico':'diseno','diseño basico':'diseno',
                }
                col_base = MODULO_MAP.get(mat_norm)
                if col_base:
                    p_col = 'p_foto' if col_base == 'fotografia' else f'p_{col_base}'
                    sets, vals = [], []
                    for pi, pv in enumerate([p1n, p2n, p3n, p4n], 1):
                        if pv:
                            sets.append(f"{col_base}_p{pi} = CASE WHEN {col_base}_p{pi} IS NULL OR {col_base}_p{pi}=0 THEN ? ELSE {col_base}_p{pi} END")
                            vals.append(pv)
                    if sets:
                        sets.append(f"{p_col} = ?")
                        vals.append(prom)
                        vals.append(est_id)
                        conn.execute(f"UPDATE estudiantes SET {', '.join(sets)} WHERE id=?", vals)
                        # Recalcular prom_modulos
                        e2 = conn.execute(
                            "SELECT p_foto, p_lv, p_diseno FROM estudiantes WHERE id=?", (est_id,)
                        ).fetchone()
                        if e2:
                            mods = [x for x in e2 if x and x > 0]
                            pm2  = round(sum(mods) / len(mods), 1) if mods else None
                            conn.execute("UPDATE estudiantes SET prom_modulos=? WHERE id=?", (pm2, est_id))

                # ── Guardar asistencia por período
                if col_asist_p1:
                    a1 = limpiar_num(ws.cell(row, col_asist_p1).value)
                    a2 = limpiar_num(ws.cell(row, col_asist_p1 + 1).value)
                    a3 = limpiar_num(ws.cell(row, col_asist_p1 + 2).value)
                    a4 = limpiar_num(ws.cell(row, col_asist_p1 + 3).value)

                    asist_sets, asist_vals = [], []
                    for pi, pv in enumerate([a1, a2, a3, a4], 1):
                        if pv:
                            asist_sets.append(
                                f"asistencia_p{pi} = CASE WHEN asistencia_p{pi} IS NULL OR asistencia_p{pi}=0 "
                                f"THEN ? ELSE asistencia_p{pi} END"
                            )
                            asist_vals.append(pv)
                    if asist_sets:
                        asist_vals.append(est_id)
                        conn.execute(
                            f"UPDATE estudiantes SET {', '.join(asist_sets)} WHERE id=?",
                            asist_vals
                        )

                # Recalcular p_acad del estudiante
                try:
                    proms = conn.execute(
                        "SELECT promedio FROM materias_calificaciones WHERE estudiante_id=? AND promedio>0",
                        (est_id,)
                    ).fetchall()
                    if proms:
                        p_acad_nuevo = round(sum(r[0] for r in proms) / len(proms), 1)
                        conn.execute("UPDATE estudiantes SET p_acad=?, tiene_notas=1 WHERE id=?",
                                     (p_acad_nuevo, est_id))
                except Exception: pass

                sheet_ok += 1

            conn.commit()
            total_ok += sheet_ok
            total_no.extend(sheet_no)
            resultados.append({
                "hoja": sheet_name,
                "materia": materia,
                "profesor": profesor,
                "cargados": sheet_ok,
                "no_encontrados": sheet_no,
            })

    cache_bust()
    return jsonify({
        "status": "success",
        "hojas_procesadas": len(resultados),
        "total_cargados": total_ok,
        "no_encontrados": total_no,
        "detalle": resultados,
        "mensaje": f"{total_ok} estudiantes actualizados en {len(resultados)} hoja(s)."
            + (f" {len(total_no)} no encontrados." if total_no else "")
    })



# ── COMPARATIVA POR MENCIÓN ──────────────────────────────────────────────────


@estudiantes_bp.route("/api/comparativa-mencion")
@login_required
def comparativa_mencion():
    """Devuelve promedios académicos, conductuales y de riesgo agrupados por mención."""
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT
                curso,
                AVG(CASE WHEN p_acad  > 0 THEN p_acad  END) as avg_acad,
                AVG(CASE WHEN p_cond  > 0 THEN p_cond  END) as avg_cond,
                AVG(CASE WHEN p_auto  > 0 THEN p_auto  END) as avg_auto,
                AVG(CASE WHEN indice_riesgo > 0 THEN indice_riesgo END) as avg_riesgo,
                AVG(CASE WHEN prom_modulos  > 0 THEN prom_modulos  END) as avg_modulos,
                COUNT(*) as total,
                SUM(CASE WHEN p_acad > 0 THEN 1 ELSE 0 END) as con_notas,
                SUM(CASE WHEN indice_riesgo >= 50 THEN 1 ELSE 0 END) as riesgo_alto,
                SUM(CASE WHEN categoria = 'ALERTA DE REPROBACIÓN' THEN 1 ELSE 0 END) as alertas
            FROM estudiantes
            WHERE condicion IS NULL OR condicion != 'RETIRADO'
            GROUP BY curso
            ORDER BY curso
        """).fetchall()

    MENCION_LABEL = {
        'multimedia': 'Multimedia',
        'teatro':     'Teatro',
        'musica':     'Música',
        'visual':     'Artes Visuales',
    }
    MENCION_COLOR = {
        'multimedia': '#c8f060',
        'teatro':     '#60b8f0',
        'musica':     '#ff9f60',
        'visual':     '#c060f0',
    }

    result = []
    for r in rows:
        curso = (r['curso'] or '').lower()
        mencion_key = next((k for k in MENCION_LABEL if k in curso), None)
        if not mencion_key:
            continue
        result.append({
            'curso':       r['curso'],
            'mencion':     MENCION_LABEL[mencion_key],
            'color':       MENCION_COLOR[mencion_key],
            'total':       r['total'],
            'con_notas':   r['con_notas'],
            'riesgo_alto': r['riesgo_alto'],
            'alertas':     r['alertas'],
            'avg_acad':    round(r['avg_acad'] or 0, 1),
            'avg_cond':    round(r['avg_cond'] or 0, 1),
            'avg_auto':    round(r['avg_auto'] or 0, 1),
            'avg_riesgo':  round(r['avg_riesgo'] or 0, 1),
            'avg_modulos': round(r['avg_modulos'] or 0, 1),
        })

    # Aggregate per mención (sum groups like "4to A Multimedia", "4to B Multimedia")
    merged = {}
    for r in result:
        m = r['mencion']
        if m not in merged:
            merged[m] = {**r, '_cnt': 1}
        else:
            ex = merged[m]
            ex['total']       += r['total']
            ex['con_notas']   += r['con_notas']
            ex['riesgo_alto'] += r['riesgo_alto']
            ex['alertas']     += r['alertas']
            # Running average
            n = ex['_cnt']
            for k in ('avg_acad','avg_cond','avg_auto','avg_riesgo','avg_modulos'):
                ex[k] = round((ex[k] * n + r[k]) / (n + 1), 1)
            ex['_cnt'] += 1

    final = []
    for m, d in merged.items():
        d.pop('_cnt', None)
        final.append(d)

    return jsonify(final)


# ── FOTO DE PERFIL ───────────────────────────────────────────────────────────


@estudiantes_bp.route("/api/estudiante/<int:id>/foto", methods=["POST"])
@login_required
def subir_foto(id):
    """Sube foto de perfil de un estudiante."""
    f = request.files.get("foto")
    if not f or not f.filename:
        return jsonify({"error": "No se recibió archivo"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        return jsonify({"error": "Formato no válido. Usa JPG, PNG o WEBP"}), 400
    datos = f.read()
    if len(datos) > 5 * 1024 * 1024:  # 5 MB máx para fotos
        return jsonify({"error": "La foto no puede superar 5 MB"}), 413
    if not _validar_magic_imagen(datos, ext):
        return jsonify({"error": "El archivo no es una imagen válida"}), 400
    filename = f"est_{id}{ext}"
    path     = os.path.join(FOTOS_DIR, filename)
    with open(path, "wb") as _fh:
        _fh.write(datos)
    # URL autenticada — funciona en Render (/data/fotos) y en local (static/fotos)
    url = f"/api/foto/{id}"
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("UPDATE estudiantes SET foto_path=? WHERE id=?", (url, id))
        conn.commit()
    return jsonify({"ok": True, "url": url})


@estudiantes_bp.route("/api/estudiante/<int:id>/foto", methods=["DELETE"])
@login_required
def borrar_foto(id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT foto_path FROM estudiantes WHERE id=?", (id,)).fetchone()
        if row and row["foto_path"]:
            try:
                # Reconstruir ruta segura: solo se permite borrar dentro de FOTOS_DIR
                filename = os.path.basename(row["foto_path"])
                safe_path = os.path.abspath(os.path.join(FOTOS_DIR, filename))
                allowed_dir = os.path.abspath(FOTOS_DIR)
                if safe_path.startswith(allowed_dir + os.sep):
                    os.remove(safe_path)
            except Exception as _e:
                logger.warning(f"[estudiantes] Excepción silenciada")
        conn.execute("UPDATE estudiantes SET foto_path=NULL WHERE id=?", (id,))
        conn.commit()
    return jsonify({"ok": True})


# ── PORTAL DEL PROFESOR ───────────────────────────────────────────────────────

# portal_profesor → see full implementation below


@estudiantes_bp.route("/api/debug-boletin", methods=["POST"])
@login_required
def debug_boletin():
    """
    Diagnóstico del parser de boletín.
    Muestra las primeras filas de cada hoja y lo que detecta sin guardar nada.
    """
    if "file" not in request.files:
        return jsonify({"error": "Sin archivo"}), 400
    file_bytes = request.files["file"].read()

    from openpyxl import load_workbook
    import io as _io

    wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    resultado = {}

    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        # Primeras 15 filas para diagnóstico
        preview = []
        for i, r in enumerate(rows[:15]):
            preview.append({
                "fila": i + 1,
                "c0": str(r[0] or '')[:40] if len(r) > 0 else '',
                "c1": str(r[1] or '')[:40] if len(r) > 1 else '',
                "c2": str(r[2] or '')[:40] if len(r) > 2 else '',
                "c3": str(r[3] or '')[:20] if len(r) > 3 else '',
                "c6": str(r[6] or '')[:40] if len(r) > 6 else '',
            })
        # Buscar filas ALUMNO/A
        alumno_rows = []
        for i, r in enumerate(rows):
            v = str(r[2] or '').strip().upper() if len(r) > 2 else ''
            if 'ALUMNO' in v:
                alumno_rows.append({
                    "fila": i + 1,
                    "c2": v,
                    "c6": str(r[6] or '')[:50] if len(r) > 6 else ''
                })
            if len(alumno_rows) >= 3:
                break

        resultado[sheet_name] = {
            "total_filas": len(rows),
            "preview_15": preview,
            "alumno_rows": alumno_rows
        }

    wb.close()
    # También correr el parser real
    try:
        parsed = _parsear_boletin_bj(file_bytes)
        resultado["__parser_result"] = {
            "total_estudiantes": len(parsed),
            "primeros_2": parsed[:2] if parsed else []
        }
    except Exception as ex:
        resultado["__parser_error"] = str(ex)

    return jsonify(resultado)



# ══════════════════════════════════════════════════════════════════════════════
#  ÍNDICE CONDUCTUAL Y BIENESTAR EMOCIONAL — Cálculo dinámico MINERD
#  Base 100, descuenta por reportes y entradas del cuaderno anecdótico
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/expediente/<int:est_id>", methods=["GET"])
@login_required
def get_expediente(est_id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        _recalcular_indicadores(conn, est_id)
        conn.commit()
        reportes = conn.execute("""
            SELECT id,'reporte' as fuente,tipo,subtipo,titulo,descripcion,
                   severidad,reportado_por,fecha,estado,seguimiento,fecha_cierre,
                   NULL as registrado_por
            FROM reportes WHERE estudiante_id=? ORDER BY fecha DESC
        """, (est_id,)).fetchall()
        logros = []
        try:
            logros = conn.execute("""
                SELECT id,'logro' as fuente,tipo,NULL as subtipo,titulo,descripcion,
                       NULL as severidad,NULL as reportado_por,fecha,'Activo' as estado,
                       NULL as seguimiento,NULL as fecha_cierre,registrado_por
                FROM logros WHERE estudiante_id=? ORDER BY fecha DESC
            """, (est_id,)).fetchall()
        except Exception as _e:
            logger.warning(f"[estudiantes] Excepción silenciada")
        try:
            indicadores = conn.execute(
                "SELECT ind_conducta,ind_psico,ind_academico,ind_logros FROM estudiantes WHERE id=?",
                (est_id,)).fetchone()
        except Exception:
            indicadores = None
    eventos = [dict(r) for r in reportes] + [dict(r) for r in logros]
    eventos.sort(key=lambda x: x.get('fecha','') or '', reverse=True)
    return jsonify({"eventos": eventos, "indicadores": dict(indicadores) if indicadores else {}})


@estudiantes_bp.route("/api/logros/<int:est_id>", methods=["GET"])
@login_required
def get_logros(est_id):
    try:
        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT * FROM logros WHERE estudiante_id=? ORDER BY fecha DESC", (est_id,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception:
        return jsonify([])


@estudiantes_bp.route("/api/logros", methods=["POST"])
@login_required
def crear_logro():
    d = request.get_json(silent=True) or {}
    est_id = d.get("estudiante_id")
    tipo   = d.get("tipo","reconocimiento")
    titulo = d.get("titulo","").strip()
    descripcion  = d.get("descripcion","").strip()
    fecha        = d.get("fecha","")
    registrado_por = d.get("registrado_por", session.get("nombre",""))
    if not est_id or not titulo:
        return jsonify({"error":"estudiante_id y titulo son requeridos"}), 400
    try:
        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.execute("INSERT INTO logros(estudiante_id,tipo,titulo,descripcion,fecha,registrado_por) VALUES(?,?,?,?,?,?)",
                         (est_id, tipo, titulo, descripcion, fecha, registrado_por))
            conn.commit()
            _recalcular_indicadores(conn, est_id)
            conn.commit()
        cache_bust()
        return jsonify({"ok":True})
    except Exception as ex:
        logger.error(f"[ESTUDIANTES] Error registrando logro: {ex}")
        return jsonify({"error": "Error al registrar el logro. Intenta de nuevo."}), 500


@estudiantes_bp.route("/api/logros/<int:logro_id>", methods=["DELETE"])
@login_required
def eliminar_logro(logro_id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        logro = conn.execute("SELECT estudiante_id FROM logros WHERE id=?", (logro_id,)).fetchone()
        if logro:
            conn.execute("DELETE FROM logros WHERE id=?", (logro_id,))
            conn.commit()
            _recalcular_indicadores(conn, logro['estudiante_id'])
            conn.commit()
    cache_bust()
    return jsonify({"ok":True})


@estudiantes_bp.route("/api/cargar-boletin", methods=["POST"])
@login_required
@rate_limited(max_calls=10, window=3600)
def cargar_boletin():
    """
    Carga el Boletín oficial del C.E. Benito Juárez.
    Soporta:
      - Boletín Primer Ciclo (1ro-3ro): sheets BC 1RO A, BC 2DO B...
      - Boletín Segundo Ciclo (4to-6to): sheets BC 4TO A MÚSICA...
    Por cada estudiante:
      1. Busca o crea el perfil en 'estudiantes'
      2. Guarda/actualiza las notas en 'materias_calificaciones'
      3. Recalcula p_acad del estudiante
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({"error": "Formato no válido. Sube un archivo Excel (.xlsx/.xlsm)"}), 400

    try:
        file_bytes = file.read()
        logger.info(f"[cargar_boletin] Recibido: {getattr(file, 'filename', '?')}, {len(file_bytes)} bytes")
        estudiantes_parsed = _parsear_boletin_bj(file_bytes)
        logger.info(f"[cargar_boletin] Parser: {len(estudiantes_parsed)} estudiantes")

        if not estudiantes_parsed:
            # Diagnóstico: listar hojas encontradas para ayudar a depurar
            try:
                from openpyxl import load_workbook
                import io as _io2
                wb_diag = load_workbook(_io2.BytesIO(file_bytes), read_only=True, data_only=True)
                hojas = wb_diag.sheetnames
                wb_diag.close()
            except Exception:
                hojas = []
            # Detectar si es un Registro del Coordinador (hojas con patrón "CF GRADO SECC MENCIÓN")
            import re as _re
            es_coordinador = any(
                _re.match(r"CF\s+\w+\s+[A-Ea-e]\s+.+", h.strip(), _re.IGNORECASE)
                for h in hojas
            )
            if es_coordinador:
                return jsonify({
                    "error": (
                        "⚠ Este archivo es un Registro del Coordinador, no un Boletín Oficial. "
                        "Usa la tarjeta morada «Registro del Coordinador» (la tercera en la sección Excel)."
                    )
                }), 400
            return jsonify({
                "error": (
                    "No se encontraron estudiantes en el archivo. "
                    f"Hojas detectadas: {hojas}. "
                    "Verifica que sea un Boletín oficial del C.E. Benito Juárez "
                    "(4to–6to o 1ro–3ro) con el formato estándar."
                )
            }), 400

        from datetime import date as _date
        hoy = _date.today().isoformat()

        resumen        = {}
        total_est      = 0
        total_materias = 0
        creados        = 0
        errores        = []

        with sqlite3.connect(DATABASE, timeout=10) as conn:
            conn.row_factory = sqlite3.Row

            for est in estudiantes_parsed:
                try:
                    est_id = _buscar_o_crear_estudiante(
                        conn,
                        est['nombre'], est['apellido'],
                        est['grado'],  est['ciclo'],
                        est['seccion'], est['mencion']
                    )
                    if not est_id:
                        raise ValueError("No se pudo crear/encontrar el perfil del estudiante")

                    # Actualizar ciclo y sección en el perfil
                    conn.execute(
                        "UPDATE estudiantes SET ciclo=?, seccion=? WHERE id=?",
                        (est['ciclo'], est['seccion'], est_id)
                    )

                    # Insertar/actualizar cada materia
                    for mat in est['materias']:
                        notas = [mat['p1'], mat['p2'], mat['p3'], mat['p4']]
                        notas_validas = [n for n in notas if n > 0]
                        prom = round(sum(notas_validas) / len(notas_validas), 2) if notas_validas else 0

                        # Intentar INSERT completo con columnas extendidas
                        try:
                            conn.execute("""
                                INSERT INTO materias_calificaciones
                                    (estudiante_id, materia, p1, p2, p3, p4,
                                     promedio, tipo, ciclo, fecha_carga, profesor)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ON CONFLICT(estudiante_id, materia) DO UPDATE SET
                                    p1       = CASE WHEN excluded.p1 > 0 THEN excluded.p1 ELSE materias_calificaciones.p1 END,
                                    p2       = CASE WHEN excluded.p2 > 0 THEN excluded.p2 ELSE materias_calificaciones.p2 END,
                                    p3       = CASE WHEN excluded.p3 > 0 THEN excluded.p3 ELSE materias_calificaciones.p3 END,
                                    p4       = CASE WHEN excluded.p4 > 0 THEN excluded.p4 ELSE materias_calificaciones.p4 END,
                                    promedio = excluded.promedio,
                                    tipo     = excluded.tipo,
                                    ciclo    = excluded.ciclo,
                                    fecha_carga = excluded.fecha_carga,
                                    profesor = COALESCE(excluded.profesor, materias_calificaciones.profesor)
                            """, (
                                est_id, _normalizar_materia(mat['nombre']),
                                mat['p1'], mat['p2'], mat['p3'], mat['p4'],
                                prom, mat['tipo'], est['ciclo'], hoy, est['maestro']
                            ))
                        except Exception:
                            # Fallback: INSERT básico sin columnas extendidas
                            conn.execute("""
                                INSERT INTO materias_calificaciones
                                    (estudiante_id, materia, p1, p2, p3, p4, promedio, fuente)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 'boletin')
                                ON CONFLICT(estudiante_id, materia) DO UPDATE SET
                                    p1       = CASE WHEN excluded.p1 > 0 THEN excluded.p1 ELSE materias_calificaciones.p1 END,
                                    p2       = CASE WHEN excluded.p2 > 0 THEN excluded.p2 ELSE materias_calificaciones.p2 END,
                                    p3       = CASE WHEN excluded.p3 > 0 THEN excluded.p3 ELSE materias_calificaciones.p3 END,
                                    p4       = CASE WHEN excluded.p4 > 0 THEN excluded.p4 ELSE materias_calificaciones.p4 END,
                                    promedio = excluded.promedio
                            """, (
                                est_id, _normalizar_materia(mat['nombre']),
                                mat['p1'], mat['p2'], mat['p3'], mat['p4'], prom
                            ))
                        total_materias += 1

                    # Recalcular p_acad del estudiante
                    proms = conn.execute("""
                        SELECT promedio FROM materias_calificaciones
                        WHERE estudiante_id=? AND tipo='académico' AND promedio > 0
                    """, (est_id,)).fetchall()
                    if proms:
                        p_acad_nuevo = round(sum(r[0] for r in proms) / len(proms), 2)
                        # Calcular p_acad_p1 y p_acad_p2 para dashboard
                        p1s = conn.execute("""
                            SELECT p1 FROM materias_calificaciones
                            WHERE estudiante_id=? AND tipo='académico' AND p1 > 0
                        """, (est_id,)).fetchall()
                        p2s = conn.execute("""
                            SELECT p2 FROM materias_calificaciones
                            WHERE estudiante_id=? AND tipo='académico' AND p2 > 0
                        """, (est_id,)).fetchall()
                        acad_p1 = round(sum(r[0] for r in p1s)/len(p1s), 2) if p1s else 0
                        acad_p2 = round(sum(r[0] for r in p2s)/len(p2s), 2) if p2s else 0

                        # Safe update — handles old DBs missing tiene_notas column
                        try:
                            conn.execute("""
                                UPDATE estudiantes
                                   SET p_acad=?, acad_p1=?, acad_p2=?, tiene_notas=1
                                 WHERE id=?
                            """, (p_acad_nuevo, acad_p1, acad_p2, est_id))
                        except Exception:
                            conn.execute("""
                                UPDATE estudiantes SET p_acad=?, acad_p1=?, acad_p2=?
                                WHERE id=?
                            """, (p_acad_nuevo, acad_p1, acad_p2, est_id))
                        # Update seccion/ciclo from boletín data
                        try:
                            conn.execute(
                                "UPDATE estudiantes SET seccion=?, ciclo=? WHERE id=?",
                                (est.get('seccion', 'A'), est.get('ciclo', 'segundo_ciclo'), est_id)
                            )
                        except Exception: pass

                    total_est += 1
                    clave = f"{est['grado']} {est['mencion']}".strip()
                    resumen[clave] = resumen.get(clave, 0) + 1

                except Exception as ex_est:
                    import traceback as _tb
                    errores.append(
                        f"{est.get('nombre','')} {est.get('apellido','')}: "
                        f"{ex_est} | {_tb.format_exc().splitlines()[-1]}"
                    )

            conn.commit()

        cache_bust()

        if total_est == 0:
            return jsonify({
                "ok": False,
                "error": (
                    errores[0] if errores else
                    "No se pudo procesar ningún estudiante. "
                    "Verifica que el archivo sea un Boletín oficial del C.E. Benito Juárez."
                ),
                "errores": errores[:10],
                "estudiantes_encontrados_en_archivo": len(estudiantes_parsed),
            }), 400

        cache_bust()

        if total_est == 0 and errores:
            return jsonify({
                "ok":    False,
                "error": f"No se guardó ningún estudiante. Error: {errores[0]}",
                "errores": errores[:5]
            }), 400

        return jsonify({
            "ok": True,
            "estudiantes_procesados": total_est,
            "materias_guardadas":     total_materias,
            "resumen_por_grupo":      resumen,
            "errores":                errores[:10],
            "mensaje": (
                f"✓ {total_est} estudiantes — {total_materias} calificaciones cargadas. "
                + (f"({len(errores)} advertencias)" if errores else "")
            )
        })

    except Exception as ex:
        logger.error(f"[cargar_boletin] ERROR: {ex}", exc_info=True)
        return jsonify({"error": "Error al procesar el boletín. Intenta de nuevo."}), 500



# ══════════════════════════════════════════════════════════════════════════════
#  CUADERNO ANECDÓTICO
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/cuaderno/<int:est_id>", methods=["GET"])
@login_required
def get_cuaderno(est_id):
    u = get_usuario()
    rol_n = _normalizar_rol(u.get("rol", ""))
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        # Verificar acceso al estudiante
        est = conn.execute("SELECT ciclo FROM estudiantes WHERE id=?", (est_id,)).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404
        ciclo_est = est["ciclo"] or "segundo_ciclo"
        ciclo_acc = u.get("ciclo_acceso")
        if ciclo_acc and ciclo_acc != ciclo_est and not u.get("es_directora"):
            return jsonify({"error": "Sin acceso a este estudiante"}), 403

        # Reglas de visibilidad según rol:
        # - Directora / coordinadores / psicólogas: ven TODO incluyendo privadas
        # - Profesor: ve solo sus propias entradas + las no-privadas de otros
        # - Cualquier otro: solo entradas no-privadas con visible_en_perfil=1
        puede_ver_privadas = (
            u.get("es_directora") or
            u.get("es_coord") or
            u.get("es_psicologa") or
            rol_n in ROLES_COORD or
            rol_n in ROLES_PSICOLOGA
        )

        if puede_ver_privadas:
            rows = conn.execute("""
                SELECT ca.*, u.nombre as autor_nombre, u.rol as autor_rol
                FROM cuaderno_anecdotico ca
                JOIN usuarios u ON ca.autor_id = u.id
                WHERE ca.estudiante_id = ?
                ORDER BY ca.fecha DESC, ca.creado_en DESC
            """, (est_id,)).fetchall()
        elif rol_n == "profesor":
            # Profesor ve sus entradas + las públicas de otros
            rows = conn.execute("""
                SELECT ca.*, u.nombre as autor_nombre, u.rol as autor_rol
                FROM cuaderno_anecdotico ca
                JOIN usuarios u ON ca.autor_id = u.id
                WHERE ca.estudiante_id = ?
                  AND (ca.autor_id = ? OR ca.privado = 0)
                ORDER BY ca.fecha DESC, ca.creado_en DESC
            """, (est_id, u["id"])).fetchall()
        else:
            rows = conn.execute("""
                SELECT ca.*, u.nombre as autor_nombre, u.rol as autor_rol
                FROM cuaderno_anecdotico ca
                JOIN usuarios u ON ca.autor_id = u.id
                WHERE ca.estudiante_id = ?
                  AND ca.privado = 0
                  AND ca.visible_en_perfil = 1
                ORDER BY ca.fecha DESC, ca.creado_en DESC
            """, (est_id,)).fetchall()

    return jsonify([dict(r) for r in rows])


@estudiantes_bp.route("/api/cuaderno", methods=["POST"])
@login_required
def crear_entrada_cuaderno():
    u = get_usuario()
    import json as _json
    data = request.get_json() or {}
    est_id      = data.get("estudiante_id")
    tipo        = data.get("tipo", "conductual")
    descripcion = (data.get("descripcion") or "").strip()
    seguimiento = (data.get("seguimiento") or "").strip()
    fecha       = data.get("fecha") or __import__("datetime").date.today().isoformat()
    privado     = 1 if data.get("privado") else 0
    polaridad   = data.get("polaridad", "neutro")
    if polaridad not in ("positivo", "negativo", "neutro"):
        polaridad = "neutro"
    tags_raw    = data.get("tags", [])
    tags_json   = _json.dumps(tags_raw if isinstance(tags_raw, list) else [], ensure_ascii=False)

    if not est_id or not descripcion:
        return jsonify({"error": "Faltan campos obligatorios"}), 400

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        est = conn.execute("SELECT ciclo FROM estudiantes WHERE id=?", (est_id,)).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        conn.execute("""
            INSERT INTO cuaderno_anecdotico
                (estudiante_id, autor_id, fecha, tipo, descripcion,
                 seguimiento, privado, visible_en_perfil, polaridad, tags)
            VALUES (?,?,?,?,?,?,?,1,?,?)
        """, (est_id, u["id"], fecha, tipo, descripcion, seguimiento, privado,
              polaridad, tags_json))
        conn.commit()
    cache_bust()
    return jsonify({"ok": True, "mensaje": "Entrada registrada en el cuaderno anecdótico"})


@estudiantes_bp.route("/api/cuaderno/<int:entrada_id>", methods=["PATCH"])
@login_required
def editar_entrada_cuaderno(entrada_id):
    u = get_usuario()
    data = request.get_json() or {}
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        entrada = conn.execute(
            "SELECT * FROM cuaderno_anecdotico WHERE id=?", (entrada_id,)
        ).fetchone()
        if not entrada:
            return jsonify({"error": "Entrada no encontrada"}), 404
        # Solo el autor o roles superiores pueden editar
        rol_n = _normalizar_rol(u.get("rol",""))
        if entrada["autor_id"] != u["id"] and rol_n not in {"directora","coordinador_general","coordinador_primer_ciclo","coordinador_segundo_ciclo"}:
            return jsonify({"error": "Solo el autor puede editar esta entrada"}), 403

        import json as _json
        pol_new = data.get("polaridad", entrada["polaridad"] or "neutro")
        if pol_new not in ("positivo", "negativo", "neutro"):
            pol_new = "neutro"
        tags_new = data.get("tags")
        if tags_new is None:
            tags_json = entrada["tags"] or "[]"
        else:
            tags_json = _json.dumps(tags_new if isinstance(tags_new, list) else [],
                                    ensure_ascii=False)
        conn.execute("""
            UPDATE cuaderno_anecdotico
               SET descripcion=?, seguimiento=?, tipo=?, privado=?, polaridad=?, tags=?
             WHERE id=?
        """, (
            data.get("descripcion", entrada["descripcion"]),
            data.get("seguimiento", entrada["seguimiento"]),
            data.get("tipo", entrada["tipo"]),
            1 if data.get("privado") else 0,
            pol_new,
            tags_json,
            entrada_id
        ))
        conn.commit()
    return jsonify({"ok": True})


@estudiantes_bp.route("/api/cuaderno/<int:entrada_id>/convertir-reporte", methods=["POST"])
@login_required
def convertir_reporte_cuaderno(entrada_id):
    """
    Escala una entrada del cuaderno anecdótico a reporte formal.
    Opcionalmente crea un Caso y envía notificaciones a psicóloga/coordinación.

    Body JSON:
      severidad     : "Leve" | "Media" | "Grave" | "Muy Grave"  (default: "Media")
      tipo_reporte  : "conducta" | "psicologico" | "academico"   (auto si no viene)
      notas         : texto adicional del que escala
      crear_caso    : bool — si true crea también un registro en casos
    """
    u = get_usuario()
    rol_n = _normalizar_rol(u.get("rol", ""))

    ROLES_ESCALA = ROLES_COORD | ROLES_PSICOLOGA | ROLES_DIRECTORA
    puede_escalar = (
        u.get("es_directora") or u.get("es_coord") or u.get("es_psicologa") or
        rol_n in ROLES_ESCALA
    )
    # Profesores también pueden escalar desde su cuaderno (solo crean reporte, no caso)
    es_profesor = rol_n == "profesor"

    if not puede_escalar and not es_profesor:
        return jsonify({"error": "Sin permisos para escalar entradas del cuaderno"}), 403

    d = request.get_json(silent=True) or {}
    severidad         = d.get("severidad", "Media")
    notas_adicionales = (d.get("notas", "") or "").strip()
    crear_caso        = bool(d.get("crear_caso", False)) and puede_escalar

    if severidad not in ("Leve", "Media", "Grave", "Muy Grave"):
        severidad = "Media"

    TIPO_MAP = {
        "conductual": "conducta",
        "emocional":  "psicologico",
        "académico":  "academico",
        "familiar":   "psicologico",
        "otro":       "conducta",
    }

    with sqlite3.connect(DATABASE, timeout=15) as conn:
        conn.row_factory = sqlite3.Row
        entrada = conn.execute(
            "SELECT * FROM cuaderno_anecdotico WHERE id=?", (entrada_id,)
        ).fetchone()
        if not entrada:
            return jsonify({"error": "Entrada no encontrada"}), 404
        if entrada["convertido_reporte"]:
            return jsonify({"error": "Esta entrada ya fue convertida a reporte", "ya_existe": True}), 409

        est = conn.execute(
            "SELECT id, nombre, apellido, grado, seccion, ciclo FROM estudiantes WHERE id=?",
            (entrada["estudiante_id"],)
        ).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        tipo_reporte = d.get("tipo_reporte") or TIPO_MAP.get(entrada["tipo"], "conducta")
        if tipo_reporte not in ("conducta", "psicologico", "academico"):
            tipo_reporte = "conducta"

        # ── Descripción compuesta ──────────────────────────────────────────
        desc_reporte = entrada["descripcion"]
        if notas_adicionales:
            desc_reporte += f"\n\n[Nota al escalar — {u['nombre']}]: {notas_adicionales}"
        if entrada["seguimiento"]:
            desc_reporte += f"\n\n[Seguimiento original]: {entrada['seguimiento']}"

        titulo_reporte = (
            f"[{severidad.upper()}] Escalado desde cuaderno — "
            f"{entrada['tipo'].capitalize()} · {est['nombre']} {est['apellido']}"
        )

        # ── Crear reporte ──────────────────────────────────────────────────
        conn.execute("""
            INSERT INTO reportes
                (estudiante_id, tipo, titulo, descripcion,
                 severidad, reportado_por, estado, fecha)
            VALUES (?,?,?,?,?,?,?,?)
        """, (
            est["id"], tipo_reporte, titulo_reporte, desc_reporte,
            severidad, u["nombre"], "Abierto", entrada["fecha"]
        ))
        reporte_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # ── Crear Caso si se solicitó ──────────────────────────────────────
        caso_id = None
        if crear_caso:
            NIVEL_MAP = {"Leve": 1, "Media": 2, "Grave": 3, "Muy Grave": 4}
            conn.execute("""
                INSERT INTO casos
                    (estudiante_id, abierto_por, tipo, titulo, descripcion,
                     estado, nivel_escala, origen_tipo, origen_id, creado_en)
                VALUES (?,?,?,?,?,?,?,?,?,datetime('now'))
            """, (
                est["id"], u["id"], tipo_reporte,
                titulo_reporte, desc_reporte,
                "Abierto", NIVEL_MAP.get(severidad, 2),
                "reporte", reporte_id
            ))
            caso_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            # Actualizar reporte con caso_id
            conn.execute("UPDATE reportes SET caso_id=? WHERE id=?", (caso_id, reporte_id))

        # ── Marcar entrada como convertida ────────────────────────────────
        conn.execute("""
            UPDATE cuaderno_anecdotico
               SET convertido_reporte=1, visible_en_perfil=1
             WHERE id=?
        """, (entrada_id,))

        # ── Notificaciones → psicóloga y coordinadores del ciclo ──────────
        ciclo_est = est["ciclo"] or "segundo_ciclo"
        if "primer" in ciclo_est:
            roles_notif = ("psicologa_primer_ciclo", "coordinador_primer_ciclo", "coordinador_general")
        else:
            roles_notif = ("psicologa_segundo_ciclo", "coordinador_segundo_ciclo", "coordinador_general")

        destinatarios = conn.execute("""
            SELECT id FROM usuarios
            WHERE rol IN ({})
              AND id != ?
        """.format(",".join("?" * len(roles_notif))),
            (*roles_notif, u["id"])
        ).fetchall()

        sev_ico = {"Leve": "🟡", "Media": "🟠", "Grave": "🔴", "Muy Grave": "🚨"}.get(severidad, "⚠️")
        notif_titulo = (
            f"{sev_ico} Reporte {severidad}: {est['nombre']} {est['apellido']} "
            f"({est['grado']}{'·' + est['seccion'] if est['seccion'] else ''})"
        )
        notif_cuerpo = (
            f"Escalado por {u['nombre']} desde cuaderno anecdótico.\n"
            f"Tipo: {tipo_reporte} · Severidad: {severidad}\n"
            + (f"Caso #{caso_id} abierto." if caso_id else "")
        )
        for dest in destinatarios:
            conn.execute("""
                INSERT INTO notificaciones
                    (destinatario_id, origen_tipo, origen_id, estudiante_id, titulo, cuerpo)
                VALUES (?,?,?,?,?,?)
            """, (dest["id"], "reporte", reporte_id, est["id"],
                  notif_titulo, notif_cuerpo))

        conn.commit()

        # ── Recalcular indicadores ────────────────────────────────────────
        _recalcular_indicadores(conn, est["id"])
        conn.commit()

    cache_bust()
    return jsonify({
        "ok":        True,
        "reporte_id": reporte_id,
        "caso_id":    caso_id,
        "notificados": len(destinatarios),
        "mensaje": (
            f"Reporte formal #{reporte_id} creado con severidad {severidad}."
            + (f" Caso #{caso_id} abierto." if caso_id else "")
            + f" {len(destinatarios)} usuario(s) notificado(s)."
        ),
    })


@estudiantes_bp.route("/api/cuaderno/<int:entrada_id>", methods=["DELETE"])
@login_required
def eliminar_entrada_cuaderno(entrada_id):
    u = get_usuario()
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        entrada = conn.execute(
            "SELECT autor_id FROM cuaderno_anecdotico WHERE id=?", (entrada_id,)
        ).fetchone()
        if not entrada:
            return jsonify({"error": "No encontrada"}), 404
        rol_n = _normalizar_rol(u.get("rol",""))
        if entrada["autor_id"] != u["id"] and rol_n not in {"directora","coordinador_general"}:
            return jsonify({"error": "Sin permisos"}), 403
        conn.execute("DELETE FROM cuaderno_anecdotico WHERE id=?", (entrada_id,))
        conn.commit()
    return jsonify({"ok": True})


@estudiantes_bp.route("/api/cuaderno/bandeja", methods=["GET"])
@login_required
def bandeja_cuaderno():
    """
    Bandeja de entradas privadas del cuaderno anecdótico para psicóloga/coordinación.
    Devuelve entradas donde privado=1 (marcadas para psicóloga/coord) del ciclo
    correspondiente al rol del usuario, ordenadas de más reciente a más antigua.
    """
    u = get_usuario()
    rol_n = _normalizar_rol(u.get("rol", ""))
    puede_ver = (
        u.get("es_psicologa") or u.get("es_coord") or u.get("es_directora") or
        rol_n in ROLES_PSICOLOGA or rol_n in ROLES_COORD
    )
    if not puede_ver:
        return jsonify({"error": "Sin acceso a la bandeja"}), 403

    ciclo = u.get("ciclo_acceso")   # None = directora ve todo

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        if ciclo:
            rows = conn.execute("""
                SELECT ca.id, ca.tipo, ca.descripcion, ca.seguimiento,
                       ca.fecha, ca.privado, ca.convertido_reporte,
                       ca.polaridad, ca.tags, ca.creado_en,
                       e.id as estudiante_id, e.nombre, e.apellido,
                       e.grado, e.seccion, e.ciclo,
                       u2.nombre as autor_nombre, u2.rol as autor_rol
                FROM cuaderno_anecdotico ca
                JOIN estudiantes e ON ca.estudiante_id = e.id
                JOIN usuarios u2   ON ca.autor_id = u2.id
                WHERE ca.privado = 1
                  AND e.ciclo = ?
                  AND ca.convertido_reporte = 0
                ORDER BY ca.fecha DESC, ca.creado_en DESC
                LIMIT 100
            """, (ciclo,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT ca.id, ca.tipo, ca.descripcion, ca.seguimiento,
                       ca.fecha, ca.privado, ca.convertido_reporte,
                       ca.polaridad, ca.tags, ca.creado_en,
                       e.id as estudiante_id, e.nombre, e.apellido,
                       e.grado, e.seccion, e.ciclo,
                       u2.nombre as autor_nombre, u2.rol as autor_rol
                FROM cuaderno_anecdotico ca
                JOIN estudiantes e ON ca.estudiante_id = e.id
                JOIN usuarios u2   ON ca.autor_id = u2.id
                WHERE ca.privado = 1
                  AND ca.convertido_reporte = 0
                ORDER BY ca.fecha DESC, ca.creado_en DESC
                LIMIT 100
            """).fetchall()

    return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════════════════════
#  CALENDARIO ESCOLAR
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/progreso/<int:est_id>", methods=["GET"])
@login_required
def get_progreso_estudiante(est_id):
    """
    Retorna un resumen de progreso de un estudiante:
    notas por período, tendencia, asistencia mensual, cuaderno anecdótico (resumen).
    """
    u = get_usuario()
    anio_esc = request.args.get("anio_escolar", "2025-2026")

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row

        # Datos básicos del estudiante
        est = conn.execute(
            "SELECT * FROM estudiantes WHERE id=?", (est_id,)
        ).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        # Materias y notas por período
        materias = conn.execute("""
            SELECT materia, tipo, p1, p2, p3, p4, promedio, profesor
            FROM materias_calificaciones
            WHERE estudiante_id=?
            ORDER BY tipo, materia
        """, (est_id,)).fetchall()

        # Asistencia mensual
        asistencia = conn.execute("""
            SELECT am.mes, am.anio, am.materia, am.porcentaje,
                   am.dias_asistio, am.dias_clase_impartidos, am.validado,
                   u.nombre as profesor_nombre
            FROM asistencia_mensual am
            JOIN usuarios u ON am.profesor_id = u.id
            WHERE am.estudiante_id=?
            ORDER BY am.anio DESC, am.mes DESC
        """, (est_id,)).fetchall()

        # Cuaderno anecdótico — solo conteo y últimas 3 entradas
        cuaderno_cnt = conn.execute(
            "SELECT COUNT(*) FROM cuaderno_anecdotico WHERE estudiante_id=?",
            (est_id,)
        ).fetchone()[0]
        cuaderno_reciente = conn.execute("""
            SELECT ca.fecha, ca.tipo, ca.descripcion, u.nombre as autor
            FROM cuaderno_anecdotico ca
            JOIN usuarios u ON ca.autor_id = u.id
            WHERE ca.estudiante_id=? AND ca.visible_en_perfil=1
            ORDER BY ca.fecha DESC LIMIT 3
        """, (est_id,)).fetchall()

        # Evaluaciones narrativas del período actual
        narrativas = conn.execute("""
            SELECT en.periodo, en.texto, u.nombre as profesor_nombre
            FROM evaluaciones_narrativas en
            JOIN usuarios u ON en.profesor_id = u.id
            WHERE en.estudiante_id=? AND en.anio_escolar=?
            ORDER BY en.periodo
        """, (est_id, anio_esc)).fetchall()

    materias_filtradas = _rls.filtrar_materias_profesor([dict(m) for m in materias])

    return jsonify({
        "estudiante":  dict(est),
        "materias":    materias_filtradas,
        "asistencia":  [dict(a) for a in asistencia],
        "cuaderno": {
            "total_entradas": cuaderno_cnt,
            "recientes": [dict(c) for c in cuaderno_reciente]
        },
        "evaluaciones_narrativas": [dict(n) for n in narrativas]
    })


# ══════════════════════════════════════════════════════════════════════════════
#  GESTIÓN DE USUARIOS (ampliada para nuevos roles)
# ══════════════════════════════════════════════════════════════════════════════


@estudiantes_bp.route("/api/cargar-plantilla-bj", methods=["POST"])
@login_required
@rate_limited(max_calls=10, window=3600)
def cargar_plantilla_bj_alias():
    """Alias para compatibilidad: redirige al nuevo endpoint cargar_boletin."""
    return cargar_boletin()


