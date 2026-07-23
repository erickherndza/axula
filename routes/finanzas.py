# -*- coding: utf-8 -*-
"""Blueprint: finanzas — Portal de Contabilidad y Finanzas"""

import sqlite3
import logging
import json as _json
import os
from datetime import datetime, date, timedelta
from flask import (
    Blueprint, render_template, request, jsonify, session,
    redirect, url_for, Response,
)

from core.constants import *
from core.database import get_db, cache_get, cache_set, cache_bust, cache_delete
from core.auth import (
    _normalizar_rol, login_required, get_usuario,
    _csrf_token, _csrf_check, rate_limited,
)
from core.helpers import *
from core.helpers import _anio_escolar_actual, _audit, _validar_magic_excel

logger = logging.getLogger("axula")

finanzas_bp = Blueprint("finanzas_bp", __name__)

def _finanzas_required(f):
    """Solo auxiliar_contabilidad y directora."""
    import functools
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect("/login")
        rol = _normalizar_rol(session.get("rol", ""))
        permitidos = {"auxiliar_contabilidad", "directora", "superusuario"}
        if rol not in permitidos:
            return jsonify({"error": "Sin permisos para esta sección"}), 403
        return f(*args, **kwargs)
    return decorated


@finanzas_bp.route("/finanzas")
@login_required
def portal_finanzas():
    u = get_usuario()
    rol = _normalizar_rol(u.get("rol", ""))
    if rol not in {"auxiliar_contabilidad", "directora", "superusuario"}:
        return redirect("/")
    return render_template("finanzas.html", usuario=u, current_user=u)


# ── CATEGORÍAS DE GASTO ──────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/categorias", methods=["GET"])
@login_required
@_finanzas_required
def listar_categorias():
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM categorias_gasto WHERE activa=1 ORDER BY nombre"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/categorias", methods=["POST"])
@login_required
@_finanzas_required
def crear_categoria():
    d = request.get_json(silent=True) or {}
    nombre = d.get("nombre", "").strip()
    tipo = d.get("tipo", "gasto")
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        try:
            conn.execute(
                "INSERT INTO categorias_gasto (nombre, tipo) VALUES (?,?)",
                (nombre, tipo))
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({"error": "Categoría ya existe"}), 400
    return jsonify({"ok": True})


# ── MOVIMIENTOS FINANCIEROS (ingresos/gastos) ────────────────────────────────

@finanzas_bp.route("/api/finanzas/movimientos", methods=["GET"])
@login_required
@_finanzas_required
def listar_movimientos():
    tipo = request.args.get("tipo", "")
    categoria = request.args.get("categoria", "")
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = """
            SELECT m.*, c.nombre as categoria_nombre, u.nombre as registrado_nombre
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            LEFT JOIN usuarios u ON u.id = m.registrado_por
            WHERE m.estado != 'anulado'
        """
        params = []
        if tipo:
            sql += " AND m.tipo=?"
            params.append(tipo)
        if categoria:
            sql += " AND m.categoria_id=?"
            params.append(int(categoria))
        if desde:
            sql += " AND m.fecha >= ?"
            params.append(desde)
        if hasta:
            sql += " AND m.fecha <= ?"
            params.append(hasta)
        sql += " ORDER BY m.fecha DESC, m.creado_en DESC LIMIT 500"
        rows = conn.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/movimientos", methods=["POST"])
@login_required
@_finanzas_required
@rate_limited(max_calls=30, window=60)
def crear_movimiento():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    concepto = d.get("concepto", "").strip()
    monto = d.get("monto")
    tipo = d.get("tipo", "gasto")

    if not concepto or monto is None:
        return jsonify({"error": "Concepto y monto son requeridos"}), 400
    try:
        monto = float(monto)
    except (ValueError, TypeError):
        return jsonify({"error": "Monto debe ser numérico"}), 400

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            INSERT INTO movimientos_financieros
            (tipo, categoria_id, concepto, descripcion, monto, fecha,
             comprobante, proveedor, registrado_por, observaciones)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            tipo, d.get("categoria_id"), concepto,
            d.get("descripcion", ""), monto,
            d.get("fecha", date.today().isoformat()),
            d.get("comprobante", ""), d.get("proveedor", ""),
            u["id"], d.get("observaciones", ""),
        ))
        mov_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()
        cache_delete("api_datos_all")

    _audit("movimiento_creado", f"{tipo}: {concepto} RD${monto:,.2f}",
           "movimientos_financieros", mov_id)
    return jsonify({"ok": True, "id": mov_id})


@finanzas_bp.route("/api/finanzas/movimientos/<int:mov_id>", methods=["DELETE"])
@login_required
@_finanzas_required
def anular_movimiento(mov_id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute(
            "UPDATE movimientos_financieros SET estado='anulado' WHERE id=?",
            (mov_id,))
        conn.commit()
        cache_delete("api_datos_all")
    _audit("movimiento_anulado", f"Movimiento #{mov_id} anulado",
           "movimientos_financieros", mov_id)
    return jsonify({"ok": True})


# ── PRESUPUESTOS ─────────────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/presupuestos", methods=["GET"])
@login_required
@_finanzas_required
def listar_presupuestos():
    anio = request.args.get("anio", _anio_escolar_actual())
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT p.*, c.nombre as categoria_nombre
            FROM presupuestos p
            LEFT JOIN categorias_gasto c ON c.id = p.categoria_id
            WHERE p.anio_escolar = ? AND p.estado = 'activo'
            ORDER BY c.nombre
        """, (anio,)).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/presupuestos", methods=["POST"])
@login_required
@_finanzas_required
def crear_presupuesto():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            INSERT INTO presupuestos
            (anio_escolar, categoria_id, concepto, monto_asignado, creado_por)
            VALUES (?,?,?,?,?)
        """, (
            d.get("anio_escolar", _anio_escolar_actual()),
            d.get("categoria_id"), d.get("concepto", ""),
            d.get("monto_asignado", 0), u["id"],
        ))
        conn.commit()
    return jsonify({"ok": True})


# ── ESTADO FINANCIERO / RESUMEN ──────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/resumen")
@login_required
@_finanzas_required
def resumen_financiero():
    anio = request.args.get("anio", str(datetime.now().year))
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row

        # Total ingresos
        ingresos = conn.execute("""
            SELECT COALESCE(SUM(monto), 0) FROM movimientos_financieros
            WHERE tipo='ingreso' AND estado='registrado' AND fecha LIKE ?
        """, (f"{anio}%",)).fetchone()[0]

        # Total gastos
        gastos = conn.execute("""
            SELECT COALESCE(SUM(monto), 0) FROM movimientos_financieros
            WHERE tipo='gasto' AND estado='registrado' AND fecha LIKE ?
        """, (f"{anio}%",)).fetchone()[0]

        # Gastos por categoría
        por_categoria = conn.execute("""
            SELECT c.nombre, m.tipo, SUM(m.monto) as total
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            WHERE m.estado='registrado' AND m.fecha LIKE ?
            GROUP BY c.nombre, m.tipo
            ORDER BY total DESC
        """, (f"{anio}%",)).fetchall()

        # Por mes
        por_mes = conn.execute("""
            SELECT substr(fecha,1,7) as mes, tipo, SUM(monto) as total
            FROM movimientos_financieros
            WHERE estado='registrado' AND fecha LIKE ?
            GROUP BY mes, tipo ORDER BY mes
        """, (f"{anio}%",)).fetchall()

        # Presupuesto vs ejecutado
        presupuesto = conn.execute("""
            SELECT p.concepto, c.nombre as categoria, p.monto_asignado,
                   COALESCE((SELECT SUM(m.monto) FROM movimientos_financieros m
                             WHERE m.categoria_id = p.categoria_id
                             AND m.tipo='gasto' AND m.estado='registrado'
                             AND m.fecha LIKE ?), 0) as ejecutado
            FROM presupuestos p
            LEFT JOIN categorias_gasto c ON c.id = p.categoria_id
            WHERE p.anio_escolar LIKE ? AND p.estado='activo'
        """, (f"{anio}%", f"{anio}%")).fetchall()

        # Últimos 10 movimientos
        recientes = conn.execute("""
            SELECT m.*, c.nombre as categoria_nombre
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            WHERE m.estado='registrado'
            ORDER BY m.fecha DESC, m.creado_en DESC LIMIT 10
        """).fetchall()

    return jsonify({
        "ingresos": ingresos,
        "gastos": gastos,
        "balance": ingresos - gastos,
        "por_categoria": [dict(r) for r in por_categoria],
        "por_mes": [dict(r) for r in por_mes],
        "presupuesto": [dict(r) for r in presupuesto],
        "recientes": [dict(r) for r in recientes],
    })


# ── RELACIÓN DE GASTOS (para imprimir) ───────────────────────────────────────

@finanzas_bp.route("/api/finanzas/relacion-gastos")
@login_required
@_finanzas_required
def relacion_gastos():
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    categoria = request.args.get("categoria", "")

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = """
            SELECT m.fecha, m.concepto, m.descripcion, m.monto, m.comprobante,
                   m.proveedor, c.nombre as categoria, m.tipo
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            WHERE m.estado='registrado'
        """
        params = []
        if desde:
            sql += " AND m.fecha >= ?"
            params.append(desde)
        if hasta:
            sql += " AND m.fecha <= ?"
            params.append(hasta)
        if categoria:
            sql += " AND m.categoria_id=?"
            params.append(int(categoria))
        sql += " ORDER BY m.fecha, m.creado_en"
        rows = conn.execute(sql, params).fetchall()

    items = [dict(r) for r in rows]
    total_ingresos = sum(r["monto"] for r in items if r["tipo"] == "ingreso")
    total_gastos = sum(r["monto"] for r in items if r["tipo"] == "gasto")

    return jsonify({
        "items": items,
        "total_ingresos": total_ingresos,
        "total_gastos": total_gastos,
        "balance": total_ingresos - total_gastos,
    })


# ── SEED CATEGORÍAS POR DEFECTO ─────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/seed-categorias", methods=["POST"])
@login_required
@_finanzas_required
def seed_categorias():
    """Crea las categorías de gasto por defecto si no existen."""
    categorias = [
        ("Inscripciones", "ingreso"),
        ("Cuotas mensuales", "ingreso"),
        ("Donaciones", "ingreso"),
        ("Actividades escolares", "ingreso"),
        ("Otros ingresos", "ingreso"),
        ("Nómina docente", "gasto"),
        ("Nómina administrativa", "gasto"),
        ("Servicios públicos", "gasto"),
        ("Materiales didácticos", "gasto"),
        ("Equipos y tecnología", "gasto"),
        ("Mantenimiento", "gasto"),
        ("Mobiliario", "gasto"),
        ("Alimentación escolar", "gasto"),
        ("Transporte", "gasto"),
        ("Actividades culturales", "gasto"),
        ("Seguros", "gasto"),
        ("Papelería e impresión", "gasto"),
        ("Otros gastos", "gasto"),
    ]
    creadas = 0
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        for nombre, tipo in categorias:
            try:
                conn.execute(
                    "INSERT INTO categorias_gasto (nombre, tipo) VALUES (?,?)",
                    (nombre, tipo))
                creadas += 1
            except sqlite3.IntegrityError:
                pass
        conn.commit()
    return jsonify({"ok": True, "creadas": creadas})


# ═══════════════════════════════════════════════════════════════════════════
#  NUEVAS FUNCIONALIDADES — Gráficos, Exportar, Recibos, Pagos inscripción
# ═══════════════════════════════════════════════════════════════════════════

# ── EXPORTAR RELACIÓN DE GASTOS A EXCEL ──────────────────────────────────────

@finanzas_bp.route("/api/finanzas/exportar-excel")
@login_required
@_finanzas_required
def exportar_excel_finanzas():
    """Exporta movimientos financieros a Excel."""
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = """
            SELECT m.fecha, m.tipo, c.nombre as categoria, m.concepto,
                   m.descripcion, m.monto, m.comprobante, m.proveedor
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            WHERE m.estado='registrado'
        """
        params = []
        if desde:
            sql += " AND m.fecha >= ?"
            params.append(desde)
        if hasta:
            sql += " AND m.fecha <= ?"
            params.append(hasta)
        sql += " ORDER BY m.fecha"
        rows = conn.execute(sql, params).fetchall()

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relación Financiera"

    # Header
    headers = ["Fecha", "Tipo", "Categoría", "Concepto", "Descripción", "Monto (RD$)", "Comprobante", "Proveedor"]
    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data
    total_ingresos = 0
    total_gastos = 0
    for i, r in enumerate(rows, 2):
        vals = [r["fecha"], r["tipo"], r["categoria"] or "", r["concepto"],
                r["descripcion"] or "", r["monto"], r["comprobante"] or "", r["proveedor"] or ""]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col == 6:  # monto
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True, color="107C10" if r["tipo"] == "ingreso" else "D13438")

        if r["tipo"] == "ingreso":
            total_ingresos += r["monto"]
        else:
            total_gastos += r["monto"]

    # Totals row
    last = len(rows) + 2
    ws.cell(row=last, column=5, value="TOTAL INGRESOS:").font = Font(bold=True)
    ws.cell(row=last, column=6, value=total_ingresos).font = Font(bold=True, color="107C10")
    ws.cell(row=last, column=6).number_format = '#,##0.00'
    ws.cell(row=last+1, column=5, value="TOTAL GASTOS:").font = Font(bold=True)
    ws.cell(row=last+1, column=6, value=total_gastos).font = Font(bold=True, color="D13438")
    ws.cell(row=last+1, column=6).number_format = '#,##0.00'
    ws.cell(row=last+2, column=5, value="BALANCE:").font = Font(bold=True, size=12)
    ws.cell(row=last+2, column=6, value=total_ingresos - total_gastos).font = Font(bold=True, size=12)
    ws.cell(row=last+2, column=6).number_format = '#,##0.00'

    # Column widths
    for col, width in enumerate([12, 10, 18, 30, 25, 14, 15, 20], 1):
        ws.column_dimensions[chr(64+col)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"relacion_financiera_{desde or 'todo'}_{hasta or 'hoy'}.xlsx",
        as_attachment=True)


# ── RECIBO DE PAGO PDF ───────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/recibo/<int:mov_id>")
@login_required
@_finanzas_required
def recibo_pdf(mov_id):
    """Genera recibo de pago imprimible en PDF."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from io import BytesIO

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        mov = conn.execute("""
            SELECT m.*, c.nombre as categoria_nombre, u.nombre as registrado_nombre
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            LEFT JOIN usuarios u ON u.id = m.registrado_por
            WHERE m.id=?
        """, (mov_id,)).fetchone()

    if not mov:
        return jsonify({"error": "Movimiento no encontrado"}), 404

    from core.helpers import _get_config_centro
    cfg = _get_config_centro()
    centro = cfg.get("nombre", "Centro Educativo")

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    w, h = letter

    # Header
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(w/2, h - 0.7*inch, centro)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(w/2, h - 1.1*inch,
        "RECIBO DE " + ("INGRESO" if mov["tipo"] == "ingreso" else "PAGO"))

    c.setStrokeColorRGB(0, 0.47, 0.83)
    c.setLineWidth(2)
    c.line(1*inch, h - 1.3*inch, w - 1*inch, h - 1.3*inch)

    # Recibo number and date
    c.setFont("Helvetica-Bold", 11)
    c.drawString(1*inch, h - 1.7*inch, f"Recibo No: {mov['comprobante'] or f'REC-{mov_id:06d}'}")
    c.drawRightString(w - 1*inch, h - 1.7*inch, f"Fecha: {mov['fecha']}")

    # Details box
    y = h - 2.2*inch
    c.setFont("Helvetica", 11)

    fields = [
        ("Concepto:", mov["concepto"]),
        ("Categoría:", mov["categoria_nombre"] or "Sin categoría"),
        ("Descripción:", mov["descripcion"] or "—"),
        ("Proveedor:", mov["proveedor"] or "—"),
    ]
    for label, value in fields:
        c.setFont("Helvetica-Bold", 10)
        c.drawString(1*inch, y, label)
        c.setFont("Helvetica", 10)
        c.drawString(2.5*inch, y, str(value))
        y -= 18

    # Monto grande
    y -= 20
    c.setFont("Helvetica-Bold", 14)
    c.drawString(1*inch, y, "MONTO:")
    c.setFont("Helvetica-Bold", 20)
    monto_fmt = f"RD$ {mov['monto']:,.2f}"
    c.drawString(2.5*inch, y, monto_fmt)

    # Firma
    y -= 80
    c.setFont("Helvetica", 10)
    c.setLineWidth(0.5)
    c.line(1.5*inch, y, 4*inch, y)
    c.drawString(1.5*inch, y - 14, "Recibido por")
    c.line(4.5*inch, y, 7*inch, y)
    c.drawString(4.5*inch, y - 14, "Autorizado por")

    # Footer
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawCentredString(w/2, 0.5*inch, f"Registrado por: {mov['registrado_nombre'] or 'Sistema'} — Axula")

    c.save()
    buf.seek(0)

    return send_file(buf, mimetype="application/pdf",
        download_name=f"recibo_{mov_id}.pdf", as_attachment=False)


# ── CONTROL DE PAGOS DE INSCRIPCIÓN ──────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/pagos-inscripcion", methods=["GET"])
@login_required
@_finanzas_required
def listar_pagos_inscripcion():
    """Lista estado de pago de inscripciones."""
    anio = request.args.get("anio", _anio_escolar_actual())
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT i.id, i.estudiante_id, i.grado, i.mencion, i.monto_inscripcion,
                   i.pagado, i.fecha_inscripcion, i.tipo,
                   e.nombre, e.apellido, e.cedula
            FROM inscripciones i
            LEFT JOIN estudiantes e ON e.id = i.estudiante_id
            WHERE i.anio_escolar=? AND i.estado='activa'
            ORDER BY i.pagado ASC, e.apellido, e.nombre
        """, (anio,)).fetchall()

    total = len(rows)
    pagados = sum(1 for r in rows if r["pagado"])
    pendientes = total - pagados
    monto_total = sum(r["monto_inscripcion"] or 0 for r in rows)
    monto_pagado = sum(r["monto_inscripcion"] or 0 for r in rows if r["pagado"])

    return jsonify({
        "inscripciones": [dict(r) for r in rows],
        "resumen": {
            "total": total,
            "pagados": pagados,
            "pendientes": pendientes,
            "monto_total": monto_total,
            "monto_pagado": monto_pagado,
            "monto_pendiente": monto_total - monto_pagado,
        }
    })


@finanzas_bp.route("/api/finanzas/pagos-inscripcion/<int:insc_id>/pagar", methods=["POST"])
@login_required
@_finanzas_required
@rate_limited(max_calls=20, window=60)
def registrar_pago_inscripcion(insc_id):
    """Marca una inscripción como pagada y registra el ingreso."""
    u = get_usuario()
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        insc = conn.execute("SELECT * FROM inscripciones WHERE id=?", (insc_id,)).fetchone()
        if not insc:
            return jsonify({"error": "Inscripción no encontrada"}), 404

        if insc["pagado"]:
            return jsonify({"error": "Ya está marcada como pagada"}), 400

        est = conn.execute(
            "SELECT nombre, apellido FROM estudiantes WHERE id=?",
            (insc["estudiante_id"],)
        ).fetchone()
        nombre_est = f"{est['nombre']} {est['apellido']}" if est else f"Est#{insc['estudiante_id']}"

        # Mark as paid
        conn.execute("UPDATE inscripciones SET pagado=1 WHERE id=?", (insc_id,))

        # Register income
        monto = insc["monto_inscripcion"] or 0
        if monto > 0:
            # Find or create Inscripciones category
            cat = conn.execute(
                "SELECT id FROM categorias_gasto WHERE nombre='Inscripciones'"
            ).fetchone()
            cat_id = cat[0] if cat else None

            conn.execute("""
                INSERT INTO movimientos_financieros
                (tipo, categoria_id, concepto, monto, registrado_por, comprobante)
                VALUES (?,?,?,?,?,?)
            """, ("ingreso", cat_id, f"Pago inscripción — {nombre_est}",
                  monto, u["id"], f"INSC-{insc_id}"))

        conn.commit()
        cache_delete("api_datos_all")

    _audit("pago_inscripcion", f"Pago inscripción {nombre_est} RD${monto:,.2f}",
           "inscripciones", insc_id)
    return jsonify({"ok": True})


# ── REPORTE FINANCIERO MENSUAL ───────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/reporte-mensual")
@login_required
@_finanzas_required
def reporte_mensual():
    """Estado financiero mes a mes del año."""
    anio = request.args.get("anio", str(datetime.now().year))

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row

        # Monthly breakdown
        meses = conn.execute("""
            SELECT substr(fecha,1,7) as mes,
                   SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END) as ingresos,
                   SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END) as gastos,
                   COUNT(*) as transacciones
            FROM movimientos_financieros
            WHERE estado='registrado' AND fecha LIKE ?
            GROUP BY mes ORDER BY mes
        """, (f"{anio}%",)).fetchall()

        # Category breakdown
        por_categoria = conn.execute("""
            SELECT c.nombre as categoria, m.tipo,
                   SUM(m.monto) as total, COUNT(*) as count
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id = m.categoria_id
            WHERE m.estado='registrado' AND m.fecha LIKE ?
            GROUP BY c.nombre, m.tipo
            ORDER BY total DESC
        """, (f"{anio}%",)).fetchall()

        # Year totals
        totales = conn.execute("""
            SELECT
                SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END) as total_ingresos,
                SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END) as total_gastos,
                COUNT(*) as total_transacciones
            FROM movimientos_financieros
            WHERE estado='registrado' AND fecha LIKE ?
        """, (f"{anio}%",)).fetchone()

    meses_nombres = {
        "01":"Enero","02":"Febrero","03":"Marzo","04":"Abril",
        "05":"Mayo","06":"Junio","07":"Julio","08":"Agosto",
        "09":"Septiembre","10":"Octubre","11":"Noviembre","12":"Diciembre"
    }

    meses_data = []
    for m in meses:
        mes_num = m["mes"].split("-")[1]
        meses_data.append({
            "mes": m["mes"],
            "nombre": meses_nombres.get(mes_num, mes_num),
            "ingresos": m["ingresos"],
            "gastos": m["gastos"],
            "balance": m["ingresos"] - m["gastos"],
            "transacciones": m["transacciones"],
        })

    return jsonify({
        "anio": anio,
        "meses": meses_data,
        "por_categoria": [dict(r) for r in por_categoria],
        "totales": {
            "ingresos": totales["total_ingresos"] or 0,
            "gastos": totales["total_gastos"] or 0,
            "balance": (totales["total_ingresos"] or 0) - (totales["total_gastos"] or 0),
            "transacciones": totales["total_transacciones"] or 0,
        }
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  MÓDULOS NUEVOS — Importar, Presupuesto Export/Import, Caja Chica,
#  Proveedores, Cuentas por Cobrar, Nómina, Cierre de Período
# ═══════════════════════════════════════════════════════════════════════════════

from flask import send_file

# ── IMPORTAR MOVIMIENTOS DESDE EXCEL ─────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/importar-excel", methods=["POST"])
@login_required
@_finanzas_required
def importar_excel():
    """Importa movimientos desde un archivo Excel (.xlsx/.xls/.csv)."""
    u = get_usuario()
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"error": "No se recibió archivo"}), 400

    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls", "csv"}:
        return jsonify({"error": "Solo se aceptan archivos .xlsx, .xls o .csv"}), 400

    try:
        from io import BytesIO
        import openpyxl

        data = archivo.read()
        if not _validar_magic_excel(data, ext):
            return jsonify({"error": "El archivo no es un Excel o CSV válido"}), 400
        buf = BytesIO(data)

        if ext == "csv":
            import csv, io
            texto = data.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(texto))
            filas = [row for row in reader]
        else:
            wb = openpyxl.load_workbook(buf, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            filas = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    filas.append(dict(zip(headers, row)))

    except Exception as e:
        logger.warning(f"[FINANZAS] Error leyendo Excel para importación: {e}")
        return jsonify({"error": "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)."}), 400

    # Mapeo flexible de columnas
    def _get(row, *keys):
        for k in keys:
            for rk in row:
                if rk and k in str(rk).lower():
                    v = row[rk]
                    return str(v).strip() if v is not None else ""
        return ""

    importados = 0
    errores = []
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        for i, row in enumerate(filas, 2):
            try:
                concepto = _get(row, "concepto", "descripcion", "detalle", "concept")
                monto_str = _get(row, "monto", "importe", "valor", "amount")
                tipo = _get(row, "tipo", "type")
                fecha = _get(row, "fecha", "date", "fec")
                comprobante = _get(row, "comprobante", "ncf", "factura", "receipt")
                proveedor = _get(row, "proveedor", "supplier", "suplidor")
                categoria_nom = _get(row, "categoria", "category", "cat")

                if not concepto or not monto_str:
                    continue

                monto = float(str(monto_str).replace(",", "").replace("RD$", "").replace("$", "").strip())
                if monto <= 0:
                    continue

                tipo = "ingreso" if "ing" in tipo.lower() else "gasto"
                if not fecha or fecha == "None":
                    fecha = date.today().isoformat()
                else:
                    # Intentar parsear fecha
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                        try:
                            fecha = datetime.strptime(str(fecha)[:10], fmt).strftime("%Y-%m-%d")
                            break
                        except:
                            pass

                cat_id = None
                if categoria_nom:
                    cat = conn.execute(
                        "SELECT id FROM categorias_gasto WHERE LOWER(nombre) LIKE LOWER(?)",
                        (f"%{categoria_nom}%",)
                    ).fetchone()
                    if cat:
                        cat_id = cat["id"]

                conn.execute("""
                    INSERT INTO movimientos_financieros
                    (tipo, categoria_id, concepto, monto, fecha, comprobante,
                     proveedor, registrado_por, observaciones)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (tipo, cat_id, concepto, monto, fecha, comprobante,
                      proveedor, u["id"], "Importado desde Excel"))
                importados += 1
            except Exception as e:
                logger.warning(f"[FINANZAS] Error importando fila {i}: {e}")
                errores.append(f"Fila {i}: dato inválido o formato incorrecto")

        conn.commit()
        cache_delete("api_datos_all")

    return jsonify({"ok": True, "importados": importados, "errores": errores[:10]})


# ── EXPORTAR PRESUPUESTO A EXCEL ──────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/presupuestos/exportar")
@login_required
@_finanzas_required
def exportar_presupuesto():
    anio = request.args.get("anio", _anio_escolar_actual())
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT p.concepto, c.nombre as categoria, p.monto_asignado, p.anio_escolar,
                   COALESCE((SELECT SUM(m.monto) FROM movimientos_financieros m
                    WHERE m.categoria_id=p.categoria_id AND m.tipo='gasto'
                    AND m.estado='registrado' AND m.fecha LIKE ?), 0) as ejecutado
            FROM presupuestos p
            LEFT JOIN categorias_gasto c ON c.id=p.categoria_id
            WHERE p.anio_escolar=? AND p.estado='activo'
            ORDER BY c.nombre
        """, (f"{anio}%", anio)).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    hf = PatternFill("solid", fgColor="0078D4")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    borde = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    hdrs = ["Concepto", "Categoría", "Monto Asignado", "Ejecutado", "Disponible", "% Ejecución"]
    anchos = [30, 22, 18, 18, 18, 14]
    for i, (h, w) in enumerate(zip(hdrs, anchos), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont; c.fill = hf
        c.alignment = Alignment(horizontal="center")
        c.border = borde
        ws.column_dimensions[chr(64+i)].width = w

    total_asig = total_ejec = 0
    for ri, r in enumerate(rows, 2):
        ejec = r["ejecutado"] or 0
        asig = r["monto_asignado"] or 0
        disp = asig - ejec
        pct = round(ejec / asig * 100, 1) if asig else 0
        total_asig += asig; total_ejec += ejec
        vals = [r["concepto"], r["categoria"] or "Sin categoría",
                asig, ejec, disp, f"{pct}%"]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = borde
            if ci in (3, 4, 5):
                cell.number_format = "#,##0.00"
            if ci == 5:
                cell.font = Font(bold=True,
                    color="107C10" if disp >= 0 else "D13438")

    # Totals
    tr = len(rows) + 2
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(bold=True, size=11)
    ws.cell(row=tr, column=3, value=total_asig).font = Font(bold=True)
    ws.cell(row=tr, column=3).number_format = "#,##0.00"
    ws.cell(row=tr, column=4, value=total_ejec).font = Font(bold=True)
    ws.cell(row=tr, column=4).number_format = "#,##0.00"
    ws.cell(row=tr, column=5, value=total_asig - total_ejec).font = Font(bold=True)
    ws.cell(row=tr, column=5).number_format = "#,##0.00"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"presupuesto_{anio}.xlsx", as_attachment=True)


# ── IMPORTAR PRESUPUESTO DESDE EXCEL ─────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/presupuestos/importar", methods=["POST"])
@login_required
@_finanzas_required
def importar_presupuesto():
    u = get_usuario()
    archivo = request.files.get("archivo")
    anio = request.form.get("anio", _anio_escolar_actual())
    if not archivo:
        return jsonify({"error": "No se recibió archivo"}), 400
    try:
        import openpyxl
        from io import BytesIO
        data_pre = archivo.read()
        ext_pre = archivo.filename.rsplit(".", 1)[-1].lower()
        if not _validar_magic_excel(data_pre, ext_pre):
            return jsonify({"error": "El archivo no es un Excel válido"}), 400
        wb = openpyxl.load_workbook(BytesIO(data_pre), data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        filas = [dict(zip(headers, [c for c in row])) for row in ws.iter_rows(min_row=2, values_only=True)]
    except Exception as e:
        return jsonify({"error": f"Error leyendo archivo: {e}"}), 400

    importados = 0
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        for row in filas:
            concepto = str(row.get("concepto") or row.get("concept") or "").strip()
            monto = row.get("monto asignado") or row.get("monto") or row.get("asignado") or 0
            cat_nom = str(row.get("categoría") or row.get("categoria") or row.get("category") or "").strip()
            if not concepto:
                continue
            try:
                monto = float(str(monto).replace(",","").replace("RD$","").strip())
            except:
                monto = 0
            cat_id = None
            if cat_nom:
                cat = conn.execute(
                    "SELECT id FROM categorias_gasto WHERE LOWER(nombre) LIKE LOWER(?)",
                    (f"%{cat_nom}%",)
                ).fetchone()
                if cat:
                    cat_id = cat["id"]
            conn.execute("""
                INSERT INTO presupuestos (anio_escolar, categoria_id, concepto, monto_asignado, creado_por)
                VALUES (?,?,?,?,?)
            """, (anio, cat_id, concepto, monto, u["id"]))
            importados += 1
        conn.commit()
    return jsonify({"ok": True, "importados": importados})


# ── PROVEEDORES ───────────────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/proveedores", methods=["GET"])
@login_required
@_finanzas_required
def listar_proveedores():
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM proveedores WHERE activo=1 ORDER BY nombre"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/proveedores", methods=["POST"])
@login_required
@_finanzas_required
def crear_proveedor():
    d = request.get_json(silent=True) or {}
    nombre = d.get("nombre", "").strip()
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            INSERT INTO proveedores (nombre, rnc, contacto, telefono, email, direccion, categoria, notas)
            VALUES (?,?,?,?,?,?,?,?)
        """, (nombre, d.get("rnc",""), d.get("contacto",""), d.get("telefono",""),
              d.get("email",""), d.get("direccion",""), d.get("categoria","general"), d.get("notas","")))
        conn.commit()
    return jsonify({"ok": True})


@finanzas_bp.route("/api/finanzas/proveedores/<int:prov_id>", methods=["DELETE"])
@login_required
@_finanzas_required
def eliminar_proveedor(prov_id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("UPDATE proveedores SET activo=0 WHERE id=?", (prov_id,))
        conn.commit()
    return jsonify({"ok": True})


@finanzas_bp.route("/api/finanzas/proveedores/<int:prov_id>/historial")
@login_required
@_finanzas_required
def historial_proveedor(prov_id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        prov = conn.execute("SELECT nombre FROM proveedores WHERE id=?", (prov_id,)).fetchone()
        if not prov:
            return jsonify({"error": "No encontrado"}), 404
        movs = conn.execute("""
            SELECT fecha, concepto, monto, comprobante
            FROM movimientos_financieros
            WHERE proveedor_id=? AND estado='registrado'
            ORDER BY fecha DESC LIMIT 50
        """, (prov_id,)).fetchall()
        total = conn.execute(
            "SELECT COALESCE(SUM(monto),0) FROM movimientos_financieros WHERE proveedor_id=? AND estado='registrado'",
            (prov_id,)
        ).fetchone()[0]
    return jsonify({"proveedor": prov["nombre"], "movimientos": [dict(r) for r in movs], "total": total})


# ── CAJA CHICA ────────────────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/caja-chica", methods=["GET"])
@login_required
@_finanzas_required
def listar_cajas():
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        cajas = conn.execute("""
            SELECT cc.*, u.nombre as responsable_nombre
            FROM caja_chica cc
            LEFT JOIN usuarios u ON u.id=cc.responsable_id
            ORDER BY cc.creado_en DESC
        """).fetchall()
    return jsonify([dict(c) for c in cajas])


@finanzas_bp.route("/api/finanzas/caja-chica", methods=["POST"])
@login_required
@_finanzas_required
def crear_caja():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    monto = float(d.get("monto_inicial", 0))
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            INSERT INTO caja_chica (nombre, monto_inicial, monto_actual, responsable_id, creado_por, observaciones)
            VALUES (?,?,?,?,?,?)
        """, (d.get("nombre","Caja Chica"), monto, monto,
              d.get("responsable_id") or u["id"], u["id"], d.get("observaciones","")))
        cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Registrar como movimiento de ingreso
        conn.execute("""
            INSERT INTO movimientos_financieros
            (tipo, concepto, monto, fecha, registrado_por, caja_chica_id, observaciones)
            VALUES ('ingreso', ?, ?, date('now'), ?, ?, 'Apertura de caja chica')
        """, (f"Apertura Caja Chica: {d.get('nombre','')}", monto, u["id"], cid))
        conn.commit()
    return jsonify({"ok": True, "id": cid})


@finanzas_bp.route("/api/finanzas/caja-chica/<int:caja_id>/movimientos", methods=["GET"])
@login_required
@_finanzas_required
def movimientos_caja(caja_id):
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        caja = conn.execute("SELECT * FROM caja_chica WHERE id=?", (caja_id,)).fetchone()
        if not caja:
            return jsonify({"error": "Caja no encontrada"}), 404
        movs = conn.execute("""
            SELECT ccm.*, u.nombre as registrado_nombre
            FROM caja_chica_movimientos ccm
            LEFT JOIN usuarios u ON u.id=ccm.registrado_por
            WHERE ccm.caja_id=?
            ORDER BY ccm.fecha DESC, ccm.creado_en DESC
        """, (caja_id,)).fetchall()
    return jsonify({"caja": dict(caja), "movimientos": [dict(m) for m in movs]})


@finanzas_bp.route("/api/finanzas/caja-chica/<int:caja_id>/movimientos", methods=["POST"])
@login_required
@_finanzas_required
def agregar_movimiento_caja(caja_id):
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    tipo = d.get("tipo", "gasto")
    monto = float(d.get("monto", 0))
    concepto = d.get("concepto", "").strip()
    if not concepto or monto <= 0:
        return jsonify({"error": "Concepto y monto requeridos"}), 400
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        caja = conn.execute("SELECT * FROM caja_chica WHERE id=? AND estado='abierta'", (caja_id,)).fetchone()
        if not caja:
            return jsonify({"error": "Caja no encontrada o cerrada"}), 404
        nuevo_monto = caja["monto_actual"] + (monto if tipo == "reintegro" else -monto)
        if nuevo_monto < 0:
            return jsonify({"error": f"Fondos insuficientes. Disponible: RD${caja['monto_actual']:,.2f}"}), 400
        conn.execute("""
            INSERT INTO caja_chica_movimientos
            (caja_id, tipo, concepto, monto, fecha, comprobante, registrado_por)
            VALUES (?,?,?,?,?,?,?)
        """, (caja_id, tipo, concepto, monto,
              d.get("fecha", date.today().isoformat()),
              d.get("comprobante",""), u["id"]))
        conn.execute("UPDATE caja_chica SET monto_actual=? WHERE id=?", (nuevo_monto, caja_id))
        conn.commit()
    return jsonify({"ok": True, "saldo_actual": nuevo_monto})


@finanzas_bp.route("/api/finanzas/caja-chica/<int:caja_id>/cerrar", methods=["POST"])
@login_required
@_finanzas_required
def cerrar_caja(caja_id):
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            UPDATE caja_chica SET estado='cerrada', fecha_cierre=date('now'), observaciones=?
            WHERE id=? AND estado='abierta'
        """, (d.get("observaciones",""), caja_id))
        conn.commit()
    return jsonify({"ok": True})


# ── CUENTAS POR COBRAR ────────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/cuentas-cobrar", methods=["GET"])
@login_required
@_finanzas_required
def listar_cuentas_cobrar():
    estado = request.args.get("estado", "")
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = """
            SELECT cc.*, e.nombre, e.apellido, e.grado, e.curso
            FROM cuentas_cobrar cc
            LEFT JOIN estudiantes e ON e.id=cc.estudiante_id
            WHERE 1=1
        """
        params = []
        if estado:
            sql += " AND cc.estado=?"
            params.append(estado)
        sql += " ORDER BY cc.fecha_vencimiento ASC, cc.creado_en DESC"
        rows = conn.execute(sql, params).fetchall()

    hoy = date.today().isoformat()
    result = []
    for r in rows:
        d = dict(r)
        d["vencida"] = (d.get("fecha_vencimiento") or "") < hoy and d["estado"] == "pendiente"
        result.append(d)

    total_pendiente = sum(r["monto"] for r in result if r["estado"] == "pendiente")
    total_vencido = sum(r["monto"] for r in result if r.get("vencida"))
    return jsonify({"items": result, "total_pendiente": total_pendiente, "total_vencido": total_vencido})


@finanzas_bp.route("/api/finanzas/cuentas-cobrar", methods=["POST"])
@login_required
@_finanzas_required
def crear_cuenta_cobrar():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    concepto = d.get("concepto","").strip()
    monto = d.get("monto", 0)
    if not concepto or not monto:
        return jsonify({"error": "Concepto y monto requeridos"}), 400
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.execute("""
            INSERT INTO cuentas_cobrar
            (estudiante_id, concepto, monto, fecha_emision, fecha_vencimiento, registrado_por, observaciones)
            VALUES (?,?,?,?,?,?,?)
        """, (d.get("estudiante_id"), concepto, float(monto),
              d.get("fecha_emision", date.today().isoformat()),
              d.get("fecha_vencimiento",""), u["id"], d.get("observaciones","")))
        conn.commit()
    return jsonify({"ok": True})


@finanzas_bp.route("/api/finanzas/cuentas-cobrar/<int:cc_id>/cobrar", methods=["POST"])
@login_required
@_finanzas_required
def cobrar_cuenta(cc_id):
    u = get_usuario()
    d = request.get_json(silent=True) or {}
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        cc = conn.execute("SELECT * FROM cuentas_cobrar WHERE id=?", (cc_id,)).fetchone()
        if not cc:
            return jsonify({"error": "No encontrada"}), 404
        if cc["estado"] != "pendiente":
            return jsonify({"error": "Ya está cobrada"}), 400
        conn.execute("""
            UPDATE cuentas_cobrar SET estado='cobrado', pagado_en=date('now') WHERE id=?
        """, (cc_id,))
        # Registrar ingreso
        cat = conn.execute("SELECT id FROM categorias_gasto WHERE nombre='Cuotas mensuales'").fetchone()
        conn.execute("""
            INSERT INTO movimientos_financieros
            (tipo, categoria_id, concepto, monto, fecha, registrado_por, comprobante)
            VALUES ('ingreso',?,?,?,date('now'),?,?)
        """, (cat["id"] if cat else None, cc["concepto"], cc["monto"],
              u["id"], d.get("comprobante","")))
        conn.commit()
        cache_delete("api_datos_all")
    return jsonify({"ok": True})


# ── NÓMINA ────────────────────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/nomina", methods=["GET"])
@login_required
@_finanzas_required
def listar_nomina():
    periodo = request.args.get("periodo", "")
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = """
            SELECT n.*, u.nombre as empleado_nombre, u.rol as empleado_rol
            FROM nomina_pagos n
            LEFT JOIN usuarios u ON u.id=n.empleado_id
            WHERE 1=1
        """
        params = []
        if periodo:
            sql += " AND n.periodo=?"
            params.append(periodo)
        sql += " ORDER BY n.periodo DESC, u.nombre"
        rows = conn.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/nomina", methods=["POST"])
@login_required
@_finanzas_required
def crear_nomina():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    empleado_id = d.get("empleado_id")
    periodo = d.get("periodo","").strip()
    bruto = float(d.get("monto_bruto", 0))
    descuentos = float(d.get("descuentos", 0))
    neto = bruto - descuentos
    if not empleado_id or not periodo or bruto <= 0:
        return jsonify({"error": "Empleado, período y monto son requeridos"}), 400
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        emp = conn.execute("SELECT nombre FROM usuarios WHERE id=?", (empleado_id,)).fetchone()
        conn.execute("""
            INSERT INTO nomina_pagos
            (empleado_id, periodo, monto_bruto, descuentos, monto_neto, metodo_pago,
             fecha_pago, estado, registrado_por, observaciones)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (empleado_id, periodo, bruto, descuentos, neto,
              d.get("metodo_pago","efectivo"), d.get("fecha_pago", date.today().isoformat()),
              d.get("estado","pagado"), u["id"], d.get("observaciones","")))
        nom_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Registrar como gasto de nómina
        if d.get("registrar_gasto", True) and d.get("estado","pagado") == "pagado":
            cat = conn.execute("SELECT id FROM categorias_gasto WHERE nombre LIKE '%Nómina%' LIMIT 1").fetchone()
            conn.execute("""
                INSERT INTO movimientos_financieros
                (tipo, categoria_id, concepto, monto, fecha, registrado_por)
                VALUES ('gasto',?,?,?,?,?)
            """, (cat["id"] if cat else None,
                  f"Salario {emp['nombre'] if emp else ''} — {periodo}",
                  neto, d.get("fecha_pago", date.today().isoformat()), u["id"]))
        conn.commit()
        cache_delete("api_datos_all")
    return jsonify({"ok": True, "id": nom_id})


@finanzas_bp.route("/api/finanzas/nomina/exportar")
@login_required
@_finanzas_required
def exportar_nomina():
    periodo = request.args.get("periodo", "")
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT n.*, u.nombre as empleado_nombre, u.rol
            FROM nomina_pagos n
            LEFT JOIN usuarios u ON u.id=n.empleado_id
            WHERE (? = '' OR n.periodo = ?)
            ORDER BY u.nombre
        """, (periodo, periodo)).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina"
    hf = PatternFill("solid", fgColor="1E3A5F")
    hfont = Font(bold=True, color="FFFFFF")
    borde = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    hdrs = ["Empleado", "Rol", "Período", "Salario Bruto", "Descuentos", "Salario Neto",
            "Método Pago", "Fecha Pago", "Estado"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont; c.fill = hf
        c.alignment = Alignment(horizontal="center")
        c.border = borde
        ws.column_dimensions[chr(64+i)].width = [22,18,12,15,12,15,12,12,10][i-1]

    total_neto = 0
    for ri, r in enumerate(rows, 2):
        vals = [r["empleado_nombre"], r["rol"], r["periodo"],
                r["monto_bruto"], r["descuentos"], r["monto_neto"],
                r["metodo_pago"], r["fecha_pago"], r["estado"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = borde
            if ci in (4, 5, 6):
                cell.number_format = "#,##0.00"
        total_neto += r["monto_neto"] or 0

    tr = len(rows) + 2
    ws.cell(row=tr, column=5, value="TOTAL NETO:").font = Font(bold=True)
    ws.cell(row=tr, column=6, value=total_neto).font = Font(bold=True, color="107C10")
    ws.cell(row=tr, column=6).number_format = "#,##0.00"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"nomina_{periodo or 'completa'}.xlsx", as_attachment=True)


@finanzas_bp.route("/api/finanzas/personal-lista")
@login_required
@_finanzas_required
def personal_para_nomina():
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, nombre, rol FROM usuarios WHERE activo=1 ORDER BY nombre"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/nomina/importar", methods=["POST"])
@login_required
@_finanzas_required
def importar_nomina():
    """Importa pagos de nómina desde Excel.
    Columnas: empleado, periodo (YYYY-MM), monto_bruto, descuentos, metodo_pago, fecha_pago, estado, observaciones
    """
    u = get_usuario()
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"error": "No se recibió archivo"}), 400
    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls", "csv"}:
        return jsonify({"error": "Solo se aceptan .xlsx, .xls o .csv"}), 400

    try:
        from io import BytesIO
        import openpyxl
        data = archivo.read()
        if not _validar_magic_excel(data, ext):
            return jsonify({"error": "El archivo no es un Excel o CSV válido"}), 400
        buf = BytesIO(data)
        if ext == "csv":
            import csv, io
            texto = data.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(texto))
            filas = [row for row in reader]
        else:
            wb = openpyxl.load_workbook(buf, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            filas = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)
                     if any(v is not None for v in row)]
    except Exception as e:
        return jsonify({"error": f"Error leyendo archivo: {e}"}), 400

    def _get(row, *keys):
        for k in keys:
            for rk in row:
                if rk and k in str(rk).lower():
                    v = row[rk]
                    return str(v).strip() if v is not None else ""
        return ""

    importados = 0
    errores = []
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        empleados = {e["nombre"].lower(): e["id"]
                     for e in conn.execute("SELECT id, nombre FROM usuarios WHERE activo=1").fetchall()}
        for i, row in enumerate(filas, 2):
            try:
                emp_nom = _get(row, "empleado", "nombre", "employee", "personal")
                periodo = _get(row, "periodo", "period", "mes", "month")
                bruto_str = _get(row, "bruto", "monto_bruto", "salario", "gross")
                desc_str = _get(row, "descuento", "deduccion", "deduction") or "0"
                metodo = _get(row, "metodo", "method", "pago") or "efectivo"
                fecha_pago = _get(row, "fecha", "date") or date.today().isoformat()
                estado = _get(row, "estado", "status") or "pagado"
                obs = _get(row, "obs", "nota", "observacion")

                if not emp_nom or not periodo:
                    continue

                emp_id = empleados.get(emp_nom.lower())
                if not emp_id:
                    # Búsqueda parcial
                    for k, v in empleados.items():
                        if emp_nom.lower() in k or k in emp_nom.lower():
                            emp_id = v
                            break
                if not emp_id:
                    errores.append(f"Fila {i}: empleado '{emp_nom}' no encontrado")
                    continue

                bruto = float(str(bruto_str).replace(",", "").replace("RD$", "").strip())
                desc = float(str(desc_str).replace(",", "").replace("RD$", "").strip())
                neto = bruto - desc

                if bruto <= 0:
                    continue

                # Normalizar fecha
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        fecha_pago = datetime.strptime(str(fecha_pago)[:10], fmt).strftime("%Y-%m-%d")
                        break
                    except:
                        pass

                conn.execute("""
                    INSERT INTO nomina_pagos
                    (empleado_id, periodo, monto_bruto, descuentos, monto_neto,
                     metodo_pago, fecha_pago, estado, registrado_por, observaciones)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (emp_id, periodo, bruto, desc, neto,
                      metodo, fecha_pago, estado, u["id"],
                      obs or "Importado desde Excel"))
                # Registrar gasto si está pagado
                if estado == "pagado":
                    emp_row = conn.execute("SELECT nombre FROM usuarios WHERE id=?", (emp_id,)).fetchone()
                    cat = conn.execute(
                        "SELECT id FROM categorias_gasto WHERE nombre LIKE '%Nómina%' LIMIT 1"
                    ).fetchone()
                    conn.execute("""
                        INSERT INTO movimientos_financieros
                        (tipo, categoria_id, concepto, monto, fecha, registrado_por)
                        VALUES ('gasto',?,?,?,?,?)
                    """, (cat["id"] if cat else None,
                          f"Salario {emp_row['nombre'] if emp_row else emp_nom} — {periodo}",
                          neto, fecha_pago, u["id"]))
                importados += 1
            except Exception as e:
                errores.append(f"Fila {i}: {str(e)[:80]}")
        conn.commit()
        cache_delete("api_datos_all")

    return jsonify({"ok": True, "importados": importados, "errores": errores[:10]})


# ── CIERRE DE PERÍODO ─────────────────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/cierres", methods=["GET"])
@login_required
@_finanzas_required
def listar_cierres():
    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT cp.*, u.nombre as cerrado_nombre
            FROM cierres_periodo cp
            LEFT JOIN usuarios u ON u.id=cp.cerrado_por
            ORDER BY cp.fecha_cierre DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@finanzas_bp.route("/api/finanzas/cierres", methods=["POST"])
@login_required
@_finanzas_required
def crear_cierre():
    d = request.get_json(silent=True) or {}
    u = get_usuario()
    periodo = d.get("periodo","").strip()
    tipo = d.get("tipo","mensual")
    if not periodo:
        return jsonify({"error": "Período requerido"}), 400

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        # Calcular totales del período
        filtro = f"{periodo}%"
        totales = conn.execute("""
            SELECT
                SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END) as ingresos,
                SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END) as gastos,
                COUNT(*) as trans
            FROM movimientos_financieros
            WHERE estado='registrado' AND fecha LIKE ?
        """, (filtro,)).fetchone()

        ing = totales["ingresos"] or 0
        gas = totales["gastos"] or 0
        conn.execute("""
            INSERT INTO cierres_periodo
            (periodo, tipo, total_ingresos, total_gastos, balance,
             total_transacciones, cerrado_por, observaciones)
            VALUES (?,?,?,?,?,?,?,?)
        """, (periodo, tipo, ing, gas, ing-gas,
              totales["trans"] or 0, u["id"], d.get("observaciones","")))
        conn.commit()
    return jsonify({"ok": True, "balance": ing - gas})


@finanzas_bp.route("/api/finanzas/cierres/<int:cierre_id>/pdf")
@login_required
@_finanzas_required
def pdf_cierre(cierre_id):
    """Genera PDF del cierre de período para imprimir."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import inch
    from io import BytesIO

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        cierre = conn.execute("""
            SELECT cp.*, u.nombre as cerrado_nombre
            FROM cierres_periodo cp
            LEFT JOIN usuarios u ON u.id=cp.cerrado_por
            WHERE cp.id=?
        """, (cierre_id,)).fetchone()
        if not cierre:
            return jsonify({"error": "No encontrado"}), 404
        movs = conn.execute("""
            SELECT m.fecha, m.tipo, c.nombre as categoria, m.concepto, m.monto
            FROM movimientos_financieros m
            LEFT JOIN categorias_gasto c ON c.id=m.categoria_id
            WHERE m.estado='registrado' AND m.fecha LIKE ?
            ORDER BY m.fecha
        """, (f"{cierre['periodo']}%",)).fetchall()

    buf = BytesIO()
    cv = rl_canvas.Canvas(buf, pagesize=letter)
    w, h = letter

    from core.helpers import _get_config_centro as _gcc_fin
    _cfg_fin = _gcc_fin()
    _nombre_centro_fin = _cfg_fin.get("nombre", "Centro Educativo")
    cv.setFont("Helvetica-Bold", 16)
    cv.drawCentredString(w/2, h-0.8*inch, _nombre_centro_fin)
    cv.setFont("Helvetica-Bold", 13)
    cv.drawCentredString(w/2, h-1.15*inch, f"CIERRE DE PERÍODO — {cierre['periodo']}")
    cv.setFont("Helvetica", 9)
    cv.drawCentredString(w/2, h-1.4*inch, f"Tipo: {cierre['tipo']} | Cerrado por: {cierre['cerrado_nombre'] or '—'} | {cierre['fecha_cierre']}")

    cv.setStrokeColorRGB(0,0.47,0.83)
    cv.setLineWidth(1.5)
    cv.line(0.75*inch, h-1.6*inch, w-0.75*inch, h-1.6*inch)

    y = h - 1.95*inch
    cv.setFont("Helvetica-Bold", 11)
    cv.setFillColorRGB(0.06,0.49,0.08)
    cv.drawString(0.75*inch, y, f"Total Ingresos: RD$ {cierre['total_ingresos']:,.2f}")
    y -= 0.3*inch
    cv.setFillColorRGB(0.82,0.08,0.08)
    cv.drawString(0.75*inch, y, f"Total Gastos:   RD$ {cierre['total_gastos']:,.2f}")
    y -= 0.3*inch
    cv.setFillColorRGB(0,0.3,0.7)
    cv.drawString(0.75*inch, y, f"Balance:        RD$ {cierre['balance']:,.2f}")
    y -= 0.4*inch

    cv.setFillColorRGB(0,0,0)
    cv.setFont("Helvetica-Bold", 9)
    cv.drawString(0.75*inch, y, "Fecha")
    cv.drawString(1.5*inch, y, "Tipo")
    cv.drawString(2.3*inch, y, "Categoría")
    cv.drawString(4.2*inch, y, "Concepto")
    cv.drawRightString(w-0.75*inch, y, "Monto")
    y -= 0.18*inch
    cv.line(0.75*inch, y, w-0.75*inch, y)
    y -= 0.15*inch

    cv.setFont("Helvetica", 8)
    for mov in movs:
        if y < 1*inch:
            cv.showPage()
            y = h - 0.75*inch
            cv.setFont("Helvetica", 8)
        color = (0.06,0.49,0.08) if mov["tipo"]=="ingreso" else (0.82,0.08,0.08)
        cv.setFillColorRGB(*color)
        cv.drawString(0.75*inch, y, str(mov["fecha"] or "")[:10])
        cv.setFillColorRGB(0,0,0)
        cv.drawString(1.5*inch, y, mov["tipo"])
        cv.drawString(2.3*inch, y, (mov["categoria"] or "Sin cat.")[:22])
        cv.drawString(4.2*inch, y, (mov["concepto"] or "")[:28])
        cv.drawRightString(w-0.75*inch, y, f"RD$ {mov['monto']:,.2f}")
        y -= 0.17*inch

    # Firmas
    y = max(y - 0.5*inch, 0.8*inch)
    cv.setLineWidth(0.5)
    cv.line(0.75*inch, y, 3*inch, y)
    cv.drawString(0.75*inch, y-0.15*inch, "Preparado por")
    cv.line(4*inch, y, w-0.75*inch, y)
    cv.drawString(4*inch, y-0.15*inch, "Aprobado por Directora")

    cv.save()
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
        download_name=f"cierre_{cierre['periodo']}.pdf", as_attachment=False)


# ── CONCILIACIÓN BANCARIA SIMPLE ──────────────────────────────────────────────

@finanzas_bp.route("/api/finanzas/conciliacion", methods=["POST"])
@login_required
@_finanzas_required
def conciliacion_bancaria():
    """
    Recibe movimientos del banco (del OCR) y los cruza con los registrados.
    Devuelve: conciliados, solo_banco, solo_sistema.
    """
    d = request.get_json(silent=True) or {}
    movimientos_banco = d.get("movimientos", [])   # [{fecha, descripcion, monto}]
    desde = d.get("desde", "")
    hasta = d.get("hasta", "")

    with sqlite3.connect(DATABASE, timeout=10) as conn:
        conn.row_factory = sqlite3.Row
        sql = "SELECT id, fecha, concepto, monto, tipo FROM movimientos_financieros WHERE estado='registrado'"
        params = []
        if desde:
            sql += " AND fecha >= ?"; params.append(desde)
        if hasta:
            sql += " AND fecha <= ?"; params.append(hasta)
        sistema = [dict(r) for r in conn.execute(sql, params).fetchall()]

    def _normalizar(s):
        return str(s).lower().strip()

    conciliados = []
    solo_banco = []
    usados_sistema = set()

    for mb in movimientos_banco:
        mejor = None
        mejor_score = 0
        mb_monto = float(str(mb.get("debito") or mb.get("credito") or mb.get("monto") or 0)
                        .replace(",","").replace("RD$","") or 0)
        for i, ms in enumerate(sistema):
            if i in usados_sistema:
                continue
            diff_monto = abs((ms["monto"] or 0) - mb_monto)
            if diff_monto < 1:  # diferencia menor a 1 peso
                mejor = (i, ms)
                mejor_score = 1
                break
        if mejor:
            usados_sistema.add(mejor[0])
            conciliados.append({"banco": mb, "sistema": mejor[1]})
        else:
            solo_banco.append(mb)

    solo_sistema = [s for i, s in enumerate(sistema) if i not in usados_sistema]

    return jsonify({
        "conciliados": conciliados,
        "solo_banco": solo_banco,
        "solo_sistema": solo_sistema,
        "resumen": {
            "total_conciliados": len(conciliados),
            "total_solo_banco": len(solo_banco),
            "total_solo_sistema": len(solo_sistema),
        }
    })
