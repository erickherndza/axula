#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cargar_listado_estudiantes.py
==============================
Carga (o actualiza) el roster de `estudiantes` a partir de CUALQUIER
"Listado de Estudiantes" oficial que entregue la coordinación — UNA hoja
por archivo, con un bloque de encabezado tipo:

    DATOS DEL ALUMNO 2026-2027          GRADO:  4TO B     ÁREA:  MÚSICA

y debajo una tabla de alumnos. Las únicas columnas que se asumen siempre
presentes son No./Número, Nombre, Apellido, Edad e ID — el resto (Sexo,
Fecha de Nacimiento, Teléfono, etc.) es opcional y su posición/orden no
importa: el algoritmo NO asume columnas fijas, busca la fila de encabezados
de la tabla y ubica cada columna por su nombre. Así admite listados de
cualquier grado/mención/sección sin tocar el script.

- "GRADO:" puede traer sección pegada (ej. "4TO B" → grado=4to, sección=B).
- "ÁREA:"/"MENCIÓN:" es opcional (no existe en 1er ciclo: 1ro-3ro).
- La etiqueta y su valor pueden estar en la misma celda ("GRADO: 4TO B")
  o en celdas separadas (label en una celda, valor en la siguiente).

Esto NO es el mismo formato que `/api/cargar-listado` (que espera el
LISTADO institucional multi-hoja con varias menciones por hoja) — por eso
un alumno de este tipo de archivo puede no aparecer si se subió por esa vía.

Empareja por cédula (columna ID) primero, luego por nombre+apellido exacto,
luego por fuzzy-match (SequenceMatcher ≥0.82) dentro del mismo grado+mención
ya existente en BD — así no duplica a quien ya esté cargado. Actualiza
identidad (nombre/apellido/cédula/grado/curso/mención/sección/edad) de los
que ya existen; inserta los que faltan. Nunca borra ni toca notas/KPIs, y
nunca reactiva solo a un estudiante RETIRADO/TRANSFERIDO (eso lo decide una
persona).

Uso:
    # Dry-run (no escribe en BD) — SIEMPRE correr esto primero
    python3 scripts/cargar_listado_estudiantes.py "Listado_Estudiantes_4TO_B_Musica.xlsx"

    # Carga real
    python3 scripts/cargar_listado_estudiantes.py "Listado_Estudiantes_4TO_B_Musica.xlsx" --commit
"""
import os
import sys
import sqlite3
import unicodedata
from difflib import SequenceMatcher

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_en_render = os.path.exists("/data")
DB_PATH = os.environ.get(
    "DATABASE_PATH",
    "/data/database.db" if _en_render else os.path.join(BASE_DIR, "database.db"),
)

COMMIT = "--commit" in sys.argv
ARGS = [a for a in sys.argv[1:] if not a.startswith("--")]

GRADOS_VALIDOS = ["1ro", "2do", "3ro", "4to", "5to", "6to"]
_GRADO_ALIAS = {"1ero": "1ro", "1er": "1ro", "2er": "2do", "3er": "3ro"}

# nombre normalizado de encabezado → clave interna. Cubre las variantes más
# comunes; agregar aquí si aparece un listado con un nombre de columna nuevo.
HEADER_MAP = {
    "NO": "no", "NO.": "no", "NUM": "no", "NUMERO": "no", "N": "no", "#": "no",
    "NOMBRE": "nombre", "NOMBRES": "nombre",
    "APELLIDO": "apellido", "APELLIDOS": "apellido",
    "EDAD": "edad",
    "ID": "cedula", "CEDULA": "cedula", "IDENTIFICACION": "cedula", "NO IDENTIFICACION": "cedula",
    "SEXO": "sexo", "GENERO": "sexo",
    "FECHA DE NACIMIENTO": "nacimiento", "FECHA NACIMIENTO": "nacimiento", "NACIMIENTO": "nacimiento",
    "TELEFONO": "telefono", "TELEFONO CONTACTO": "telefono", "CONTACTO": "telefono", "TEL": "telefono",
}
COLUMNAS_REQUERIDAS = {"no", "nombre", "apellido"}


def norm(s):
    s = (s or "").strip().lower()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _norm_header(s):
    s = str(s or "").strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.rstrip(".").strip()


def _parsear_grado_seccion(valor):
    """'4TO' → ('4to', None) · '4TO B' → ('4to', 'B') · '1er.' → ('1ro', None)"""
    grado, seccion = None, None
    resto = []
    for tok in str(valor or "").split():
        t = tok.strip().lower().rstrip(".")
        t = _GRADO_ALIAS.get(t, t)
        if t in GRADOS_VALIDOS:
            grado = t
        else:
            resto.append(tok.strip())
    if resto:
        cand = resto[0].strip().upper().rstrip(".")
        if cand and len(cand) <= 2 and cand.isalpha():
            seccion = cand
    return grado, seccion


def _extraer_etiqueta(ws, etiquetas, max_row=15):
    """
    Busca cualquiera de `etiquetas` (ej. "GRADO:") en las primeras filas.
    Soporta 'ETIQUETA: valor' en una sola celda o etiqueta y valor en
    celdas separadas (valor a la derecha). Devuelve el string del valor o ''.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        for cell in row:
            texto = str(cell.value or "").strip()
            if not texto:
                continue
            texto_up = _norm_header(texto)
            for etq in etiquetas:
                etq_n = _norm_header(etq)
                if texto_up == etq_n:
                    # etiqueta sola → el valor está en la celda de al lado
                    vecino = ws.cell(row=cell.row, column=cell.column + 1).value
                    return str(vecino or "").strip()
                if texto_up.startswith(etq_n + ":") or texto_up.startswith(etq_n):
                    resto = texto[len(etq):].lstrip(":").strip()
                    if resto:
                        return resto
    return ""


def _detectar_fila_encabezado(ws, max_row=20):
    """
    Busca la fila que tiene los encabezados de la tabla de alumnos y arma
    {clave_interna: numero_de_columna}. No asume columnas fijas — cada
    listado puede traer las columnas en el orden/posición que sea.
    """
    mejor_fila, mejor_cols = None, {}
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        cols = {}
        for cell in row:
            clave = HEADER_MAP.get(_norm_header(cell.value))
            if clave and clave not in cols:
                cols[clave] = cell.column
        if COLUMNAS_REQUERIDAS.issubset(cols.keys()) and len(cols) > len(mejor_cols):
            mejor_fila, mejor_cols = row[0].row, cols
    if not mejor_fila:
        faltan = COLUMNAS_REQUERIDAS
        raise ValueError(
            "No se encontró una fila de encabezados con las columnas mínimas "
            f"{sorted(faltan)}. Revisa que el archivo tenga No./Nombre/Apellido "
            "como títulos de columna en alguna fila."
        )
    return mejor_fila, mejor_cols


def leer_listado(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    grado_raw   = _extraer_etiqueta(ws, ["GRADO:", "GRADO"])
    mencion_raw = _extraer_etiqueta(ws, ["ÁREA:", "AREA:", "MENCIÓN:", "MENCION:", "ÁREA", "MENCIÓN"])

    grado, seccion = _parsear_grado_seccion(grado_raw)
    if not grado:
        raise ValueError(
            f"No se pudo determinar el grado a partir de 'GRADO: {grado_raw!r}'. "
            "Verifica que el encabezado del archivo tenga 'GRADO:' seguido de "
            "1ro/2do/3ro/4to/5to/6to."
        )
    mencion = mencion_raw.strip().upper() or None

    fila_header, cols = _detectar_fila_encabezado(ws)

    def _val(row, clave):
        col = cols.get(clave)
        return row[col - 1].value if col else None

    alumnos = []
    for row in ws.iter_rows(min_row=fila_header + 1, max_row=ws.max_row):
        no = _val(row, "no")
        if not isinstance(no, (int, float)) or not (1 <= no <= 999):
            continue
        nombre   = str(_val(row, "nombre") or "").strip()
        apellido = str(_val(row, "apellido") or "").strip()
        if not nombre or not apellido:
            continue

        edad_raw = _val(row, "edad")
        try:
            edad = int(edad_raw) if edad_raw not in (None, "") else None
        except (TypeError, ValueError):
            edad = None

        cedula_raw = str(_val(row, "cedula") or "").strip()
        cedula_raw = cedula_raw.replace(".0", "").replace(",", "").replace("-", "").strip()
        cedula = cedula_raw if cedula_raw.isdigit() else None

        alumnos.append({
            "no": int(no), "nombre": nombre, "apellido": apellido,
            "edad": edad, "cedula": cedula,
        })
    return grado, seccion, mencion, alumnos


def buscar_existente(conn, alumno, grado, mencion):
    """cédula exacta → nombre+apellido exacto → fuzzy dentro del mismo grado+mención."""
    if alumno["cedula"]:
        row = conn.execute(
            "SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes WHERE cedula=?",
            (alumno["cedula"],)
        ).fetchone()
        if row:
            return row

    row = conn.execute(
        """SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes
           WHERE lower(nombre)=lower(?) AND lower(apellido)=lower(?)""",
        (alumno["nombre"], alumno["apellido"])
    ).fetchone()
    if row:
        return row

    clave = norm(alumno["nombre"] + " " + alumno["apellido"])
    if mencion:
        candidatos = conn.execute(
            """SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes
               WHERE grado=? AND (mencion=? OR mencion IS NULL OR mencion='')""",
            (grado, mencion)
        ).fetchall()
    else:
        candidatos = conn.execute(
            "SELECT id, nombre, apellido, cedula, grado, curso, condicion FROM estudiantes WHERE grado=?",
            (grado,)
        ).fetchall()
    mejor, mejor_score = None, 0.0
    for c in candidatos:
        score = SequenceMatcher(None, clave, norm(c["nombre"] + " " + c["apellido"])).ratio()
        if score > mejor_score:
            mejor, mejor_score = c, score
    if mejor and mejor_score >= 0.82:
        return mejor
    return None


def main():
    if not ARGS:
        print("Uso: python3 scripts/cargar_listado_estudiantes.py <archivo.xlsx> [--commit]")
        sys.exit(1)
    path = ARGS[0]
    if not os.path.exists(path):
        print(f"No existe el archivo: {path}")
        sys.exit(1)

    grado, seccion, mencion, alumnos = leer_listado(path)
    curso = f"{grado} {mencion}".strip() if mencion else grado
    ciclo = "primer_ciclo" if grado in ("1ro", "2do", "3ro") else "segundo_ciclo"

    print(f"Archivo: {path}")
    print(f"Grado detectado: {grado!r} · Sección: {seccion!r} · Mención: {mencion!r} · Curso: {curso!r}")
    print(f"Alumnos en el archivo: {len(alumnos)}")
    print(f"Modo: {'COMMIT (escribe en BD)' if COMMIT else 'DRY-RUN (solo muestra qué haría)'}")
    print(f"BD: {DB_PATH}\n")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    nuevos, actualizados, sin_cedula = 0, 0, []

    for alumno in alumnos:
        existente = buscar_existente(conn, alumno, grado, mencion)
        if not alumno["cedula"]:
            sin_cedula.append(f"{alumno['nombre']} {alumno['apellido']}")

        if existente:
            cambios = []
            if existente["grado"] != grado:      cambios.append(f"grado {existente['grado']!r}→{grado!r}")
            if existente["curso"] != curso:       cambios.append(f"curso {existente['curso']!r}→{curso!r}")
            if not existente["cedula"] and alumno["cedula"]:
                cambios.append(f"cedula → {alumno['cedula']}")
            if existente["condicion"] in ("RETIRADO", "TRANSFERIDO"):
                cambios.append(f"⚠ condicion actual={existente['condicion']!r} — NO se reactiva automáticamente, revisar a mano")
            print(f"  [ACTUALIZA] #{alumno['no']:>2} {alumno['nombre']} {alumno['apellido']}"
                  + (f"  ({', '.join(cambios)})" if cambios else "  (sin cambios)"))
            actualizados += 1
            if COMMIT:
                # No se toca 'condicion' en un UPDATE — si el alumno ya existía
                # como RETIRADO/TRANSFERIDO, reactivarlo es una decisión humana,
                # no algo que este script deba hacer solo.
                conn.execute(
                    """UPDATE estudiantes
                          SET nombre=?, apellido=?, grado=?, curso=?, mencion=?, ciclo=?,
                              seccion=COALESCE(?, seccion),
                              cedula=COALESCE(NULLIF(cedula,''), ?),
                              edad=COALESCE(?, edad)
                        WHERE id=?""",
                    (alumno["nombre"], alumno["apellido"], grado, curso, mencion, ciclo,
                     seccion, alumno["cedula"] or "", alumno["edad"], existente["id"])
                )
        else:
            print(f"  [NUEVO]     #{alumno['no']:>2} {alumno['nombre']} {alumno['apellido']}"
                  + (f"  (cédula {alumno['cedula']})" if alumno["cedula"] else "  (SIN cédula)"))
            nuevos += 1
            if COMMIT:
                conn.execute(
                    """INSERT INTO estudiantes
                           (id_evaluacion, cedula, nombre, apellido, curso, grado,
                            mencion, ciclo, seccion, condicion, edad)
                       VALUES (?,?,?,?,?,?,?,?,COALESCE(?,'A'),'ACTIVO',?)""",
                    (alumno["cedula"] or f"PROV_{alumno['nombre'].split()[0]}_{alumno['apellido'].split()[0]}",
                     alumno["cedula"] or "", alumno["nombre"], alumno["apellido"],
                     curso, grado, mencion, ciclo, seccion, alumno["edad"])
                )

    if COMMIT:
        conn.commit()
        print(f"\n✓ Guardado: {nuevos} nuevo(s), {actualizados} actualizado(s).")
    else:
        print(f"\n(dry-run) Se crearían {nuevos} nuevo(s), se actualizarían {actualizados}.")
        print("Corre de nuevo con --commit para aplicar los cambios.")

    if sin_cedula:
        print(f"\n⚠ {len(sin_cedula)} alumno(s) sin ID/cédula en el archivo (quedan con ID provisional):")
        for n in sin_cedula:
            print(f"   - {n}")

    conn.close()


if __name__ == "__main__":
    main()
