# -*- coding: utf-8 -*-
# MultimediaTrack — C.E. Benito Juárez · Modalidad Artes

from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, g, send_from_directory
import sqlite3, hashlib, functools, os

app = Flask(__name__)
DATABASE  = os.environ.get("DATABASE", "database.db")
FOTOS_DIR = os.path.join("static", "fotos")
os.makedirs(FOTOS_DIR, exist_ok=True)

_secret = os.environ.get("SECRET_KEY")
if not _secret:
    raise RuntimeError("SECRET_KEY no está definida en .env")
app.secret_key = _secret

# ── GROQ CLIENT ──────────────────────────────────────────────────────────────
def _get_groq_client():
    """Lazy Groq client — solo se importa cuando se llama la IA."""
    from groq import Groq
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        raise RuntimeError("GROQ_API_KEY no está en .env")
    return Groq(api_key=api_key)

groq_client = None


# ── AUTH / SESIONES ──────────────────────────────────────────────────────────

def _hash(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# ── CACHÉ SIMPLE EN MEMORIA ──────────────────────────────────────────────────
import time as _time
_CACHE   = {}
_CACHE_TTL = 90  # segundos — suficiente para navegación fluida, actualiza tras carga

def cache_get(key):
    entry = _CACHE.get(key)
    if entry and (_time.time() - entry['ts']) < _CACHE_TTL:
        return entry['val']
    return None

def cache_set(key, val):
    _CACHE[key] = {'val': val, 'ts': _time.time()}

def cache_bust():
    """Llama esto después de cualquier escritura a la BD."""
    _CACHE.clear()


# ── JERARQUÍA DE ROLES ──────────────────────────────────────────────────────
# directora                  → acceso total (CEO)
# coordinador_general        → gestiona coordinadores de ambos ciclos
# coordinador_primer_ciclo   → 1ro–3ro únicamente
# coordinador_segundo_ciclo  → 4to–6to únicamente
# psicologa_primer_ciclo     → perfiles + cuaderno 1er ciclo
# psicologa_segundo_ciclo    → perfiles + cuaderno 2do ciclo
# profesor                   → sus estudiantes (ambos ciclos si da básicas)

ROLES_ADMIN = {
    "directora", "coordinador_general",
    "coordinador_primer_ciclo", "coordinador_segundo_ciclo"
}
ROLES_COORD = {
    "directora", "coordinador_general",
    "coordinador_primer_ciclo", "coordinador_segundo_ciclo"
}
ROLES_SUPER = {"directora", "coordinador_general"}  # ven ambos ciclos
ROLES_PSICOLOGA = {"psicologa_primer_ciclo", "psicologa_segundo_ciclo"}
ROLES_ADMINISTRATIVO = {
    "secretaria", "secretaria_docente", "digitador", "auxiliar_contabilidad"
}

# Dominios institucionales aceptados para login
DOMINIOS_INSTITUCIONALES = {"educacion.edu.do", "minerd.gob.do"}

def _validar_email_institucional(email):
    """Valida que el email sea de dominio institucional permitido."""
    if not email or "@" not in email:
        return False, "El correo debe tener formato válido (ej: nombre@educacion.edu.do)"
    dominio = email.split("@")[-1].lower().strip()
    if dominio not in DOMINIOS_INSTITUCIONALES:
        return False, "Solo se aceptan correos @educacion.edu.do o @minerd.gob.do"
    return True, ""

# Compatibilidad: el rol antiguo "coordinador" se trata como coordinador_general
def _normalizar_rol(rol):
    if rol == "coordinador":
        return "coordinador_general"
    return rol or "profesor"

def _ciclo_del_rol(rol):
    """Retorna el ciclo al que tiene acceso un rol, o None si ve todo."""
    rol = _normalizar_rol(rol)
    if rol in ROLES_SUPER or rol == "directora":
        return None  # ambos ciclos
    if "primer" in rol:
        return "primer_ciclo"
    if "segundo" in rol:
        return "segundo_ciclo"
    return None  # profesor — se maneja por asignación

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def coord_required(f):
    """Coordinadores, coordinador_general y directora."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login_page"))
        rol = _normalizar_rol(session.get("rol", ""))
        if rol not in ROLES_COORD:
            return jsonify({"error": "Sin permisos"}), 403
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Solo directora y coordinador_general."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login_page"))
        rol = _normalizar_rol(session.get("rol", ""))
        if rol not in ROLES_SUPER:
            return jsonify({"error": "Sin permisos de administración"}), 403
        return f(*args, **kwargs)
    return decorated

def directora_required(f):
    """Solo directora."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login_page"))
        if _normalizar_rol(session.get("rol","")) != "directora":
            return jsonify({"error": "Acceso exclusivo para la Dirección"}), 403
        return f(*args, **kwargs)
    return decorated

def get_usuario():
    """Retorna dict del usuario logueado o None."""
    if not session.get("user_id"):
        return None
    rol = _normalizar_rol(session.get("rol", ""))
    return {
        "id":       session["user_id"],
        "username": session["username"],
        "nombre":   session["nombre"],
        "rol":      rol,
        "rol_raw":  session.get("rol", ""),
        "materia":  session.get("materia", ""),
        "grado":    session.get("grado", ""),
        "mencion":  session.get("mencion", ""),
        "ciclo_acceso": _ciclo_del_rol(rol),
        "es_admin": rol in ROLES_SUPER,
        "es_coord": rol in ROLES_COORD,
        "es_psicologa": rol in ROLES_PSICOLOGA,
        "es_directora": rol == "directora",
    }

@app.context_processor
def inject_usuario():
    return {"current_user": get_usuario()}

def _seed_admin():
    """
    Crea usuarios por defecto si no existen.
    Migra el rol 'coordinador' → 'coordinador_general'.
    """
    with sqlite3.connect(DATABASE) as conn:
        # ── Migrar rol antiguo 'coordinador' → 'coordinador_general' ────────
        conn.execute("""
            UPDATE usuarios SET rol='coordinador_general'
            WHERE rol='coordinador'
        """)
        conn.commit()

        n = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
        if n == 0:
            # Coordinador general por defecto
            conn.execute("""
                INSERT INTO usuarios (username, password, nombre, rol)
                VALUES (?,?,?,?)
            """, ("admin", _hash("admin123"), "Coordinador General", "coordinador_general"))
            conn.commit()
            print("  ── Usuario admin creado (admin / admin123)")

        # ── Crear directora si no existe ─────────────────────────────────────
        existe_directora = conn.execute(
            "SELECT id FROM usuarios WHERE rol='directora' LIMIT 1"
        ).fetchone()
        if not existe_directora:
            conn.execute("""
                INSERT INTO usuarios (username, password, nombre, rol, activo)
                VALUES (?,?,?,?,1)
            """, ("directora", _hash("administradorgral123"),
                  "Dirección General", "directora"))
            conn.commit()
            print("  ── Directora creada (directora / administradorgral123)")

# ── LOGIN ROUTES ──────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET"])
def login_page():
    if session.get("user_id"):
        return redirect("/")
    error = request.args.get("error", "")
    return render_template("login.html", error=error)

@app.route("/login", methods=["POST"])
def login_post():
    username = request.form.get("username","").strip().lower()
    password = request.form.get("password","").strip()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        # Buscar por username (que puede ser un correo institucional)
        u = conn.execute(
            "SELECT * FROM usuarios WHERE lower(username)=? AND activo=1", (username,)
        ).fetchone()
    if not u or u["password"] != _hash(password):
        return redirect("/login?error=Credenciales incorrectas")
    session.permanent = True
    session["user_id"]  = u["id"]
    session["username"] = u["username"]
    session["nombre"]   = u["nombre"]
    session["rol"]      = u["rol"]
    session["materia"]  = u["materia"] or ""
    session["grado"]    = u["grado"]   or ""
    session["mencion"]  = u["mencion"] or ""
    # Redirigir según rol
    rol_norm = _normalizar_rol(u["rol"])
    if rol_norm == "profesor":
        return redirect("/profesor")
    return redirect("/")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/api/usuarios", methods=["GET"])
@coord_required
def listar_usuarios():
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id,username,nombre,rol,materia,grado,mencion,asignaturas,activo,creado FROM usuarios ORDER BY rol,nombre"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/usuarios", methods=["POST"])
@coord_required
def crear_usuario():
    d = request.get_json(silent=True) or {}
    username = d.get("username","").strip()
    password = d.get("password","").strip()
    nombre   = d.get("nombre","").strip()
    rol         = d.get("rol","profesor").strip()
    materia     = d.get("materia","").strip()
    grado       = d.get("grado","").strip()
    mencion     = d.get("mencion","").strip()
    asignaturas = d.get("asignaturas","").strip()
    # Usar email como username si tiene formato de correo
    if not username and "@" in nombre:
        username = nombre
    if not username or not password or not nombre:
        return jsonify({"error": "correo, password y nombre son requeridos"}), 400
    # Validar dominio institucional
    ok_email, msg_email = _validar_email_institucional(username)
    if not ok_email:
        return jsonify({"error": msg_email}), 400
    try:
        with sqlite3.connect(DATABASE) as conn:
            conn.execute(
                """INSERT INTO usuarios (username,password,nombre,rol,materia,grado,mencion,asignaturas)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (username.lower(), _hash(password), nombre, rol, materia, grado, mencion, asignaturas)
            )
            conn.commit()
        return jsonify({"ok": True})
    except sqlite3.IntegrityError:
        return jsonify({"error": "Ya existe un usuario con ese correo"}), 400

@app.route("/api/usuarios/<int:uid>", methods=["PATCH"])
@coord_required
def editar_usuario(uid):
    d = request.get_json(silent=True) or {}
    allowed = ["nombre","rol","materia","grado","mencion","asignaturas","activo"]
    updates = {k: d[k] for k in allowed if k in d}
    if "password" in d and d["password"]:
        updates["password"] = _hash(d["password"])
    if not updates:
        return jsonify({"error": "Sin campos"}), 400
    sets = ", ".join(f"{k}=?" for k in updates)
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(f"UPDATE usuarios SET {sets} WHERE id=?", list(updates.values())+[uid])
        conn.commit()
    return jsonify({"ok": True})

@app.route("/api/usuarios/<int:uid>", methods=["DELETE"])
@coord_required
def borrar_usuario(uid):
    if uid == session.get("user_id"):
        return jsonify({"error": "No puedes eliminarte a ti mismo"}), 400
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM usuarios WHERE id=?", (uid,))
        conn.commit()
    return jsonify({"ok": True})


# ── UTILIDADES ──────────────────────────────────────────────────────────────

def limpiar_v(v):
    if v is None:
        return 0.0
    try:
        # Evitar importar pandas solo para isna
        import math
        if isinstance(v, float) and math.isnan(v):
            return 0.0
    except Exception:
        pass
    try:
        return round(float(str(v).replace('%', '').replace(',', '.').strip()), 1)
    except Exception:
        return 0.0


def calcular_proyeccion(actual, anterior):
    """Proyección optimista/pesimista basada en la tendencia entre períodos."""
    if anterior == 0:
        return actual
    tendencia = actual - anterior
    proyectado = actual + (tendencia * 1.2)
    return round(min(max(proyectado, 0), 100), 1)


def calcular_promedio_modulos(p_foto, p_lv, p_diseno):
    """
    Calcula el promedio real de los 3 módulos técnicos multimedia.
    Solo promedia los módulos que tienen datos (mayor a 0).
    """
    valores = [v for v in [p_foto, p_lv, p_diseno] if v > 0]
    if not valores:
        return 0.0
    return round(sum(valores) / len(valores), 1)


def analizar_perfil_maestro(d, historico=None):
    # Promedios principales — nombres exactos del Excel
    acad      = limpiar_v(d.get('Promedio_Academico', 0))
    cond      = limpiar_v(d.get('Promedio_Conductual', 0))
    emocional = limpiar_v(d.get('Promedio_Emocional', 0))
    auto      = limpiar_v(d.get('Autoestima', 0))

    # Módulos técnicos multimedia — nombres exactos del Excel
    p_foto  = limpiar_v(d.get('Promedio_Fotografia', 0))
    p_lv    = limpiar_v(d.get('Promedio_Lenguaje_Visual', 0))
    p_diseno= limpiar_v(d.get('Promedio_Diseño', 0))

    # Otros indicadores del Excel
    asistencia  = limpiar_v(d.get('Asistencia', 0))
    motivacion  = limpiar_v(d.get('Motivacion', 0))
    indice_riesgo = limpiar_v(d.get('Indice_Riesgo', 0))
    nivel_riesgo  = str(d.get('Nivel_Riesgo', '')).strip()

    # Promedio real de módulos técnicos
    prom_modulos = calcular_promedio_modulos(p_foto, p_lv, p_diseno)

    # Proyección basada en historial
    acad_anterior = historico['p_acad'] if historico else acad
    proyeccion = calcular_proyeccion(acad, acad_anterior)

    res = {
        "es_caso_silencioso": False,
        "nivel": "Estable",
        "color": "#28a745",
        "reporte": "",
        "proyeccion": proyeccion,
        "tendencia": "igual",
        "prom_modulos": prom_modulos
    }

    if proyeccion > acad:
        res["tendencia"] = "subiendo"
    elif proyeccion < acad:
        res["tendencia"] = "bajando"

    # Alerta por proyección
    if proyeccion < 70:
        res["nivel"] = "ALERTA DE REPROBACIÓN"
        res["color"] = "#fd7e14"
        modulo_bajo = ""
        if p_foto > 0 and p_foto < 70:
            modulo_bajo += "Fotografía"
        if p_lv > 0 and p_lv < 70:
            modulo_bajo += (", " if modulo_bajo else "") + "Lenguaje Visual"
        if p_diseno > 0 and p_diseno < 70:
            modulo_bajo += (", " if modulo_bajo else "") + "Diseño"
        detalle = f" Módulos críticos: {modulo_bajo}." if modulo_bajo else ""
        res["reporte"] = (
            f"PLANIFICACIÓN URGENTE: La tendencia indica un promedio final de {proyeccion}.{detalle} "
            "Se requiere intervención inmediata."
        )

    # Erick's Rule — alto rendimiento + bienestar emocional bajo
    if acad >= 80 and (auto < 70 or auto == 0):
        res["es_caso_silencioso"] = True
        res["nivel"] = "CASO SILENCIOSO"
        res["color"] = "#17a2b8"
        res["reporte"] = "Rendimiento alto con bienestar emocional crítico. Programar Entrevista de Bienestar."

    return res


def construir_prompt(e):
    """
    Construye el prompt pedagógico para el LLM.
    Incluye todos los módulos técnicos correctos del Excel.
    """
    # Módulos técnicos
    modulos = []
    if e.get('p_foto'):   modulos.append(f"Fotografía: {e['p_foto']}%")
    if e.get('p_lv'):     modulos.append(f"Lenguaje Visual: {e['p_lv']}%")
    if e.get('p_diseno'): modulos.append(f"Diseño: {e['p_diseno']}%")
    modulos_txt = ", ".join(modulos) if modulos else "No registrados aún"

    silencioso_txt = (
        "SÍ — Rendimiento alto con autoestima crítica (Caso Silencioso)"
        if e.get('silencioso') else "No"
    )

    return f"""Eres un orientador pedagógico experto en educación artística en República Dominicana,
especializado en la Modalidad de Artes del bachillerato dominicano, específicamente en el área de Multimedia.

Analiza el siguiente perfil y genera un plan pedagógico estructurado.

═══════════════════════════════════════
DATOS DEL ESTUDIANTE
═══════════════════════════════════════
Curso: {e.get('curso', '4to Multimedia')}
Estado actual: {e.get('categoria', 'Estable')}
Tendencia académica: {e.get('tendencia', 'igual')}

MÉTRICAS PRINCIPALES:
- Promedio Académico General: {e.get('p_acad', 0)}%
- Promedio Conductual: {e.get('p_cond', 0) or 'No registrado'}%
- Promedio Emocional: {e.get('p_emocional', 0) or 'No registrado'}%
- Índice de Autoestima: {e.get('p_auto', 0) or 'No registrado'}%
- Motivación: {e.get('motivacion', 0) or 'No registrado'}%
- Asistencia: {e.get('asistencia', 0) or 'No registrado'}%
- Proyección cierre de año (IA): {e.get('proyeccion', 0)}%

MÓDULOS TÉCNICOS MULTIMEDIA:
{modulos_txt}
Promedio módulos técnicos: {e.get('prom_modulos', 0)}%

Índice de Riesgo: {e.get('indice_riesgo', 0)} — Nivel: {e.get('nivel_riesgo', 'N/D')}

¿Es Caso Silencioso?: {silencioso_txt}
═══════════════════════════════════════

Genera una respuesta con exactamente estas 4 secciones:

1. DIAGNÓSTICO (2-3 oraciones): Análisis del perfil académico, técnico y emocional actual.

2. FORTALEZAS (2-3 puntos): Aspectos positivos identificados en el perfil.

3. PLAN DE ACCIÓN (3-4 acciones concretas): Estrategias pedagógicas específicas para
   la Modalidad de Artes Multimedia, mencionando los módulos de Fotografía,
   Lenguaje Visual y Diseño donde aplique.

4. SEGUIMIENTO (1-2 oraciones): Cuándo y cómo verificar el progreso.

Usa lenguaje profesional pero accesible para un docente de bachillerato.
Sé específico con las artes visuales y multimedia. Máximo 300 palabras."""


# ── RUTAS ───────────────────────────────────────────────────────────────────


@app.route("/api/usuarios/bulk", methods=["POST"])
@coord_required
def crear_usuarios_bulk():
    """Crea multiples usuarios desde JSON. Retorna resultados."""
    data = request.get_json(silent=True) or []
    if not isinstance(data, list):
        return jsonify({"error": "Se esperaba lista"}), 400
    resultados = []
    for u in data:
        username    = str(u.get("username", "")).strip()
        password    = str(u.get("password", "")).strip()
        nombre      = str(u.get("nombre", "")).strip()
        rol         = str(u.get("rol", "profesor")).strip()
        grado       = str(u.get("grado", "")).strip()
        mencion     = str(u.get("mencion", "")).strip()
        asignaturas = str(u.get("asignaturas", "")).strip()
        materia     = asignaturas.split(",")[0].strip() if asignaturas else ""
        if not username or not password or not nombre:
            resultados.append({"username": username, "ok": False, "error": "Faltan campos"})
            continue
        try:
            with sqlite3.connect(DATABASE) as conn:
                conn.execute(
                    "INSERT INTO usuarios (username,password,nombre,rol,materia,grado,mencion,asignaturas) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (username, _hash(password), nombre, rol, materia, grado, mencion, asignaturas)
                )
                conn.commit()
            resultados.append({"username": username, "ok": True})
        except sqlite3.IntegrityError:
            resultados.append({"username": username, "ok": False, "error": "Usuario ya existe"})
    ok_count = sum(1 for r in resultados if r["ok"])
    return jsonify({"resultados": resultados, "ok": ok_count, "errores": len(resultados) - ok_count})


@app.route("/api/asistencia/lote", methods=["POST"])
@login_required
def registrar_asistencia_lote():
    """Registra asistencia de multiples estudiantes en un POST.
    Body: { fecha, materia, grado, periodo, horas_clase,
            registros: [{estudiante_id, estado, observacion}] }
    Valida materia/grado contra perfil del profesor.
    """
    d = request.get_json(silent=True) or {}
    fecha     = d.get("fecha", "")
    materia   = d.get("materia", "").strip()
    periodo   = int(d.get("periodo", 1))
    horas     = int(d.get("horas_clase", 1))
    registros = d.get("registros", [])
    grado     = d.get("grado", "").strip()
    if not fecha or not materia or not registros:
        return jsonify({"error": "fecha, materia y registros son requeridos"}), 400
    prof = _get_profesor()
    ok, msg = _validar_materia_profesor(materia, grado + " MULTIMEDIA", prof)
    if not ok:
        return jsonify({"error": msg}), 403
    prof_id = session.get("user_id")
    creados = 0
    with sqlite3.connect(DATABASE) as conn:
        for r in registros:
            est_id = r.get("estudiante_id")
            estado = r.get("estado", "presente")
            obs    = r.get("observacion", "")
            if not est_id:
                continue
            existing = conn.execute(
                "SELECT id FROM asistencia WHERE estudiante_id=? AND materia=? AND fecha=? AND profesor_id=?",
                (est_id, materia, fecha, prof_id)
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE asistencia SET estado=?, horas_clase=?, observacion=? WHERE id=?",
                    (estado, horas, obs, existing[0])
                )
            else:
                conn.execute(
                    "INSERT INTO asistencia "
                    "(estudiante_id,profesor_id,materia,fecha,periodo,estado,horas_clase,observacion) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (est_id, prof_id, materia, fecha, periodo, estado, horas, obs)
                )
            creados += 1
        conn.commit()
    return jsonify({"ok": True, "registros": creados})


@app.route("/api/asistencia/resumen-mensual/<int:prof_id>")
@login_required
def resumen_mensual_asistencia(prof_id):
    """Resumen mensual de asistencia para un profesor.
    Calculo MINERD: (horas presentes / horas totales impartidas) * 100.
    Alerta si < 75% (riesgo de reprobacion por inasistencia segun Ordenanza 1-96 Art 51).
    """
    mes     = request.args.get("mes", "")
    materia = request.args.get("materia", "")
    if session.get("rol") != "coordinador" and session.get("user_id") != prof_id:
        return jsonify({"error": "Sin permisos"}), 403
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        q = (
            "SELECT a.estudiante_id, e.nombre AS est_nombre, e.grado, a.materia,"
            " COUNT(*) AS total_sesiones,"
            " SUM(a.horas_clase) AS horas_totales,"
            " SUM(CASE WHEN a.estado='presente' THEN a.horas_clase ELSE 0 END) AS horas_presentes,"
            " SUM(CASE WHEN a.estado='tardanza' THEN 1 ELSE 0 END) AS tardanzas,"
            " SUM(CASE WHEN a.estado='ausente' THEN 1 ELSE 0 END) AS ausencias"
            " FROM asistencia a"
            " JOIN estudiantes e ON e.id=a.estudiante_id"
            " WHERE a.profesor_id=?"
        )
        params = [prof_id]
        if mes:
            q += " AND strftime('%Y-%m',a.fecha)=?"
            params.append(mes)
        if materia:
            q += " AND a.materia=?"
            params.append(materia)
        q += " GROUP BY a.estudiante_id,a.materia ORDER BY e.nombre"
        rows = conn.execute(q, params).fetchall()
    result = []
    for r in rows:
        ht  = r["horas_totales"] or 0
        hp  = r["horas_presentes"] or 0
        pct = round((hp / ht * 100), 1) if ht > 0 else 0
        result.append({
            "estudiante_id":         r["estudiante_id"],
            "nombre":                r["est_nombre"],
            "grado":                 r["grado"],
            "materia":               r["materia"],
            "total_sesiones":        r["total_sesiones"],
            "horas_totales":         ht,
            "horas_presentes":       hp,
            "ausencias":             r["ausencias"],
            "tardanzas":             r["tardanzas"],
            "porcentaje_asistencia": pct,
            "alerta_inasistencia":   pct < 75 if ht > 0 else False,
        })
    return jsonify(result)


@app.route("/")
@login_required
def index():
    u = get_usuario()
    rol = u.get("rol","")
    return render_template(
        "index.html",
        usuario=u,
        current_user=u,
        es_profesor=(rol == "profesor"),
        es_coord=(rol in ROLES_COORD),
        es_admin=(rol in ROLES_SUPER),
        prof_grado=(u.get("grado") or ""),
        prof_mencion=(u.get("mencion") or ""),
        prof_asignaturas=(u.get("asignaturas") or ""),
    )


@app.route("/cargar", methods=["POST"])
def cargar():
    """
    Carga la Plantilla MultimediaTrack (cualquier version).
    Detecta header buscando la celda 'Nombre' en cols 2-6, filas 1-8.
    UPSERT: preserva todos los estudiantes del LISTADO.
    """
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "No se recibio ningún archivo"}), 400

    try:
        from openpyxl import load_workbook
        import io as _io, unicodedata

        raw = file.read()
        wb  = load_workbook(_io.BytesIO(raw), data_only=True)
        hoja = next((s for s in wb.sheetnames
                     if 'SEGUIMIENTO' in s.upper() or 'ESTUDIANTE' in s.upper()),
                    wb.sheetnames[0])
        ws = wb[hoja]

        # ── Detectar fila de encabezado ───────────────────────────────────
        # Buscamos la celda exacta que diga 'Nombre' en cols 2-6
        header_row = None
        nombre_col = None
        for row in range(1, 9):
            for col in range(2, 7):
                v = str(ws.cell(row=row, column=col).value or '').strip()
                if v == 'Nombre':
                    header_row = row
                    nombre_col = col
                    break
            if header_row:
                break

        if not header_row:
            return jsonify({
                "error": "No se encontro la columna 'Nombre' en las primeras 8 filas. "
                         "Verifica que estas subiendo la Plantilla MultimediaTrack."
            }), 400

        # Columnas basicas relativas al nombre_col
        col_nombre   = nombre_col        # 'Nombre'
        col_apellido = nombre_col + 1    # 'Apellido'
        col_curso    = nombre_col + 2    # 'Curso'
        col_grado    = nombre_col + 3    # 'Grado'
        col_condicion= nombre_col + 4    # 'Condición'
        col_edad     = nombre_col + 5    # 'Edad'
        col_id       = nombre_col - 2    # 'ID' (2 antes de Nombre)

        data_start = header_row + 1

        # ── Detectar grupos de columnas buscando en fila anterior ────────
        grupo_row = header_row - 1
        col_foto = col_lv = col_diseno = col_acad_prom = None
        col_conductual = col_emocional = col_analisis = None

        for col in range(1, ws.max_column + 1):
            gv = str(ws.cell(row=grupo_row, column=col).value or '').upper().strip()
            if not gv:
                continue
            if 'FOTOGRAF' in gv:                    col_foto       = col
            elif 'LENGUAJE' in gv:                  col_lv         = col
            elif 'DISE' in gv:                      col_diseno     = col
            elif 'PROM' in gv and 'ACAD' in gv:    col_acad_prom  = col
            elif 'CONDUCTUAL' in gv:                col_conductual = col
            elif 'EMOCIONAL' in gv:                 col_emocional  = col
            elif 'AN' in gv and ('LISIS' in gv or 'ÁLISIS' in gv): col_analisis = col

        # Fallback por version si no se detectaron grupos
        if col_foto is None:
            if header_row == 3:   # MultimediaTrack
                col_foto=29; col_lv=33; col_diseno=37; col_acad_prom=41
                col_conductual=18; col_emocional=24; col_analisis=48
            else:                  # Modelo v3 (header_row == 4)
                col_foto=14; col_lv=18; col_diseno=22; col_acad_prom=30
                col_conductual=35; col_emocional=42; col_analisis=48

        def fl(v):
            try:
                f = float(v)
                return round(f, 2) if f == f else None
            except Exception:
                return None

        def norm(s):
            return ''.join(c for c in unicodedata.normalize('NFD', str(s).lower())
                           if unicodedata.category(c) != 'Mn')

        def gc(base, offset):
            if base is None: return None
            return fl(ws.cell(row=row, column=base + offset).value)

        # ── Leer estudiantes ─────────────────────────────────────────────
        students = []
        for row in range(data_start, ws.max_row + 1):
            nombre = ws.cell(row=row, column=col_nombre).value
            if not nombre or str(nombre).strip() in ('', 'nan', 'None', 'Nombre'):
                continue

            id_bj = str(ws.cell(row=row, column=col_id).value or '').strip()

            # Conductual base (Puntualidad-Rendimiento, cols 9-13 en ambas versiones)
            punt = fl(ws.cell(row=row, column=9).value)
            tar  = fl(ws.cell(row=row, column=10).value)
            part = fl(ws.cell(row=row, column=11).value)
            comp = fl(ws.cell(row=row, column=12).value)
            rend = fl(ws.cell(row=row, column=13).value)

            foto_p1=gc(col_foto,0); foto_p2=gc(col_foto,1)
            foto_p3=gc(col_foto,2); foto_p4=gc(col_foto,3)
            lv_p1=gc(col_lv,0);    lv_p2=gc(col_lv,1)
            lv_p3=gc(col_lv,2);    lv_p4=gc(col_lv,3)
            dis_p1=gc(col_diseno,0); dis_p2=gc(col_diseno,1)
            dis_p3=gc(col_diseno,2); dis_p4=gc(col_diseno,3)

            acad_p1=gc(col_acad_prom,0); acad_p2=gc(col_acad_prom,1)
            acad_p3=gc(col_acad_prom,2); acad_p4=gc(col_acad_prom,3)
            p_acad = gc(col_acad_prom,4)  # Final

            intr   = gc(col_conductual,0)
            uso_cel= str(ws.cell(row=row, column=col_conductual+1).value or 'No') if col_conductual else 'No'
            conf   = gc(col_conductual,2)
            desaf  = gc(col_conductual,3)
            distr  = gc(col_conductual,4)
            falt   = gc(col_conductual,5)
            p_cond = gc(col_conductual,6)

            motiv  = gc(col_emocional,0)
            p_auto = gc(col_emocional,1)
            est_em = gc(col_emocional,2)
            int_fut= gc(col_emocional,3)
            apoy   = gc(col_emocional,4)
            p_emoc = gc(col_emocional,5)

            idx_r  = gc(col_analisis,0)
            niv_r  = str(ws.cell(row=row, column=col_analisis+1).value or 'N/D') if col_analisis else 'N/D'
            tend   = str(ws.cell(row=row, column=col_analisis+2).value or 'igual') if col_analisis else 'igual'
            proy   = gc(col_analisis,3)

            fotos=[x for x in [foto_p1,foto_p2,foto_p3,foto_p4] if x]
            lvs  =[x for x in [lv_p1,lv_p2,lv_p3,lv_p4] if x]
            diss =[x for x in [dis_p1,dis_p2,dis_p3,dis_p4] if x]
            p_foto   = round(sum(fotos)/len(fotos),2) if fotos else None
            p_lv     = round(sum(lvs)/len(lvs),2)     if lvs   else None
            p_diseno = round(sum(diss)/len(diss),2)   if diss  else None
            mods=[x for x in [p_foto,p_lv,p_diseno] if x]
            prom_mod = round(sum(mods)/len(mods),2) if mods else None
            if not p_acad and prom_mod: p_acad = prom_mod

            students.append({
                'id_bj':id_bj,'nombre':str(nombre).strip(),
                'apellido':str(ws.cell(row=row,column=col_apellido).value or '').strip(),
                'curso':str(ws.cell(row=row,column=col_curso).value or '').strip(),
                'grado':str(ws.cell(row=row,column=col_grado).value or '4to').strip(),
                'condicion':str(ws.cell(row=row,column=col_condicion).value or 'ACTIVO'),
                'edad':fl(ws.cell(row=row,column=col_edad).value),
                'puntualidad':punt,'tareas':tar,'participacion':part,
                'comprension':comp,'rendimiento':rend,
                'interrupciones':intr,'uso_celular':uso_cel,'conflictos':conf,
                'desafia_autoridad':desaf,'distraccion':distr,'falta_respeto':falt,
                'motivacion':motiv,'p_auto':p_auto,'estado_emocional':est_em,
                'interes_futuro':int_fut,'apoyo_familiar':apoy,
                'fotografia_p1':foto_p1,'fotografia_p2':foto_p2,
                'fotografia_p3':foto_p3,'fotografia_p4':foto_p4,
                'lv_p1':lv_p1,'lv_p2':lv_p2,'lv_p3':lv_p3,'lv_p4':lv_p4,
                'diseno_p1':dis_p1,'diseno_p2':dis_p2,
                'diseno_p3':dis_p3,'diseno_p4':dis_p4,
                'acad_p1':acad_p1,'acad_p2':acad_p2,'acad_p3':acad_p3,'acad_p4':acad_p4,
                'p_acad':p_acad,'p_cond':p_cond,'p_emocional':p_emoc,
                'p_foto':p_foto,'p_lv':p_lv,'p_diseno':p_diseno,'prom_modulos':prom_mod,
                'indice_riesgo':idx_r,'nivel_riesgo':niv_r,'tendencia':tend,'proyeccion':proy,
            })

        # ── UPSERT ───────────────────────────────────────────────────────
        actualizados=0; nuevos=0

        with sqlite3.connect(DATABASE) as conn:
            conn.row_factory = sqlite3.Row
            conn.create_function("norm", 1, norm)

            for s in students:
                est = None
                if s['id_bj'] and s['id_bj'].startswith('BJ'):
                    est = conn.execute(
                        "SELECT id FROM estudiantes WHERE cedula=? OR id_evaluacion=? LIMIT 1",
                        (s['id_bj'], s['id_bj'])
                    ).fetchone()
                if not est:
                    nom1 = norm(s['nombre']).split()[0]
                    ape1 = norm(s['apellido']).split()[0] if s['apellido'] else ''
                    est  = conn.execute(
                        "SELECT id FROM estudiantes "
                        "WHERE norm(nombre) LIKE ? AND norm(apellido) LIKE ? LIMIT 1",
                        (f"%{nom1}%", f"%{ape1}%")
                    ).fetchone()
                if not est:
                    conn.execute(
                        "INSERT INTO estudiantes "
                        "(id_evaluacion,cedula,nombre,apellido,curso,grado,condicion) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (s['id_bj'],s['id_bj'],s['nombre'],s['apellido'],
                         s['curso'],s['grado'],s['condicion'])
                    )
                    est_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    nuevos += 1
                else:
                    est_id = est['id']

                # ── Leer valores existentes del estudiante para no pisar notas ya cargadas
                ex = conn.execute("SELECT * FROM estudiantes WHERE id=?", (est_id,)).fetchone()
                ex = dict(ex) if ex else {}

                def _keep(new_val, old_key):
                    """Usa new_val si tiene dato real; si no, conserva el existente."""
                    if new_val is not None and new_val != 0:
                        return new_val
                    return ex.get(old_key)

                conn.execute("""
                    UPDATE estudiantes SET
                        id_evaluacion=?,cedula=?,nombre=?,apellido=?,
                        curso=?,grado=?,condicion=?,edad=?,
                        puntualidad=?,tareas=?,participacion=?,comprension=?,
                        rendimiento=?,interrupciones=?,uso_celular=?,
                        conflictos=?,desafia_autoridad=?,distraccion=?,falta_respeto=?,
                        motivacion=?,p_auto=?,estado_emocional=?,
                        interes_futuro=?,apoyo_familiar=?,
                        fotografia_p1=?,fotografia_p2=?,fotografia_p3=?,fotografia_p4=?,
                        lv_p1=?,lv_p2=?,lv_p3=?,lv_p4=?,
                        diseno_p1=?,diseno_p2=?,diseno_p3=?,diseno_p4=?,
                        acad_p1=?,acad_p2=?,acad_p3=?,acad_p4=?,
                        p_acad=?,p_cond=?,p_emocional=?,
                        p_foto=?,p_lv=?,p_diseno=?,prom_modulos=?,
                        indice_riesgo=?,nivel_riesgo=?,tendencia=?,proyeccion=?
                    WHERE id=?
                """, (
                    s['id_bj'],s['id_bj'],s['nombre'],s['apellido'],
                    s['curso'],s['grado'],s['condicion'],s['edad'],
                    s['puntualidad'],s['tareas'],s['participacion'],s['comprension'],
                    s['rendimiento'],s['interrupciones'],s['uso_celular'],
                    s['conflictos'],s['desafia_autoridad'],s['distraccion'],s['falta_respeto'],
                    s['motivacion'],s['p_auto'],s['estado_emocional'],
                    s['interes_futuro'],s['apoyo_familiar'],
                    _keep(s['fotografia_p1'],'fotografia_p1'),
                    _keep(s['fotografia_p2'],'fotografia_p2'),
                    _keep(s['fotografia_p3'],'fotografia_p3'),
                    _keep(s['fotografia_p4'],'fotografia_p4'),
                    _keep(s['lv_p1'],'lv_p1'), _keep(s['lv_p2'],'lv_p2'),
                    _keep(s['lv_p3'],'lv_p3'), _keep(s['lv_p4'],'lv_p4'),
                    _keep(s['diseno_p1'],'diseno_p1'), _keep(s['diseno_p2'],'diseno_p2'),
                    _keep(s['diseno_p3'],'diseno_p3'), _keep(s['diseno_p4'],'diseno_p4'),
                    _keep(s['acad_p1'],'acad_p1'), _keep(s['acad_p2'],'acad_p2'),
                    _keep(s['acad_p3'],'acad_p3'), _keep(s['acad_p4'],'acad_p4'),
                    s['p_acad'],s['p_cond'],s['p_emocional'],
                    s['p_foto'],s['p_lv'],s['p_diseno'],s['prom_modulos'],
                    s['indice_riesgo'],s['nivel_riesgo'],s['tendencia'],s['proyeccion'],
                    est_id
                ))
                actualizados += 1

            conn.commit()

        with sqlite3.connect(DATABASE) as c2:
            total     = c2.execute("SELECT COUNT(*) FROM estudiantes").fetchone()[0]
            con_notas = c2.execute(
                "SELECT COUNT(*) FROM estudiantes WHERE p_acad>0 OR acad_p1>0"
            ).fetchone()[0]

        return jsonify({
            "status":"success",
            "version_detectada": f"header fila {header_row}, Nombre en C{nombre_col}",
            "actualizados":actualizados,"nuevos":nuevos,
            "total_sistema":total,"con_notas":con_notas,
            "mensaje":(f"{actualizados} estudiantes actualizados. "
                       f"Total en sistema: {total} ({con_notas} con notas).")
        })

    except Exception as ex:
        import traceback
        return jsonify({"error": str(ex), "detalle": traceback.format_exc()}), 500


@app.route("/api/datos")
def api_datos():
    """
    Devuelve todos los estudiantes con tiene_notas=True si tienen
    calificaciones en la Plantilla O en materias_calificaciones dinamicas.
    Soporta filtros: grado, mencion, ciclo (primer_ciclo/segundo_ciclo).
    """
    grado      = request.args.get("grado", "").strip()
    mencion    = request.args.get("mencion", "").strip()
    ciclo      = request.args.get("ciclo", "").strip()
    solo_notas = request.args.get("solo_con_notas", "0") == "1"

    # ── Aplicar restricciones de acceso por rol ──────────────────────────
    _usr = _get_profesor()
    if _usr:
        rol_n = _normalizar_rol(_usr.get("rol",""))
        ciclo_usr = _ciclo_del_rol(rol_n)
        # Profesor: filtrar por grado y mención asignados
        if rol_n == "profesor":
            if _usr.get("grado") and not grado:
                grado = _usr["grado"]
            if _usr.get("mencion") and not mencion:
                mencion = _usr["mencion"]
        # Coordinador/psicóloga de ciclo: restringir a su ciclo
        elif ciclo_usr and not ciclo:
            ciclo = ciclo_usr

    # Cache solo cuando no hay filtros activos (y es coordinador)
    if not grado and not mencion and not solo_notas and not ciclo:
        cached = cache_get('api_datos')
        if cached is not None:
            return jsonify(cached)

    seccion = request.args.get("seccion", "").strip()

    query  = "SELECT * FROM estudiantes WHERE 1=1"
    params = []
    if grado:
        query  += " AND upper(grado) LIKE upper(?)"
        params.append(f"%{grado}%")
    if mencion:
        query  += " AND upper(curso) LIKE upper(?)"
        params.append(f"%{mencion}%")
    if ciclo:
        query  += " AND ciclo=?"
        params.append(ciclo)
    if seccion:
        query  += " AND upper(seccion)=upper(?)"
        params.append(seccion)

    query += " ORDER BY apellido, nombre"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(query, params).fetchall()

            # IDs con materias dinamicas cargadas
            ids_con_materias = set(
                r[0] for r in conn.execute(
                    "SELECT DISTINCT estudiante_id FROM materias_calificaciones"
                ).fetchall()
            )

            resultado = []
            for r in rows:
                d = dict(r)
                tiene_plantilla = bool(
                    (d.get('p_acad') or 0) > 0 or
                    (d.get('acad_p1') or 0) > 0 or
                    (d.get('fotografia_p1') or 0) > 0
                )
                tiene_materias = d.get('id') in ids_con_materias
                d['tiene_notas']    = tiene_plantilla or tiene_materias
                d['tiene_plantilla'] = tiene_plantilla
                d['tiene_materias']  = tiene_materias

                if solo_notas and not d['tiene_notas']:
                    continue
                resultado.append(d)

            # ── Calcular categoría de alerta dinámica ─────────────
            for d in resultado:
                p = float(d.get('p_acad') or 0)
                if not d.get('tiene_notas') or p == 0:
                    d['categoria']    = 'SIN DATOS'
                    d['alerta_nivel'] = 0
                    d['alerta_color'] = ''
                elif p < 70:
                    d['categoria']    = 'ALERTA CRÍTICA'
                    d['alerta_nivel'] = 2
                    d['alerta_color'] = 'rojo'
                elif p < 80:
                    d['categoria']    = 'ESTUDIANTE EN OBSERVACIÓN'
                    d['alerta_nivel'] = 1
                    d['alerta_color'] = 'naranja'
                elif p >= 85:
                    d['categoria']    = 'EXCELENTE'
                    d['alerta_nivel'] = 0
                    d['alerta_color'] = 'verde'
                else:
                    d['categoria']    = 'REGULAR'
                    d['alerta_nivel'] = 0
                    d['alerta_color'] = ''

            # Ordenar: alertas críticas primero, luego observación, luego el resto
            resultado.sort(key=lambda x: (
                0 if x['tiene_notas'] else 1,
                -(x.get('alerta_nivel') or 0),
                -(x.get('p_acad') or 0),
                x.get('apellido','')
            ))
            cache_set('api_datos', resultado)
            return jsonify(resultado)
        except Exception as ex:
            print(f"api_datos error: {ex}")
            return jsonify([])


@app.route("/perfil/<int:id>")
@login_required
def perfil_estudiante(id):
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        estudiante = conn.execute(
            "SELECT * FROM estudiantes WHERE id = ?", (id,)
        ).fetchone()
    if estudiante:
        e = dict(estudiante)

        # ── Completar notas desde materias_calificaciones si campos directos están vacíos ──
        with sqlite3.connect(DATABASE) as conn2:
            conn2.row_factory = sqlite3.Row
            mats = conn2.execute("""
                SELECT materia, p1, p2, p3, p4, promedio, fecha_carga, profesor
                FROM materias_calificaciones
                WHERE estudiante_id = ?
            """, (id,)).fetchall()

        if mats:
            MODULO_MAP_PERFIL = {
                # Fotografía — todas las variantes del boletín
                'fotografía': 'fotografia', 'fotografia': 'fotografia', 'foto': 'fotografia',
                'fotografía digital': 'fotografia', 'fotografia digital': 'fotografia',
                'introducción fotografía digital': 'fotografia',
                'introduccion fotografia digital': 'fotografia',
                'introducción a la fotografía digital': 'fotografia',
                'introduccion a la fotografia digital': 'fotografia',
                # Lenguaje Visual — todas las variantes del boletín
                'lenguaje visual': 'lv', 'lenguaje_visual': 'lv', 'lv': 'lv',
                'lenguaje visual dibujo y creación personajes': 'lv',
                'lenguaje visual dibujo y creacion personajes': 'lv',
                'lenguaje visual y principios diseño': 'lv',
                'lenguaje visual y principios diseno': 'lv',
                'lenguaje visual artesanal': 'lv',
                # Diseño — todas las variantes del boletín
                'diseño': 'diseno', 'diseno': 'diseno',
                'diseño básico': 'diseno', 'diseño basico': 'diseno',
                'diseño básico y expresión visual': 'diseno',
                'diseno basico y expresion visual': 'diseno',
                'diseño básico y expresion visual': 'diseno',
            }
            ASISTENCIA_KEYS = {'asistencia', 'asistencias'}
            promedios_acad = []

            for mat in mats:
                mn  = mat['materia'].lower().strip()
                # Buscar en el mapa — exacto primero, luego por subcadena
                col = MODULO_MAP_PERFIL.get(mn)
                if not col:
                    for k, v in MODULO_MAP_PERFIL.items():
                        if k in mn or mn in k:
                            col = v
                            break

                if col:  # Módulo técnico Multimedia
                    p_col = 'p_foto' if col == 'fotografia' else f'p_{col}'
                    for pi in [1,2,3,4]:
                        key = f'{col}_p{pi}'
                        if not e.get(key):
                            e[key] = mat[f'p{pi}'] or 0
                    if not e.get(p_col) and mat['promedio']:
                        e[p_col] = mat['promedio']
                elif mn in ASISTENCIA_KEYS:
                    for pi in [1,2,3,4]:
                        key = f'asistencia_p{pi}'
                        if not e.get(key):
                            e[key] = mat[f'p{pi}'] or 0
                else:
                    # Materia regular → contribuye al promedio académico
                    if mat['promedio'] and mat['promedio'] > 0:
                        promedios_acad.append(mat['promedio'])
                    # Sync periodos a acad_p1..p4 si están vacíos
                    for pi in [1,2,3,4]:
                        key = f'acad_p{pi}'
                        if not e.get(key) and mat[f'p{pi}']:
                            e[key] = mat[f'p{pi}']

            # Recalcular prom_modulos si fue completado desde Excel
            mods = [e.get(k) or 0 for k in ['p_foto','p_lv','p_diseno'] if (e.get(k) or 0) > 0]
            if mods and not e.get('prom_modulos'):
                e['prom_modulos'] = round(sum(mods)/len(mods), 1)

            # Calcular p_acad desde materias regulares si estaba en 0
            if not e.get('p_acad') and promedios_acad:
                e['p_acad'] = round(sum(promedios_acad)/len(promedios_acad), 1)

        # Campos de texto con fallback
        texto_defaults = {
            'uso_celular': 'No', 'nivel_riesgo': 'N/D',
            'grado': '4to', 'condicion': 'ACTIVO',
            'tendencia': 'igual', 'categoria': '',
            'reporte': '', 'color': '', 'ia_analisis': None,
            'nombre': '', 'apellido': '', 'curso': '',
        }
        for k, v in texto_defaults.items():
            if e.get(k) is None:
                e[k] = v

        # Todos los campos numéricos: None → 0.0
        # Esto evita el error ">= no compatible con NoneType"
        campos_numericos = [
            'p_acad','p_foto','p_lv','p_diseno','p_cond','p_auto',
            'p_emocional','prom_modulos','asistencia','proyeccion',
            'puntualidad','tareas','participacion','comprension','rendimiento',
            'interrupciones','conflictos','desafia_autoridad','distraccion',
            'falta_respeto','motivacion','estado_emocional','interes_futuro',
            'apoyo_familiar','indice_riesgo','edad','silencioso',
            'fotografia_p1','fotografia_p2','fotografia_p3','fotografia_p4',
            'lv_p1','lv_p2','lv_p3','lv_p4',
            'diseno_p1','diseno_p2','diseno_p3','diseno_p4',
            'asistencia_p1','asistencia_p2','asistencia_p3','asistencia_p4',
            'acad_p1','acad_p2','acad_p3','acad_p4',
        ]
        for k in campos_numericos:
            if e.get(k) is None:
                e[k] = 0.0

        # ── Calcular proyección dinamicamente ─────────────────────────────
        # No depende de la columna de la Plantilla (que el maestro no llena)
        # Usa los periodos reales para proyectar la tendencia al cierre del año
        acad_p1 = e.get('acad_p1') or 0
        acad_p2 = e.get('acad_p2') or 0
        acad_p3 = e.get('acad_p3') or 0
        acad_p4 = e.get('acad_p4') or 0
        p_acad  = e.get('p_acad')  or 0

        periodos_con_nota = [p for p in [acad_p1,acad_p2,acad_p3,acad_p4] if p > 0]

        if len(periodos_con_nota) >= 2:
            # Tendencia entre los dos últimos períodos disponibles
            ultimo    = periodos_con_nota[-1]
            penultimo = periodos_con_nota[-2]
            delta     = ultimo - penultimo
            # Proyectar con amortiguación: la tendencia se modera con más períodos
            amort     = 0.6 if len(periodos_con_nota) >= 3 else 0.8
            proyeccion_calc = round(min(max(ultimo + delta * amort, 0), 100), 1)
        elif len(periodos_con_nota) == 1:
            # Solo un período: proyectar igual + ajuste emocional/conductual
            base = periodos_con_nota[0]
            p_cond  = e.get('p_cond') or 0
            p_auto  = e.get('p_auto') or 0
            # Si conducta y autoestima son buenas, proyectar ligeramente al alza
            boost = 0
            if p_cond  >= 75: boost += 1.5
            if p_auto  >= 70: boost += 1.0
            if p_cond  <  60: boost -= 2.0
            proyeccion_calc = round(min(max(base + boost, 0), 100), 1)
        elif p_acad > 0:
            # Tiene promedio final pero sin desglose por período
            proyeccion_calc = p_acad
        else:
            # Sin datos académicos — proyectar con módulos técnicos si existen
            mods = [e.get(m) or 0 for m in ['p_foto','p_lv','p_diseno'] if (e.get(m) or 0) > 0]
            proyeccion_calc = round(sum(mods)/len(mods), 1) if mods else 0.0

        # Solo sobreescribir si la DB tiene 0 o None
        if not e.get('proyeccion') or e.get('proyeccion', 0) == 0:
            e['proyeccion'] = proyeccion_calc

        # Recalcular tendencia
        if len(periodos_con_nota) >= 2:
            if periodos_con_nota[-1] > periodos_con_nota[-2]:
                e['tendencia'] = 'subiendo'
            elif periodos_con_nota[-1] < periodos_con_nota[-2]:
                e['tendencia'] = 'bajando'
            else:
                e['tendencia'] = 'igual'

        # Marcar si tiene notas para que el template pueda mostrar aviso
        e['tiene_notas'] = bool(e.get('p_acad', 0) > 0 or e.get('acad_p1', 0) > 0)

        return render_template("perfil.html", e=e, current_user=get_usuario())
    return "Estudiante no encontrado", 404


# ── GROQ: Generar análisis pedagógico IA ────────────────────────────────────

@app.route("/api/analisis-ia/<int:id>", methods=["POST"])
def generar_analisis_ia(id):
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        e = conn.execute(
            "SELECT * FROM estudiantes WHERE id = ?", (id,)
        ).fetchone()

    if not e:
        return jsonify({"error": "Estudiante no encontrado"}), 404

    e = dict(e)

    if e.get("ia_analisis"):
        return jsonify({"analisis": e["ia_analisis"], "cached": True})

    try:
        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Eres un orientador pedagógico experto en la Modalidad de Artes "
                        "del bachillerato dominicano. Respondes siempre en español, "
                        "de forma estructurada y profesional."
                    )
                },
                {
                    "role": "user",
                    "content": construir_prompt(e)
                }
            ],
            temperature=0.6,
            max_tokens=600,
            top_p=0.9
        )

        analisis = completion.choices[0].message.content.strip()

        with sqlite3.connect(DATABASE) as conn:
            conn.execute(
                "UPDATE estudiantes SET ia_analisis = ? WHERE id = ?",
                (analisis, id)
            )
            conn.commit()

        return jsonify({"analisis": analisis, "cached": False})

    except Exception as ex:
        return jsonify({"error": f"Error al contactar Groq: {str(ex)}"}), 500


@app.route("/api/analisis-ia/<int:id>", methods=["DELETE"])
def limpiar_analisis_ia(id):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "UPDATE estudiantes SET ia_analisis = NULL WHERE id = ?", (id,)
        )
        conn.commit()
    return jsonify({"status": "cleared"})


# ── SERVICIO DE PLANIFICACIÓN DOCENTE ───────────────────────────────────────

# Base curricular oficial MINERD extraída del documento
CURRICULUM_MULTIMEDIA = {
    "Fotografía": {
        "competencia": "DCM.1.2 — Diseño y Creatividad Multimedia",
        "descripcion": "Ejecuta el proceso del uso de la cámara fotográfica en proyectos de expresión artística y como medio de comunicación aplicando técnicas novedosas y creativas.",
        "saberes_conceptuales": [
            "Historia de la fotografía", "Componentes de la cámara fotográfica",
            "Diafragma, obturador e ISO", "Profundidad de campo", "Tipos de objetivos",
            "Iluminación de estudio", "Ley del horizonte, de la mirada y tercios",
            "Planos fotográficos", "Ángulos fotográficos y significados psicológicos",
            "Producción fotográfica", "Manipulación de imagen digital",
            "El dispositivo móvil como medio fotográfico"
        ],
        "saberes_procedimentales": [
            "Uso correcto de la cámara fotográfica", "Manejo del diafragma, obturador e ISO",
            "Evaluación del ambiente a fotografiar", "Fotografía como medio de expresión",
            "Realización de producciones fotográficas", "Planeación de proyectos fotográficos comunicativos"
        ],
        "saberes_actitudinales": [
            "Actitud positiva", "Fomento a la creatividad", "Honestidad",
            "Responsabilidad", "Puntualidad", "Emprendedurismo"
        ],
        "indicadores_logro": [
            "Explica antecedentes y evolución de la fotografía",
            "Realiza fotografías técnicamente correctas",
            "Manipula la cámara de manera correcta",
            "Identifica mecanismos según tipo de iluminación",
            "Valora la fotografía como expresión artística y comunicación",
            "Maneja la profundidad de campo",
            "Distingue el uso de distintos objetivos",
            "Evalúa la iluminación del ambiente",
            "Domina técnicas y leyes fotográficas",
            "Utiliza el dispositivo móvil como medio fotográfico"
        ],
        "evidencias": "Exposición fotográfica",
        "horas": 4
    },
    "Lenguaje_Visual": {
        "competencia": "AA.1.1 — Competencia Animación Artística",
        "descripcion": "Comunica a través del lenguaje visual las características de la personalidad conceptualizada utilizando medios tradicionales y digitales de dibujo, textura, color y profundidad.",
        "saberes_conceptuales": [
            "Introducción a las artes visuales y artesanías",
            "Formas, estructuras, colores, texturas", "Figura-fondo, composición, luz y sombras",
            "Tamaños, escalas, equilibrios, unidad, variedad espacial",
            "Aspectos connotativos y denotativos de imágenes",
            "Diseño de personajes", "Model sheet", "Uso de tableta gráfica y Photoshop"
        ],
        "saberes_procedimentales": [
            "Argumenta el rol del lenguaje visual en la historia",
            "Identifica elementos del lenguaje visual",
            "Analiza imágenes de distintos estilos artísticos",
            "Elabora obras artísticas visuales y artesanales",
            "Diseña personajes con formas geométricas básicas",
            "Digitaliza personajes con color, texturas y profundidad"
        ],
        "saberes_actitudinales": [
            "Creatividad", "Respeto a la diversidad artística",
            "Trabajo colaborativo", "Identidad cultural dominicana"
        ],
        "indicadores_logro": [
            "Expone orígenes y funciones de las artes visuales",
            "Identifica elementos del lenguaje visual: forma, color, textura",
            "Analiza aspectos connotativos y denotativos de imágenes",
            "Elabora obras artísticas visuales con técnicas básicas",
            "Describe características de personajes existentes",
            "Diseña estructura interna de personajes",
            "Digitaliza personajes con tableta gráfica en Photoshop"
        ],
        "evidencias": "Portafolio de obras / Proyecto final de digitalización de personaje",
        "horas": 5
    },
    "Diseño": {
        "competencia": "DCM.1.1 — Diseño Básico y Expresión Visual",
        "descripcion": "Emplea atributos físicos y visuales de la forma en la creación de mensajes gráficos, aplicando modelos de resolución de problemas.",
        "saberes_conceptuales": [
            "Concepto y clasificación de diseño", "Función del diseño",
            "El proceso de diseño como método de solución de problemas",
            "Áreas del diseño gráfico", "Elementos conceptuales del diseño",
            "Punto, línea, plano, color, textura, tamaño, proporción",
            "Leyes de la Gestalt sobre percepción visual", "Señalética y sus aplicaciones"
        ],
        "saberes_procedimentales": [
            "Utiliza modelo de solución de problemas para producir formas creativas",
            "Crea comunicaciones efectivas por medio de figuras y formas",
            "Expresa ideas a través del dibujo",
            "Elabora propuestas de proyecto de comunicación gráfica"
        ],
        "saberes_actitudinales": [
            "Observación y análisis de formas", "Planificación y conceptualización",
            "Trabajo en equipo", "Creatividad", "Emprendedurismo"
        ],
        "indicadores_logro": [
            "Identifica conceptos básicos de diseño",
            "Comunica de forma gráfica mensajes e ideas",
            "Utiliza modelo de solución de problemas en proyecto de diseño",
            "Crea señales y mensajes gráficos efectivos",
            "Aplica elementos visuales: punto, línea, plano, color",
            "Reconoce leyes de la Gestalt en diseños"
        ],
        "evidencias": "Portafolio de proyectos de diseño / Creación de logo o señalética",
        "horas": 4
    }
}


def construir_prompt_planificacion(materia, grado, tema, duracion_clases, nivel_grupo):
    """Genera un prompt para planificación de clase basado en el currículo MINERD."""
    curr = CURRICULUM_MULTIMEDIA.get(materia, {})

    saberes_c = "\n".join([f"  • {s}" for s in curr.get("saberes_conceptuales", [])[:6]])
    saberes_p = "\n".join([f"  • {s}" for s in curr.get("saberes_procedimentales", [])[:4]])
    indicadores = "\n".join([f"  • {s}" for s in curr.get("indicadores_logro", [])[:5]])

    return f"""Eres un experto en planificación curricular para la Modalidad de Artes, mención Multimedia del bachillerato dominicano (MINERD).

Genera una planificación de clases completa y detallada basada en el currículo oficial.

═══════════════════════════════════════
DATOS DE LA PLANIFICACIÓN
═══════════════════════════════════════
Asignatura: {materia.replace('_', ' ')}
Competencia oficial: {curr.get('competencia', '')}
Grado: {grado}
Tema específico a trabajar: {tema}
Duración: {duracion_clases} clases de 45 minutos
Nivel del grupo: {nivel_grupo}

SABERES CONCEPTUALES DEL CURRÍCULO OFICIAL:
{saberes_c}

SABERES PROCEDIMENTALES:
{saberes_p}

INDICADORES DE LOGRO OFICIALES:
{indicadores}

Evidencia de desempeño esperada: {curr.get('evidencias', '')}
═══════════════════════════════════════

Genera la planificación con estas secciones exactas:

1. COMPETENCIAS A DESARROLLAR
Menciona las competencias fundamentales y específicas del MINERD que aplican.

2. OBJETIVO DE APRENDIZAJE
Una oración clara y medible usando verbos de acción.

3. MOTIVACIÓN INICIAL (10 min)
Actividad creativa y atractiva para activar saberes previos.

4. DESARROLLO DE LA CLASE
Describe {duracion_clases} momentos de clase con:
- Actividad principal
- Recursos necesarios (materiales, tecnología disponible en RD)
- Tiempo estimado
- Rol del docente y del estudiante

5. ACTIVIDAD PRÁCTICA
Proyecto concreto alineado a la evidencia de desempeño oficial.

6. EVALUACIÓN
- Técnica de evaluación (según el MINERD)
- Criterios observables
- Instrumento sugerido (rúbrica, portafolio, lista de cotejo)

7. RECURSOS Y MATERIALES
Lista de recursos accesibles en centros educativos dominicanos.

8. INTEGRACIÓN CON OTRAS ASIGNATURAS
Cómo conectar este tema con las otras materias de la mención Multimedia.

Usa un lenguaje práctico, motivador y accesible para docentes de bachillerato dominicano.
Máximo 600 palabras."""


def construir_prompt_rubrica(materia, indicador, nivel):
    """Genera un prompt para crear una rúbrica de evaluación oficial."""
    curr = CURRICULUM_MULTIMEDIA.get(materia, {})
    todos_indicadores = curr.get("indicadores_logro", [])

    return f"""Eres un experto en evaluación educativa para la Modalidad de Artes Multimedia del MINERD de República Dominicana.

Crea una rúbrica de evaluación completa y lista para usar en el aula.

═══════════════════════════════════════
DATOS DE LA RÚBRICA
═══════════════════════════════════════
Asignatura: {materia.replace('_', ' ')}
Competencia: {curr.get('competencia', '')}
Indicador de logro a evaluar: {indicador}
Grado: {nivel}
Evidencia de desempeño: {curr.get('evidencias', '')}

TODOS LOS INDICADORES DE ESTA ASIGNATURA:
{chr(10).join([f'• {i}' for i in todos_indicadores])}
═══════════════════════════════════════

Genera una rúbrica con EXACTAMENTE este formato:

RÚBRICA DE EVALUACIÓN
Asignatura: {materia.replace('_', ' ')} | Indicador: {indicador}

TABLA DE CRITERIOS (4 niveles: Excelente/Bueno/En desarrollo/Inicio):
Crea entre 4 y 5 criterios de evaluación observables, con descripción para cada nivel.
Asigna un peso porcentual a cada criterio que sume 100%.

ESCALA DE CALIFICACIÓN:
Excelente (90-100): descripción
Bueno (75-89): descripción
En desarrollo (60-74): descripción
Inicio (0-59): descripción

INDICACIONES PARA EL DOCENTE:
Máximo 3 recomendaciones prácticas para aplicar esta rúbrica.

Usa lenguaje claro, criterios observables y medibles.
Máximo 400 palabras."""


def construir_prompt_estrategia(materia, problema, perfil_grupo):
    """Genera estrategias didácticas para superar dificultades específicas."""
    curr = CURRICULUM_MULTIMEDIA.get(materia, {})

    return f"""Eres un orientador pedagógico especializado en la Modalidad de Artes Multimedia del bachillerato dominicano.

Un docente necesita estrategias didácticas para superar una dificultad específica en el aula.

═══════════════════════════════════════
SITUACIÓN
═══════════════════════════════════════
Asignatura: {materia.replace('_', ' ')}
Competencia: {curr.get('competencia', '')}
Dificultad o problema identificado: {problema}
Perfil del grupo: {perfil_grupo}
═══════════════════════════════════════

Genera una respuesta con estas 3 secciones:

1. DIAGNÓSTICO DE LA DIFICULTAD
Analiza por qué puede estar ocurriendo este problema en el contexto de la Modalidad de Artes dominicana.

2. ESTRATEGIAS DIDÁCTICAS (3 estrategias concretas)
Para cada estrategia indica:
- Nombre de la estrategia
- Cómo implementarla paso a paso
- Materiales necesarios (accesibles en RD)
- Tiempo estimado

3. SEGUIMIENTO Y VERIFICACIÓN
Cómo saber si la estrategia funcionó, con indicadores observables.

Sé específico y práctico para el contexto dominicano. Máximo 400 palabras."""


@app.route("/planificacion")
def planificacion():
    """Dashboard del Asistente de Planificación Docente."""
    return render_template("planificacion.html", current_user=get_usuario())


@app.route("/api/planificacion/generar", methods=["POST"])
def generar_planificacion():
    """Genera una planificación de clase con LLaMA 3."""
    data = request.json
    materia    = data.get("materia", "")
    grado      = data.get("grado", "4to")
    tema       = data.get("tema", "")
    duracion   = data.get("duracion", 2)
    nivel      = data.get("nivel_grupo", "Intermedio")

    if not materia or not tema:
        return jsonify({"error": "Materia y tema son requeridos"}), 400

    try:
        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Eres un experto en planificación curricular para la Modalidad de Artes "
                        "del bachillerato dominicano (C.E. Benito Juárez). "
                        "Respondes siempre en español con planificaciones prácticas, "
                        "creativas y alineadas al currículo oficial del MINERD."
                    )
                },
                {
                    "role": "user",
                    "content": construir_prompt_planificacion(materia, grado, tema, duracion, nivel)
                }
            ],
            temperature=0.7,
            max_tokens=1200,
            top_p=0.9
        )
        resultado = completion.choices[0].message.content.strip()
        return jsonify({"resultado": resultado, "tipo": "planificacion"})

    except Exception as ex:
        return jsonify({"error": f"Error Groq: {str(ex)}"}), 500


@app.route("/api/planificacion/rubrica", methods=["POST"])
def generar_rubrica():
    """Genera una rúbrica de evaluación alineada al currículo MINERD."""
    data = request.json
    materia    = data.get("materia", "")
    indicador  = data.get("indicador", "")
    nivel      = data.get("nivel", "4to")

    if not materia or not indicador:
        return jsonify({"error": "Materia e indicador son requeridos"}), 400

    try:
        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Eres un experto en evaluación educativa para la Modalidad de Artes "
                        "del bachillerato dominicano. Creas rúbricas prácticas y alineadas al MINERD."
                    )
                },
                {
                    "role": "user",
                    "content": construir_prompt_rubrica(materia, indicador, nivel)
                }
            ],
            temperature=0.5,
            max_tokens=800,
            top_p=0.9
        )
        resultado = completion.choices[0].message.content.strip()
        return jsonify({"resultado": resultado, "tipo": "rubrica"})

    except Exception as ex:
        return jsonify({"error": f"Error Groq: {str(ex)}"}), 500


@app.route("/api/planificacion/estrategia", methods=["POST"])
def generar_estrategia():
    """Genera estrategias didácticas para superar dificultades específicas."""
    data = request.json
    materia = data.get("materia", "")
    problema = data.get("problema", "")
    perfil   = data.get("perfil_grupo", "")

    if not materia or not problema:
        return jsonify({"error": "Materia y problema son requeridos"}), 400

    try:
        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Eres un orientador pedagógico experto en la Modalidad de Artes "
                        "del bachillerato dominicano. Ofreces estrategias didácticas prácticas "
                        "y contextualizadas para el sistema educativo de RD."
                    )
                },
                {
                    "role": "user",
                    "content": construir_prompt_estrategia(materia, problema, perfil)
                }
            ],
            temperature=0.7,
            max_tokens=700,
            top_p=0.9
        )
        resultado = completion.choices[0].message.content.strip()
        return jsonify({"resultado": resultado, "tipo": "estrategia"})

    except Exception as ex:
        return jsonify({"error": f"Error Groq: {str(ex)}"}), 500


@app.route("/api/planificacion/guardar", methods=["POST"])
@login_required
def guardar_planificacion():
    """Guarda una planificación generada en el historial."""
    u = get_usuario()
    d = request.get_json(silent=True) or {}
    contenido = (d.get("contenido") or "").strip()
    materia   = (d.get("materia") or "").strip()
    grado     = (d.get("grado") or "").strip()
    tema      = (d.get("tema") or "").strip()
    if not contenido:
        return jsonify({"error": "Sin contenido para guardar"}), 400
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            INSERT INTO historial_planificaciones
                (profesor_id, materia, grado, tema, nivel_grupo, contenido,
                 estudiante_id, nombre_estudiante, fecha)
            VALUES (?,?,?,?,?,?,?,?,date('now'))
        """, (
            u["id"], materia, grado, tema,
            d.get("nivel_grupo",""),
            contenido,
            d.get("estudiante_id"),
            d.get("nombre_estudiante","")
        ))
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"ok": True, "id": rid})


@app.route("/api/planificacion/historial", methods=["GET"])
@login_required
def historial_planificaciones():
    """Devuelve el historial de planificaciones del usuario."""
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, materia, grado, tema, nivel_grupo,
                   nombre_estudiante, fecha,
                   substr(contenido,1,120) as preview
            FROM historial_planificaciones
            WHERE profesor_id=?
            ORDER BY fecha DESC, id DESC
            LIMIT 50
        """, (u["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/planificacion/historial/<int:pid>", methods=["GET"])
@login_required
def get_planificacion(pid):
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM historial_planificaciones WHERE id=? AND profesor_id=?",
            (pid, u["id"])
        ).fetchone()
    if not row:
        return jsonify({"error": "No encontrado"}), 404
    return jsonify(dict(row))


@app.route("/api/planificacion/historial/<int:pid>", methods=["DELETE"])
@login_required
def eliminar_planificacion(pid):
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "DELETE FROM historial_planificaciones WHERE id=? AND profesor_id=?",
            (pid, u["id"])
        )
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/planificacion/curriculo/<materia>")
def obtener_curriculo(materia):
    """Devuelve los datos curriculares oficiales de una materia."""
    datos = CURRICULUM_MULTIMEDIA.get(materia, {
        "nombre": materia.replace("_", " "),
        "descripcion": f"Asignatura: {materia.replace('_', ' ')}",
        "indicadores": [],
        "temas": [],
        "horas_semana": 0
    })
    return jsonify(datos)

# ── MIGRACIÓN AUTOMÁTICA DE BASE DE DATOS ───────────────────────────────────
#
# CÓMO AGREGAR UNA FUNCIÓN NUEVA SIN BORRAR LA BASE DE DATOS:
#
#   1. Nueva columna en 'estudiantes':
#      → Agrégala a COLUMNAS_ESTUDIANTES abajo. Flask la crea al reiniciar.
#
#   2. Nueva tabla completa:
#      → Agrégala a TABLAS_NUEVAS abajo con su CREATE TABLE IF NOT EXISTS.
#
#   3. Cambio de tipo o renombrar columna (SQLite no lo permite directo):
#      → Agrega la entrada a MIGRACIONES_ESPECIALES con tu lógica SQL.
#
#   Nunca borres database.db. Solo reinicia Flask y listo.
# ────────────────────────────────────────────────────────────────────────────

# ── 1. COLUMNAS de la tabla estudiantes ──────────────────────────────────────
# Formato: (nombre_columna, tipo_sql, valor_default)
COLUMNAS_ESTUDIANTES = [
    # — Identificación —
    ("id_evaluacion",     "TEXT",    "''"),
    ("cedula",            "TEXT",    "''"),
    ("grado",             "TEXT",    "'4to'"),
    ("condicion",         "TEXT",    "'ACTIVO'"),
    ("edad",              "REAL",    0),
    # — Comportamiento (indicadores académicos) —
    ("puntualidad",       "REAL",    0),
    ("tareas",            "REAL",    0),
    ("participacion",     "REAL",    0),
    ("comprension",       "REAL",    0),
    ("rendimiento",       "REAL",    0),
    # — Fotografía por período —
    ("fotografia_p1",     "REAL",    0), ("fotografia_p2", "REAL", 0),
    ("fotografia_p3",     "REAL",    0), ("fotografia_p4", "REAL", 0),
    # — Lenguaje Visual por período —
    ("lv_p1",             "REAL",    0), ("lv_p2",         "REAL", 0),
    ("lv_p3",             "REAL",    0), ("lv_p4",         "REAL", 0),
    # — Diseño por período —
    ("diseno_p1",         "REAL",    0), ("diseno_p2",     "REAL", 0),
    ("diseno_p3",         "REAL",    0), ("diseno_p4",     "REAL", 0),
    # — Asistencia por período —
    ("asistencia_p1",     "REAL",    0), ("asistencia_p2", "REAL", 0),
    ("asistencia_p3",     "REAL",    0), ("asistencia_p4", "REAL", 0),
    # — Promedios académicos por período —
    ("acad_p1",           "REAL",    0), ("acad_p2",       "REAL", 0),
    ("acad_p3",           "REAL",    0), ("acad_p4",       "REAL", 0),
    # — Promedios finales —
    ("p_acad",            "REAL",    0),
    ("p_cond",            "REAL",    0),
    ("p_auto",            "REAL",    0),
    ("p_emocional",       "REAL",    0),
    ("p_foto",            "REAL",    0),
    ("p_lv",              "REAL",    0),
    ("p_diseno",          "REAL",    0),
    ("prom_modulos",      "REAL",    0),
    ("asistencia",        "REAL",    0),
    # — Conductuales —
    ("interrupciones",    "REAL",    0),
    ("uso_celular",       "TEXT",    "''"),
    ("conflictos",        "REAL",    0),
    ("desafia_autoridad", "REAL",    0),
    ("distraccion",       "REAL",    0),
    ("falta_respeto",     "REAL",    0),
    # — Emocionales —
    ("motivacion",        "REAL",    0),
    ("estado_emocional",  "REAL",    0),
    ("interes_futuro",    "REAL",    0),
    ("apoyo_familiar",    "REAL",    0),
    # — Análisis y riesgo —
    ("indice_riesgo",     "REAL",    0),
    ("nivel_riesgo",      "TEXT",    "''"),
    ("categoria",         "TEXT",    "''"),
    ("reporte",           "TEXT",    "''"),
    ("color",             "TEXT",    "''"),
    ("silencioso",        "INTEGER", 0),
    ("proyeccion",        "REAL",    0),
    ("tendencia",         "TEXT",    "''"),
    ("ia_analisis",       "TEXT",    "NULL"),
    ("cluster_id",        "INTEGER", "NULL"),
    ("cluster_label",     "TEXT",    "NULL"),
    ("cluster_color",     "TEXT",    "NULL"),
    ("cluster_score",     "REAL",    "NULL"),
    # — Foto de perfil —
    ("foto_path",         "TEXT",    "NULL"),
    # — Notas del coordinador —
    ("notas_coord",       "TEXT",    "NULL"),
    # — Ciclo y sección (primer ciclo 1ro-3ro, segundo ciclo 4to-6to) —
    ("ciclo",             "TEXT",    "'segundo_ciclo'"),
    ("seccion",           "TEXT",    "'A'"),
    # ← AGREGA COLUMNAS NUEVAS AQUÍ — se crean solas al reiniciar Flask
]

# ── 2. TABLAS NUEVAS (se crean si no existen) ────────────────────────────────
# Agrega el CREATE TABLE IF NOT EXISTS de cada tabla nueva aquí.
TABLAS_NUEVAS = [
    # — Registro oficial del liceo (identidad por cédula) —
    """
    CREATE TABLE IF NOT EXISTS registro_liceo (
        cedula          TEXT PRIMARY KEY,
        nombre          TEXT NOT NULL,
        apellido        TEXT NOT NULL,
        sexo            TEXT,
        nacimiento      TEXT,
        edad            INTEGER,
        telefono        TEXT,
        grado           TEXT,
        mencion         TEXT,
        es_provisional  INTEGER DEFAULT 0,
        fecha_carga     TEXT DEFAULT (date('now'))
    )
    """,
    # — Historial de planificaciones docentes —
    """
    CREATE TABLE IF NOT EXISTS historial_planificaciones (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha           TEXT DEFAULT (date('now')),
        materia         TEXT,
        periodo         TEXT,
        contenido       TEXT,
        estudiante_id   INTEGER,
        tipo            TEXT
    )
    """,
    # — Calificaciones dinámicas por materia (cualquier materia, cualquier grado) —
    """
    CREATE TABLE IF NOT EXISTS materias_calificaciones (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id INTEGER NOT NULL,
        cedula        TEXT,
        materia       TEXT NOT NULL,
        grado         TEXT,
        p1            REAL DEFAULT 0,
        p2            REAL DEFAULT 0,
        p3            REAL DEFAULT 0,
        p4            REAL DEFAULT 0,
        promedio      REAL DEFAULT 0,
        tipo          TEXT DEFAULT 'académico',
        ciclo         TEXT DEFAULT 'segundo_ciclo',
        fuente        TEXT,
        profesor      TEXT,
        fecha_carga   TEXT DEFAULT (date('now')),
        UNIQUE(estudiante_id, materia)
    )
    """,
    # — Mapeos de columnas recordados por archivo/maestro —
    """
    CREATE TABLE IF NOT EXISTS mapeos_excel (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre_archivo  TEXT,
        materia         TEXT,
        col_nombre      TEXT,
        col_apellido    TEXT,
        col_nombre_completo TEXT,
        col_p1          TEXT,
        col_p2          TEXT,
        col_p3          TEXT,
        col_p4          TEXT,
        fecha_uso   TEXT DEFAULT (date('now')),
        UNIQUE(nombre_archivo)
    )
    """,
    # ── REPORTES (conducta, psicológico, académico, incidente) ──────────────
    """
    CREATE TABLE IF NOT EXISTS reportes (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id   INTEGER NOT NULL,
        tipo            TEXT NOT NULL,
        subtipo         TEXT,
        titulo          TEXT,
        descripcion     TEXT NOT NULL,
        severidad       TEXT DEFAULT 'Media',
        reportado_por   TEXT,
        fecha           TEXT DEFAULT (date('now')),
        estado          TEXT DEFAULT 'Abierto',
        seguimiento     TEXT,
        fecha_cierre    TEXT,
        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id)
    )
    """,
    # ── PATRONES ML ──────────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS ml_clusters (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_calculo   TEXT DEFAULT (datetime('now')),
        n_clusters      INTEGER,
        features_usadas TEXT,
        resumen         TEXT
    )
    """,
    # ── USUARIOS (login) ─────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS usuarios (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        username    TEXT UNIQUE NOT NULL,
        password    TEXT NOT NULL,
        nombre      TEXT NOT NULL,
        rol         TEXT NOT NULL DEFAULT 'profesor',
        materia     TEXT,
        grado       TEXT,
        mencion     TEXT,
        asignaturas TEXT,
        activo      INTEGER DEFAULT 1,
        creado      TEXT DEFAULT (date('now'))
    )
    """,
    # ── ASISTENCIA ──────────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS asistencia (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id   INTEGER NOT NULL,
        profesor_id     INTEGER NOT NULL,
        materia         TEXT NOT NULL,
        fecha           TEXT NOT NULL,
        periodo         INTEGER NOT NULL DEFAULT 1,
        estado          TEXT NOT NULL DEFAULT 'presente',
        horas_clase     INTEGER NOT NULL DEFAULT 1,
        observacion     TEXT,
        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id),
        FOREIGN KEY (profesor_id) REFERENCES usuarios(id)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS calificaciones_periodo (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id   INTEGER NOT NULL,
        profesor_id     INTEGER NOT NULL,
        materia         TEXT NOT NULL,
        periodo         TEXT NOT NULL,
        calificacion    REAL NOT NULL,
        anio_escolar    TEXT NOT NULL,
        observacion     TEXT,
        creado          TEXT DEFAULT (datetime('now')),
        actualizado     TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id),
        FOREIGN KEY (profesor_id)   REFERENCES usuarios(id),
        UNIQUE (estudiante_id, materia, periodo, anio_escolar)
    )
    """,
    # ← AGREGA CREATE TABLE IF NOT EXISTS DE TABLAS NUEVAS AQUÍ
    # ── CUADERNO ANECDÓTICO ──────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS cuaderno_anecdotico (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id       INTEGER NOT NULL,
        autor_id            INTEGER NOT NULL,
        fecha               TEXT NOT NULL DEFAULT (date('now')),
        tipo                TEXT NOT NULL DEFAULT 'conductual',
        descripcion         TEXT NOT NULL,
        seguimiento         TEXT,
        convertido_reporte  INTEGER DEFAULT 0,
        visible_en_perfil   INTEGER DEFAULT 1,
        privado             INTEGER DEFAULT 0,
        creado_en           TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id),
        FOREIGN KEY (autor_id)      REFERENCES usuarios(id)
    )
    """,
    # ── CALENDARIO ESCOLAR ───────────────────────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS calendario_escolar (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha       TEXT NOT NULL UNIQUE,
        tipo        TEXT NOT NULL DEFAULT 'feriado',
        descripcion TEXT,
        anio_escolar TEXT DEFAULT '2025-2026',
        creado_por  INTEGER,
        creado_en   TEXT DEFAULT (datetime('now'))
    )
    """,
    # ── ASISTENCIA MENSUAL (resumen validado por el maestro) ─────────────────
    """
    CREATE TABLE IF NOT EXISTS asistencia_mensual (
        id                      INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id           INTEGER NOT NULL,
        profesor_id             INTEGER NOT NULL,
        materia                 TEXT NOT NULL,
        mes                     INTEGER NOT NULL,
        anio                    INTEGER NOT NULL,
        dias_habiles            INTEGER DEFAULT 0,
        dias_clase_impartidos   INTEGER DEFAULT 0,
        dias_asistio            INTEGER DEFAULT 0,
        porcentaje              REAL DEFAULT 0,
        validado                INTEGER DEFAULT 0,
        fecha_validacion        TEXT,
        observacion             TEXT,
        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id),
        FOREIGN KEY (profesor_id)   REFERENCES usuarios(id),
        UNIQUE (estudiante_id, profesor_id, materia, mes, anio)
    )
    """,
    # ── EVALUACIONES NARRATIVAS DEL PROFESOR ─────────────────────────────────
    """
    CREATE TABLE IF NOT EXISTS evaluaciones_narrativas (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        estudiante_id   INTEGER NOT NULL,
        profesor_id     INTEGER NOT NULL,
        periodo         INTEGER NOT NULL DEFAULT 1,
        anio_escolar    TEXT DEFAULT '2025-2026',
        texto           TEXT NOT NULL,
        creado_en       TEXT DEFAULT (datetime('now')),
        actualizado_en  TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id),
        FOREIGN KEY (profesor_id)   REFERENCES usuarios(id),
        UNIQUE (estudiante_id, profesor_id, periodo, anio_escolar)
    )
    """,
]

# ── 3. MIGRACIONES ESPECIALES (renombrar, convertir tipo, etc.) ───────────────
# Cada entrada: ("descripcion", "SQL a ejecutar")
# Solo se ejecuta si la columna/condición indicada no existe aún.
# Formato: ("descripcion", "sql", "verificar_col_ausente", "en_tabla")
MIGRACIONES_ESPECIALES = [
    ("Agregar columna profesor a materias_calificaciones",
     "ALTER TABLE materias_calificaciones ADD COLUMN profesor TEXT",
     "profesor", "materias_calificaciones"),
    ("Agregar columna grado a usuarios",
     "ALTER TABLE usuarios ADD COLUMN grado TEXT",
     "grado", "usuarios"),
    ("Agregar columna mencion a usuarios",
     "ALTER TABLE usuarios ADD COLUMN mencion TEXT",
     "mencion", "usuarios"),
    ("Agregar columna asignaturas a usuarios",
     "ALTER TABLE usuarios ADD COLUMN asignaturas TEXT",
     "asignaturas", "usuarios"),
    ("Agregar columna tipo a materias_calificaciones",
     "ALTER TABLE materias_calificaciones ADD COLUMN tipo TEXT DEFAULT 'académico'",
     "tipo", "materias_calificaciones"),
    ("Agregar columna ciclo a materias_calificaciones",
     "ALTER TABLE materias_calificaciones ADD COLUMN ciclo TEXT DEFAULT 'segundo_ciclo'",
     "ciclo", "materias_calificaciones"),
    ("Agregar columna ciclo a usuarios",
     "ALTER TABLE usuarios ADD COLUMN ciclo TEXT",
     "ciclo", "usuarios"),
    ("Agregar columna nombre_estudiante a historial_planificaciones",
     "ALTER TABLE historial_planificaciones ADD COLUMN nombre_estudiante TEXT DEFAULT ''",
     "nombre_estudiante", "historial_planificaciones"),
    ("Agregar columna grado a historial_planificaciones",
     "ALTER TABLE historial_planificaciones ADD COLUMN grado TEXT DEFAULT ''",
     "grado", "historial_planificaciones"),
    # ← AGREGA MIGRACIONES PUNTUALES AQUÍ
]


def migrar_bd():
    """
    Ejecuta toda la migración al arrancar Flask.
    - Crea tablas que no existen
    - Agrega columnas nuevas a tablas existentes
    - Ejecuta migraciones especiales puntuales
    Nunca borra datos. Siempre seguro de correr.
    """
    print("  ── Migrando base de datos...")

    with sqlite3.connect(DATABASE) as conn:

        # ── Paso 1: Crear tablas nuevas ───────────────────────────────
        for sql in TABLAS_NUEVAS:
            try:
                conn.execute(sql)
            except Exception as ex:
                print(f"  ⚠ Error creando tabla: {ex}")
        conn.commit()

        # ── Paso 2: Crear tabla 'estudiantes' si no existe ─────────
        # Construimos el CREATE TABLE dinámicamente desde COLUMNAS_ESTUDIANTES
        # para que siempre esté sincronizado con la lista maestra.
        cols_sql = ",\n                    ".join(
            f"{col} {tipo} DEFAULT {default}"
            for col, tipo, default in COLUMNAS_ESTUDIANTES
        )
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS estudiantes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre      TEXT,
                apellido    TEXT,
                curso       TEXT,
                {cols_sql}
            )
        """)
        conn.commit()

        # ── Paso 2b: Agregar columnas nuevas si la tabla ya existía ──
        try:
            cols_existentes = {
                row[1] for row in
                conn.execute("PRAGMA table_info(estudiantes)").fetchall()
            }
            # Columnas base que ya están en el CREATE TABLE
            cols_base = {'id', 'nombre', 'apellido', 'curso'}
            agregadas = []
            for col, tipo, default in COLUMNAS_ESTUDIANTES:
                if col not in cols_existentes and col not in cols_base:
                    try:
                        conn.execute(
                            f"ALTER TABLE estudiantes ADD COLUMN {col} {tipo} DEFAULT {default}"
                        )
                        agregadas.append(col)
                    except Exception as ex:
                        print(f"  ⚠ No se pudo agregar '{col}': {ex}")
            conn.commit()
            if agregadas:
                print(f"  ✓ Columnas nuevas agregadas: {', '.join(agregadas)}")
            else:
                print("  ✓ Esquema estudiantes — sin cambios")
        except Exception as ex:
            print(f"  ⚠ Error verificando columnas: {ex}")

        # ── Paso 3: Migraciones especiales ───────────────────────────
        for desc, sql, verificar_col, en_tabla in MIGRACIONES_ESPECIALES:
            try:
                cols = {
                    row[1] for row in
                    conn.execute(f"PRAGMA table_info({en_tabla})").fetchall()
                }
                if verificar_col in cols:
                    # Columna ya existe, saltar
                    continue
                conn.execute(sql)
                conn.commit()
                print(f"  ✓ Migración especial ejecutada: {desc}")
            except Exception as ex:
                print(f"  ⚠ Error en migración '{desc}': {ex}")

    print("  ── Migración completada ✓")


# Ejecutar migración al arrancar
migrar_bd()

# ── EVALUACIÓN POR COMPETENCIAS ──────────────────────────────────────────────

# ── Manejadores de error globales — siempre devuelven JSON ───────────────────
@app.errorhandler(404)
def error_404(e):
    return jsonify({"error": f"Ruta no encontrada: {request.path}"}), 404

@app.errorhandler(500)
def error_500(e):
    return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

@app.errorhandler(Exception)
def error_excepcion(e):
    import traceback
    return jsonify({"error": str(e), "detalle": traceback.format_exc()[-300:]}), 500


@app.route("/api/competencias-ia/<int:id>", methods=["POST"])
def evaluar_competencias_ia(id):
    """Genera evaluación por competencias usando datos reales de indicadores."""
    try:
        with sqlite3.connect(DATABASE) as conn:
            conn.row_factory = sqlite3.Row
            e = conn.execute("SELECT * FROM estudiantes WHERE id=?", (id,)).fetchone()

        if not e:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        e = dict(e)

        # Construir resumen de notas por período
        def pv(campo):
            v = e.get(campo)
            return round(float(v), 1) if v and float(v) > 0 else None

        bloques = []

        # Fotografía
        f_vals = [(f"P{i+1}", pv(f"fotografia_p{i+1}")) for i in range(4)]
        f_con  = [(l,v) for l,v in f_vals if v]
        if f_con:
            bloques.append("FOTOGRAFÍA: " + " | ".join(f"{l}={v}" for l,v in f_con))

        # Lenguaje Visual
        lv_vals = [(f"P{i+1}", pv(f"lv_p{i+1}")) for i in range(4)]
        lv_con  = [(l,v) for l,v in lv_vals if v]
        if lv_con:
            bloques.append("LENGUAJE VISUAL: " + " | ".join(f"{l}={v}" for l,v in lv_con))

        # Diseño
        d_vals = [(f"P{i+1}", pv(f"diseno_p{i+1}")) for i in range(4)]
        d_con  = [(l,v) for l,v in d_vals if v]
        if d_con:
            bloques.append("DISEÑO: " + " | ".join(f"{l}={v}" for l,v in d_con))

        # Asistencia
        a_vals = [(f"P{i+1}", pv(f"asistencia_p{i+1}")) for i in range(4)]
        a_con  = [(l,v) for l,v in a_vals if v]
        if a_con:
            bloques.append("ASISTENCIA: " + " | ".join(f"{l}={v}%" for l,v in a_con))

        # Promedio Académico por período
        ac_vals = [(f"P{i+1}", pv(f"acad_p{i+1}")) for i in range(4)]
        ac_con  = [(l,v) for l,v in ac_vals if v]
        if ac_con:
            bloques.append("PROM. ACADÉMICO: " + " | ".join(f"{l}={v}" for l,v in ac_con))

        # Comportamiento
        comp_items = [
            ("Puntualidad", pv("puntualidad")), ("Tareas", pv("tareas")),
            ("Participación", pv("participacion")), ("Comprensión", pv("comprension")),
            ("Rendimiento", pv("rendimiento")),
        ]
        comp_con = [(l,v) for l,v in comp_items if v]
        if comp_con:
            bloques.append("COMPORTAMIENTO ACADÉMICO: " + " | ".join(f"{l}={v}" for l,v in comp_con))

        # Emocional
        emoc_items = [
            ("Motivación", pv("motivacion")), ("Autoestima", pv("p_auto")),
            ("Estado Emocional", pv("estado_emocional")), ("Interés Futuro", pv("interes_futuro")),
            ("Apoyo Familiar", pv("apoyo_familiar")),
        ]
        emoc_con = [(l,v) for l,v in emoc_items if v]
        if emoc_con:
            bloques.append("INDICADORES EMOCIONALES: " + " | ".join(f"{l}={v}" for l,v in emoc_con))

        if not bloques:
            return jsonify({"error": "Este estudiante no tiene datos de calificaciones cargados aún."}), 400

        datos_txt = "\n".join(bloques)
        grado = e.get("grado", "4to")
        nombre_completo = f"{e.get('nombre','')} {e.get('apellido','')}".strip()

        prompt = f"""Eres un especialista en evaluación por competencias del bachillerato dominicano,
Modalidad de Artes, mención Multimedia. Genera un informe de competencias basado en datos reales.

ESTUDIANTE: {nombre_completo}
GRADO: {grado} Multimedia
CATEGORÍA: {e.get('categoria','Estable')} | Riesgo: {e.get('nivel_riesgo','Bajo')}
PROMEDIO FINAL: {round(float(e.get('p_acad') or 0), 1)} | Conductual: {round(float(e.get('p_cond') or 0), 1)} | Emocional: {round(float(e.get('p_emocional') or 0), 1)}

CALIFICACIONES:
{datos_txt}

Genera un informe con EXACTAMENTE estas 4 secciones:

1. NIVEL DE COMPETENCIA ALCANZADO
Por cada materia presente, indica el nivel real: Inicial / En desarrollo / Competente / Destacado.
Basa el nivel en los promedios reales. Menciona los datos específicos.

2. PROGRESIÓN ENTRE PERÍODOS
Analiza la evolución P1→P2 (y P3/P4 si existen). ¿Qué mejoró? ¿Qué bajó?
¿La tendencia es positiva o negativa? Sé específico con los números.

3. INDICADORES CRÍTICOS
Lista los 2-3 aspectos con notas más bajas o mayor preocupación.
Para cada uno sugiere una acción concreta y práctica para el docente.

4. PROYECCIÓN P3 Y P4
Con base en la tendencia actual, ¿qué puede lograr al final del año?
¿Qué debe priorizar el docente en los próximos períodos?

Usa lenguaje técnico-pedagógico. Sé preciso con los datos. Máximo 380 palabras."""

        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "Eres un evaluador pedagógico experto en la Modalidad de Artes del bachillerato dominicano. Respondes en español con análisis precisos basados en datos reales."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5,
            max_tokens=700
        )
        resultado = completion.choices[0].message.content.strip()
        return jsonify({"resultado": resultado})

    except Exception as ex:
        return jsonify({"error": f"Error al generar análisis: {str(ex)}"}), 500


@app.route("/api/promover/<int:id>", methods=["POST"])
def promover_estudiante(id):
    """Cambia el grado del estudiante (4to ↔ 5to)."""
    try:
        data = request.json or {}
        nuevo_grado = data.get("grado", "5to")
        with sqlite3.connect(DATABASE) as conn:
            conn.execute("UPDATE estudiantes SET grado=? WHERE id=?", (nuevo_grado, id))
            conn.commit()
        return jsonify({"ok": True, "grado": nuevo_grado})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500



# ── CARGA DE LISTADO MAESTRO DEL LICEO ───────────────────────────────────────

@app.route("/api/cargar-listado", methods=["POST"])
def cargar_listado():
    """
    Carga el Listado oficial del liceo (LISTADO-AÑO-XXXX.xlsx).
    - Lee hojas 4TO, 5TO, 6TO con todas sus menciones
    - Guarda identidad en registro_liceo (cédula, nombre, apellido, grado, mención)
    - Crea perfil vacío en 'estudiantes' para TODOS los alumnos
      (aparecen en dashboard aunque no tengan notas todavía)
    - Sin cédula → ID provisional, sistema no se detiene
    - Carga segura: re-ejecutable sin duplicar ni borrar datos
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({"error": "Formato no válido. Sube el archivo LISTADO Excel del liceo"}), 400

    # ── Filtro de perfil de profesor ─────────────────────────────────────────
    # Si el usuario es profesor, el sistema solo cargará hojas/alumnos
    # correspondientes al grado y mención asignados a su perfil.
    _prof_listado = _get_profesor()
    _prof_grado   = ""
    _prof_mencion = ""
    if _prof_listado and _prof_listado.get("rol") == "profesor":
        _prof_grado   = (_prof_listado.get("grado") or "").upper()
        _prof_mencion = (_prof_listado.get("mencion") or "").upper()

    try:
        from openpyxl import load_workbook
        from datetime import datetime as dt
        import io

        wb = load_workbook(io.BytesIO(file.read()), data_only=True)

        # Detectar hojas del listado — segundo ciclo (4to-6to) y primer ciclo (1ro-3ro)
        grados_objetivo = [s for s in wb.sheetnames
                           if any(g in s.upper() for g in
                                  ['4TO', '5TO', '6TO', '1RO', '2DO', '3RO', '1ER', '2DO', '3ER'])]

        if not grados_objetivo:
            return jsonify({"error": "No se encontraron hojas de grados reconocidos. "
                                     "Verifica que es el Listado correcto del liceo."}), 400

        # Si es profesor, filtrar solo las hojas de su grado y mención
        if _prof_grado:
            grados_filtrados = [s for s in grados_objetivo if _prof_grado in s.upper()]
            if not grados_filtrados:
                return jsonify({
                    "error": f"Esta lista no contiene hojas del grado asignado a tu perfil ({_prof_grado}). "
                             f"Solo puedes cargar listas de tu grado. Verifica el archivo."
                }), 403
            grados_objetivo = grados_filtrados
        if _prof_mencion:
            grados_filtrados = [s for s in grados_objetivo if _prof_mencion in s.upper()]
            if grados_filtrados:  # Only restrict if matching sheets exist
                grados_objetivo = grados_filtrados

        resumen = {}
        sin_cedula_lista = []

        with sqlite3.connect(DATABASE) as conn:
            conn.row_factory = sqlite3.Row

            for sheet_name in grados_objetivo:
                ws    = wb[sheet_name]
                grado = sheet_name.replace('(2)', '').replace('(3)', '').strip()
                mencion_actual = 'GENERAL'
                conteo = 0

                for row in range(1, ws.max_row + 1):
                    col1 = ws.cell(row=row, column=1).value
                    col8 = ws.cell(row=row, column=8).value

                    # ── Detectar cambio de mención ────────────────────────
                    if isinstance(col8, str) and len(col8) > 2:
                        palabras_clave = ['MUSICA', 'MÚSICA', 'MULTIMEDIA',
                                          'TEATRO', 'VISUAL', 'DANZA',
                                          'ARTES', 'BACHILLER']
                        if any(p in col8.upper() for p in palabras_clave):
                            mencion_actual = col8.strip()
                            continue

                    # ── Fila de estudiante ────────────────────────────────
                    if not (isinstance(col1, (int, float)) and 1 <= col1 <= 60):
                        continue

                    nombre   = str(ws.cell(row=row, column=2).value or '').strip()
                    apellido = str(ws.cell(row=row, column=3).value or '').strip()
                    if not nombre or nombre.startswith('='):
                        continue

                    sexo     = str(ws.cell(row=row, column=4).value or '').strip()
                    nac_raw  = ws.cell(row=row, column=5).value
                    ced_raw  = ws.cell(row=row, column=7).value
                    telefono = str(ws.cell(row=row, column=8).value or '').strip()

                    # Fecha y edad
                    nac_str   = ''
                    edad_calc = None
                    if isinstance(nac_raw, dt):
                        nac_str   = nac_raw.strftime('%d/%m/%Y')
                        edad_calc = (dt.now() - nac_raw).days // 365
                    elif nac_raw:
                        nac_str = str(nac_raw)

                    # Cédula real o provisional
                    es_provisional = 0
                    ced_str = str(ced_raw or '').replace('.0','').replace(',','').strip()
                    if ced_str.isdigit():
                        cedula = ced_str
                    else:
                        nom1 = nombre.split()[0]  if nombre  else 'X'
                        ape1 = apellido.split()[0] if apellido else 'X'
                        cedula = f"PROV_{nom1}_{ape1}"
                        es_provisional = 1
                        sin_cedula_lista.append(f"{nombre} {apellido} ({grado} {mencion_actual})")

                    # ── UPSERT en registro_liceo ──────────────────────────
                    existing = conn.execute(
                        "SELECT cedula FROM registro_liceo WHERE cedula=?", (cedula,)
                    ).fetchone()

                    if existing:
                        conn.execute("""
                            UPDATE registro_liceo
                               SET nombre=?, apellido=?, sexo=?, nacimiento=?,
                                   edad=?, telefono=?, grado=?, mencion=?, es_provisional=?
                             WHERE cedula=?
                        """, (nombre, apellido, sexo, nac_str,
                              edad_calc, telefono, grado, mencion_actual,
                              es_provisional, cedula))
                    else:
                        conn.execute("""
                            INSERT INTO registro_liceo
                                (cedula, nombre, apellido, sexo, nacimiento,
                                 edad, telefono, grado, mencion, es_provisional)
                            VALUES (?,?,?,?,?,?,?,?,?,?)
                        """, (cedula, nombre, apellido, sexo, nac_str,
                              edad_calc, telefono, grado, mencion_actual,
                              es_provisional))

                    # ── Crear perfil vacío en 'estudiantes' si no existe ──
                    # Así aparecen en el dashboard aunque no tengan notas
                    ya_existe = conn.execute(
                        """SELECT id FROM estudiantes
                           WHERE cedula=?
                              OR (lower(nombre)=lower(?) AND lower(apellido)=lower(?))
                           LIMIT 1""",
                        (cedula, nombre, apellido)
                    ).fetchone()

                    # Detectar ciclo según grado
                    _ciclo_listado = 'segundo_ciclo'
                    for _g1c in ['1RO', '2DO', '3RO', '1ER', '2ER', '3ER']:
                        if _g1c in grado.upper():
                            _ciclo_listado = 'primer_ciclo'
                            break

                    if not ya_existe:
                        conn.execute("""
                            INSERT INTO estudiantes
                                (id_evaluacion, cedula, nombre, apellido,
                                 curso, grado, condicion, edad, ciclo)
                            VALUES (?, ?, ?, ?, ?, ?, 'ACTIVO', ?, ?)
                        """, (
                            cedula,          # id_evaluacion = cédula hasta tener BJ
                            cedula,
                            nombre,
                            apellido,
                            f"{grado} {mencion_actual}",
                            grado,
                            edad_calc or 0,
                            _ciclo_listado
                        ))
                    else:
                        # Actualizar cédula si el perfil existía con BJ* o vacío
                        conn.execute("""
                            UPDATE estudiantes
                               SET cedula=?, grado=?, curso=?, edad=?
                             WHERE (cedula IS NULL OR cedula='' OR cedula LIKE 'BJ%')
                               AND lower(nombre)=lower(?)
                               AND lower(apellido)=lower(?)
                        """, (cedula, grado, f"{grado} {mencion_actual}",
                              edad_calc or 0, nombre, apellido))

                    conteo += 1

                resumen[f"{grado} {mencion_actual}"] = conteo

            conn.commit()

        # Totales
        with sqlite3.connect(DATABASE) as conn2:
            total_liceo  = conn2.execute("SELECT COUNT(*) FROM registro_liceo").fetchone()[0]
            total_perfiles = conn2.execute("SELECT COUNT(*) FROM estudiantes").fetchone()[0]
            con_notas    = conn2.execute(
                "SELECT COUNT(*) FROM estudiantes WHERE p_acad > 0 OR acad_p1 > 0"
            ).fetchone()[0]

        return jsonify({
            "ok": True,
            "resumen_por_seccion": resumen,
            "totales": {
                "en_registro_liceo":    total_liceo,
                "perfiles_en_sistema":  total_perfiles,
                "con_notas_cargadas":   con_notas,
                "solo_identidad":       total_perfiles - con_notas
            },
            "sin_cedula": sin_cedula_lista,
            "mensaje": (
                f"{total_liceo} estudiantes en el registro del liceo. "
                f"{total_perfiles} perfiles en el sistema "
                f"({con_notas} con notas, {total_perfiles - con_notas} solo identidad)."
                + (f" {len(sin_cedula_lista)} sin cédula (ID provisional)."
                   if sin_cedula_lista else "")
            )
        })

    except Exception as ex:
        import traceback
        return jsonify({
            "error": f"Error al procesar listado: {str(ex)}",
            "detalle": traceback.format_exc()
        }), 500


@app.route("/api/registro-liceo", methods=["GET"])
def get_registro_liceo():
    """Devuelve el registro de todos los estudiantes del liceo para consulta."""
    grado  = request.args.get("grado", "")
    mencion = request.args.get("mencion", "")

    query = "SELECT * FROM registro_liceo WHERE 1=1"
    params = []
    if grado:
        query += " AND upper(grado) LIKE upper(?)"
        params.append(f"%{grado}%")
    if mencion:
        query += " AND upper(mencion) LIKE upper(?)"
        params.append(f"%{mencion}%")
    query += " ORDER BY grado, mencion, apellido"

    try:
        with sqlite3.connect(DATABASE) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(query, params).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500



# ══════════════════════════════════════════════════════════════════════════════
# CARGA DINÁMICA DE MATERIAS — LLaMA interpreta cualquier Excel de maestro
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/interpretar-excel", methods=["POST"])
def interpretar_excel():
    """
    Analiza el Excel del maestro con lógica inteligente en 3 pasos:
    1. Busca zona de resumen (Periodo I/II/III/IV) → lectura directa sin IA
    2. Si no encuentra, usa LLaMA para interpretar columnas
    3. Devuelve mapeo para confirmación del usuario
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file     = request.files["file"]
    materia  = request.form.get("materia", "").strip()
    if not materia:
        return jsonify({"error": "Debes indicar el nombre de la materia"}), 400

    nombre_archivo = file.filename
    raw_bytes = file.read()

    # ── Verificar mapeo guardado ──────────────────────────────────
    with sqlite3.connect(DATABASE) as conn:
        mapeo_guardado = conn.execute(
            "SELECT * FROM mapeos_excel WHERE nombre_archivo=?", (nombre_archivo,)
        ).fetchone()
    if mapeo_guardado:
        m = dict(mapeo_guardado)
        return jsonify({
            "mapeo": {
                "col_nombre": m["col_nombre"], "col_apellido": m["col_apellido"],
                "col_nombre_completo": m["col_nombre_completo"],
                "col_p1": m["col_p1"], "col_p2": m["col_p2"],
                "col_p3": m["col_p3"], "col_p4": m["col_p4"],
                "modo": m.get("col_nombre_completo") or "openpyxl",
            },
            "materia": materia, "archivo": nombre_archivo,
            "confianza": "guardado",
            "mensaje": f"Mapeo recordado de carga anterior.",
            "columnas_disponibles": [], "muestra": [],
            "modo_lectura": "guardado"
        })

    # ── MODO 1: Lector inteligente openpyxl ──────────────────────
    # Detecta zona de resumen (Periodo I/II/III/IV) automáticamente
    try:
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(raw_bytes), data_only=True)

        hojas_validas = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            col_p1=col_p2=col_p3=col_p4=col_nom=col_ape=None
            header_row=None

            # Buscar fila con Periodo I, II, III, IV
            for row in range(1, min(20, ws.max_row+1)):
                for col in range(1, ws.max_column+1):
                    v = str(ws.cell(row=row, column=col).value or '').strip().upper()
                    v = v.replace('\xa0',' ').strip()
                    if v in ('PERIODO I','PERIODO 1') or v.endswith('PERIODO I'):
                        col_p1 = col; header_row = row
                    elif v in ('PERIODO II','PERIODO 2'):
                        col_p2 = col
                    elif v in ('PERIODO III','PERIODO 3'):
                        col_p3 = col
                    elif v in ('PERIODO IV','PERIODO 4'):
                        col_p4 = col

            if not header_row:
                continue

            # Buscar Nombres/Apellidos en zona del resumen
            for row in range(max(1,header_row-2), header_row+4):
                for col in range(max(1,(col_p1 or 1)-15), ws.max_column+1):
                    v = str(ws.cell(row=row, column=col).value or '').strip().upper()
                    if 'NOMBRE' in v and col_nom is None: col_nom = col
                    if 'APELLIDO' in v and col_ape is None: col_ape = col

            # Encontrar fila de inicio de datos
            data_start = None
            for row in range(header_row+1, ws.max_row+1):
                v = ws.cell(row=row, column=col_nom or 3).value
                if v and str(v).strip() not in ('','None','nan','Nombres','NOMBRES'):
                    # Verificar que no es otra fila de encabezado
                    if not any(kw in str(v).upper() for kw in ['NOMBRE','ALUMNO','#']):
                        data_start = row
                        break

            if col_nom and data_start:
                hojas_validas.append({
                    "hoja": sname, "header_row": header_row,
                    "data_start": data_start,
                    "col_nom": col_nom, "col_ape": col_ape,
                    "col_p1": col_p1, "col_p2": col_p2,
                    "col_p3": col_p3, "col_p4": col_p4,
                })

        if hojas_validas:
            h = hojas_validas[0]
            muestra = []
            ws0 = wb[h["hoja"]]
            for row in range(h["data_start"], min(h["data_start"]+4, ws0.max_row+1)):
                nom = ws0.cell(row=row, column=h["col_nom"]).value
                ape = ws0.cell(row=row, column=h["col_ape"]).value if h["col_ape"] else ''
                p1  = ws0.cell(row=row, column=h["col_p1"]).value if h["col_p1"] else None
                if nom:
                    muestra.append([str(nom), str(ape or ''), str(p1 or '-')])

            n_hojas = len(hojas_validas)
            return jsonify({
                "mapeo": {
                    "col_nombre":          h["col_nom"],
                    "col_apellido":        h["col_ape"],
                    "col_nombre_completo": None,
                    "col_p1": h["col_p1"], "col_p2": h["col_p2"],
                    "col_p3": h["col_p3"], "col_p4": h["col_p4"],
                    "header_row":   h["header_row"],
                    "data_start":   h["data_start"],
                    "fila_inicio_datos": h["data_start"],
                    "modo":         "openpyxl_resumen",
                    "hojas":        [x["hoja"] for x in hojas_validas],
                    "notas": f"Zona de resumen detectada automaticamente. "
                             f"{n_hojas} hoja(s): {', '.join(x['hoja'] for x in hojas_validas)}. "
                             f"Periodo I→IV mapeados directamente."
                },
                "materia": materia, "archivo": nombre_archivo,
                "confianza": "alta",
                "mensaje": f"Estructura de RAEs detectada. {n_hojas} seccion(es) encontrada(s).",
                "columnas_disponibles": ["(auto - zona resumen)"],
                "muestra": muestra,
                "modo_lectura": "openpyxl_resumen",
                "hojas_info": hojas_validas
            })

    except Exception as ex_openpyxl:
        print(f"openpyxl fallback: {ex_openpyxl}")

    # ── MODO 2: LLaMA interpreta columnas simples ─────────────────
    try:
        import pandas as pd, io as _io, json as _json
        df = pd.read_excel(_io.BytesIO(raw_bytes), header=None, nrows=12)
        mejor_fila = max(range(min(8,len(df))),
                        key=lambda i: sum(1 for v in df.iloc[i]
                                          if isinstance(v,str) and len(str(v).strip())>0))
        df2 = pd.read_excel(_io.BytesIO(raw_bytes), header=mejor_fila)
        columnas = [str(c) for c in df2.columns if str(c).strip() and str(c)!='nan']
        muestra  = df2.head(4).fillna('').astype(str).values.tolist()

        prompt = f"""Eres un asistente que analiza archivos Excel de maestros dominicanos.
Materia: {materia} | Archivo: {nombre_archivo}
Columnas: {columnas}
Primeras filas: {muestra[:3]}

Responde SOLO con JSON:
{{"col_nombre":null,"col_apellido":null,"col_nombre_completo":null,
"col_p1":null,"col_p2":null,"col_p3":null,"col_p4":null,
"fila_inicio_datos":1,"confianza":"media","notas":""}}

Reglas: P1/P2 pueden llamarse Periodo 1, Trimestre 1, Parcial 1, PROM P1, etc.
Si nombre+apellido van juntos usa col_nombre_completo."""

        completion = _get_groq_client().chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role":"system","content":"Responde SOLO con JSON válido."},
                {"role":"user","content":prompt}
            ],
            temperature=0.1, max_tokens=400,
        )
        resp = completion.choices[0].message.content.strip()
        resp = resp.replace("```json","").replace("```","").strip()
        mapeo = _json.loads(resp)
        mapeo["modo"] = "llama"

        return jsonify({
            "mapeo": mapeo, "materia": materia, "archivo": nombre_archivo,
            "confianza": mapeo.get("confianza","media"),
            "mensaje": mapeo.get("notas","LLaMA interpreto el archivo."),
            "columnas_disponibles": columnas, "muestra": muestra[:3],
            "modo_lectura": "llama"
        })

    except Exception as ex_llama:
        return jsonify({
            "mapeo": None, "materia": materia, "archivo": nombre_archivo,
            "confianza": "manual",
            "mensaje": "No se pudo interpretar automaticamente. Selecciona las columnas manualmente.",
            "columnas_disponibles": [], "muestra": [], "modo_lectura": "manual",
            "error_ia": str(ex_llama)
        })


@app.route("/api/cargar-materia", methods=["POST"])
def cargar_materia():
    """
    Carga notas de una materia usando el mapeo confirmado.
    Modo openpyxl_resumen: lee zona Periodo I/II/III/IV en todas las hojas.
    Modo pandas: lee con nombres de columnas (LLaMA o manual).
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibio archivo"}), 400

    import io as _io, json as _json

    file      = request.files["file"]
    materia   = request.form.get("materia","").strip()
    mapeo_str = request.form.get("mapeo","{}")
    raw_bytes = file.read()

    try:
        mapeo = _json.loads(mapeo_str)
    except Exception:
        return jsonify({"error": "Mapeo JSON invalido"}), 400

    if not materia:
        return jsonify({"error": "Materia requerida"}), 400

    # ── VALIDACIÓN DE PERFIL DE PROFESOR ────────────────────────────────────
    # Si el usuario autenticado es un profesor (no coordinador),
    # verificar que la materia y el curso correspondan a su perfil asignado.
    _prof_actual = _get_profesor()
    if _prof_actual and _prof_actual.get("rol") == "profesor":
        _curso_archivo = request.form.get("grado", "") + " " + request.form.get("mencion", "")
        _ok_val, _msg_val = _validar_materia_profesor(materia, _curso_archivo, _prof_actual)
        if not _ok_val:
            return jsonify({"error": _msg_val, "tipo": "validacion_perfil"}), 403

    # Nombre del profesor — desde form o extraído del nombre del archivo
    profesor_nombre = request.form.get("profesor", "").strip()
    if not profesor_nombre and file:
        import re as _re2
        fn = file.filename or ""
        fn_clean = _re2.sub(r"\.(xlsx?|xlsm)$", "", fn, flags=_re2.I)
        fn_clean = _re2.sub(r"(6to|5to|4to|[AB]_|grado|teatro|musica|multimedia|artes|diseno|diseño)",
                            " ", fn_clean, flags=_re2.I)
        fn_clean = fn_clean.replace("_", " ").replace("-", " ").strip()
        if fn_clean:
            profesor_nombre = fn_clean[:50]

    modo = mapeo.get("modo","llama")

    def limpiar_num(v):
        try:
            if v is None or str(v).strip() in ('','nan','None','-'):
                return None
            return round(float(v), 1)
        except Exception:
            return None

    registros = []

    # ── MODO openpyxl_resumen ─────────────────────────────────────
    if modo == "openpyxl_resumen":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(raw_bytes), data_only=True)
            hojas_raw  = mapeo.get("hojas", wb.sheetnames)
            hojas_usar = hojas_raw if isinstance(hojas_raw, list) else wb.sheetnames

            for sname in hojas_usar:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                col_p1=col_p2=col_p3=col_p4=col_nom=col_ape=None
                header_row=None

                for row in range(1, min(20, ws.max_row+1)):
                    for col in range(1, ws.max_column+1):
                        v = str(ws.cell(row=row,column=col).value or '').strip().upper()
                        if v in ('PERIODO I','PERIODO 1') or v.endswith('PERIODO I'):
                            col_p1=col; header_row=row
                        elif v in ('PERIODO II','PERIODO 2'): col_p2=col
                        elif v in ('PERIODO III','PERIODO 3'): col_p3=col
                        elif v in ('PERIODO IV','PERIODO 4'): col_p4=col

                if not header_row:
                    continue

                for row in range(max(1,header_row-2), header_row+4):
                    for col in range(max(1,(col_p1 or 1)-15), ws.max_column+1):
                        v = str(ws.cell(row=row,column=col).value or '').strip().upper()
                        if 'NOMBRE' in v and col_nom is None: col_nom=col
                        if 'APELLIDO' in v and col_ape is None: col_ape=col

                if not col_nom:
                    continue

                data_start = None
                for row in range(header_row+1, ws.max_row+1):
                    v = ws.cell(row=row, column=col_nom).value
                    vs = str(v or '').strip()
                    if vs and vs not in ('None','nan','Nombres','NOMBRES'):
                        if not any(kw in vs.upper() for kw in ['NOMBRE','ALUMNO','#']):
                            data_start=row; break

                if not data_start:
                    continue

                for row in range(data_start, ws.max_row+1):
                    nom = ws.cell(row=row, column=col_nom).value
                    if not nom or str(nom).strip() in ('','None','nan'):
                        continue
                    ape = ws.cell(row=row, column=col_ape).value if col_ape else ''
                    p1  = limpiar_num(ws.cell(row=row,column=col_p1).value) if col_p1 else None
                    p2  = limpiar_num(ws.cell(row=row,column=col_p2).value) if col_p2 else None
                    p3  = limpiar_num(ws.cell(row=row,column=col_p3).value) if col_p3 else None
                    p4  = limpiar_num(ws.cell(row=row,column=col_p4).value) if col_p4 else None
                    registros.append({
                        "nombre":   str(nom).strip(),
                        "apellido": str(ape or '').strip(),
                        "p1":p1,"p2":p2,"p3":p3,"p4":p4
                    })

        except Exception as ex:
            return jsonify({"error": f"Error leyendo Excel: {str(ex)}"}), 400

    # ── MODO pandas ───────────────────────────────────────────────
    else:
        import pandas as pd
        col_nombre          = mapeo.get("col_nombre")
        col_apellido        = mapeo.get("col_apellido")
        col_nombre_completo = mapeo.get("col_nombre_completo")
        col_p1 = mapeo.get("col_p1")
        col_p2 = mapeo.get("col_p2")
        col_p3 = mapeo.get("col_p3")
        col_p4 = mapeo.get("col_p4")
        fila_h = max(0, int(mapeo.get("fila_inicio_datos",1))-1)

        try:
            df = pd.read_excel(_io.BytesIO(raw_bytes), header=fila_h)
            df.columns = [str(c).strip() for c in df.columns]
        except Exception as ex:
            return jsonify({"error": f"Error leyendo Excel: {str(ex)}"}), 400

        for _, fila in df.iterrows():
            d = fila.to_dict()
            if col_nombre_completo and col_nombre_completo in d:
                nc = str(d[col_nombre_completo] or '').strip()
                partes = nc.split()
                nom = ' '.join(partes[:max(1,len(partes)//2)])
                ape = ' '.join(partes[max(1,len(partes)//2):])
            else:
                nom = str(d.get(col_nombre,'') or '').strip()
                ape = str(d.get(col_apellido,'') or '').strip()
            if not nom or nom.lower() in ('nan','none',''):
                continue
            registros.append({
                "nombre":nom,"apellido":ape,
                "p1":limpiar_num(d.get(col_p1)) if col_p1 else None,
                "p2":limpiar_num(d.get(col_p2)) if col_p2 else None,
                "p3":limpiar_num(d.get(col_p3)) if col_p3 else None,
                "p4":limpiar_num(d.get(col_p4)) if col_p4 else None,
            })

    # ── Vincular a estudiantes y guardar ─────────────────────────
    cargados = 0
    no_encontrados = []

    # Filtro de grado y mención: evita vincular "María García 4to" con "María García 5to"
    filtro_grado   = request.form.get("grado",   "").strip()
    filtro_mencion = request.form.get("mencion", "").strip()

    # Búsqueda delegada al helper de módulo _buscar_estudiante_bd
    def buscar_estudiante(conn, nom1, ape1):
        return _buscar_estudiante_bd(conn, nom1, ape1,
                                     filtro_grado=filtro_grado,
                                     filtro_mencion=filtro_mencion)
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row

        for reg in registros:
            nom1 = reg["nombre"].lower().split()[0]  if reg["nombre"]  else ''
            ape1 = reg["apellido"].lower().split()[0] if reg["apellido"] else ''
            if not nom1:
                continue

            est = _buscar_estudiante_bd(conn, reg["nombre"], reg["apellido"])

            if not est:
                no_encontrados.append(f"{reg['nombre']} {reg['apellido']}")
                continue

            notas = [x for x in [reg["p1"],reg["p2"],reg["p3"],reg["p4"]] if x is not None]
            prom  = round(sum(notas)/len(notas),1) if notas else None

            conn.execute("""
                INSERT INTO materias_calificaciones
                    (estudiante_id,cedula,materia,p1,p2,p3,p4,promedio,fuente,profesor)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(estudiante_id,materia) DO UPDATE SET
                    p1      = CASE WHEN excluded.p1 IS NOT NULL AND excluded.p1 > 0
                                   THEN excluded.p1 ELSE materias_calificaciones.p1 END,
                    p2      = CASE WHEN excluded.p2 IS NOT NULL AND excluded.p2 > 0
                                   THEN excluded.p2 ELSE materias_calificaciones.p2 END,
                    p3      = CASE WHEN excluded.p3 IS NOT NULL AND excluded.p3 > 0
                                   THEN excluded.p3 ELSE materias_calificaciones.p3 END,
                    p4      = CASE WHEN excluded.p4 IS NOT NULL AND excluded.p4 > 0
                                   THEN excluded.p4 ELSE materias_calificaciones.p4 END,
                    promedio= excluded.promedio,
                    fuente  = excluded.fuente,
                    profesor= CASE WHEN excluded.profesor IS NOT NULL AND excluded.profesor != ''
                                   THEN excluded.profesor ELSE materias_calificaciones.profesor END,
                    fecha_carga = date('now')
            """, (est["id"],est["cedula"],materia,
                  reg["p1"],reg["p2"],reg["p3"],reg["p4"],prom,
                  file.filename, profesor_nombre))
            cargados += 1

            # ── Sincronizar módulos técnicos Multimedia y asistencia ──────────
            # Si la materia coincide con un módulo de Multimedia, actualizar
            # también los campos directos en la tabla estudiantes
            mat_norm = materia.lower().strip()

            # Mapeo de nombre de materia → columna base en estudiantes
            MODULO_MAP = {
                'fotografía': 'fotografia', 'fotografia': 'fotografia',
                'foto':       'fotografia',
                'lenguaje visual': 'lv', 'lenguaje_visual': 'lv', 'lv': 'lv',
                'diseño': 'diseno', 'diseno': 'diseno', 'diseño básico': 'diseno',
                'diseño basico': 'diseno',
            }
            ASISTENCIA_KEYS = {'asistencia', 'asistencias', 'attend', 'attendance'}

            col_base = MODULO_MAP.get(mat_norm)
            es_asistencia = mat_norm in ASISTENCIA_KEYS

            if col_base or es_asistencia:
                def _v(val):
                    return val if (val is not None and val != 0) else None

                if col_base:
                    # Módulo técnico → actualizar p1-p4 y promedio del módulo
                    p_col = f'p_{col_base}' if col_base != 'fotografia' else 'p_foto'
                    notas_mod = [x for x in [reg['p1'],reg['p2'],reg['p3'],reg['p4']] if x]
                    prom_mod  = round(sum(notas_mod)/len(notas_mod),1) if notas_mod else None

                    # Build SET clause only for non-null values
                    sets, vals = [], []
                    for pi, pv in enumerate([reg['p1'],reg['p2'],reg['p3'],reg['p4']], 1):
                        if _v(pv) is not None:
                            sets.append(f"{col_base}_p{pi} = CASE WHEN {col_base}_p{pi} IS NULL OR {col_base}_p{pi}=0 THEN ? ELSE {col_base}_p{pi} END")
                            vals.append(pv)
                    if prom_mod and sets:
                        sets.append(f"{p_col} = ?")
                        vals.append(prom_mod)
                    if sets:
                        vals.append(est['id'])
                        conn.execute(
                            f"UPDATE estudiantes SET {', '.join(sets)} WHERE id=?", vals
                        )
                        # Recalculate prom_modulos
                        e2 = conn.execute(
                            "SELECT p_foto,p_lv,p_diseno FROM estudiantes WHERE id=?",
                            (est['id'],)
                        ).fetchone()
                        if e2:
                            mods = [x for x in e2 if x and x > 0]
                            pm2  = round(sum(mods)/len(mods),1) if mods else None
                            conn.execute(
                                "UPDATE estudiantes SET prom_modulos=? WHERE id=?",
                                (pm2, est['id'])
                            )

                elif es_asistencia:
                    # Asistencia → asistencia_p1..p4 y asistencia general
                    sets, vals = [], []
                    for pi, pv in enumerate([reg['p1'],reg['p2'],reg['p3'],reg['p4']], 1):
                        if _v(pv) is not None:
                            sets.append(f"asistencia_p{pi} = CASE WHEN asistencia_p{pi} IS NULL OR asistencia_p{pi}=0 THEN ? ELSE asistencia_p{pi} END")
                            vals.append(pv)
                    if sets:
                        vals.append(est['id'])
                        conn.execute(
                            f"UPDATE estudiantes SET {', '.join(sets)} WHERE id=?", vals
                        )

        conn.commit()

        # Guardar mapeo
        try:
            conn.execute("""
                INSERT INTO mapeos_excel
                    (nombre_archivo,materia,col_nombre,col_apellido,
                     col_nombre_completo,col_p1,col_p2,col_p3,col_p4)
                VALUES (?,?,?,?,?,?,?,?,?)
                ON CONFLICT(nombre_archivo) DO UPDATE SET
                    materia=excluded.materia,
                    col_nombre_completo=excluded.col_nombre_completo,
                    fecha_uso=date('now')
            """, (file.filename, materia,
                  str(mapeo.get("col_nombre","") or ""),
                  str(mapeo.get("col_apellido","") or ""),
                  str(mapeo.get("modo","llama")),
                  str(mapeo.get("col_p1","") or ""),
                  str(mapeo.get("col_p2","") or ""),
                  str(mapeo.get("col_p3","") or ""),
                  str(mapeo.get("col_p4","") or "")))
            conn.commit()
        except Exception:
            pass

    return jsonify({
        "ok":             True,
        "materia":        materia,
        "cargados":       cargados,
        "total_excel":    len(registros),
        "no_encontrados": no_encontrados,
        "mensaje": (f"{cargados} de {len(registros)} estudiantes actualizados en '{materia}'." +
                    (f" {len(no_encontrados)} no encontrados." if no_encontrados else ""))
    })


@app.route("/api/materias/<int:estudiante_id>")
def get_materias_estudiante(estudiante_id):
    """Devuelve todas las materias cargadas para un estudiante, incluyendo tipo y profesor."""
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT materia, p1, p2, p3, p4, promedio, fecha_carga, fuente,
                   COALESCE(tipo, 'académico') as tipo,
                   COALESCE(profesor, '') as profesor
            FROM materias_calificaciones
            WHERE estudiante_id = ?
            ORDER BY CASE COALESCE(tipo,'académico')
                          WHEN 'académico' THEN 0 ELSE 1 END,
                     materia
        """, (estudiante_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/materias-disponibles")
def get_materias_disponibles():
    """Lista todas las materias que han sido cargadas al sistema."""
    grado = request.args.get("grado", "")
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT DISTINCT mc.materia, COUNT(DISTINCT mc.estudiante_id) as total_estudiantes,
                   mc.fecha_carga
            FROM materias_calificaciones mc
            JOIN estudiantes e ON e.id = mc.estudiante_id
            WHERE (? = '' OR e.grado LIKE ?)
            GROUP BY mc.materia
            ORDER BY mc.materia
        """, (grado, f"%{grado}%")).fetchall()
    return jsonify([dict(r) for r in rows])

# ── Indicadores por período (para gráficas del perfil) ───────────────────────

@app.route("/api/indicadores/materias/<int:estudiante_id>")
def get_indicadores_materias(estudiante_id):
    """Lista de materias disponibles para el estudiante con promedio."""
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT materia,
                   COALESCE(p1,0) p1, COALESCE(p2,0) p2,
                   COALESCE(p3,0) p3, COALESCE(p4,0) p4,
                   COALESCE(promedio,0) promedio, fecha_carga
            FROM materias_calificaciones
            WHERE estudiante_id = ?
            ORDER BY materia
        """, (estudiante_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/indicadores/<int:estudiante_id>")
def get_indicadores(estudiante_id):
    """Indicadores de una materia específica por período."""
    materia = request.args.get("materia", "")
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("""
            SELECT materia, p1, p2, p3, p4, promedio, fecha_carga
            FROM materias_calificaciones
            WHERE estudiante_id=? AND materia=?
        """, (estudiante_id, materia)).fetchone()
    if not row:
        return jsonify({"indicadores": [], "materia": materia, "promedio": 0})
    r = dict(row)
    # Build period rows for chart
    periodos = []
    for i, (p, label) in enumerate([("p1","P1"),("p2","P2"),("p3","P3"),("p4","P4")], 1):
        val = r.get(p)
        if val and val > 0:
            periodos.append({
                "periodo": label,
                "indicador_texto": f"Período {i}",
                "p1": val if i==1 else None,
                "p2": val if i==2 else None,
                "p3": val if i==3 else None,
                "p4": val if i==4 else None,
            })
    return jsonify({
        "materia":    r["materia"],
        "promedio":   r["promedio"] or 0,
        "indicadores": periodos,
        "p1": r["p1"], "p2": r["p2"], "p3": r["p3"], "p4": r["p4"],
    })


# ══════════════════════════════════════════════════════════════════════════════
# CRUD DE PERFILES — Eliminar, Retirar, Agregar, Editar campos manuales
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/estudiante/<int:id>", methods=["DELETE"])
def eliminar_estudiante(id):
    """Elimina un perfil completo y sus materias dinámicas."""
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM materias_calificaciones WHERE estudiante_id=?", (id,))
        conn.execute("DELETE FROM estudiantes WHERE id=?", (id,))
        conn.commit()
    return jsonify({"ok": True, "mensaje": "Perfil eliminado."})


@app.route("/api/estudiante/<int:id>/condicion", methods=["POST"])
def cambiar_condicion(id):
    """Cambia la condición del estudiante: ACTIVO / RETIRADO / GRADUADO."""
    data      = request.get_json(silent=True) or {}
    condicion = str(data.get("condicion", "ACTIVO")).upper().strip()
    if condicion not in ("ACTIVO", "RETIRADO", "GRADUADO"):
        return jsonify({"error": "Condición inválida. Usa ACTIVO, RETIRADO o GRADUADO"}), 400
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("UPDATE estudiantes SET condicion=? WHERE id=?", (condicion, id))
        conn.commit()
    return jsonify({"ok": True, "condicion": condicion})


@app.route("/api/estudiante", methods=["POST"])
def agregar_estudiante():
    """Crea un perfil nuevo manualmente."""
    data = request.get_json(silent=True) or {}
    nombre   = str(data.get("nombre",   "")).strip()
    apellido = str(data.get("apellido", "")).strip()
    if not nombre or not apellido:
        return jsonify({"error": "Nombre y apellido son requeridos"}), 400

    grado   = str(data.get("grado",   "4to")).strip()
    curso   = str(data.get("curso",   "")).strip()
    cedula  = str(data.get("cedula",  "")).strip()
    edad    = data.get("edad")

    if not cedula:
        cedula = f"MANUAL_{nombre[:4].upper()}_{apellido[:4].upper()}"

    with sqlite3.connect(DATABASE) as conn:
        existing = conn.execute(
            "SELECT id FROM estudiantes WHERE cedula=? OR (lower(nombre)=lower(?) AND lower(apellido)=lower(?))",
            (cedula, nombre, apellido)
        ).fetchone()
        if existing:
            return jsonify({"error": "Ya existe un estudiante con ese nombre o cédula"}), 409

        conn.execute("""
            INSERT INTO estudiantes
                (id_evaluacion, cedula, nombre, apellido, curso, grado, condicion, edad)
            VALUES (?,?,?,?,?,?,?,?)
        """, (cedula, cedula, nombre, apellido, curso, grado, "ACTIVO",
              float(edad) if edad else None))
        conn.commit()
        nuevo_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    return jsonify({"ok": True, "id": nuevo_id,
                    "mensaje": f"Perfil de {nombre} {apellido} creado."})


@app.route("/api/estudiante/<int:id>", methods=["PATCH"])
def editar_estudiante(id):
    """
    Actualiza campos específicos de un estudiante.
    Acepta cualquier campo numérico o de texto de la tabla estudiantes.
    Usado para ingresar datos manuales (conductual, emocional, módulos).
    """
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"error": "Sin datos"}), 400

    # Campos permitidos (seguridad: no se puede cambiar id ni cedula via PATCH)
    CAMPOS_TEXTO = {
        'nombre','apellido','curso','grado','condicion','uso_celular',
        'nivel_riesgo','tendencia','reporte','color','ia_analisis'
    }
    CAMPOS_NUM = {
        'edad','p_acad','p_cond','p_auto','p_emocional','prom_modulos',
        'puntualidad','tareas','participacion','comprension','rendimiento',
        'interrupciones','conflictos','desafia_autoridad','distraccion',
        'falta_respeto','motivacion','estado_emocional','interes_futuro',
        'apoyo_familiar','indice_riesgo','proyeccion','asistencia',
        'fotografia_p1','fotografia_p2','fotografia_p3','fotografia_p4',
        'lv_p1','lv_p2','lv_p3','lv_p4',
        'diseno_p1','diseno_p2','diseno_p3','diseno_p4',
        'acad_p1','acad_p2','acad_p3','acad_p4',
        'asistencia_p1','asistencia_p2','asistencia_p3','asistencia_p4',
        'silencioso',
    }

    set_parts  = []
    set_values = []

    for k, v in data.items():
        if k in CAMPOS_NUM:
            try:
                set_parts.append(f"{k}=?")
                set_values.append(round(float(v), 2) if v not in (None, '') else None)
            except (ValueError, TypeError):
                pass
        elif k in CAMPOS_TEXTO:
            set_parts.append(f"{k}=?")
            set_values.append(str(v).strip() if v is not None else None)

    if not set_parts:
        return jsonify({"error": "Ningún campo válido para actualizar"}), 400

    # Recalcular promedios derivados si se actualizaron módulos
    campos_recibidos = set(data.keys())
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            f"UPDATE estudiantes SET {', '.join(set_parts)} WHERE id=?",
            set_values + [id]
        )

        # ── Releer estudiante actualizado para recalcular promedios ──────────────
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM estudiantes WHERE id=?", (id,)).fetchone()
        e   = dict(row) if row else {}

        def _pm(vals):
            v = [x for x in vals if x and float(x) > 0]
            return round(sum(float(x) for x in v) / len(v), 2) if v else None

        updates = {}

        # Recalcular p_foto, p_lv, p_diseno, prom_modulos
        if campos_recibidos & {'fotografia_p1','fotografia_p2','fotografia_p3','fotografia_p4',
                                'lv_p1','lv_p2','lv_p3','lv_p4',
                                'diseno_p1','diseno_p2','diseno_p3','diseno_p4'}:
            pf   = _pm([e.get('fotografia_p1'),e.get('fotografia_p2'),e.get('fotografia_p3'),e.get('fotografia_p4')])
            pl   = _pm([e.get('lv_p1'),e.get('lv_p2'),e.get('lv_p3'),e.get('lv_p4')])
            pd   = _pm([e.get('diseno_p1'),e.get('diseno_p2'),e.get('diseno_p3'),e.get('diseno_p4')])
            mods = [x for x in [pf, pl, pd] if x]
            pm2  = round(sum(mods)/len(mods), 2) if mods else None
            updates.update({'p_foto': pf, 'p_lv': pl, 'p_diseno': pd, 'prom_modulos': pm2})

        # Recalcular p_cond desde campos conductuales individuales
        CAMPOS_COND = ['puntualidad','tareas','participacion','comprension','rendimiento',
                       'interrupciones','conflictos','desafia_autoridad','distraccion','falta_respeto']
        if campos_recibidos & set(CAMPOS_COND):
            vals_cond = [e.get(c) for c in CAMPOS_COND]
            updates['p_cond'] = _pm(vals_cond)

        # Recalcular p_auto / p_emocional desde campos emocionales
        CAMPOS_EMOC = ['motivacion','p_auto','estado_emocional','interes_futuro','apoyo_familiar']
        if campos_recibidos & set(CAMPOS_EMOC):
            vals_emoc = [e.get(c) for c in CAMPOS_EMOC]
            updates['p_emocional'] = _pm(vals_emoc)
            # p_auto viene directo del campo, no recalculado

        # También recalcular indice_riesgo básico si hay datos suficientes
        if updates.get('p_cond') is not None or updates.get('p_emocional') is not None:
            p_cond_v = updates.get('p_cond') or e.get('p_cond') or 0
            p_emoc_v = updates.get('p_emocional') or e.get('p_emocional') or 0
            p_acad_v = e.get('p_acad') or 0
            if p_cond_v > 0 or p_emoc_v > 0:
                # Riesgo = inverso ponderado: mala conducta y bajo desempeño emocional suben el riesgo
                riesgo = round(
                    max(0, min(100,
                        (100 - p_cond_v) * 0.5 +
                        (100 - p_emoc_v) * 0.3 +
                        (100 - p_acad_v) * 0.2
                    )), 1)
                updates['indice_riesgo'] = riesgo
                updates['nivel_riesgo']  = (
                    'Alto'  if riesgo >= 50 else
                    'Medio' if riesgo >= 30 else 'Bajo'
                )

        if updates:
            sets = ', '.join(f"{k}=?" for k in updates)
            conn.execute(f"UPDATE estudiantes SET {sets} WHERE id=?",
                         list(updates.values()) + [id])

        conn.commit()

    cache_bust()
    return jsonify({"ok": True, "mensaje": "Perfil actualizado."})




# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO REPORTES — Conducta · Psicológico · Académico · Incidente
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/reportes")
@login_required
def vista_reportes():
    return render_template("reportes.html", current_user=get_usuario())

@app.route("/api/reportes", methods=["GET"])
def listar_reportes():
    tipo      = request.args.get("tipo", "")
    estado    = request.args.get("estado", "")
    est_id    = request.args.get("estudiante_id", "")
    severidad = request.args.get("severidad", "")

    q = """
        SELECT r.*, e.nombre, e.apellido, e.grado, e.curso
        FROM reportes r
        JOIN estudiantes e ON e.id = r.estudiante_id
        WHERE 1=1
    """
    params = []
    if tipo:      q += " AND r.tipo=?";             params.append(tipo)
    if estado:    q += " AND r.estado=?";           params.append(estado)
    if est_id:    q += " AND r.estudiante_id=?";    params.append(int(est_id))
    if severidad: q += " AND r.severidad=?";        params.append(severidad)
    q += " ORDER BY r.fecha DESC, r.id DESC"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/reportes", methods=["POST"])
def crear_reporte():
    d = request.get_json(silent=True) or {}
    est_id      = d.get("estudiante_id")
    tipo        = d.get("tipo", "").strip()
    descripcion = d.get("descripcion", "").strip()
    if not est_id or not tipo or not descripcion:
        return jsonify({"error": "estudiante_id, tipo y descripcion son requeridos"}), 400

    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            INSERT INTO reportes
                (estudiante_id, tipo, subtipo, titulo, descripcion,
                 severidad, reportado_por, estado)
            VALUES (?,?,?,?,?,?,?,?)
        """, (est_id, tipo,
              d.get("subtipo",""),
              d.get("titulo",""),
              descripcion,
              d.get("severidad","Media"),
              d.get("reportado_por",""),
              "Abierto"))
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"ok": True, "id": rid})


@app.route("/api/reportes/<int:rid>", methods=["PATCH"])
def actualizar_reporte(rid):
    d = request.get_json(silent=True) or {}
    campos = {}
    for k in ["estado","seguimiento","fecha_cierre","severidad","titulo","descripcion"]:
        if k in d: campos[k] = d[k]
    if not campos:
        return jsonify({"error": "Sin campos"}), 400
    sets   = ", ".join(f"{k}=?" for k in campos)
    vals   = list(campos.values()) + [rid]
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(f"UPDATE reportes SET {sets} WHERE id=?", vals)
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/reportes/<int:rid>", methods=["DELETE"])
def eliminar_reporte(rid):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM reportes WHERE id=?", (rid,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/reportes/resumen")
def resumen_reportes():
    with sqlite3.connect(DATABASE) as conn:
        total  = conn.execute("SELECT COUNT(*) FROM reportes").fetchone()[0]
        abiertos= conn.execute("SELECT COUNT(*) FROM reportes WHERE estado='Abierto'").fetchone()[0]
        graves  = conn.execute("SELECT COUNT(*) FROM reportes WHERE tipo='incidente_grave'").fetchone()[0]
        por_tipo= conn.execute("""
            SELECT tipo, COUNT(*) n FROM reportes GROUP BY tipo ORDER BY n DESC
        """).fetchall()
        recientes = conn.execute("""
            SELECT r.id, r.tipo, r.titulo, r.severidad, r.fecha, r.estado,
                   e.nombre, e.apellido
            FROM reportes r JOIN estudiantes e ON e.id=r.estudiante_id
            ORDER BY r.fecha DESC, r.id DESC LIMIT 10
        """).fetchall()
    return jsonify({
        "total": total, "abiertos": abiertos, "graves": graves,
        "por_tipo": [{"tipo":r[0],"n":r[1]} for r in por_tipo],
        "recientes": [dict(zip(
            ["id","tipo","titulo","severidad","fecha","estado","nombre","apellido"], r
        )) for r in recientes]
    })


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO ML — Clustering K-Means + Patrones de comportamiento
# ══════════════════════════════════════════════════════════════════════════════

# Etiquetas y colores para cada cluster (se asignan dinámicamente)
CLUSTER_META = [
    {"label": "Alto rendimiento estable",    "color": "#4dffb4", "icon": "⭐",
     "desc": "Buen desempeño académico y emocional. Bajo riesgo.",
     "accion": "Mantener seguimiento regular. Oportunidades de liderazgo."},
    {"label": "Rendimiento medio, conducta variable", "color": "#c8f060", "icon": "📊",
     "desc": "Notas aceptables pero con irregularidades conductuales.",
     "accion": "Refuerzo de hábitos. Entrevista trimestral."},
    {"label": "Riesgo conductual",            "color": "#ffc94d", "icon": "⚠️",
     "desc": "Conflictos frecuentes o bajo autocontrol. Académico variable.",
     "accion": "Intervención conductual. Coordinación con psicología."},
    {"label": "Caso silencioso",              "color": "#60b8f0", "icon": "🔇",
     "desc": "Alto rendimiento pero señales de malestar emocional.",
     "accion": "Entrevista de bienestar. Seguimiento emocional discreto."},
    {"label": "En crisis – intervención urgente", "color": "#ff6b6b", "icon": "🚨",
     "desc": "Múltiples indicadores de riesgo simultáneos.",
     "accion": "Intervención inmediata. Reporte a coordinación y psicología."},
]


def _features_para_clustering(conn):
    """Extrae features numéricas de estudiantes con datos suficientes."""
    rows = conn.execute("""
        SELECT id,
               COALESCE(p_acad,0)         as acad,
               COALESCE(p_cond,0)         as cond,
               COALESCE(p_auto,0)         as auto_e,
               COALESCE(p_emocional,0)    as emoc,
               COALESCE(motivacion,0)     as motiv,
               COALESCE(apoyo_familiar,0) as apoyo,
               COALESCE(indice_riesgo,0)  as riesgo,
               COALESCE(interrupciones,0) as interr,
               COALESCE(conflictos,0)     as conflic,
               COALESCE(falta_respeto,0)  as faltaR,
               COALESCE(distraccion,0)    as distr,
               COALESCE(prom_modulos,0)   as modulos
        FROM estudiantes
        WHERE (p_acad > 0 OR p_cond > 0 OR prom_modulos > 0
               OR motivacion > 0 OR indice_riesgo > 0)
    """).fetchall()
    return rows




@app.route("/api/ml/calcular", methods=["POST"])
def calcular_clusters():
    """K-Means puro con numpy — compatible con cualquier Python."""
    try:
        import numpy as np

        def kmeans_numpy(X, k, max_iter=300, n_init=10, seed=42):
            """K-Means implementado con numpy puro."""
            rng = np.random.RandomState(seed)
            best_inertia = float('inf')
            best_labels  = None
            best_centers = None

            for _ in range(n_init):
                # Inicialización K-Means++
                idx = [rng.randint(0, len(X))]
                for _ in range(k - 1):
                    dists = np.array([min(np.sum((x - X[i])**2) for i in idx) for x in X])
                    probs = dists / dists.sum()
                    idx.append(rng.choice(len(X), p=probs))
                centers = X[idx].copy()

                labels = np.zeros(len(X), dtype=int)
                for iteration in range(max_iter):
                    # Asignar clusters
                    dists  = np.array([[np.sum((x - c)**2) for c in centers] for x in X])
                    new_labels = np.argmin(dists, axis=1)
                    if np.all(new_labels == labels):
                        break
                    labels = new_labels
                    # Actualizar centros
                    for ci in range(k):
                        mask = labels == ci
                        if mask.sum() > 0:
                            centers[ci] = X[mask].mean(axis=0)

                inertia = sum(np.sum((X[labels == ci] - centers[ci])**2)
                              for ci in range(k) if (labels == ci).sum() > 0)
                if inertia < best_inertia:
                    best_inertia  = inertia
                    best_labels   = labels.copy()
                    best_centers  = centers.copy()

            return best_labels, best_centers, best_inertia

        def silhouette(X, labels):
            """Silhouette score simplificado."""
            n = len(X)
            if n < 4:
                return 0.0
            scores = []
            unique = list(set(labels))
            if len(unique) < 2:
                return 0.0
            for i in range(n):
                same  = X[labels == labels[i]]
                a     = np.mean([np.sqrt(np.sum((X[i]-x)**2)) for x in same if not np.all(x==X[i])]) if len(same) > 1 else 0
                other_means = []
                for c in unique:
                    if c != labels[i]:
                        grp = X[labels == c]
                        if len(grp) > 0:
                            other_means.append(np.mean([np.sqrt(np.sum((X[i]-x)**2)) for x in grp]))
                b = min(other_means) if other_means else 0
                m = max(a, b)
                scores.append((b - a) / m if m > 0 else 0)
            return float(np.mean(scores))

        with sqlite3.connect(DATABASE) as conn:
            conn.row_factory = sqlite3.Row
            rows = _features_para_clustering(conn)

        if len(rows) < 10:
            return jsonify({"error": f"Solo {len(rows)} estudiantes con datos. "
                                      "Se necesitan al menos 10 para clustering."}), 400

        ids   = [r[0] for r in rows]
        X_raw = np.array([[r[1],r[2],r[3],r[4],r[5],r[6],r[7],r[8],r[9],r[10],r[11],r[12]]
                           for r in rows], dtype=float)

        # Estandarizar (Z-score manual)
        mean = X_raw.mean(axis=0)
        std  = X_raw.std(axis=0)
        std[std == 0] = 1
        Xs = (X_raw - mean) / std

        # Buscar k óptimo (2-5)
        max_k  = min(5, len(rows) // 5)
        max_k  = max(2, max_k)
        best_k = 2
        best_s = -1
        best_labels = None
        best_centers = None

        for k in range(2, max_k + 1):
            lbs, ctrs, _ = kmeans_numpy(Xs, k)
            if len(set(lbs.tolist())) > 1:
                s = silhouette(Xs, lbs)
                if s > best_s:
                    best_s = s; best_k = k
                    best_labels  = lbs
                    best_centers = ctrs

        if best_labels is None:
            best_labels, best_centers, _ = kmeans_numpy(Xs, 2)
            best_k = 2

        # Ordenar clusters por riesgo compuesto
        cluster_risk = {}
        for ci in range(best_k):
            mask = best_labels == ci
            if mask.sum() == 0:
                cluster_risk[ci] = 0
                continue
            avg_riesgo = float(X_raw[mask, 6].mean())
            avg_acad   = float(X_raw[mask, 0].mean())
            avg_auto   = float(X_raw[mask, 2].mean())
            cluster_risk[ci] = avg_riesgo - avg_acad * 0.3 - avg_auto * 0.2

        sorted_clusters = sorted(cluster_risk.keys(), key=lambda c: cluster_risk[c])
        cluster_map     = {orig: pos for pos, orig in enumerate(sorted_clusters)}

        # Calcular distancia al centroide para score
        distances = np.array([[np.sqrt(np.sum((Xs[i] - best_centers[ci])**2))
                                for ci in range(best_k)]
                               for i in range(len(Xs))])

        meta_map = {}
        for orig_ci in range(best_k):
            mask = best_labels == orig_ci
            if mask.sum() == 0:
                continue
            avg_acad   = float(X_raw[mask, 0].mean())
            avg_cond   = float(X_raw[mask, 1].mean())
            avg_auto   = float(X_raw[mask, 2].mean())
            avg_riesgo = float(X_raw[mask, 6].mean())
            avg_conflic= float(X_raw[mask, 8].mean())

            if avg_acad >= 75 and avg_auto >= 70 and avg_riesgo < 25:
                meta_idx = 0
            elif avg_acad >= 75 and avg_auto < 60:
                meta_idx = 3
            elif avg_riesgo >= 45 or avg_conflic >= 50:
                meta_idx = 4
            elif avg_conflic >= 35 or avg_cond < 55:
                meta_idx = 2
            else:
                meta_idx = 1

            meta = CLUSTER_META[min(meta_idx, len(CLUSTER_META) - 1)]
            meta_map[orig_ci] = {
                "label":       meta["label"],
                "color":       meta["color"],
                "icon":        meta["icon"],
                "accion":      meta["accion"],
                "desc":        meta["desc"],
                "n":           int(mask.sum()),
                "avg_acad":    round(avg_acad, 1),
                "avg_cond":    round(avg_cond, 1),
                "avg_auto":    round(avg_auto, 1),
                "avg_riesgo":  round(avg_riesgo, 1),
            }

        # Guardar en DB
        with sqlite3.connect(DATABASE) as conn:
            for i, (est_id, orig_ci) in enumerate(zip(ids, best_labels.tolist())):
                meta      = meta_map.get(int(orig_ci), {})
                dist_min  = float(distances[i][orig_ci])
                score     = round(max(0, 100 - dist_min * 10), 1)
                conn.execute("""
                    UPDATE estudiantes
                    SET cluster_id=?, cluster_label=?, cluster_color=?, cluster_score=?
                    WHERE id=?
                """, (int(cluster_map[int(orig_ci)]), meta.get("label",""),
                      meta.get("color","#888"), score, est_id))
            conn.execute("""
                INSERT INTO ml_clusters (n_clusters, features_usadas, resumen)
                VALUES (?,?,?)
            """, (best_k,
                  "p_acad,p_cond,p_auto,p_emocional,motivacion,apoyo_familiar,"
                  "indice_riesgo,interrupciones,conflictos,falta_respeto,distraccion,prom_modulos",
                  str({v["label"]: v["n"] for v in meta_map.values()})))
            conn.commit()

        return jsonify({
            "ok":    True,
            "k":     best_k,
            "silhouette": round(best_s, 3),
            "estudiantes_analizados": len(ids),
            "clusters": [meta_map[ci] for ci in sorted(meta_map.keys(),
                          key=lambda c: cluster_risk[c])]
        })

    except Exception as ex:
        import traceback
        return jsonify({"error": str(ex), "detalle": traceback.format_exc()}), 500


@app.route("/api/ml/patrones")
def get_patrones():
    """Devuelve resumen de clusters + estudiantes por cluster."""
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        clusters = conn.execute("""
            SELECT cluster_id, cluster_label, cluster_color,
                   COUNT(*) n,
                   ROUND(AVG(p_acad),1)       avg_acad,
                   ROUND(AVG(p_cond),1)       avg_cond,
                   ROUND(AVG(p_auto),1)       avg_auto,
                   ROUND(AVG(indice_riesgo),1) avg_riesgo,
                   ROUND(AVG(cluster_score),1) avg_score
            FROM estudiantes
            WHERE cluster_id IS NOT NULL
            GROUP BY cluster_id
            ORDER BY cluster_id
        """).fetchall()

        estudiantes = conn.execute("""
            SELECT id, nombre, apellido, grado, curso,
                   cluster_id, cluster_label, cluster_color, cluster_score,
                   p_acad, p_cond, indice_riesgo
            FROM estudiantes
            WHERE cluster_id IS NOT NULL
            ORDER BY cluster_id, indice_riesgo DESC
        """).fetchall()

        ultimo = conn.execute("""
            SELECT fecha_calculo, n_clusters FROM ml_clusters
            ORDER BY id DESC LIMIT 1
        """).fetchone()

    # Añadir accion/desc/icon de CLUSTER_META por label
    label_to_meta = {m["label"]: m for m in CLUSTER_META}

    clusters_out = []
    for c in clusters:
        d = dict(c)
        meta = label_to_meta.get(d["cluster_label"], {})
        d["icon"]   = meta.get("icon",  "📌")
        d["accion"] = meta.get("accion","")
        d["desc"]   = meta.get("desc",  "")
        clusters_out.append(d)

    return jsonify({
        "clusters":     clusters_out,
        "estudiantes":  [dict(e) for e in estudiantes],
        "ultimo_calculo": dict(ultimo) if ultimo else None
    })


@app.route("/api/ml/similar/<int:est_id>")
def estudiantes_similares(est_id):
    """Devuelve los 5 estudiantes más similares al perfil dado."""
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        target = conn.execute(
            "SELECT * FROM estudiantes WHERE id=?", (est_id,)
        ).fetchone()
        if not target:
            return jsonify([])
        t = dict(target)

        same_cluster = conn.execute("""
            SELECT id, nombre, apellido, grado, curso,
                   p_acad, p_cond, indice_riesgo, cluster_label, cluster_score
            FROM estudiantes
            WHERE cluster_id=? AND id!=?
            ORDER BY ABS(COALESCE(p_acad,0) - ?)
                   + ABS(COALESCE(indice_riesgo,0) - ?)
            LIMIT 5
        """, (t.get("cluster_id"), est_id,
              t.get("p_acad") or 0, t.get("indice_riesgo") or 0)).fetchall()

    return jsonify([dict(e) for e in same_cluster])


@app.route("/patrones")
@login_required
def vista_patrones():
    return render_template("patrones.html", current_user=get_usuario())


# ── Búsqueda rápida de estudiantes (autocomplete) ────────────────────────────
@app.route("/api/buscar-estudiantes")
def buscar_estudiantes():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    like = f"%{q}%"
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, nombre, apellido, grado, curso
            FROM estudiantes
            WHERE nombre LIKE ? OR apellido LIKE ?
               OR (nombre || ' ' || apellido) LIKE ?
            ORDER BY apellido, nombre LIMIT 12
        """, (like, like, like)).fetchall()
    return jsonify([dict(r) for r in rows])



# ══════════════════════════════════════════════════════════════════════════════
# HELPER DE BÚSQUEDA DE ESTUDIANTES — nivel módulo
# Usado por cargar_registro y cargar_boletin
# ══════════════════════════════════════════════════════════════════════════════

def _buscar_estudiante_bd(conn, nombre, apellido, filtro_grado="", filtro_mencion=""):
    """
    Busca un estudiante por nombre/apellido con normalización unicode.
    Cuatro intentos en cascada: exacto → solo nombre → sin filtro grado → fuzzy.
    Retorna sqlite3.Row o None.
    NOTA: Solo selecciona columnas que siempre existen (sin cedula por compatibilidad).
    """
    import unicodedata

    def norm(s):
        if not s: return ""
        s = str(s).lower().strip()
        return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")

    # Registrar función norm en esta conexión
    try:
        conn.create_function("norm", 1, norm)
    except Exception:
        pass

    nom1 = norm(nombre.split()[0] if nombre else "")
    ape1 = norm(apellido.split()[0] if apellido else "")

    # Detectar columnas disponibles (BD nueva puede no tener cedula aún)
    cols_disponibles = {row[1] for row in conn.execute("PRAGMA table_info(estudiantes)").fetchall()}
    safe_cols = "id, nombre, apellido, grado, curso"
    if "cedula" in cols_disponibles:
        safe_cols = "id, cedula, nombre, apellido, grado, curso"

    extra_cond   = ""
    extra_params = []
    if filtro_grado:
        extra_cond   += " AND upper(grado) LIKE upper(?)"
        extra_params.append(f"%{filtro_grado}%")
    if filtro_mencion:
        extra_cond   += " AND upper(curso) LIKE upper(?)"
        extra_params.append(f"%{filtro_mencion}%")

    # Intento 1: nombre + apellido + filtros
    est = conn.execute(
        f"SELECT {safe_cols} FROM estudiantes "
        f"WHERE norm(nombre) LIKE ? AND norm(apellido) LIKE ? {extra_cond} LIMIT 1",
        [f"%{nom1}%", f"%{ape1}%"] + extra_params
    ).fetchone()

    # Intento 2 eliminado — causaba matches incorrectos por nombre parcial

    # Intento 3: nombre + apellido sin filtros
    if not est and ape1:
        est = conn.execute(
            f"SELECT {safe_cols} FROM estudiantes "
            "WHERE norm(nombre) LIKE ? AND norm(apellido) LIKE ? LIMIT 1",
            [f"%{nom1}%", f"%{ape1}%"]
        ).fetchone()

    # Intento 4: fuzzy con fuzzywuzzy
    if not est and nom1:
        try:
            from fuzzywuzzy import fuzz
            candidatos = conn.execute(
                f"SELECT {safe_cols} FROM estudiantes LIMIT 500"
            ).fetchall()
            full_buscado = f"{nom1} {ape1}".strip()
            mejor = None
            mejor_score = 0
            for c in candidatos:
                full_c = norm(f"{c['nombre']} {c['apellido']}")
                score = fuzz.token_sort_ratio(full_buscado, full_c)
                if score > mejor_score:
                    mejor_score = score
                    mejor = c
            if mejor_score >= 82:
                est = mejor
        except Exception:
            pass

    return est


# ══════════════════════════════════════════════════════════════════════════════
# CARGA DIRECTA — Registro de Calificaciones Oficial (formato MINERD)
# Lee múltiples hojas, detecta profesor, carga notas+asistencia
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/cargar-registro", methods=["POST"])
def cargar_registro():
    """
    Carga el 'Registro de Calificaciones' oficial (formato MINERD).
    Detecta automáticamente:
      - Nombre del profesor (fila 4)
      - Materia / área (fila 3)
      - Columnas de calificaciones por período (sección resumen derecha)
      - Columnas de porcentaje de asistencia por período
    Funciona con múltiples hojas (ej: 6TO A, 6TO B).
    No sobreescribe notas que ya existan.
    """
    import openpyxl, re as _re

    if "file" not in request.files:
        return jsonify({"error": "Sin archivo"}), 400

    file   = request.files["file"]
    raw    = file.read()
    materia_override = request.form.get("materia", "").strip()

    try:
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, keep_vba=False)
    except Exception as ex:
        return jsonify({"error": f"No se pudo leer el archivo: {ex}"}), 400

    def limpiar_num(v):
        try:
            if v is None or str(v).strip() in ('', 'nan', 'None', '-', '0'):
                return None
            f = float(v)
            return round(f, 1) if f > 0 else None
        except Exception:
            return None

    def limpiar_str(v):
        return str(v).strip() if v else ''

    resultados = []   # por hoja
    total_ok = 0
    total_no = []

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # ── 1. Detectar profesor (fila 4, col 2)
            cell_prof = limpiar_str(ws.cell(4, 2).value)
            profesor  = _re.sub(r'(?i)maestro[/a]*\s*:\s*', '', cell_prof).strip()
            if not profesor:
                profesor = file.filename

            # ── 2. Detectar materia (fila 3, col 2)
            cell_mat = limpiar_str(ws.cell(3, 2).value)
            materia  = materia_override or _re.sub(r'(?i)área[/técnica]*\s*:\s*', '', cell_mat).strip()
            if not materia:
                materia = sheet_name

            # ── 3. Localizar columna "PORCENTAJE DE ASISTENCIA" (fila 6)
            # y columnas del resumen de períodos (fila 8: "Periodo I/II/III/IV")
            col_asist_p1 = None
            col_resumen_p1 = None

            for c in range(1, ws.max_column + 1):
                v6 = limpiar_str(ws.cell(6, c).value).lower()
                if 'porcentaje' in v6 and 'asistencia' in v6:
                    col_asist_p1 = c   # P1 asistencia, P2=c+1, P3=c+2, P4=c+3

                v8 = limpiar_str(ws.cell(8, c).value).lower().strip()
                # Exacto: debe ser "periodo i" (no "periodo ii", "periodo iv"...)
                if v8 in ('periodo i', 'período i', 'periodo  i'):
                    col_resumen_p1 = c  # P1, P2=c+1, P3=c+2, P4=c+3

            sheet_ok  = 0
            sheet_no  = []

            # ── 4. Recorrer estudiantes (fila 10 en adelante, col 3=nombre, 4=apellido)
            for row in range(10, ws.max_row + 1):
                nombre_raw   = ws.cell(row, 3).value
                apellido_raw = ws.cell(row, 4).value

                if not nombre_raw or str(nombre_raw).strip() in ('', '0', 'Nombres'):
                    continue

                nombre   = limpiar_str(nombre_raw).split()[0].lower()
                apellido = limpiar_str(apellido_raw).split()[0].lower() if apellido_raw else ''

                est = _buscar_estudiante_bd(conn, nombre, apellido)
                if not est:
                    sheet_no.append(f"{limpiar_str(nombre_raw)} {limpiar_str(apellido_raw)}")
                    continue

                est_id  = est['id']
                cedula  = est['cedula']

                # ── Leer calificaciones de la sección resumen
                p1n = p2n = p3n = p4n = None
                if col_resumen_p1:
                    p1n = limpiar_num(ws.cell(row, col_resumen_p1).value)
                    p2n = limpiar_num(ws.cell(row, col_resumen_p1 + 1).value)
                    p3n = limpiar_num(ws.cell(row, col_resumen_p1 + 2).value)
                    p4n = limpiar_num(ws.cell(row, col_resumen_p1 + 3).value)

                notas = [x for x in [p1n, p2n, p3n, p4n] if x]
                if not notas:
                    sheet_no.append(f"{limpiar_str(nombre_raw)} (sin notas)")
                    continue

                prom = round(sum(notas) / len(notas), 1)

                # ── Guardar en materias_calificaciones (sin sobreescribir)
                conn.execute("""
                    INSERT INTO materias_calificaciones
                        (estudiante_id, cedula, materia, p1, p2, p3, p4, promedio, fuente, profesor)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(estudiante_id, materia) DO UPDATE SET
                        p1       = CASE WHEN excluded.p1 IS NOT NULL THEN excluded.p1
                                        ELSE materias_calificaciones.p1 END,
                        p2       = CASE WHEN excluded.p2 IS NOT NULL THEN excluded.p2
                                        ELSE materias_calificaciones.p2 END,
                        p3       = CASE WHEN excluded.p3 IS NOT NULL THEN excluded.p3
                                        ELSE materias_calificaciones.p3 END,
                        p4       = CASE WHEN excluded.p4 IS NOT NULL THEN excluded.p4
                                        ELSE materias_calificaciones.p4 END,
                        promedio = excluded.promedio,
                        fuente   = excluded.fuente,
                        profesor = CASE WHEN excluded.profesor IS NOT NULL AND excluded.profesor != ''
                                        THEN excluded.profesor
                                        ELSE materias_calificaciones.profesor END,
                        fecha_carga = date('now')
                """, (est_id, cedula, materia, p1n, p2n, p3n, p4n, prom, file.filename, profesor))

                # ── Sincronizar módulos Multimedia si aplica
                mat_norm = materia.lower().strip()
                MODULO_MAP = {
                    'fotografía':'fotografia','fotografia':'fotografia','foto':'fotografia',
                    'lenguaje visual':'lv','lv':'lv',
                    'diseño':'diseno','diseno':'diseno','diseño básico':'diseno','diseño basico':'diseno',
                }
                col_base = MODULO_MAP.get(mat_norm)
                if col_base:
                    p_col = 'p_foto' if col_base == 'fotografia' else f'p_{col_base}'
                    sets, vals = [], []
                    for pi, pv in enumerate([p1n, p2n, p3n, p4n], 1):
                        if pv:
                            sets.append(f"{col_base}_p{pi} = CASE WHEN {col_base}_p{pi} IS NULL OR {col_base}_p{pi}=0 THEN ? ELSE {col_base}_p{pi} END")
                            vals.append(pv)
                    if sets:
                        sets.append(f"{p_col} = ?")
                        vals.append(prom)
                        vals.append(est_id)
                        conn.execute(f"UPDATE estudiantes SET {', '.join(sets)} WHERE id=?", vals)
                        # Recalcular prom_modulos
                        e2 = conn.execute(
                            "SELECT p_foto, p_lv, p_diseno FROM estudiantes WHERE id=?", (est_id,)
                        ).fetchone()
                        if e2:
                            mods = [x for x in e2 if x and x > 0]
                            pm2  = round(sum(mods) / len(mods), 1) if mods else None
                            conn.execute("UPDATE estudiantes SET prom_modulos=? WHERE id=?", (pm2, est_id))

                # ── Guardar asistencia por período
                if col_asist_p1:
                    a1 = limpiar_num(ws.cell(row, col_asist_p1).value)
                    a2 = limpiar_num(ws.cell(row, col_asist_p1 + 1).value)
                    a3 = limpiar_num(ws.cell(row, col_asist_p1 + 2).value)
                    a4 = limpiar_num(ws.cell(row, col_asist_p1 + 3).value)

                    asist_sets, asist_vals = [], []
                    for pi, pv in enumerate([a1, a2, a3, a4], 1):
                        if pv:
                            asist_sets.append(
                                f"asistencia_p{pi} = CASE WHEN asistencia_p{pi} IS NULL OR asistencia_p{pi}=0 "
                                f"THEN ? ELSE asistencia_p{pi} END"
                            )
                            asist_vals.append(pv)
                    if asist_sets:
                        asist_vals.append(est_id)
                        conn.execute(
                            f"UPDATE estudiantes SET {', '.join(asist_sets)} WHERE id=?",
                            asist_vals
                        )

                sheet_ok += 1

            conn.commit()
            total_ok += sheet_ok
            total_no.extend(sheet_no)
            resultados.append({
                "hoja": sheet_name,
                "materia": materia,
                "profesor": profesor,
                "cargados": sheet_ok,
                "no_encontrados": sheet_no,
            })

    cache_bust()
    return jsonify({
        "status": "success",
        "hojas_procesadas": len(resultados),
        "total_cargados": total_ok,
        "no_encontrados": total_no,
        "detalle": resultados,
        "mensaje": f"{total_ok} estudiantes actualizados en {len(resultados)} hoja(s)."
            + (f" {len(total_no)} no encontrados." if total_no else "")
    })



# ── COMPARATIVA POR MENCIÓN ──────────────────────────────────────────────────

@app.route("/api/comparativa-mencion")
def comparativa_mencion():
    """Devuelve promedios académicos, conductuales y de riesgo agrupados por mención."""
    conn = get_db()
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT
            curso,
            AVG(CASE WHEN p_acad  > 0 THEN p_acad  END) as avg_acad,
            AVG(CASE WHEN p_cond  > 0 THEN p_cond  END) as avg_cond,
            AVG(CASE WHEN p_auto  > 0 THEN p_auto  END) as avg_auto,
            AVG(CASE WHEN indice_riesgo > 0 THEN indice_riesgo END) as avg_riesgo,
            AVG(CASE WHEN prom_modulos  > 0 THEN prom_modulos  END) as avg_modulos,
            COUNT(*) as total,
            SUM(CASE WHEN p_acad > 0 THEN 1 ELSE 0 END) as con_notas,
            SUM(CASE WHEN indice_riesgo >= 50 THEN 1 ELSE 0 END) as riesgo_alto,
            SUM(CASE WHEN categoria = 'ALERTA DE REPROBACIÓN' THEN 1 ELSE 0 END) as alertas
        FROM estudiantes
        WHERE condicion IS NULL OR condicion != 'RETIRADO'
        GROUP BY curso
        ORDER BY curso
    """).fetchall()
    conn.close()

    MENCION_LABEL = {
        'multimedia': 'Multimedia',
        'teatro':     'Teatro',
        'musica':     'Música',
        'visual':     'Artes Visuales',
    }
    MENCION_COLOR = {
        'multimedia': '#c8f060',
        'teatro':     '#60b8f0',
        'musica':     '#ff9f60',
        'visual':     '#c060f0',
    }

    result = []
    for r in rows:
        curso = (r['curso'] or '').lower()
        mencion_key = next((k for k in MENCION_LABEL if k in curso), None)
        if not mencion_key:
            continue
        result.append({
            'curso':       r['curso'],
            'mencion':     MENCION_LABEL[mencion_key],
            'color':       MENCION_COLOR[mencion_key],
            'total':       r['total'],
            'con_notas':   r['con_notas'],
            'riesgo_alto': r['riesgo_alto'],
            'alertas':     r['alertas'],
            'avg_acad':    round(r['avg_acad'] or 0, 1),
            'avg_cond':    round(r['avg_cond'] or 0, 1),
            'avg_auto':    round(r['avg_auto'] or 0, 1),
            'avg_riesgo':  round(r['avg_riesgo'] or 0, 1),
            'avg_modulos': round(r['avg_modulos'] or 0, 1),
        })

    # Aggregate per mención (sum groups like "4to A Multimedia", "4to B Multimedia")
    merged = {}
    for r in result:
        m = r['mencion']
        if m not in merged:
            merged[m] = {**r, '_cnt': 1}
        else:
            ex = merged[m]
            ex['total']       += r['total']
            ex['con_notas']   += r['con_notas']
            ex['riesgo_alto'] += r['riesgo_alto']
            ex['alertas']     += r['alertas']
            # Running average
            n = ex['_cnt']
            for k in ('avg_acad','avg_cond','avg_auto','avg_riesgo','avg_modulos'):
                ex[k] = round((ex[k] * n + r[k]) / (n + 1), 1)
            ex['_cnt'] += 1

    final = []
    for m, d in merged.items():
        d.pop('_cnt', None)
        final.append(d)

    return jsonify(final)


# ── FOTO DE PERFIL ───────────────────────────────────────────────────────────

@app.route("/api/estudiante/<int:id>/foto", methods=["POST"])
@login_required
def subir_foto(id):
    """Sube foto de perfil de un estudiante."""
    f = request.files.get("foto")
    if not f or not f.filename:
        return jsonify({"error": "No se recibió archivo"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        return jsonify({"error": "Formato no válido"}), 400
    filename = f"est_{id}{ext}"
    path     = os.path.join(FOTOS_DIR, filename)
    f.save(path)
    url = f"/static/fotos/{filename}"
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("UPDATE estudiantes SET foto_path=? WHERE id=?", (url, id))
        conn.commit()
    return jsonify({"ok": True, "url": url})

@app.route("/api/estudiante/<int:id>/foto", methods=["DELETE"])
@login_required
def borrar_foto(id):
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT foto_path FROM estudiantes WHERE id=?", (id,)).fetchone()
        if row and row["foto_path"]:
            try:
                os.remove(row["foto_path"].lstrip("/"))
            except Exception:
                pass
        conn.execute("UPDATE estudiantes SET foto_path=NULL WHERE id=?", (id,))
        conn.commit()
    return jsonify({"ok": True})


# ── PORTAL DEL PROFESOR ───────────────────────────────────────────────────────

# portal_profesor → see full implementation below

@app.route("/api/profesor/mis-estudiantes")
@login_required
def mis_estudiantes():
    """Devuelve solo los estudiantes que tienen materias del profesor logueado."""
    u = get_usuario()
    materia = u.get("materia", "")
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        if materia:
            ids = [r[0] for r in conn.execute(
                "SELECT DISTINCT estudiante_id FROM materias_calificaciones WHERE LOWER(profesor) LIKE LOWER(?)",
                (f"%{u['nombre']}%",)
            ).fetchall()]
            # Also by materia name
            ids2 = [r[0] for r in conn.execute(
                "SELECT DISTINCT estudiante_id FROM materias_calificaciones WHERE LOWER(materia) LIKE LOWER(?)",
                (f"%{materia}%",)
            ).fetchall()]
            all_ids = list(set(ids + ids2))
        else:
            all_ids = []
        if not all_ids:
            # If no filter, return students of same grado/course as a fallback
            rows = conn.execute(
                "SELECT id,nombre,apellido,curso,grado,p_acad,puntualidad,participacion,asistencia,indice_riesgo,nivel_riesgo FROM estudiantes ORDER BY apellido,nombre"
            ).fetchall()
        else:
            placeholders = ",".join("?" * len(all_ids))
            rows = conn.execute(
                f"SELECT id,nombre,apellido,curso,grado,p_acad,puntualidad,participacion,asistencia,indice_riesgo,nivel_riesgo FROM estudiantes WHERE id IN ({placeholders}) ORDER BY apellido,nombre",
                all_ids
            ).fetchall()
        materias_map = {}
        for r in conn.execute(
            "SELECT estudiante_id, materia, p1, p2, p3, p4, promedio FROM materias_calificaciones WHERE LOWER(materia) LIKE LOWER(?)",
            (f"%{materia}%",) if materia else ("%",)
        ).fetchall():
            materias_map.setdefault(r[0], []).append(dict(r))
    result = []
    for r in rows:
        d = dict(r)
        d["materias"] = materias_map.get(d["id"], [])
        result.append(d)
    return jsonify(result)


# ── EXPORTAR REPORTES (Excel) ─────────────────────────────────────────────────

@app.route("/api/reportes/exportar-xlsx")
@login_required
def exportar_reportes_xlsx():
    """Descarga los reportes filtrados como archivo Excel."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    tipo      = request.args.get("tipo","")
    estado    = request.args.get("estado","")
    severidad = request.args.get("severidad","")

    q = """
        SELECT r.fecha, e.nombre||' '||e.apellido AS estudiante,
               e.grado, r.tipo, r.subtipo, r.titulo, r.descripcion,
               r.severidad, r.estado, r.reportado_por,
               r.seguimiento, r.fecha_cierre
        FROM reportes r JOIN estudiantes e ON e.id=r.estudiante_id
        WHERE 1=1
    """
    params = []
    if tipo:      q += " AND r.tipo=?";      params.append(tipo)
    if estado:    q += " AND r.estado=?";    params.append(estado)
    if severidad: q += " AND r.severidad=?"; params.append(severidad)
    q += " ORDER BY r.fecha DESC"

    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(q, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ["Fecha","Estudiante","Grado","Tipo","Subtipo","Título",
               "Descripción","Severidad","Estado","Reportado por","Seguimiento","Fecha cierre"]

    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="C8F060", size=10)

    sev_colors = {"Alta":"FF6B6B","Media":"FFC44D","Baja":"4DFFB4"}

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val or "")
        sev = row[7] or ""
        if sev in sev_colors:
            fill = PatternFill("solid", fgColor=sev_colors[sev])
            ws.cell(row=ri, column=8).fill = fill

    col_widths = [12,25,12,14,16,28,40,10,14,20,30,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import Response
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reportes_multimediatrack.xlsx"}
    )


# ── GESTIÓN DE USUARIOS (página) ─────────────────────────────────────────────

@app.route("/usuarios")
@login_required
def vista_usuarios():
    u = get_usuario()
    if u["rol"] != "coordinador":
        return redirect("/")
    return render_template("usuarios.html", usuario=u)


# ── EXPORTAR REPORTES PDF ────────────────────────────────────────────────────

@app.route("/api/reportes/exportar-pdf")
@login_required
def exportar_reportes_pdf():
    """Genera un PDF printable de los reportes filtrados usando HTML → PDF."""
    from flask import make_response
    import datetime

    tipo      = request.args.get("tipo", "")
    estado    = request.args.get("estado", "")
    severidad = request.args.get("severidad", "")

    q = """
        SELECT r.fecha, e.nombre||' '||e.apellido AS estudiante,
               e.grado, r.tipo, r.titulo, r.descripcion,
               r.severidad, r.estado, r.reportado_por, r.seguimiento
        FROM reportes r JOIN estudiantes e ON e.id=r.estudiante_id
        WHERE 1=1
    """
    params = []
    if tipo:      q += " AND r.tipo=?";      params.append(tipo)
    if estado:    q += " AND r.estado=?";    params.append(estado)
    if severidad: q += " AND r.severidad=?"; params.append(severidad)
    q += " ORDER BY r.fecha DESC"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]

    tipo_label = {
        "conducta": "Conducta", "psicologico": "Psicológico",
        "academico": "Académico", "incidente_grave": "Incidente Grave"
    }
    sev_color = {"Alta": "#ff6b6b", "Media": "#ffc44d", "Baja": "#4dffb4"}

    filas_html = ""
    for r in rows:
        sev_c = sev_color.get(r.get("severidad",""), "#888")
        filas_html += f"""
        <tr>
          <td>{r.get('fecha','')}</td>
          <td><b>{r.get('estudiante','')}</b><br><small>{r.get('grado','')}</small></td>
          <td>{tipo_label.get(r.get('tipo',''), r.get('tipo',''))}</td>
          <td>{r.get('titulo','')}</td>
          <td style="color:{sev_c};font-weight:700;">{r.get('severidad','')}</td>
          <td>{r.get('estado','')}</td>
          <td style="font-size:10px;">{(r.get('descripcion') or '')[:120]}</td>
          <td style="font-size:10px;">{r.get('reportado_por','')}</td>
        </tr>"""

    now = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11pt; color: #111; margin: 20px; }}
  h1   {{ font-size: 16pt; margin-bottom: 2px; }}
  .sub {{ font-size: 10pt; color: #666; margin-bottom: 16px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  thead th {{ background: #1a1a1a; color: #c8f060; padding: 6px 8px;
              font-size: 9pt; text-align: left; }}
  tbody tr:nth-child(even) {{ background: #f9f9f9; }}
  td {{ padding: 5px 8px; vertical-align: top; border-bottom: 1px solid #eee; font-size: 9pt; }}
  @page {{ margin: 15mm; size: A4 landscape; }}
  @media print {{ button {{ display:none; }} }}
</style>
</head><body>
<h1>📋 Reportes — MultimediaTrack</h1>
<div class="sub">Generado: {now} · {len(rows)} registros{' · Tipo: '+tipo_label.get(tipo,tipo) if tipo else ''}{' · Estado: '+estado if estado else ''}{' · Severidad: '+severidad if severidad else ''}</div>
<button onclick="window.print()" style="margin-bottom:14px;padding:7px 18px;background:#1a1a1a;
  color:#c8f060;border:1px solid #333;border-radius:6px;cursor:pointer;font-size:12px;">
  🖨 Imprimir / Guardar PDF
</button>
<table>
  <thead>
    <tr>
      <th>Fecha</th><th>Estudiante</th><th>Tipo</th><th>Título</th>
      <th>Severidad</th><th>Estado</th><th>Descripción</th><th>Reportado por</th>
    </tr>
  </thead>
  <tbody>{filas_html}</tbody>
</table>
</body></html>"""

    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp



# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO PROFESORES — Asignación, Validación y Asistencia
# Normativa: Ordenanza 1'96 mod. 1'98 + 04-2023 MINERD
# • Escala 0-100, mínimo aprobatorio 70 puntos
# • Art. 51 mod. 1'98: reprueba con >20% inasistencias injustificadas
# • Asistencia: (horas_presentes / horas_totales) × 100
# ══════════════════════════════════════════════════════════════════════════════

# Plan de estudio oficial MINERD — Bachillerato Arte Multimedia (por grado)
PLAN_MULTIMEDIA = {
    "4to": [
        ("Identidad, Cultura y Emprendimiento", 2),
        ("Historia del Arte Universal y la Estética Digital", 4),
        ("Lenguaje Musical", 2),
        ("Lenguaje Danzario y Teatral", 2),
        ("Lenguaje Visual, Dibujo y Creación de Personajes", 5),
        ("Diseño Básico y Expresión Visual", 4),
        ("Fotografía", 4),
        ("Lengua Española", 3),
        ("Inglés", 4),
        ("Matemática", 3),
        ("Ciencias Sociales", 2),
        ("Ciencias de la Naturaleza", 3),
        ("Formación Integral Humana y Religiosa", 1),
        ("Educación Física", 1),
    ],
    "5to": [
        ("Diseño Web", 6),
        ("Diseño Gráfico", 4),
        ("Publicidad y Creatividad", 3),
        ("Operación de Cámara de Video", 4),
        ("Guión", 4),
        ("Medios de Comunicación", 2),
        ("Lengua Española", 3),
        ("Inglés", 4),
        ("Matemática", 3),
        ("Ciencias Sociales", 2),
        ("Ciencias de la Naturaleza", 3),
        ("Formación Integral Humana y Religiosa", 1),
        ("Educación Física", 1),
    ],
    "6to": [
        ("Redes Sociales", 2),
        ("Producción Audiovisual", 4),
        ("Videoarte", 5),
        ("Animación", 4),
        ("Edición, Sonido y Musicalización", 4),
        ("Producción de Proyecto Emprendedor", 4),
        ("Lengua Española", 3),
        ("Inglés", 4),
        ("Matemática", 3),
        ("Ciencias Sociales", 2),
        ("Ciencias de la Naturaleza", 3),
        ("Formación Integral Humana y Religiosa", 1),
        ("Educación Física", 1),
    ],
}

def _get_profesor():
    """Retorna dict del usuario profesor autenticado o None."""
    uid = session.get("user_id")
    if not uid:
        return None
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        u = conn.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    return dict(u) if u else None

def _validar_materia_profesor(nombre_materia, nombre_curso, profesor):
    """
    Valida si una materia/curso corresponde al perfil asignado del profesor.
    Retorna (ok: bool, mensaje: str)
    """
    if not profesor or profesor.get("rol") == "coordinador":
        return True, ""

    prof_grado   = (profesor.get("grado") or "").lower()
    prof_mencion = (profesor.get("mencion") or "").lower()
    prof_asigs   = [a.strip().lower() for a in (profesor.get("asignaturas") or "").split(",") if a.strip()]

    curso_lower  = (nombre_curso or "").lower()
    materia_lower = (nombre_materia or "").lower()

    # Verificar grado
    if prof_grado and prof_grado not in curso_lower:
        return False, (
            f"Esta nota no corresponde a tu grado asignado ({profesor.get('grado').upper()}). "
            f"El archivo parece ser del curso '{nombre_curso}'. "
            "Verifica que estás subiendo el archivo correcto."
        )

    # Verificar mención
    if prof_mencion and prof_mencion not in curso_lower:
        # Allow if mencion is in subject name (some subjects span menciones)
        if not any(prof_mencion in a for a in [materia_lower, curso_lower]):
            return False, (
                f"La nota no corresponde a la mención/modalidad asignada a tu perfil ({profesor.get('mencion')}). "
                f"El archivo indica el curso '{nombre_curso}'. "
                "Solo puedes cargar notas de tu mención."
            )

    # Verificar materia
    if prof_asigs:
        match = any(
            asig in materia_lower or materia_lower in asig
            for asig in prof_asigs
        )
        if not match:
            return False, (
                f"La materia '{nombre_materia}' no está en tu lista de asignaturas asignadas "
                f"({profesor.get('asignaturas')}). "
                "Si crees que hay un error, contacta al coordinador."
            )

    return True, ""


@app.route("/api/plan-estudio")
def plan_estudio():
    """Retorna el plan de estudio oficial MINERD para Multimedia."""
    grado = request.args.get("grado", "")
    if grado in PLAN_MULTIMEDIA:
        return jsonify({
            "grado": grado,
            "asignaturas": [{"nombre": n, "horas_semana": h} for n, h in PLAN_MULTIMEDIA[grado]]
        })
    return jsonify({
        "todos": {g: [{"nombre": n, "horas_semana": h} for n, h in lst]
                  for g, lst in PLAN_MULTIMEDIA.items()}
    })


# ── ASISTENCIA ─────────────────────────────────────────────────────────────

@app.route("/api/asistencia", methods=["GET"])
@login_required
def listar_asistencia():
    prof = _get_profesor()
    est_id    = request.args.get("estudiante_id", "")
    materia   = request.args.get("materia", "")
    periodo   = request.args.get("periodo", "")
    prof_id   = request.args.get("profesor_id", "")
    fecha_ini = request.args.get("fecha_ini", "")
    fecha_fin = request.args.get("fecha_fin", "")

    q = """
        SELECT a.*, e.nombre, e.apellido, e.curso
        FROM asistencia a
        JOIN estudiantes e ON e.id = a.estudiante_id
        WHERE 1=1
    """
    params = []

    # Profesores solo ven su propia asistencia
    if prof and prof.get("rol") == "profesor":
        q += " AND a.profesor_id=?"; params.append(prof["id"])
    elif prof_id:
        q += " AND a.profesor_id=?"; params.append(int(prof_id))

    if est_id:   q += " AND a.estudiante_id=?"; params.append(int(est_id))
    if materia:  q += " AND a.materia=?";        params.append(materia)
    if periodo:  q += " AND a.periodo=?";         params.append(int(periodo))
    if fecha_ini: q += " AND a.fecha>=?";          params.append(fecha_ini)
    if fecha_fin: q += " AND a.fecha<=?";          params.append(fecha_fin)

    q += " ORDER BY a.fecha DESC, e.apellido"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/asistencia", methods=["POST"])
@login_required
def registrar_asistencia():
    """Registra asistencia masiva para una clase (lista de estudiantes)."""
    prof = _get_profesor()
    d = request.get_json(silent=True) or {}

    materia  = d.get("materia", "").strip()
    periodo  = d.get("periodo", 1)
    fecha    = d.get("fecha", "")
    horas    = d.get("horas_clase", 1)
    registros= d.get("registros", [])  # [{estudiante_id, estado, observacion}]

    if not materia or not fecha or not registros:
        return jsonify({"error": "materia, fecha y registros son requeridos"}), 400

    prof_id = prof["id"] if prof else 0

    with sqlite3.connect(DATABASE) as conn:
        for r in registros:
            est_id = r.get("estudiante_id")
            if not est_id:
                continue
            # Upsert: si ya existe esa fecha+estudiante+materia, actualiza
            existing = conn.execute(
                "SELECT id FROM asistencia WHERE estudiante_id=? AND materia=? AND fecha=? AND profesor_id=?",
                (est_id, materia, fecha, prof_id)
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE asistencia SET estado=?, horas_clase=?, observacion=?, periodo=? WHERE id=?",
                    (r.get("estado","presente"), horas, r.get("observacion",""), periodo, existing[0])
                )
            else:
                conn.execute(
                    """INSERT INTO asistencia
                       (estudiante_id, profesor_id, materia, fecha, periodo, estado, horas_clase, observacion)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (est_id, prof_id, materia, fecha, periodo,
                     r.get("estado","presente"), horas, r.get("observacion",""))
                )
        conn.commit()
    return jsonify({"ok": True, "registrados": len(registros)})


@app.route("/api/asistencia/resumen/<int:est_id>")
@login_required
def resumen_asistencia(est_id):
    """
    Calcula resumen de asistencia por materia y período.
    Fórmula MINERD: % = (horas_presentes / horas_totales) × 100
    Umbral reprobación: <80% (margen sobre el 20% de inasistencias permitidas)
    """
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT materia, periodo,
                   SUM(horas_clase) as horas_total,
                   SUM(CASE WHEN estado='presente' THEN horas_clase ELSE 0 END) as horas_presente,
                   SUM(CASE WHEN estado='tardanza' THEN horas_clase ELSE 0 END) as horas_tardanza,
                   SUM(CASE WHEN estado='ausente' THEN horas_clase ELSE 0 END)  as horas_ausente,
                   SUM(CASE WHEN estado='justificado' THEN horas_clase ELSE 0 END) as horas_justif,
                   COUNT(*) as total_dias
            FROM asistencia
            WHERE estudiante_id=?
            GROUP BY materia, periodo
            ORDER BY materia, periodo
        """, (est_id,)).fetchall()

    resultado = []
    for r in rows:
        total  = r["horas_total"] or 1
        pres   = (r["horas_presente"] or 0) + (r["horas_tardanza"] or 0) * 0.5
        pct    = round(pres / total * 100, 1)
        inasist_pct = round((r["horas_ausente"] or 0) / total * 100, 1)
        alerta = inasist_pct > 15  # Alerta temprana (MINERD reprueba >20%)
        critico= inasist_pct >= 20  # Ya en zona de reprobación

        resultado.append({
            "materia":         r["materia"],
            "periodo":         r["periodo"],
            "horas_total":     r["horas_total"],
            "horas_presente":  r["horas_presente"],
            "horas_ausente":   r["horas_ausente"],
            "horas_tardanza":  r["horas_tardanza"],
            "horas_justif":    r["horas_justif"],
            "pct_asistencia":  pct,
            "pct_inasistencia":inasist_pct,
            "alerta":          alerta,
            "critico":         critico,
            "total_dias":      r["total_dias"],
        })
    return jsonify(resultado)


@app.route("/api/asistencia/clase")
@login_required
def asistencia_por_clase():
    """Retorna lista de estudiantes para pasar lista, filtrada por profesor."""
    prof = _get_profesor()
    materia = request.args.get("materia", "")
    grado   = request.args.get("grado", "")

    # Build filter
    q = "SELECT id, nombre, apellido, curso, grado FROM estudiantes WHERE 1=1"
    params = []

    if prof and prof.get("rol") == "profesor":
        if prof.get("grado"):
            q += " AND grado LIKE ?"; params.append(f"%{prof['grado']}%")
        if prof.get("mencion"):
            q += " AND curso LIKE ?"; params.append(f"%{prof['mencion']}%")
    elif grado:
        q += " AND grado LIKE ?"; params.append(f"%{grado}%")

    q += " ORDER BY apellido, nombre"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])


# ── IMPORTAR LISTA DE PROFESORES (Excel) ────────────────────────────────────

@app.route("/api/importar-profesores", methods=["POST"])
@coord_required
def importar_profesores():
    """
    Importa un Excel con columnas:
    nombre | usuario | password | rol | grado | mencion | asignaturas
    Crea o actualiza profesores en lote.
    """
    f = request.files.get("archivo")
    if not f:
        return jsonify({"error": "No se envió archivo"}), 400

    try:
        import pandas as pd
        df = pd.read_excel(f, dtype=str).fillna("")
    except Exception as ex:
        return jsonify({"error": f"No se pudo leer el Excel: {ex}"}), 400

    # Normalize column names
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    requeridas = {"nombre", "usuario"}
    if not requeridas.issubset(set(df.columns)):
        return jsonify({"error": f"El Excel debe tener columnas: nombre, usuario (y opcionalmente: password, rol, grado, mencion, asignaturas)"}), 400

    creados = 0
    actualizados = 0
    errores = []

    with sqlite3.connect(DATABASE) as conn:
        for _, row in df.iterrows():
            try:
                nombre   = str(row.get("nombre", "")).strip()
                username = str(row.get("usuario", "")).strip()
                password = str(row.get("password", "")).strip() or username + "123"
                rol      = str(row.get("rol", "profesor")).strip() or "profesor"
                grado    = str(row.get("grado", "")).strip()
                mencion  = str(row.get("mencion", "")).strip()
                asigs    = str(row.get("asignaturas", "")).strip()
                materia  = asigs.split(",")[0].strip() if asigs else ""

                if not nombre or not username:
                    continue

                existing = conn.execute(
                    "SELECT id FROM usuarios WHERE username=?", (username,)
                ).fetchone()

                if existing:
                    conn.execute("""
                        UPDATE usuarios SET nombre=?, rol=?, grado=?, mencion=?, asignaturas=?, materia=?
                        WHERE username=?
                    """, (nombre, rol, grado, mencion, asigs, materia, username))
                    actualizados += 1
                else:
                    conn.execute("""
                        INSERT INTO usuarios (username,password,nombre,rol,materia,grado,mencion,asignaturas)
                        VALUES (?,?,?,?,?,?,?,?)
                    """, (username, _hash(password), nombre, rol, materia, grado, mencion, asigs))
                    creados += 1
            except Exception as ex:
                errores.append(f"Fila {_}: {ex}")

        conn.commit()

    return jsonify({
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores
    })


# ── VALIDACIÓN EN CARGA (hook para cargar-materia) ──────────────────────────

@app.route("/api/validar-materia-profesor", methods=["POST"])
@login_required
def validar_materia_profesor():
    """
    Valida si una materia+curso es compatible con el perfil del profesor.
    Llamado ANTES de confirmar la carga.
    """
    prof = _get_profesor()
    d = request.get_json(silent=True) or {}
    nombre_materia = d.get("materia", "")
    nombre_curso   = d.get("curso", "")

    ok, msg = _validar_materia_profesor(nombre_materia, nombre_curso, prof)
    return jsonify({"ok": ok, "mensaje": msg})


# ── PORTAL PROFESOR (página) ─────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS CALIFICACIONES
# ═══════════════════════════════════════════════════════════════════════════════

def _anio_escolar_actual():
    """Detecta automáticamente el año escolar según la fecha.
    El año escolar dominicano comienza en agosto.
    Si estamos en agosto-diciembre → 'YYYY-(YYY+1)'
    Si estamos en enero-julio  → '(YYYY-1)-YYYY'
    """
    from datetime import date
    hoy = date.today()
    if hoy.month >= 8:
        return f"{hoy.year}-{hoy.year + 1}"
    else:
        return f"{hoy.year - 1}-{hoy.year}"


def _periodo_actual():
    """Detecta el período activo por fecha.
    P1: agosto–octubre
    P2: noviembre–enero
    P3: febrero–abril
    P4: mayo–julio
    """
    from datetime import date
    m = date.today().month
    if m in (8, 9, 10):     return "P1"
    if m in (11, 12, 1):    return "P2"
    if m in (2, 3, 4):      return "P3"
    return "P4"  # 5, 6, 7


def _nota_estado(nota):
    """Clasifica la nota según MINERD."""
    if nota is None: return "sin_nota"
    if nota >= 70:   return "aprobado"
    if nota >= 50:   return "completiva"
    return "reprobado"


def _color_nota(nota):
    if nota is None: return "#555"
    if nota >= 70:   return "#4dffb4"
    if nota >= 50:   return "#f7b731"
    return "#ff4d4d"


# ═══════════════════════════════════════════════════════════════════════════════
#  ENDPOINTS CALIFICACIONES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/calificaciones", methods=["GET"])
@login_required
def listar_calificaciones():
    """
    Lista calificaciones con filtros opcionales.
    Profesores ven solo las que ellos registraron.
    Coordinadores ven todas.
    ?estudiante_id=  &materia=  &periodo=  &anio=
    """
    prof = _get_profesor()
    est_id  = request.args.get("estudiante_id", "")
    materia = request.args.get("materia", "")
    periodo = request.args.get("periodo", "")
    anio    = request.args.get("anio", "")

    q = """
        SELECT cp.*, e.nombre, e.apellido, e.curso, e.grado,
               u.nombre AS profesor_nombre
        FROM calificaciones_periodo cp
        JOIN estudiantes e ON e.id = cp.estudiante_id
        JOIN usuarios u    ON u.id = cp.profesor_id
        WHERE 1=1
    """
    params = []

    if prof and prof.get("rol") == "profesor":
        q += " AND cp.profesor_id=?"; params.append(prof["id"])

    if est_id:  q += " AND cp.estudiante_id=?"; params.append(int(est_id))
    if materia: q += " AND cp.materia=?";        params.append(materia)
    if periodo: q += " AND cp.periodo=?";        params.append(periodo)
    if anio:    q += " AND cp.anio_escolar=?";   params.append(anio)
    else:       q += " AND cp.anio_escolar=?";   params.append(_anio_escolar_actual())

    q += " ORDER BY e.apellido, cp.materia, cp.periodo"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(q, params).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route("/api/calificaciones", methods=["POST"])
@login_required
def registrar_calificacion():
    """
    Registra o actualiza la nota de un período.
    UPSERT por (estudiante_id + materia + periodo + anio_escolar).
    Body: {estudiante_id, materia, periodo, calificacion, observacion?, anio_escolar?}
    Puede recibir una lista (batch) o un objeto único.
    """
    prof = _get_profesor()
    if not prof:
        return jsonify({"error": "No autenticado"}), 401

    data = request.get_json(silent=True) or {}

    # Soporte batch (lista) o single (objeto)
    registros = data if isinstance(data, list) else [data]

    prof_id = prof["id"]
    anio    = _anio_escolar_actual()
    guardados = 0
    errores   = []

    with sqlite3.connect(DATABASE) as conn:
        for item in registros:
            est_id  = item.get("estudiante_id")
            materia = (item.get("materia") or "").strip()
            periodo = (item.get("periodo") or "").strip().upper()
            nota    = item.get("calificacion")
            obs     = item.get("observacion", "")
            item_anio = item.get("anio_escolar", anio)

            if not est_id or not materia or not periodo or nota is None:
                errores.append(f"Datos incompletos: {item}")
                continue

            # Validar rango
            try:
                nota = float(nota)
                if nota < 0 or nota > 100:
                    raise ValueError
            except (ValueError, TypeError):
                errores.append(f"Calificación inválida ({nota}) para est {est_id}")
                continue

            # Validar período
            if periodo not in ("P1", "P2", "P3", "P4"):
                errores.append(f"Período inválido: {periodo}")
                continue

            existing = conn.execute(
                "SELECT id FROM calificaciones_periodo "
                "WHERE estudiante_id=? AND materia=? AND periodo=? AND anio_escolar=?",
                (est_id, materia, periodo, item_anio)
            ).fetchone()

            if existing:
                conn.execute(
                    "UPDATE calificaciones_periodo "
                    "SET calificacion=?, observacion=?, profesor_id=?, actualizado=datetime('now') "
                    "WHERE id=?",
                    (nota, obs, prof_id, existing[0])
                )
            else:
                conn.execute(
                    "INSERT INTO calificaciones_periodo "
                    "(estudiante_id,profesor_id,materia,periodo,calificacion,anio_escolar,observacion) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (est_id, prof_id, materia, periodo, nota, item_anio, obs)
                )
            guardados += 1
        conn.commit()

    return jsonify({"ok": True, "guardados": guardados, "errores": errores})


@app.route("/api/calificaciones/resumen/<int:est_id>")
@login_required
def resumen_calificaciones(est_id):
    """
    Devuelve nota por período (P1-P4) y calcula:
    - Nota final anual = promedio de los períodos disponibles
    - Nota semestral 1er = (P1+P2)/2, 2do = (P3+P4)/2
    - Estado: aprobado / completiva / reprobado / sin_nota
    - Incluye también resumen de asistencia MINERD por materia
    """
    anio = request.args.get("anio", _anio_escolar_actual())

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row

        # Notas
        notas_rows = conn.execute(
            "SELECT materia, periodo, calificacion, observacion, profesor_id, actualizado "
            "FROM calificaciones_periodo "
            "WHERE estudiante_id=? AND anio_escolar=? "
            "ORDER BY materia, periodo",
            (est_id, anio)
        ).fetchall()

        # Asistencia MINERD
        asist_rows = conn.execute("""
            SELECT materia, periodo,
                   SUM(horas_clase) AS horas_total,
                   SUM(CASE WHEN estado IN ('P','presente') THEN horas_clase ELSE 0 END) AS horas_presente,
                   SUM(CASE WHEN estado IN ('J','justificado') THEN horas_clase ELSE 0 END) AS horas_justif,
                   SUM(CASE WHEN estado IN ('A','ausente') THEN horas_clase ELSE 0 END) AS horas_ausente,
                   SUM(CASE WHEN estado IN ('T','tardanza') THEN horas_clase * 0.5 ELSE 0 END) AS horas_tardanza_peso
            FROM asistencia
            WHERE estudiante_id=?
            GROUP BY materia, periodo
        """, (est_id,)).fetchall()

    # Organizar notas por materia
    from collections import defaultdict
    materias = defaultdict(lambda: {"P1": None, "P2": None, "P3": None, "P4": None})
    for r in notas_rows:
        materias[r["materia"]][r["periodo"]] = r["calificacion"]

    # Organizar asistencia por materia (acumular todos los períodos)
    asist_por_materia = defaultdict(lambda: {"horas_total": 0, "horas_ausente": 0, "horas_tardanza_peso": 0})
    for r in asist_rows:
        m = r["materia"]
        asist_por_materia[m]["horas_total"]          += (r["horas_total"] or 0)
        asist_por_materia[m]["horas_ausente"]         += (r["horas_ausente"] or 0)
        asist_por_materia[m]["horas_tardanza_peso"]   += (r["horas_tardanza_peso"] or 0)

    resultado = []
    for materia, periodos in materias.items():
        valores = [v for v in periodos.values() if v is not None]
        nota_final = round(sum(valores) / len(valores), 1) if valores else None

        # Semestral
        s1_vals = [periodos["P1"], periodos["P2"]]
        s2_vals = [periodos["P3"], periodos["P4"]]
        s1 = round(sum(v for v in s1_vals if v is not None) / len([v for v in s1_vals if v is not None]), 1) if any(v is not None for v in s1_vals) else None
        s2 = round(sum(v for v in s2_vals if v is not None) / len([v for v in s2_vals if v is not None]), 1) if any(v is not None for v in s2_vals) else None

        # Asistencia
        ai = asist_por_materia.get(materia, {})
        ht = ai.get("horas_total", 0)
        ha_injustif = (ai.get("horas_ausente", 0) or 0) + (ai.get("horas_tardanza_peso", 0) or 0)
        pct_inasist = round(ha_injustif / ht * 100, 1) if ht > 0 else 0
        reprueba_asist = pct_inasist >= 20

        resultado.append({
            "materia":          materia,
            "anio_escolar":     anio,
            "periodos":         periodos,
            "nota_final":       nota_final,
            "semestre_1":       s1,
            "semestre_2":       s2,
            "estado":           _nota_estado(nota_final),
            "color":            _color_nota(nota_final),
            "periodos_con_nota": len(valores),
            "pct_inasistencia": pct_inasist,
            "reprueba_asistencia": reprueba_asist,
            "alerta_asistencia": pct_inasist >= 15,
        })

    resultado.sort(key=lambda x: x["materia"])
    return jsonify({
        "estudiante_id": est_id,
        "anio_escolar":  anio,
        "materias":      resultado,
        "periodo_actual": _periodo_actual(),
        "total_materias": len(resultado),
        "materias_riesgo": sum(1 for m in resultado if m["reprueba_asistencia"] or m["estado"] in ("reprobado","completiva")),
    })


@app.route("/api/calificaciones/reporte-grupo")
@login_required
def reporte_grupo_calificaciones():
    """
    Reporte de grupo: todas las notas de un grado/mención por período.
    Solo coordinadores o profesores de ese grado.
    ?grado= &mencion= &periodo= &anio=
    """
    grado   = request.args.get("grado", "").strip()
    mencion = request.args.get("mencion", "").upper().strip()
    periodo = request.args.get("periodo", "").strip().upper()
    anio    = request.args.get("anio", _anio_escolar_actual())

    prof = _get_profesor()
    # Profesores solo ven su grado
    if prof and prof.get("rol") == "profesor":
        if grado and prof.get("grado") and grado.lower() not in prof.get("grado","").lower():
            return jsonify({"error": "Sin permisos para ese grado"}), 403

    q = """
        SELECT e.id, e.nombre, e.apellido, e.curso, e.grado,
               cp.materia, cp.periodo, cp.calificacion
        FROM calificaciones_periodo cp
        JOIN estudiantes e ON e.id = cp.estudiante_id
        WHERE cp.anio_escolar=?
    """
    params = [anio]

    if grado:
        q += " AND (e.grado LIKE ? OR e.curso LIKE ?)"; params += [f"%{grado}%", f"%{grado}%"]
    if mencion:
        q += " AND UPPER(e.curso) LIKE ?"; params.append(f"%{mencion}%")
    if periodo:
        q += " AND cp.periodo=?"; params.append(periodo)

    q += " ORDER BY e.apellido, e.nombre, cp.materia, cp.periodo"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(q, params).fetchall()

    # Pivot: {estudiante: {materia: {periodo: nota}}}
    from collections import defaultdict
    pivot = defaultdict(lambda: defaultdict(dict))
    info_est = {}

    for r in rows:
        eid = r["id"]
        info_est[eid] = {
            "id": eid, "nombre": r["nombre"], "apellido": r["apellido"],
            "curso": r["curso"], "grado": r["grado"]
        }
        pivot[eid][r["materia"]][r["periodo"]] = r["calificacion"]

    resultado = []
    for eid, materias_dict in pivot.items():
        notas_est = []
        for materia, periodos in materias_dict.items():
            vals = list(periodos.values())
            nota_final = round(sum(vals)/len(vals), 1) if vals else None
            notas_est.append({
                "materia": materia,
                "periodos": periodos,
                "nota_final": nota_final,
                "estado": _nota_estado(nota_final)
            })
        resultado.append({
            "estudiante": info_est[eid],
            "notas": notas_est
        })

    resultado.sort(key=lambda x: x["estudiante"]["apellido"])
    return jsonify({
        "anio_escolar": anio,
        "grado": grado, "mencion": mencion, "periodo_filtro": periodo,
        "total_estudiantes": len(resultado),
        "estudiantes": resultado
    })


@app.route("/api/calificaciones/boletin/<int:est_id>")
@login_required
def boletin_estudiante(est_id):
    """
    Boletín completo del estudiante: notas + asistencia por materia.
    Incluye estado final, alertas MINERD y datos del estudiante.
    """
    anio = request.args.get("anio", _anio_escolar_actual())

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        est = conn.execute(
            "SELECT * FROM estudiantes WHERE id=?", (est_id,)
        ).fetchone()

    if not est:
        return jsonify({"error": "Estudiante no encontrado"}), 404

    # Reusar resumen_calificaciones
    from flask import g
    with app.test_request_context(f"/api/calificaciones/resumen/{est_id}?anio={anio}"):
        pass

    # Llamar directamente la lógica
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row

        notas_rows = conn.execute(
            "SELECT materia, periodo, calificacion FROM calificaciones_periodo "
            "WHERE estudiante_id=? AND anio_escolar=? ORDER BY materia, periodo",
            (est_id, anio)
        ).fetchall()

        asist_rows = conn.execute("""
            SELECT materia,
                   SUM(horas_clase) AS ht,
                   SUM(CASE WHEN estado IN ('A','ausente') THEN horas_clase ELSE 0 END) AS ha,
                   SUM(CASE WHEN estado IN ('T','tardanza') THEN horas_clase*0.5 ELSE 0 END) AS ht_peso,
                   SUM(CASE WHEN estado IN ('P','presente') THEN 1 ELSE 0 END) AS dias_presentes,
                   COUNT(*) AS total_dias
            FROM asistencia
            WHERE estudiante_id=?
            GROUP BY materia
        """, (est_id,)).fetchall()

    from collections import defaultdict
    materias = defaultdict(lambda: {"P1": None, "P2": None, "P3": None, "P4": None})
    for r in notas_rows:
        materias[r["materia"]][r["periodo"]] = r["calificacion"]

    asist_m = {}
    for r in asist_rows:
        ht = r["ht"] or 0
        ha_total = (r["ha"] or 0) + (r["ht_peso"] or 0)
        asist_m[r["materia"]] = {
            "horas_total": ht,
            "pct_inasist": round(ha_total / ht * 100, 1) if ht > 0 else 0,
            "dias_presentes": r["dias_presentes"],
            "total_dias": r["total_dias"],
        }

    boletin_materias = []
    for materia, periodos in materias.items():
        vals = [v for v in periodos.values() if v is not None]
        nota_final = round(sum(vals)/len(vals), 1) if vals else None
        ai = asist_m.get(materia, {"pct_inasist": 0, "horas_total": 0})
        boletin_materias.append({
            "materia": materia,
            "P1": periodos["P1"], "P2": periodos["P2"],
            "P3": periodos["P3"], "P4": periodos["P4"],
            # Aliases in lowercase for frontend compatibility
            "p1": periodos["P1"], "p2": periodos["P2"],
            "p3": periodos["P3"], "p4": periodos["P4"],
            "nota_final": nota_final,
            "promedio": nota_final,
            "estado": _nota_estado(nota_final),
            "pct_inasistencia": ai["pct_inasist"],
            "pct_inasistencia_injustificada": ai["pct_inasist"],
            "reprueba_asistencia": ai["pct_inasist"] >= 20,
            "alerta_asistencia": ai["pct_inasist"] >= 15,
        })

    boletin_materias.sort(key=lambda x: x["materia"])

    mat_reprobadas = [m for m in boletin_materias if m["estado"] == "reprobado" or m["reprueba_asistencia"]]
    mat_completiva = [m for m in boletin_materias if m["estado"] == "completiva"]

    return jsonify({
        "estudiante": dict(est),
        "anio_escolar": anio,
        "periodo_actual": _periodo_actual(),
        "materias": boletin_materias,
        "resumen": {
            "total_materias": len(boletin_materias),
            "aprobadas": sum(1 for m in boletin_materias if m["estado"] == "aprobado"),
            "completivas": len(mat_completiva),
            "reprobadas": len(mat_reprobadas),
            "promueve": len(mat_reprobadas) < 4,
            "repite": len(mat_reprobadas) >= 4,
        }
    })


@app.route("/profesor")
@login_required
def portal_profesor():
    prof = _get_profesor()
    if not prof:
        return redirect("/login")
    if prof.get("rol") == "coordinador":
        return redirect("/")

    # Get assigned students
    q = "SELECT id, nombre, apellido, curso, grado FROM estudiantes WHERE 1=1"
    params = []
    if prof.get("grado"):
        q += " AND grado LIKE ?"; params.append(f"%{prof['grado']}%")
    if prof.get("mencion"):
        q += " AND curso LIKE ?"; params.append(f"%{prof['mencion']}%")
    q += " ORDER BY apellido, nombre"

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        estudiantes = [dict(r) for r in conn.execute(q, params).fetchall()]

    # Plan de estudio for assigned grade
    grado_key = (prof.get("grado") or "4to").split()[0]
    plan = PLAN_MULTIMEDIA.get(grado_key, [])

    from datetime import date as _date
    return render_template(
        "profesor.html",
        profesor=prof,
        estudiantes=estudiantes,
        plan=plan,
        fecha_hoy=_date.today().isoformat(),
        current_user=get_usuario()
    )





# ══════════════════════════════════════════════════════════════════════════════
#  PARSER DE BOLETINES OFICIALES — C.E. BENITO JUÁREZ
#  Soporta:
#    - Boletín Primer Ciclo (1ro, 2do, 3ro): solo materias académicas, secciones A-E
#    - Boletín Segundo Ciclo (4to, 5to, 6to): académicas + técnicas por mención
#
#  Estructura del Excel (ambos ciclos):
#    - Una hoja por sección (1ro) o por mención-sección (4to)
#    - Cada estudiante ocupa un bloque de ~43 filas que se repite
#    - idx[2]='ALUMNO/A:', idx[6]=nombre completo del estudiante
#    - Académicas: idx[2]=materia, P1=idx[3], P2=idx[5], P3=idx[7], P4=idx[9]
#    - Técnicas:   idx[1]=materia, P1=idx[3], P2=idx[4]
# ══════════════════════════════════════════════════════════════════════════════

def _limpiar_nota(v):
    """
    Convierte celda de Excel a float entre 1-100, o 0 si no es válida.
    Acepta: int, float, strings '75', '73.5', '85,0' (coma decimal).
    Ignora: None, booleans, textos, errores Excel.
    """
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            return round(f, 2) if 1 <= f <= 100 else 0.0
        except Exception:
            return 0.0
    # String: limpiar espacios y convertir coma decimal a punto
    s = str(v).strip().replace(',', '.')
    if not s or s in ('#N/A', '#REF!', '#VALUE!', '#DIV/0!', '-', 'N/A', 'None', 'False', 'True'):
        return 0.0
    if s.isalpha():   # texto puro como 'P1', 'PC'
        return 0.0
    try:
        f = float(s)
        return round(f, 2) if 1 <= f <= 100 else 0.0
    except (ValueError, TypeError):
        return 0.0


def _parsear_boletin_bj(file_bytes):
    """
    Parser del Boletín Oficial del C.E. Benito Juárez.
    Estructura verificada con archivos reales:
      - Bloque de 43 filas por estudiante (4to-6to) / 41 filas (1ro-3ro)
      - col[2]=='ALUMNO/A:' → nombre en col[6]
      - col[2]=='MAESTRO/A:' → maestro en col[6]
      - Académicas: col[2]=materia, P1=col[3], P2=col[5], P3=col[7], P4=col[9]
      - Técnicas (2do ciclo): col[1]='COMPONENTE TÉCNICO' marca inicio,
        luego col[1]=materia, P1=col[3], P2=col[4]
    """
    from openpyxl import load_workbook
    import io as _io, re as _re

    SKIP_MATERIAS = {
        'promedio de grupos', 'calificación final', 'calificacion final',
        'condición', 'condicion', 'el estudiante', 'felicidades',
        'competencias', 'períodos', 'periodos', 'pc1', 'pc2', 'pc3', 'pc4',
        'componete académico', 'componete academico',
        'áreas curriculares', 'areas curriculares',
        'componente técnico', 'componete tecnico',
        'ninguno', 'yordania', 'maestro/a encargado',
        '50%', 'c.e.c', 'c.c.f', 'promedio',
    }

    wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    estudiantes = []

    for sheet_name in wb.sheetnames:
        sn = sheet_name.upper()

        # Detectar grado desde nombre de hoja
        grado = ''
        for g in ['6TO', '5TO', '4TO', '3RO', '2DO', '1RO']:
            if g in sn:
                grado = g.upper()
                break
        if not grado:
            continue

        ciclo = 'primer_ciclo' if grado in ['1RO', '2DO', '3RO'] else 'segundo_ciclo'

        # Mención (solo 2do ciclo)
        mencion = ''
        if ciclo == 'segundo_ciclo':
            for m in ['MULTIMEDIA', 'TEATRO', 'VISUALES', 'VISUAL', 'MÚSICA', 'MUSICA']:
                if m in sn:
                    mencion = m
                    break
            if 'VISUAL' in mencion:
                mencion = 'ARTES VISUALES'
            elif 'MUSICA' in mencion or 'MÚSICA' in mencion:
                mencion = 'MÚSICA'

        # Sección (letra A-E del nombre de hoja)
        seccion = 'A'
        ms = _re.search(r'\b([A-E])\b', sheet_name)
        if ms:
            seccion = ms.group(1)

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        i = 0

        while i < len(rows):
            r = rows[i]
            # Inicio de bloque de estudiante
            if len(r) > 6 and str(r[2] or '').strip().upper() in ('ALUMNO/A:', 'ALUMNO/A', 'ALUMNA/O:') and r[6]:
                nombre_completo = " ".join(str(r[6]).strip().split())

                # Maestro en la fila siguiente
                maestro = ''
                if i + 1 < len(rows) and str(rows[i + 1][2] or '').strip().upper() in ('MAESTRO/A:', 'MAESTRA/O:', 'MAESTRO/A'):
                    maestro = str(rows[i + 1][6] or '').strip()

                # Separar nombre y apellido (nombres hispanos dominicanos)
                # Patrón: [Nombre1] [Nombre2?] [Apellido1] [Apellido2?]
                partes = nombre_completo.split()
                if len(partes) >= 4:
                    # Ej: "Miguel Angel Martinez Medina" → nombre="Miguel Angel" apellido="Martinez Medina"
                    nombre   = ' '.join(partes[:2])
                    apellido = ' '.join(partes[2:])
                elif len(partes) == 3:
                    # Ambiguo: puede ser "Jose Junior Perez" (nombre compuesto + 1 apellido)
                    # o "Maria Garcia Lopez" (1 nombre + 2 apellidos)
                    # Guardamos nombre_completo para fuzzy matching flexible
                    nombre   = ' '.join(partes[:2])   # "Jose Junior"
                    apellido = partes[2]              # "Perez"
                elif len(partes) == 2:
                    nombre, apellido = partes[0], partes[1]
                else:
                    nombre, apellido = nombre_completo, ''

                materias = []
                en_tecnico = False
                j = i + 2

                while j < min(i + 52, len(rows)):
                    rj = rows[j]
                    if not rj or len(rj) < 3:
                        j += 1
                        continue

                    c0 = str(rj[0] or '').strip().upper()
                    c1 = str(rj[1] or '').strip()
                    c2 = str(rj[2] or '').strip()

                    # Fin de bloque por CONDICIÓN FINAL
                    if c0 and ('CONDICIÓN' in c0 or 'CONDICION' in c0):
                        break

                    # Inicio de siguiente estudiante
                    if str(c2).strip().upper() in ('ALUMNO/A:', 'ALUMNO/A', 'ALUMNA/O:'):
                        break

                    # Detectar sección técnica
                    if 'COMPONENTE TÉCNICO' in c1.upper() or 'COMPONETE TÉCNICO' in c1.upper():
                        en_tecnico = True
                        j += 1
                        continue

                    if not en_tecnico:
                        # — Materias académicas —
                        # Nombre en col[2], notas en cols 3,5,7,9
                        if c2 and c2 != ' ':
                            c2_low = c2.lower()
                            skip = any(s in c2_low for s in SKIP_MATERIAS)
                            if not skip and len(c2) > 1:
                                p1 = _limpiar_nota(rj[3] if len(rj) > 3 else None)
                                p2 = _limpiar_nota(rj[5] if len(rj) > 5 else None)
                                p3 = _limpiar_nota(rj[7] if len(rj) > 7 else None)
                                p4 = _limpiar_nota(rj[9] if len(rj) > 9 else None)
                                notas = [n for n in [p1, p2, p3, p4] if n > 0]
                                if notas:
                                    # Deduplicar: si ya existe la materia con notas, no sobreescribir
                                    nom_key = c2.lower().strip()
                                    ya_existe = any(m['nombre'].lower().strip() == nom_key for m in materias)
                                    if not ya_existe:
                                        materias.append({
                                            'nombre':   c2,
                                            'tipo':     'académico',
                                            'p1': p1, 'p2': p2, 'p3': p3, 'p4': p4,
                                            'promedio': round(sum(notas) / len(notas), 2)
                                        })
                    else:
                        # — Materias técnicas —
                        # Nombre en col[1], notas en cols 3 y 4
                        if c1 and c1 != ' ':
                            c1_low = c1.lower()
                            skip = any(s in c1_low for s in SKIP_MATERIAS)
                            if not skip and len(c1) > 3:
                                p1 = _limpiar_nota(rj[3] if len(rj) > 3 else None)
                                p2 = _limpiar_nota(rj[4] if len(rj) > 4 else None)
                                notas = [n for n in [p1, p2] if n > 0]
                                if notas:
                                    nom_key = c1.lower().strip()
                                    ya_existe = any(m['nombre'].lower().strip() == nom_key for m in materias)
                                    if not ya_existe:
                                        materias.append({
                                            'nombre':   c1,
                                            'tipo':     'técnico',
                                            'p1': p1, 'p2': p2, 'p3': 0.0, 'p4': 0.0,
                                            'promedio': round(sum(notas) / len(notas), 2)
                                        })
                    j += 1

                if nombre and materias:
                    estudiantes.append({
                        'nombre':   nombre.strip(),
                        'apellido': apellido.strip(),
                        'grado':    grado,
                        'seccion':  seccion,
                        'mencion':  mencion,
                        'ciclo':    ciclo,
                        'maestro':  maestro,
                        'materias': materias
                    })
                i = j
            else:
                i += 1

    wb.close()
    return estudiantes



def _buscar_o_crear_estudiante(conn, nombre, apellido, grado, ciclo, seccion, mencion):
    """
    Busca un estudiante usando _buscar_estudiante_bd.
    Intenta múltiples combinaciones de nombre/apellido para cubrir casos como
    'Jose Junior Perez' que puede estar guardado como 'Jose Junior / Perez'
    o 'Jose / Junior Perez'.
    Si no existe, lo crea. Retorna el id del estudiante.
    """
    nombre_raw   = nombre.strip()
    apellido_raw = apellido.strip()
    nombre_completo = f"{nombre_raw} {apellido_raw}".strip()

    # Generar todas las combinaciones posibles de split nombre/apellido
    partes = nombre_completo.split()
    candidatos = []
    if len(partes) >= 2:
        for split_at in range(1, len(partes)):
            n = ' '.join(partes[:split_at])
            a = ' '.join(partes[split_at:])
            candidatos.append((n, a))
    else:
        candidatos = [(nombre_raw, apellido_raw)]

    # Probar cada combinación
    for n, a in candidatos:
        est = _buscar_estudiante_bd(conn, n, a, filtro_grado=grado)
        if est:
            # Actualizar seccion y ciclo si cambiaron
            curso_nuevo = f"{grado} {mencion}".strip() if mencion else grado
            conn.execute(
                "UPDATE estudiantes SET seccion=?, ciclo=?, curso=? WHERE id=?",
                (seccion, ciclo, curso_nuevo, est['id'])
            )
            return est['id']

    # Último intento: buscar por nombre completo sin split (fuzzy global)
    est = _buscar_estudiante_bd(conn, nombre_completo, '', filtro_grado=grado)
    if est:
        conn.execute(
            "UPDATE estudiantes SET seccion=?, ciclo=? WHERE id=?",
            (seccion, ciclo, est['id'])
        )
        return est['id']

    # No encontrado — crear perfil nuevo con la separación original del parser
    curso = f"{grado} {mencion}".strip() if mencion else grado
    conn.execute(
        """INSERT INTO estudiantes (nombre, apellido, grado, curso, ciclo, seccion, condicion)
           VALUES (?, ?, ?, ?, ?, ?, 'ACTIVO')""",
        (nombre_raw, apellido_raw, grado, curso, ciclo, seccion)
    )
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return new_id


@app.route("/api/debug-boletin", methods=["POST"])
@login_required
def debug_boletin():
    """
    Diagnóstico del parser de boletín.
    Muestra las primeras filas de cada hoja y lo que detecta sin guardar nada.
    """
    if "file" not in request.files:
        return jsonify({"error": "Sin archivo"}), 400
    file_bytes = request.files["file"].read()

    from openpyxl import load_workbook
    import io as _io

    wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    resultado = {}

    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        # Primeras 15 filas para diagnóstico
        preview = []
        for i, r in enumerate(rows[:15]):
            preview.append({
                "fila": i + 1,
                "c0": str(r[0] or '')[:40] if len(r) > 0 else '',
                "c1": str(r[1] or '')[:40] if len(r) > 1 else '',
                "c2": str(r[2] or '')[:40] if len(r) > 2 else '',
                "c3": str(r[3] or '')[:20] if len(r) > 3 else '',
                "c6": str(r[6] or '')[:40] if len(r) > 6 else '',
            })
        # Buscar filas ALUMNO/A
        alumno_rows = []
        for i, r in enumerate(rows):
            v = str(r[2] or '').strip().upper() if len(r) > 2 else ''
            if 'ALUMNO' in v:
                alumno_rows.append({
                    "fila": i + 1,
                    "c2": v,
                    "c6": str(r[6] or '')[:50] if len(r) > 6 else ''
                })
            if len(alumno_rows) >= 3:
                break

        resultado[sheet_name] = {
            "total_filas": len(rows),
            "preview_15": preview,
            "alumno_rows": alumno_rows
        }

    wb.close()
    # También correr el parser real
    try:
        parsed = _parsear_boletin_bj(file_bytes)
        resultado["__parser_result"] = {
            "total_estudiantes": len(parsed),
            "primeros_2": parsed[:2] if parsed else []
        }
    except Exception as ex:
        resultado["__parser_error"] = str(ex)

    return jsonify(resultado)


@app.route("/api/cargar-boletin", methods=["POST"])
@login_required
def cargar_boletin():
    """
    Carga el Boletín oficial del C.E. Benito Juárez.
    Soporta:
      - Boletín Primer Ciclo (1ro-3ro): sheets BC 1RO A, BC 2DO B...
      - Boletín Segundo Ciclo (4to-6to): sheets BC 4TO A MÚSICA...
    Por cada estudiante:
      1. Busca o crea el perfil en 'estudiantes'
      2. Guarda/actualiza las notas en 'materias_calificaciones'
      3. Recalcula p_acad del estudiante
    """
    if "file" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({"error": "Formato no válido. Sube un archivo Excel (.xlsx/.xlsm)"}), 400

    try:
        file_bytes = file.read()
        print(f"[cargar_boletin] Recibido: {getattr(file, 'filename', '?')}, {len(file_bytes)} bytes")
        estudiantes_parsed = _parsear_boletin_bj(file_bytes)
        print(f"[cargar_boletin] Parser: {len(estudiantes_parsed)} estudiantes")

        if not estudiantes_parsed:
            # Diagnóstico: listar hojas encontradas para ayudar a depurar
            try:
                from openpyxl import load_workbook
                import io as _io2
                wb_diag = load_workbook(_io2.BytesIO(file_bytes), read_only=True, data_only=True)
                hojas = wb_diag.sheetnames
                wb_diag.close()
            except Exception:
                hojas = []
            return jsonify({
                "error": (
                    "No se encontraron estudiantes en el archivo. "
                    f"Hojas detectadas: {hojas}. "
                    "Verifica que sea un Boletín oficial del C.E. Benito Juárez "
                    "(4to–6to o 1ro–3ro) con el formato estándar."
                )
            }), 400

        from datetime import date as _date
        hoy = _date.today().isoformat()

        resumen        = {}
        total_est      = 0
        total_materias = 0
        creados        = 0
        errores        = []

        with sqlite3.connect(DATABASE) as conn:
            conn.row_factory = sqlite3.Row

            for est in estudiantes_parsed:
                try:
                    est_id = _buscar_o_crear_estudiante(
                        conn,
                        est['nombre'], est['apellido'],
                        est['grado'],  est['ciclo'],
                        est['seccion'], est['mencion']
                    )
                    if not est_id:
                        raise ValueError("No se pudo crear/encontrar el perfil del estudiante")

                    # Actualizar ciclo y sección en el perfil
                    conn.execute(
                        "UPDATE estudiantes SET ciclo=?, seccion=? WHERE id=?",
                        (est['ciclo'], est['seccion'], est_id)
                    )

                    # Insertar/actualizar cada materia
                    for mat in est['materias']:
                        notas = [mat['p1'], mat['p2'], mat['p3'], mat['p4']]
                        notas_validas = [n for n in notas if n > 0]
                        prom = round(sum(notas_validas) / len(notas_validas), 2) if notas_validas else 0

                        # Intentar INSERT completo con columnas extendidas
                        try:
                            conn.execute("""
                                INSERT INTO materias_calificaciones
                                    (estudiante_id, materia, p1, p2, p3, p4,
                                     promedio, tipo, ciclo, fecha_carga, profesor)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ON CONFLICT(estudiante_id, materia) DO UPDATE SET
                                    p1       = CASE WHEN excluded.p1 > 0 THEN excluded.p1 ELSE materias_calificaciones.p1 END,
                                    p2       = CASE WHEN excluded.p2 > 0 THEN excluded.p2 ELSE materias_calificaciones.p2 END,
                                    p3       = CASE WHEN excluded.p3 > 0 THEN excluded.p3 ELSE materias_calificaciones.p3 END,
                                    p4       = CASE WHEN excluded.p4 > 0 THEN excluded.p4 ELSE materias_calificaciones.p4 END,
                                    promedio = excluded.promedio,
                                    tipo     = excluded.tipo,
                                    ciclo    = excluded.ciclo,
                                    fecha_carga = excluded.fecha_carga,
                                    profesor = COALESCE(excluded.profesor, materias_calificaciones.profesor)
                            """, (
                                est_id, mat['nombre'],
                                mat['p1'], mat['p2'], mat['p3'], mat['p4'],
                                prom, mat['tipo'], est['ciclo'], hoy, est['maestro']
                            ))
                        except Exception:
                            # Fallback: INSERT básico sin columnas extendidas
                            conn.execute("""
                                INSERT INTO materias_calificaciones
                                    (estudiante_id, materia, p1, p2, p3, p4, promedio, fuente)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 'boletin')
                                ON CONFLICT(estudiante_id, materia) DO UPDATE SET
                                    p1       = CASE WHEN excluded.p1 > 0 THEN excluded.p1 ELSE materias_calificaciones.p1 END,
                                    p2       = CASE WHEN excluded.p2 > 0 THEN excluded.p2 ELSE materias_calificaciones.p2 END,
                                    p3       = CASE WHEN excluded.p3 > 0 THEN excluded.p3 ELSE materias_calificaciones.p3 END,
                                    p4       = CASE WHEN excluded.p4 > 0 THEN excluded.p4 ELSE materias_calificaciones.p4 END,
                                    promedio = excluded.promedio
                            """, (
                                est_id, mat['nombre'],
                                mat['p1'], mat['p2'], mat['p3'], mat['p4'], prom
                            ))
                        total_materias += 1

                    # Recalcular p_acad del estudiante
                    proms = conn.execute("""
                        SELECT promedio FROM materias_calificaciones
                        WHERE estudiante_id=? AND tipo='académico' AND promedio > 0
                    """, (est_id,)).fetchall()
                    if proms:
                        p_acad_nuevo = round(sum(r[0] for r in proms) / len(proms), 2)
                        # Calcular p_acad_p1 y p_acad_p2 para dashboard
                        p1s = conn.execute("""
                            SELECT p1 FROM materias_calificaciones
                            WHERE estudiante_id=? AND tipo='académico' AND p1 > 0
                        """, (est_id,)).fetchall()
                        p2s = conn.execute("""
                            SELECT p2 FROM materias_calificaciones
                            WHERE estudiante_id=? AND tipo='académico' AND p2 > 0
                        """, (est_id,)).fetchall()
                        acad_p1 = round(sum(r[0] for r in p1s)/len(p1s), 2) if p1s else 0
                        acad_p2 = round(sum(r[0] for r in p2s)/len(p2s), 2) if p2s else 0

                        conn.execute("""
                            UPDATE estudiantes
                               SET p_acad=?, acad_p1=?, acad_p2=?, tiene_notas=1
                             WHERE id=?
                        """, (p_acad_nuevo, acad_p1, acad_p2, est_id))
                        try:
                            conn.execute("UPDATE estudiantes SET tiene_notas=1 WHERE id=?", (est_id,))
                        except Exception:
                            pass

                    total_est += 1
                    clave = f"{est['grado']} {est['mencion']}".strip()
                    resumen[clave] = resumen.get(clave, 0) + 1

                except Exception as ex_est:
                    import traceback as _tb
                    errores.append(
                        f"{est.get('nombre','')} {est.get('apellido','')}: "
                        f"{ex_est} | {_tb.format_exc().splitlines()[-1]}"
                    )

            conn.commit()

        cache_bust()

        if total_est == 0:
            return jsonify({
                "ok": False,
                "error": (
                    errores[0] if errores else
                    "No se pudo procesar ningún estudiante. "
                    "Verifica que el archivo sea un Boletín oficial del C.E. Benito Juárez."
                ),
                "errores": errores[:10],
                "estudiantes_encontrados_en_archivo": len(estudiantes_parsed),
            }), 400

        cache_bust()

        if total_est == 0 and errores:
            return jsonify({
                "ok":    False,
                "error": f"No se guardó ningún estudiante. Error: {errores[0]}",
                "errores": errores[:5]
            }), 400

        return jsonify({
            "ok": True,
            "estudiantes_procesados": total_est,
            "materias_guardadas":     total_materias,
            "resumen_por_grupo":      resumen,
            "errores":                errores[:10],
            "mensaje": (
                f"✓ {total_est} estudiantes — {total_materias} calificaciones cargadas. "
                + (f"({len(errores)} advertencias)" if errores else "")
            )
        })

    except Exception as ex:
        import traceback
        tb = traceback.format_exc()
        print(f"[cargar_boletin] ERROR: {ex}\n{tb}")
        return jsonify({
            "error": f"Error al procesar el boletín: {str(ex)}",
            "detalle": tb
        }), 500



# ══════════════════════════════════════════════════════════════════════════════
#  CUADERNO ANECDÓTICO
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/cuaderno/<int:est_id>", methods=["GET"])
@login_required
def get_cuaderno(est_id):
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        # Verificar acceso al estudiante
        est = conn.execute("SELECT ciclo FROM estudiantes WHERE id=?", (est_id,)).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404
        ciclo_est = est["ciclo"] or "segundo_ciclo"
        ciclo_acc = u.get("ciclo_acceso")
        if ciclo_acc and ciclo_acc != ciclo_est and not u.get("es_directora"):
            return jsonify({"error": "Sin acceso a este estudiante"}), 403

        # Entradas del cuaderno
        rows = conn.execute("""
            SELECT ca.*, u.nombre as autor_nombre, u.rol as autor_rol
            FROM cuaderno_anecdotico ca
            JOIN usuarios u ON ca.autor_id = u.id
            WHERE ca.estudiante_id = ?
            ORDER BY ca.fecha DESC, ca.creado_en DESC
        """, (est_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/cuaderno", methods=["POST"])
@login_required
def crear_entrada_cuaderno():
    u = get_usuario()
    data = request.get_json() or {}
    est_id      = data.get("estudiante_id")
    tipo        = data.get("tipo", "conductual")
    descripcion = (data.get("descripcion") or "").strip()
    seguimiento = (data.get("seguimiento") or "").strip()
    fecha       = data.get("fecha") or __import__("datetime").date.today().isoformat()
    privado     = 1 if data.get("privado") else 0

    if not est_id or not descripcion:
        return jsonify({"error": "Faltan campos obligatorios"}), 400

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        est = conn.execute("SELECT ciclo FROM estudiantes WHERE id=?", (est_id,)).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        conn.execute("""
            INSERT INTO cuaderno_anecdotico
                (estudiante_id, autor_id, fecha, tipo, descripcion,
                 seguimiento, privado, visible_en_perfil)
            VALUES (?,?,?,?,?,?,?,1)
        """, (est_id, u["id"], fecha, tipo, descripcion, seguimiento, privado))
        conn.commit()
    cache_bust()
    return jsonify({"ok": True, "mensaje": "Entrada registrada en el cuaderno anecdótico"})


@app.route("/api/cuaderno/<int:entrada_id>", methods=["PATCH"])
@login_required
def editar_entrada_cuaderno(entrada_id):
    u = get_usuario()
    data = request.get_json() or {}
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        entrada = conn.execute(
            "SELECT * FROM cuaderno_anecdotico WHERE id=?", (entrada_id,)
        ).fetchone()
        if not entrada:
            return jsonify({"error": "Entrada no encontrada"}), 404
        # Solo el autor o roles superiores pueden editar
        rol_n = _normalizar_rol(u.get("rol",""))
        if entrada["autor_id"] != u["id"] and rol_n not in {"directora","coordinador_general","coordinador_primer_ciclo","coordinador_segundo_ciclo"}:
            return jsonify({"error": "Solo el autor puede editar esta entrada"}), 403

        conn.execute("""
            UPDATE cuaderno_anecdotico
               SET descripcion=?, seguimiento=?, tipo=?, privado=?
             WHERE id=?
        """, (
            data.get("descripcion", entrada["descripcion"]),
            data.get("seguimiento", entrada["seguimiento"]),
            data.get("tipo", entrada["tipo"]),
            1 if data.get("privado") else 0,
            entrada_id
        ))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/cuaderno/<int:entrada_id>/convertir-reporte", methods=["POST"])
@login_required
def convertir_reporte_cuaderno(entrada_id):
    """Convierte una entrada del cuaderno en reporte visible en el perfil."""
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        entrada = conn.execute(
            "SELECT * FROM cuaderno_anecdotico WHERE id=?", (entrada_id,)
        ).fetchone()
        if not entrada:
            return jsonify({"error": "Entrada no encontrada"}), 404
        if entrada["autor_id"] != u["id"] and not u.get("es_coord") and not u.get("es_directora"):
            return jsonify({"error": "Sin permisos"}), 403

        conn.execute("""
            UPDATE cuaderno_anecdotico
               SET convertido_reporte=1, visible_en_perfil=1
             WHERE id=?
        """, (entrada_id,))
        conn.commit()
    return jsonify({"ok": True, "mensaje": "Entrada convertida en reporte y visible en el perfil del estudiante"})


@app.route("/api/cuaderno/<int:entrada_id>", methods=["DELETE"])
@login_required
def eliminar_entrada_cuaderno(entrada_id):
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        entrada = conn.execute(
            "SELECT autor_id FROM cuaderno_anecdotico WHERE id=?", (entrada_id,)
        ).fetchone()
        if not entrada:
            return jsonify({"error": "No encontrada"}), 404
        rol_n = _normalizar_rol(u.get("rol",""))
        if entrada["autor_id"] != u["id"] and rol_n not in {"directora","coordinador_general"}:
            return jsonify({"error": "Sin permisos"}), 403
        conn.execute("DELETE FROM cuaderno_anecdotico WHERE id=?", (entrada_id,))
        conn.commit()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
#  CALENDARIO ESCOLAR
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/calendario", methods=["GET"])
@login_required
def get_calendario():
    anio = request.args.get("anio_escolar", "2025-2026")
    mes  = request.args.get("mes", "")
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        q = "SELECT * FROM calendario_escolar WHERE anio_escolar=?"
        p = [anio]
        if mes:
            q += " AND substr(fecha,1,7)=?"
            p.append(mes)
        q += " ORDER BY fecha"
        rows = conn.execute(q, p).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/calendario", methods=["POST"])
@login_required
def crear_dia_calendario():
    u = get_usuario()
    rol_n = _normalizar_rol(u.get("rol",""))
    if rol_n not in {"directora","coordinador_general","coordinador_primer_ciclo","coordinador_segundo_ciclo"}:
        return jsonify({"error": "Sin permisos"}), 403

    data = request.get_json() or {}
    fecha       = data.get("fecha","").strip()
    tipo        = data.get("tipo","feriado")
    descripcion = data.get("descripcion","").strip()
    anio        = data.get("anio_escolar","2025-2026")

    if not fecha:
        return jsonify({"error": "La fecha es requerida"}), 400

    with sqlite3.connect(DATABASE) as conn:
        try:
            conn.execute("""
                INSERT INTO calendario_escolar (fecha, tipo, descripcion, anio_escolar, creado_por)
                VALUES (?,?,?,?,?)
            """, (fecha, tipo, descripcion, anio, u["id"]))
            conn.commit()
        except Exception:
            conn.execute("""
                UPDATE calendario_escolar
                   SET tipo=?, descripcion=?, anio_escolar=?
                 WHERE fecha=?
            """, (tipo, descripcion, anio, fecha))
            conn.commit()
    return jsonify({"ok": True})


@app.route("/api/calendario/<fecha>", methods=["DELETE"])
@login_required
def eliminar_dia_calendario(fecha):
    u = get_usuario()
    rol_n = _normalizar_rol(u.get("rol",""))
    if rol_n not in {"directora","coordinador_general"}:
        return jsonify({"error": "Sin permisos"}), 403
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM calendario_escolar WHERE fecha=?", (fecha,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/calendario/dias-habiles", methods=["GET"])
@login_required
def get_dias_habiles():
    """Retorna el total de días hábiles de un mes dado el calendario escolar."""
    import calendar as _cal
    mes  = int(request.args.get("mes", 1))
    anio = int(request.args.get("anio", 2025))
    anio_esc = request.args.get("anio_escolar", "2025-2026")

    # Total de días del mes
    _, total_dias = _cal.monthrange(anio, mes)

    # Días no laborables del calendario escolar en ese mes
    mes_str = f"{anio:04d}-{mes:02d}"
    with sqlite3.connect(DATABASE) as conn:
        no_laborables = conn.execute("""
            SELECT fecha FROM calendario_escolar
            WHERE substr(fecha,1,7)=? AND anio_escolar=?
        """, (mes_str, anio_esc)).fetchall()
        fechas_no_lab = {r[0] for r in no_laborables}

    # Contar días hábiles (lunes–viernes que no estén en el calendario)
    habiles = 0
    for dia in range(1, total_dias + 1):
        import datetime as _dt
        d = _dt.date(anio, mes, dia)
        fecha_str = d.isoformat()
        if d.weekday() < 5 and fecha_str not in fechas_no_lab:  # 0=Lun, 4=Vie
            habiles += 1

    return jsonify({
        "mes": mes, "anio": anio,
        "total_dias": total_dias,
        "dias_habiles": habiles,
        "dias_no_laborables": len(fechas_no_lab),
        "fechas_no_laborables": list(fechas_no_lab)
    })


# ══════════════════════════════════════════════════════════════════════════════
#  ASISTENCIA MENSUAL
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/asistencia-mensual/calcular", methods=["POST"])
@login_required
def calcular_asistencia_mensual():
    """
    Calcula el resumen de asistencia del mes para todos los estudiantes
    del profesor. El profesor indica cuántos días dio clase ese mes.
    """
    u = get_usuario()
    data = request.get_json() or {}
    mes     = int(data.get("mes", 1))
    anio    = int(data.get("anio", 2025))
    materia = data.get("materia", "").strip()
    dias_clase_impartidos = int(data.get("dias_clase_impartidos", 0))
    anio_esc = data.get("anio_escolar", "2025-2026")

    if not materia or dias_clase_impartidos < 1:
        return jsonify({"error": "Materia y días de clase son requeridos"}), 400

    mes_str = f"{anio:04d}-{mes:02d}"
    import datetime as _dt

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        # Obtener días hábiles del mes desde calendario
        no_lab = {r[0] for r in conn.execute("""
            SELECT fecha FROM calendario_escolar
            WHERE substr(fecha,1,7)=? AND anio_escolar=?
        """, (mes_str, anio_esc)).fetchall()}

        import calendar as _cal
        _, total_dias = _cal.monthrange(anio, mes)
        habiles = sum(
            1 for d in range(1, total_dias+1)
            if _dt.date(anio, mes, d).weekday() < 5
            and f"{anio:04d}-{mes:02d}-{d:02d}" not in no_lab
        )

        # Contar asistencias por estudiante en el mes
        asist_rows = conn.execute("""
            SELECT estudiante_id,
                   SUM(CASE WHEN estado='presente' THEN 1 ELSE 0 END) as presentes,
                   COUNT(*) as total_registros
            FROM asistencia
            WHERE profesor_id=? AND materia=?
              AND substr(fecha,1,7)=?
            GROUP BY estudiante_id
        """, (u["id"], materia, mes_str)).fetchall()

        resumen = []
        for row in asist_rows:
            presentes = row["presentes"] or 0
            pct = round((presentes / dias_clase_impartidos) * 100, 1) if dias_clase_impartidos > 0 else 0

            conn.execute("""
                INSERT INTO asistencia_mensual
                    (estudiante_id, profesor_id, materia, mes, anio,
                     dias_habiles, dias_clase_impartidos, dias_asistio, porcentaje)
                VALUES (?,?,?,?,?,?,?,?,?)
                ON CONFLICT(estudiante_id, profesor_id, materia, mes, anio)
                DO UPDATE SET
                    dias_habiles=excluded.dias_habiles,
                    dias_clase_impartidos=excluded.dias_clase_impartidos,
                    dias_asistio=excluded.dias_asistio,
                    porcentaje=excluded.porcentaje,
                    validado=0
            """, (row["estudiante_id"], u["id"], materia, mes, anio,
                  habiles, dias_clase_impartidos, presentes, pct))

            resumen.append({
                "estudiante_id": row["estudiante_id"],
                "presentes": presentes,
                "porcentaje": pct
            })

        conn.commit()

    return jsonify({
        "ok": True,
        "mes": mes, "anio": anio,
        "dias_habiles": habiles,
        "dias_clase_impartidos": dias_clase_impartidos,
        "estudiantes_calculados": len(resumen),
        "resumen": resumen
    })


@app.route("/api/asistencia-mensual/validar", methods=["POST"])
@login_required
def validar_asistencia_mensual():
    """El profesor valida (cierra) el mes de asistencia."""
    u   = get_usuario()
    data = request.get_json() or {}
    mes     = int(data.get("mes"))
    anio    = int(data.get("anio"))
    materia = data.get("materia","").strip()
    import datetime as _dt
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            UPDATE asistencia_mensual
               SET validado=1, fecha_validacion=?
             WHERE profesor_id=? AND materia=? AND mes=? AND anio=?
        """, (_dt.datetime.now().isoformat(), u["id"], materia, mes, anio))
        conn.commit()
    cache_bust()
    return jsonify({"ok": True, "mensaje": f"Asistencia de {mes}/{anio} validada"})


@app.route("/api/asistencia-mensual/<int:est_id>", methods=["GET"])
@login_required
def get_asistencia_mensual_est(est_id):
    """Historial de asistencia mensual de un estudiante."""
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT am.*, u.nombre as profesor_nombre
            FROM asistencia_mensual am
            JOIN usuarios u ON am.profesor_id = u.id
            WHERE am.estudiante_id=?
            ORDER BY am.anio DESC, am.mes DESC
        """, (est_id,)).fetchall()
    return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════════════════════
#  EVALUACIONES NARRATIVAS DEL PROFESOR
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/evaluacion-narrativa", methods=["POST"])
@login_required
def guardar_evaluacion_narrativa():
    u    = get_usuario()
    data = request.get_json() or {}
    est_id   = data.get("estudiante_id")
    periodo  = int(data.get("periodo", 1))
    texto    = (data.get("texto") or "").strip()
    anio_esc = data.get("anio_escolar", "2025-2026")

    if not est_id or not texto:
        return jsonify({"error": "Faltan campos"}), 400

    import datetime as _dt
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            INSERT INTO evaluaciones_narrativas
                (estudiante_id, profesor_id, periodo, anio_escolar, texto, actualizado_en)
            VALUES (?,?,?,?,?,?)
            ON CONFLICT(estudiante_id, profesor_id, periodo, anio_escolar)
            DO UPDATE SET texto=excluded.texto, actualizado_en=excluded.actualizado_en
        """, (est_id, u["id"], periodo, anio_esc, texto,
              _dt.datetime.now().isoformat()))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/evaluacion-narrativa/<int:est_id>", methods=["GET"])
@login_required
def get_evaluaciones_narrativas(est_id):
    anio_esc = request.args.get("anio_escolar", "2025-2026")
    u = get_usuario()
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        # Profesores solo ven las suyas; coordinadores/directora ven todas
        rol_n = _normalizar_rol(u.get("rol",""))
        if rol_n == "profesor":
            rows = conn.execute("""
                SELECT en.*, u.nombre as profesor_nombre
                FROM evaluaciones_narrativas en
                JOIN usuarios u ON en.profesor_id = u.id
                WHERE en.estudiante_id=? AND en.profesor_id=? AND en.anio_escolar=?
                ORDER BY en.periodo
            """, (est_id, u["id"], anio_esc)).fetchall()
        else:
            rows = conn.execute("""
                SELECT en.*, u.nombre as profesor_nombre
                FROM evaluaciones_narrativas en
                JOIN usuarios u ON en.profesor_id = u.id
                WHERE en.estudiante_id=? AND en.anio_escolar=?
                ORDER BY en.periodo, u.nombre
            """, (est_id, anio_esc)).fetchall()
    return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════════════════════
#  PROGRESO DEL ESTUDIANTE (vista del profesor)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/progreso/<int:est_id>", methods=["GET"])
@login_required
def get_progreso_estudiante(est_id):
    """
    Retorna un resumen de progreso de un estudiante:
    notas por período, tendencia, asistencia mensual, cuaderno anecdótico (resumen).
    """
    u = get_usuario()
    anio_esc = request.args.get("anio_escolar", "2025-2026")

    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row

        # Datos básicos del estudiante
        est = conn.execute(
            "SELECT * FROM estudiantes WHERE id=?", (est_id,)
        ).fetchone()
        if not est:
            return jsonify({"error": "Estudiante no encontrado"}), 404

        # Materias y notas por período
        materias = conn.execute("""
            SELECT materia, tipo, p1, p2, p3, p4, promedio, profesor
            FROM materias_calificaciones
            WHERE estudiante_id=?
            ORDER BY tipo, materia
        """, (est_id,)).fetchall()

        # Asistencia mensual
        asistencia = conn.execute("""
            SELECT am.mes, am.anio, am.materia, am.porcentaje,
                   am.dias_asistio, am.dias_clase_impartidos, am.validado,
                   u.nombre as profesor_nombre
            FROM asistencia_mensual am
            JOIN usuarios u ON am.profesor_id = u.id
            WHERE am.estudiante_id=?
            ORDER BY am.anio DESC, am.mes DESC
        """, (est_id,)).fetchall()

        # Cuaderno anecdótico — solo conteo y últimas 3 entradas
        cuaderno_cnt = conn.execute(
            "SELECT COUNT(*) FROM cuaderno_anecdotico WHERE estudiante_id=?",
            (est_id,)
        ).fetchone()[0]
        cuaderno_reciente = conn.execute("""
            SELECT ca.fecha, ca.tipo, ca.descripcion, u.nombre as autor
            FROM cuaderno_anecdotico ca
            JOIN usuarios u ON ca.autor_id = u.id
            WHERE ca.estudiante_id=? AND ca.visible_en_perfil=1
            ORDER BY ca.fecha DESC LIMIT 3
        """, (est_id,)).fetchall()

        # Evaluaciones narrativas del período actual
        narrativas = conn.execute("""
            SELECT en.periodo, en.texto, u.nombre as profesor_nombre
            FROM evaluaciones_narrativas en
            JOIN usuarios u ON en.profesor_id = u.id
            WHERE en.estudiante_id=? AND en.anio_escolar=?
            ORDER BY en.periodo
        """, (est_id, anio_esc)).fetchall()

    return jsonify({
        "estudiante":  dict(est),
        "materias":    [dict(m) for m in materias],
        "asistencia":  [dict(a) for a in asistencia],
        "cuaderno": {
            "total_entradas": cuaderno_cnt,
            "recientes": [dict(c) for c in cuaderno_reciente]
        },
        "evaluaciones_narrativas": [dict(n) for n in narrativas]
    })


# ══════════════════════════════════════════════════════════════════════════════
#  GESTIÓN DE USUARIOS (ampliada para nuevos roles)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/roles-disponibles", methods=["GET"])
@login_required
def get_roles_disponibles():
    """Retorna los roles que el usuario actual puede asignar."""
    u   = get_usuario()
    rol = _normalizar_rol(u.get("rol",""))
    if rol == "directora":
        roles = [
            "coordinador_general", "coordinador_primer_ciclo",
            "coordinador_segundo_ciclo", "psicologa_primer_ciclo",
            "psicologa_segundo_ciclo", "profesor"
        ]
    elif rol == "coordinador_general":
        roles = [
            "coordinador_primer_ciclo", "coordinador_segundo_ciclo",
            "psicologa_primer_ciclo", "psicologa_segundo_ciclo", "profesor",
            "secretaria", "secretaria_docente", "digitador", "auxiliar_contabilidad"
        ]
    elif rol in {"coordinador_primer_ciclo", "coordinador_segundo_ciclo"}:
        roles = ["profesor"]
    else:
        roles = []
    return jsonify(roles)


@app.route("/api/cargar-plantilla-bj", methods=["POST"])
@login_required
def cargar_plantilla_bj_alias():
    """Alias para compatibilidad: redirige al nuevo endpoint cargar_boletin."""
    return cargar_boletin()


@app.route("/api/recalcular-promedios", methods=["POST"])
@login_required
def recalcular_promedios():
    """
    Recalcula p_acad, acad_p1..p4 y tiene_notas para TODOS los estudiantes
    basándose en materias_calificaciones. Útil cuando las notas existen pero
    el perfil no las refleja (ej: carga del boletín con la plantilla antigua).
    """
    actualizados = 0
    with sqlite3.connect(DATABASE) as conn:
        conn.row_factory = sqlite3.Row
        
        # Detectar columnas disponibles
        cols = {r[1] for r in conn.execute("PRAGMA table_info(estudiantes)").fetchall()}
        
        estudiantes = conn.execute("SELECT id FROM estudiantes").fetchall()
        
        for est_row in estudiantes:
            est_id = est_row['id']
            
            # Notas académicas
            acad = conn.execute("""
                SELECT p1, p2, p3, p4, promedio FROM materias_calificaciones
                WHERE estudiante_id=? AND (tipo='académico' OR tipo IS NULL)
                  AND promedio > 0
            """, (est_id,)).fetchall()
            
            if not acad:
                continue
            
            promedios = [r['promedio'] for r in acad if r['promedio'] and r['promedio'] > 0]
            p1s = [r['p1'] for r in acad if r['p1'] and r['p1'] > 0]
            p2s = [r['p2'] for r in acad if r['p2'] and r['p2'] > 0]
            
            p_acad  = round(sum(promedios)/len(promedios), 2) if promedios else 0
            acad_p1 = round(sum(p1s)/len(p1s), 2) if p1s else 0
            acad_p2 = round(sum(p2s)/len(p2s), 2) if p2s else 0
            
            # Módulos técnicos Multimedia
            MODULO_MAP_RECALC = {
                'fotografi': 'p_foto',
                'lenguaje visual': 'p_lv',
                'diseno': 'p_diseno', 'diseño': 'p_diseno',
            }
            mods_vals = {}
            mats_tecn = conn.execute("""
                SELECT materia, promedio FROM materias_calificaciones
                WHERE estudiante_id=? AND tipo='técnico' AND promedio > 0
            """, (est_id,)).fetchall()
            for mt in mats_tecn:
                mn = mt['materia'].lower()
                for key, col in MODULO_MAP_RECALC.items():
                    if key in mn:
                        mods_vals[col] = mt['promedio']
                        break
            
            try:
                set_parts = ["p_acad=?", "acad_p1=?", "acad_p2=?", "tiene_notas=1"]
                vals = [p_acad, acad_p1, acad_p2]
                
                for col, val in mods_vals.items():
                    if col in cols:
                        set_parts.append(f"{col}=?")
                        vals.append(val)
                
                vals.append(est_id)
                conn.execute(
                    f"UPDATE estudiantes SET {', '.join(set_parts)} WHERE id=?",
                    vals
                )
                actualizados += 1
            except Exception as ex:
                pass
        
        conn.commit()
    cache_bust()
    return jsonify({"ok": True, "actualizados": actualizados,
                    "mensaje": f"Promedios recalculados para {actualizados} estudiantes."})


# ── ARRANQUE ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    migrar_bd()
    _seed_admin()
    app.run(
        debug=True,       # Auto-recarga cuando guardas cualquier .py
        port=5000,
        use_reloader=True # Detecta cambios en archivos automáticamente
    )